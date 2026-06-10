# -*- coding: utf-8 -*-
"""
Comparador de precios publicados en marketplaces  →  Excel descargable.

Trae el PRECIO DE VENTA PUBLICADO de cada producto en los 5 marketplaces
(MercadoLibre, Walmart, Paris, Falabella, Ripley), los pone lado a lado y
marca dónde Walmart NO tiene el mejor precio. Pensado para alinear precios
y darle a Walmart el más competitivo.

Cómo se usa (blueprint):
    en app.py, junto a los otros registros de blueprint:
        from comparador_precios_mkt import comparador_precios_bp
        app.register_blueprint(comparador_precios_bp)

Endpoint:
    GET /admin/lusync/precios/comparar?descargar=si
        descargar=si  → devuelve el .xlsx (por defecto)
        descargar=no  → devuelve un JSON con el mismo contenido (para depurar)
        canales=walmart,paris,...  → limita a esos canales (default: los 5)

Requiere sesión de admin Lusync (mismo decorador que el resto del panel).
"""

from flask import Blueprint, request, session, redirect, jsonify, send_file

comparador_precios_bp = Blueprint("comparador_precios", __name__)

# Orden y etiquetas de los canales en el Excel. Walmart va primero (después del
# SKU/Producto) porque es el foco de la comparación.
CANALES = ["walmart", "mercadolibre", "paris", "falabella", "ripley"]

ETIQUETA_CANAL = {
    "walmart":      "Walmart",
    "mercadolibre": "MercadoLibre",
    "paris":        "Paris",
    "falabella":    "Falabella",
    "ripley":       "Ripley",
}

# Para cada canal: en qué campo viene el SKU del seller, el título y el precio.
# Es el MISMO mapeo que usa el endpoint de auto-mapeo, ya probado en producción.
CANAL_CONFIG = {
    "mercadolibre": {"sku_field": "sku_seller", "title": "title",         "price": "price"},
    "paris":        {"sku_field": "sellerSku",  "title": "name",          "price": "price"},
    "walmart":      {"sku_field": "sku",        "title": "productName",   "price": "price"},
    "falabella":    {"sku_field": "sellerSku",  "title": "name",          "price": "price"},
    "ripley":       {"sku_field": "shop_sku",   "title": "product_title", "price": "price"},
}


def _to_float(v):
    """Convierte un precio a float de forma tolerante (acepta None, str, dict)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dict):
        # algunas APIs anidan el monto
        for k in ("amount", "normal", "value", "currentPrice"):
            if k in v:
                return _to_float(v[k])
        return None
    try:
        return float(str(v).replace("$", "").replace(".", "").replace(",", "").strip())
    except Exception:
        return None


def _traer_items_por_canal(canales_pedidos, log):
    """Trae todas las publicaciones de cada marketplace.

    Reutiliza exactamente las mismas funciones que el resto del sistema, con
    la misma paginación. Devuelve {canal: [items...]}.
    """
    items_por_canal = {}

    if "mercadolibre" in canales_pedidos:
        try:
            from mercadolibre import obtener_publicaciones_meli
            todos, offset = [], 0
            for _ in range(20):
                data = obtener_publicaciones_meli(limite=50, offset=offset)
                if not data or not data.get("items"):
                    break
                todos.extend(data["items"])
                if len(data["items"]) < 50:
                    break
                offset += 50
            items_por_canal["mercadolibre"] = todos
            log.append(f"[MELI] {len(todos)} publicaciones")
        except Exception as e:
            log.append(f"[MELI] ERROR: {e}")
            items_por_canal["mercadolibre"] = []

    if "paris" in canales_pedidos:
        try:
            from paris import obtener_productos_paris
            todos, offset = [], 0
            for _ in range(20):
                data = obtener_productos_paris(limite=25, offset=offset)
                if not data:
                    break
                productos = data.get("products", []) if isinstance(data, dict) else (data or [])
                if isinstance(productos, dict):
                    productos = [productos]
                if not productos:
                    break
                for p in productos:
                    todos.append({
                        "sellerSku": p.get("sellerSku") or p.get("sku") or "",
                        "name":      p.get("name") or p.get("productName") or "",
                        "price":     (p.get("price") or {}).get("normal") if isinstance(p.get("price"), dict) else p.get("price"),
                    })
                if len(productos) < 25:
                    break
                offset += 25
            items_por_canal["paris"] = todos
            log.append(f"[Paris] {len(todos)} productos")
        except Exception as e:
            log.append(f"[Paris] ERROR: {e}")
            items_por_canal["paris"] = []

    if "walmart" in canales_pedidos:
        try:
            from walmart import obtener_productos_walmart
            todos = obtener_productos_walmart(limit=200, max_paginas=5)
            items_por_canal["walmart"] = todos or []
            log.append(f"[Walmart] {len(todos or [])} productos")
        except Exception as e:
            log.append(f"[Walmart] ERROR: {e}")
            items_por_canal["walmart"] = []

    if "falabella" in canales_pedidos:
        try:
            from falabella import obtener_productos_falabella
            todos, offset = [], 0
            for _ in range(20):
                productos = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                if not productos:
                    break
                for p in productos:
                    todos.append({
                        "sellerSku": p.get("SellerSku") or p.get("sellerSku") or "",
                        "name":      p.get("Name") or p.get("name") or "",
                        "price":     p.get("Price") or p.get("price"),
                    })
                if len(productos) < 100:
                    break
                offset += 100
            items_por_canal["falabella"] = todos
            log.append(f"[Falabella] {len(todos)} productos")
        except Exception as e:
            log.append(f"[Falabella] ERROR: {e}")
            items_por_canal["falabella"] = []

    if "ripley" in canales_pedidos:
        try:
            from ripley import obtener_productos_ripley
            todos = obtener_productos_ripley(max_paginas=15, page_size=100)
            items_por_canal["ripley"] = todos or []
            log.append(f"[Ripley] {len(todos or [])} productos")
        except Exception as e:
            log.append(f"[Ripley] ERROR: {e}")
            items_por_canal["ripley"] = []

    return items_por_canal


def _indexar_por_sku(items, sku_field, price_field, title_field):
    """Devuelve {sku_normalizado: {'precio':float, 'titulo':str}} para un canal."""
    idx = {}
    for it in (items or []):
        sku = (str(it.get(sku_field, "")) or "").strip()
        if not sku:
            continue
        idx[sku.upper()] = {
            "precio": _to_float(it.get(price_field)),
            "titulo": it.get(title_field, "") or "",
        }
    return idx


def _construir_filas(canales_pedidos):
    """Arma la matriz producto × canal con el precio publicado de cada uno.

    La columna vertebral es la tabla sku_mapeo_canal: para cada SKU Lusync
    sabemos qué SKU tiene en cada marketplace, y con eso buscamos su precio
    en el índice del canal correspondiente.
    """
    log = []
    from inventario import (cargar_productos, get_conn, release_conn,
                            init_sku_mapeo_canal)

    init_sku_mapeo_canal()

    # 1. Productos Lusync (la web mandante) con su precio de referencia.
    productos = cargar_productos()
    log.append(f"Productos Lusync: {len(productos)}")

    # 2. Mapeo SKU Lusync → SKU de cada canal (desde sku_mapeo_canal).
    #    mapeo[sku_lusync][canal] = sku_canal
    mapeo = {}
    conn = get_conn(is_admin=True)
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT sku_lusync, canal, sku_canal
            FROM sku_mapeo_canal
            WHERE activo = TRUE
        """)
        for sku_lusync, canal, sku_canal in cur.fetchall():
            canal = (canal or "").lower().strip()
            if not sku_lusync or not sku_canal:
                continue
            mapeo.setdefault(sku_lusync, {})[canal] = str(sku_canal).strip()
        cur.close()
    finally:
        release_conn(conn)
    log.append(f"SKUs con mapeo: {len(mapeo)}")

    # 3. Traer publicaciones de cada marketplace e indexarlas por SKU.
    items_por_canal = _traer_items_por_canal(canales_pedidos, log)
    idx_por_canal = {}
    for canal in canales_pedidos:
        cfg = CANAL_CONFIG[canal]
        idx_por_canal[canal] = _indexar_por_sku(
            items_por_canal.get(canal, []),
            cfg["sku_field"], cfg["price"], cfg["title"]
        )

    # 4. Una fila por producto, con el precio publicado en cada canal.
    filas = []
    for p in productos:
        sku_lusync = (p.get("sku", "") or "").strip()
        if not sku_lusync:
            continue
        fila = {
            "sku_lusync": sku_lusync,
            "nombre":     p.get("nombre", "") or "",
            "precio_web": _to_float(p.get("precio_normal") if "precio_normal" in p else p.get("precio")),
            "precios":    {},   # canal -> precio publicado (o None)
        }
        canales_mapeados = mapeo.get(sku_lusync, {})
        for canal in canales_pedidos:
            sku_canal = canales_mapeados.get(canal, "")
            precio = None
            if sku_canal:
                hit = idx_por_canal[canal].get(sku_canal.upper())
                if hit:
                    precio = hit["precio"]
            fila["precios"][canal] = precio
        filas.append(fila)

    return filas, log


def _mejor_precio(precios_dict, canales):
    """Devuelve el menor precio publicado entre los canales (ignora None)."""
    vals = [precios_dict.get(c) for c in canales if precios_dict.get(c) is not None]
    return min(vals) if vals else None


@comparador_precios_bp.route("/admin/lusync/precios/comparar", methods=["GET"])
def comparar_precios_marketplaces():
    # Mismo control de acceso que el resto del panel admin Lusync.
    # (replica la lógica de @requiere_lusync_admin sin importar app.py para
    #  evitar import circular: sesión admin o token bypass por header/query)
    import os
    autorizado = bool(session.get("is_lusync_admin"))
    if not autorizado:
        bypass_token = os.environ.get(
            "ADMIN_BYPASS_TOKEN",
            "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw",
        )
        token_recibido = request.headers.get("x-admin-token") or request.args.get("token")
        autorizado = bool(token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return redirect("/admin/lusync/login")

    canales_param = (request.args.get("canales", "") or "").strip()
    if canales_param:
        canales_pedidos = [c.strip().lower() for c in canales_param.split(",")
                           if c.strip().lower() in CANALES]
    else:
        canales_pedidos = list(CANALES)
    if not canales_pedidos:
        canales_pedidos = list(CANALES)

    descargar = (request.args.get("descargar", "si") or "si").lower() != "no"

    try:
        filas, log = _construir_filas(canales_pedidos)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ── Salida JSON (depuración) ──
    if not descargar:
        salida = []
        for f in filas:
            mejor = _mejor_precio(f["precios"], canales_pedidos)
            wm = f["precios"].get("walmart")
            salida.append({
                "sku_lusync":  f["sku_lusync"],
                "nombre":      f["nombre"],
                "precio_web":  f["precio_web"],
                "precios":     f["precios"],
                "mejor_precio": mejor,
                "walmart_es_mejor": (wm is not None and mejor is not None and wm <= mejor),
            })
        return jsonify({"log": log, "total": len(salida), "filas": salida})

    # ── Salida Excel ──
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios marketplaces"

    # Estilos
    fuente_titulo = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    relleno_cab   = PatternFill("solid", fgColor="1F4E78")
    relleno_wm    = PatternFill("solid", fgColor="2E75B6")   # Walmart destacado
    centro        = Alignment(horizontal="center", vertical="center")
    izq           = Alignment(horizontal="left", vertical="center")
    fmt_money     = '$#,##0;[Red]($#,##0);"-"'
    borde_fino    = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    relleno_alerta = PatternFill("solid", fgColor="FCE4D6")  # Walmart NO es el mejor
    relleno_ok     = PatternFill("solid", fgColor="E2EFDA")  # Walmart SÍ es el mejor
    fuente_ndisp  = Font(name="Arial", italic=True, color="808080")

    # Encabezados: SKU, Producto, Precio Web, [un precio por canal], Mejor precio, Estado
    cabeceras = ["SKU Lusync", "Producto", "Precio Web"]
    cabeceras += [f"Precio {ETIQUETA_CANAL[c]}" for c in canales_pedidos]
    cabeceras += ["Mejor precio", "¿Walmart es el mejor?", "Dif. vs mejor"]
    ws.append(cabeceras)

    for col_idx, _ in enumerate(cabeceras, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = fuente_titulo
        c.alignment = centro
        c.border = borde_fino
        # Walmart resaltado en la cabecera
        if cabeceras[col_idx - 1] == f"Precio {ETIQUETA_CANAL['walmart']}":
            c.fill = relleno_wm
        else:
            c.fill = relleno_cab

    col_walmart = 3 + 1 + canales_pedidos.index("walmart") if "walmart" in canales_pedidos else None

    fila_xl = 2
    for f in filas:
        precios = f["precios"]
        mejor = _mejor_precio(precios, canales_pedidos)
        wm = precios.get("walmart")

        valores = [f["sku_lusync"], f["nombre"], f["precio_web"]]
        for c in canales_pedidos:
            valores.append(precios.get(c))
        valores.append(mejor)

        if wm is None:
            estado = "Sin publicar en Walmart"
        elif mejor is None:
            estado = "Sin datos"
        elif wm <= mejor:
            estado = "Sí"
        else:
            estado = "No"
        valores.append(estado)

        # Diferencia de Walmart contra el mejor precio (positivo = Walmart más caro)
        dif = (wm - mejor) if (wm is not None and mejor is not None) else None
        valores.append(dif)

        ws.append(valores)

        # Formato de la fila
        for col_idx in range(1, len(valores) + 1):
            celda = ws.cell(row=fila_xl, column=col_idx)
            celda.border = borde_fino
            if col_idx == 1:
                celda.alignment = izq
                celda.font = Font(name="Arial", bold=True)
            elif col_idx == 2:
                celda.alignment = izq
            else:
                celda.alignment = centro
            # columnas de dinero (Precio Web=3 en adelante hasta Mejor precio, y Dif al final)
            es_col_dinero = (3 <= col_idx <= 3 + len(canales_pedidos) + 1) or (col_idx == len(valores))
            if es_col_dinero:
                celda.number_format = fmt_money
                if celda.value is None:
                    celda.value = None
                    celda.font = fuente_ndisp

        # Resaltar la celda de Walmart según si es o no el mejor precio
        if col_walmart and wm is not None and mejor is not None:
            celda_wm = ws.cell(row=fila_xl, column=col_walmart)
            celda_wm.fill = relleno_ok if wm <= mejor else relleno_alerta

        # Resaltar la columna "Estado" en rojo suave cuando Walmart no es el mejor
        celda_estado = ws.cell(row=fila_xl, column=len(valores) - 1)
        if estado == "No":
            celda_estado.fill = relleno_alerta
            celda_estado.font = Font(name="Arial", bold=True, color="C00000")
        elif estado == "Sí":
            celda_estado.fill = relleno_ok
            celda_estado.font = Font(name="Arial", bold=True, color="375623")

        fila_xl += 1

    # Anchos de columna
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    for i in range(3, len(cabeceras) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(len(cabeceras) - 1)].width = 20

    # Congelar cabecera y primera columna; activar autofiltro
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabeceras))}{fila_xl - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name="precios_marketplaces.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
