from flask import Flask, request, render_template, session, redirect, jsonify, send_file
import requests
import os
from datetime import datetime, timedelta
from config import *
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from walmart import (actualizar_stock_walmart, actualizar_precio_walmart,
                     obtener_ordenes_walmart, confirmar_orden_walmart,
                     verificar_conexion_walmart)
from paris import (verificar_conexion_paris, obtener_ordenes_paris_todas,
                   actualizar_stock_paris, obtener_stock_paris,
                   actualizar_precio_paris, obtener_orden_paris,
                   get_seller_id as get_paris_seller_id)
from woo import actualizar_stock_woo
# Imports para sync centralizado (los 6 marketplaces)
from mercadolibre import actualizar_stock_meli
from falabella import actualizar_stock_falabella_lusync as actualizar_stock_falabella
from ripley import actualizar_stock_ripley_lusync as actualizar_stock_ripley
from inventario import (cargar_productos, guardar_productos, guardar_producto,
                        registrar_movimiento, cargar_movimientos, cargar_movimientos_hoy,
                        init_db, orden_ya_procesada, marcar_orden_procesada, actualizar_precios,
                        get_configuracion, set_configuracion, set_lead_time, eliminar_producto,
                        orden_ya_procesada_texto, marcar_orden_procesada_texto,
                        init_devoluciones, generar_codigo_dev, crear_devolucion,
                        asignar_codigo_dev, actualizar_devolucion, listar_devoluciones,
                        get_devolucion,
                        init_audit, registrar_audit, listar_audit,
                        init_sku_mapeo, listar_sku_mapeo, guardar_sku_mapeo_fila,
                        get_sku_canal, get_plataforma_web, set_plataforma_web,
                        registrar_importacion_mapeo, listar_historial_mapeo,
                        init_alertas, crear_alerta, listar_alertas,
                        contar_alertas_no_leidas, marcar_alerta_leida,
                        marcar_todas_leidas, get_alertas_config, set_alertas_config,
                        init_meli_auth, get_meli_auth, set_meli_auth, borrar_meli_auth,
                        stats_ventas_por_canal_dia, stats_top_productos_vendidos,
                        stats_movimientos_dia, stats_distribucion_stock_canal,
                        stats_kpis_dashboard,
                        init_bodegas, listar_bodegas, stock_por_bodega,
                        get_stock_bodega, set_stock_bodega, ajustar_stock_bodega,
                        listar_stock_completo, stock_total_por_bodega,
                        determinar_bodega_para_canal,
                        actualizar_nombres_bodegas)

app = Flask(__name__)
app.secret_key = "clave_super_segura"

init_db()
init_devoluciones()
init_audit()
init_sku_mapeo()
init_alertas()
init_meli_auth()
init_bodegas()

# Registrar Blueprints (módulos de marketplaces)
from walmart import walmart_bp
from paris import paris_bp
from ripley import ripley_bp
from falabella import falabella_bp
app.register_blueprint(walmart_bp)
app.register_blueprint(paris_bp)
app.register_blueprint(ripley_bp)
app.register_blueprint(falabella_bp)

# ════════════════════════════════════════════════════════════════════════════
# SYNC CENTRALIZADO MULTI-MARKETPLACE
# ════════════════════════════════════════════════════════════════════════════
# Helper único que se llama AUTOMÁTICAMENTE en TODOS los lugares donde cambia
# stock (entrada manual, salida manual, sync de órdenes, devoluciones, etc.)
#
# Sincroniza con los 6 marketplaces conectados de forma RESILIENTE:
#  - Si un canal falla, los demás siguen ejecutándose
#  - Devuelve resumen detallado de qué funcionó y qué no
#  - No bloquea la operación principal aunque algún canal esté caído
#
# Uso típico:
#   sincronizar_stock_marketplaces(sku, nuevo_stock)
# ════════════════════════════════════════════════════════════════════════════

def sincronizar_stock_marketplaces(sku, stock, contexto="manual"):
    """Sincroniza el stock de un SKU con TODOS los marketplaces conectados.

    Args:
        sku: SKU Lusync del producto
        stock: nuevo stock total (entero)
        contexto: descripción para logs (ej: "entrada_manual", "venta_meli", etc.)

    Returns:
        dict con resultado por canal: {"meli": "ok", "falabella": "error: ...", ...}

    NUNCA lanza excepciones. Si algo falla, se registra en el log y devuelve "error".
    """
    resultado = {}
    canales = [
        ("woo",          actualizar_stock_woo),
        ("walmart",      actualizar_stock_walmart),
        ("paris",        actualizar_stock_paris),
        ("mercadolibre", actualizar_stock_meli),
        ("falabella",    actualizar_stock_falabella),
        ("ripley",       actualizar_stock_ripley),
    ]
    for nombre_canal, fn in canales:
        try:
            fn(sku, stock)
            resultado[nombre_canal] = "ok"
        except Exception as e:
            resultado[nombre_canal] = f"error: {str(e)[:80]}"
            print(f"[SyncCentral][{contexto}] {nombre_canal} falló para SKU {sku}: {e}")
    return resultado

# ── SYNC AUTOMÁTICO WALMART CADA 5 MINUTOS ──
def _sync_walmart_automatico():
    """Tarea de background: sincroniza órdenes Walmart sin requerir sesión"""
    # FIX: lock anti-overlapping igual que los demás canales
    if not hasattr(_sync_walmart_automatico, "_running"):
        _sync_walmart_automatico._running = False
    if _sync_walmart_automatico._running:
        print("[Scheduler Walmart] Ya hay un sync corriendo, salto")
        return
    _sync_walmart_automatico._running = True
    try:
        print("[Scheduler] Iniciando sync automático Walmart...")
        from inventario import obtener_sku_lusync_por_canal, descontar_venta_inteligente
        from bodegas_logic import detectar_fulfillment_walmart
        productos = cargar_productos()
        nuevas = 0
        errores = []

        # Descuenta apenas Walmart crea la orden (Created) para evitar sobreventa en otros canales.
        # Si el cliente cancela, el bloque de cancelaciones (más abajo) reintegra el stock automáticamente.
        for estado in ["Created", "Acknowledged", "Shipped", "Delivered"]:
            ordenes = obtener_ordenes_walmart(estado)
            for o in ordenes:
                order_id = o.get("purchaseOrderId")
                if not order_id:
                    continue
                customer_order_id = str(o.get("customerOrderId", order_id))
                if orden_ya_procesada_texto(customer_order_id):
                    continue

                lineas = o.get("orderLines", {}).get("orderLine", [])
                if isinstance(lineas, dict):
                    lineas = [lineas]

                # Detectar si esta orden es WFS (Walmart Fulfillment Services)
                es_wfs = detectar_fulfillment_walmart(o)
                tipo_str = "WFS" if es_wfs else "Seller"

                items_descontados = []
                for linea in lineas:
                    try:
                        sku = linea.get("item", {}).get("sku")
                        if not sku:
                            continue
                        cantidad = 1
                        qty = linea.get("orderLineQuantity", {})
                        if qty and qty.get("amount"):
                            cantidad = int(float(qty.get("amount", 1)))
                        if cantidad == 1:
                            status_qty = linea.get("statusQuantity", {})
                            if status_qty and status_qty.get("amount"):
                                cantidad = int(float(status_qty.get("amount", 1)))

                        # FIX: usar obtener_sku_lusync_por_canal (más eficiente que listar_sku_mapeo)
                        sku_lusync = obtener_sku_lusync_por_canal("walmart", sku) or sku

                        # Buscar producto y descontar de bodega correcta
                        producto_existe = any(p["sku"] == sku_lusync for p in productos)
                        if not producto_existe:
                            print(f"[Scheduler] SKU '{sku_lusync}' no encontrado en inventario")
                            continue

                        # Parsear fecha real de compra (Walmart: orderDate o createdAt en ISO)
                        fecha_compra_wm = None
                        try:
                            date_str = (o.get("orderDate") or o.get("createdAt") or
                                        o.get("orderPlacedTime") or "")
                            if date_str:
                                fecha_compra_wm = datetime.fromisoformat(
                                    str(date_str).replace("Z", "+00:00")
                                )
                        except Exception:
                            fecha_compra_wm = None

                        resultado = descontar_venta_inteligente(
                            sku=sku_lusync,
                            cantidad=cantidad,
                            canal="Walmart",
                            fulfillment=es_wfs,
                            orden_id=customer_order_id,
                            motivo=f"Venta Walmart {tipo_str}",
                            usuario="Sistema",
                            fecha_compra_marketplace=fecha_compra_wm
                        )
                        print(f"[Scheduler] {customer_order_id} {tipo_str}: {sku_lusync} -{cantidad} desde {resultado['bodega']}")

                        # Sync a otros canales SOLO si fue Seller (afectó Central)
                        if not es_wfs:
                            stock_total = resultado.get("stock_despues", 0)
                            sincronizar_stock_marketplaces(sku_lusync, stock_total, contexto="walmart_orden_bg")
                        items_descontados.append(sku_lusync)
                    except Exception as e:
                        errores.append(str(e))
                        print(f"[Scheduler] Error linea: {e}")

                # FIX: marcar solo UNA vez después de procesar (antes había doble marcar)
                marcar_orden_procesada_texto(customer_order_id)
                nuevas += 1

        # ── CANCELACIONES WALMART — devolver stock si se canceló una orden ya procesada
        try:
            canceladas = obtener_ordenes_walmart("Cancelled")
            reingresadas = 0
            for o in canceladas:
                order_id = o.get("purchaseOrderId")
                if not order_id:
                    continue
                customer_order_id = str(o.get("customerOrderId", order_id))
                cancel_key = f"CANCEL-{customer_order_id}"

                # Solo procesar si la orden fue previamente descontada Y no se reingresó antes
                if not orden_ya_procesada_texto(customer_order_id):
                    continue  # nunca se procesó, no hay stock que devolver
                if orden_ya_procesada_texto(cancel_key):
                    continue  # ya se procesó la cancelación

                lineas = o.get("orderLines", {}).get("orderLine", [])
                if isinstance(lineas, dict):
                    lineas = [lineas]

                productos = cargar_productos()
                items_cancelados = []  # para alerta consolidada
                for linea in lineas:
                    try:
                        sku = linea.get("item", {}).get("sku")
                        if not sku:
                            continue
                        cantidad = 1
                        qty = linea.get("orderLineQuantity", {})
                        if qty and qty.get("amount"):
                            cantidad = int(float(qty.get("amount", 1)))
                        if cantidad == 1:
                            status_qty = linea.get("statusQuantity", {})
                            if status_qty and status_qty.get("amount"):
                                cantidad = int(float(status_qty.get("amount", 1)))

                        for p in productos:
                            if p["sku"] == sku:
                                p["stock"] = p["stock"] + cantidad
                                guardar_producto(p)
                                registrar_movimiento("entrada", p["sku"], p["nombre"],
                                                    cantidad, "Cancelación Walmart",
                                                    usuario="Sistema", canal="Walmart",
                                                    orden_id=customer_order_id)
                                sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="walmart_cancelacion")
                                items_cancelados.append(f"{p['nombre']} (SKU: {sku}) x{cantidad}")
                                print(f"[Scheduler] CANCELACIÓN SKU:{sku} +{cantidad} Stock:{p['stock']}")
                    except Exception as e:
                        print(f"[Scheduler] Error cancelación linea: {e}")

                # Crear alerta consolidada por orden cancelada
                if items_cancelados:
                    try:
                        crear_alerta(
                            tipo="cancelacion",
                            canal="Walmart",
                            titulo=f"Orden cancelada en Walmart: {customer_order_id}",
                            mensaje="El cliente canceló la orden. Stock reintegrado automáticamente:<br><br>" +
                                    "<br>".join(f"• {it}" for it in items_cancelados),
                            orden_id=customer_order_id,
                            sku=items_cancelados[0].split("SKU: ")[1].split(")")[0] if items_cancelados else None
                        )
                    except Exception as e:
                        print(f"[Scheduler] Error creando alerta: {e}")

                marcar_orden_procesada_texto(cancel_key)
                reingresadas += 1

            if reingresadas:
                print(f"[Scheduler] Cancelaciones procesadas: {reingresadas}")
        except Exception as e:
            print(f"[Scheduler] Error procesando cancelaciones: {e}")

        # Paris ya tiene su propia función _sync_paris_automatico con job registrado
        # No se duplica aquí para evitar doble procesamiento

        print(f"[Scheduler] Sync completado — nuevas:{nuevas} errores:{len(errores)}")
    except Exception as e:
        print(f"[Scheduler] Error general: {e}")
    finally:
        _sync_walmart_automatico._running = False

scheduler = BackgroundScheduler(daemon=True)

# ════════════════════════════════════════════════════════════════════════════
# SYNC AUTOMÁTICO MULTI-MARKETPLACE
# ════════════════════════════════════════════════════════════════════════════
# Distribución escalonada para NO saturar Render ni APIs externas:
#
#   Min 0:  MELI          ← más vendido = más reactivo (cada 5 min)
#   Min 2:  Falabella     ← 2do en volumen (cada 10 min)
#   Min 4:  Walmart       ← (ya existente, cada 5 min)
#   Min 5:  MELI nuevamente
#   Min 6:  París         ← cada 10 min
#   Min 8:  Ripley        ← cada 10 min
#   Min 10: MELI nuevamente
#   Min 10: Woo (web)     ← cada 10 min, último por menor volumen
#
# Cada sync tiene LOCK anti-overlapping: si un sync tarda más de su intervalo,
# el siguiente se salta hasta que termine.
# ════════════════════════════════════════════════════════════════════════════

# Locks globales para evitar que 2 ejecuciones del mismo canal corran a la vez
_sync_locks = {
    "meli":      {"running": False},
    "falabella": {"running": False},
    "paris":     {"running": False},
    "ripley":    {"running": False},
    "woo":       {"running": False},
}


def _sync_meli_automatico():
    """Sync órdenes + cancelaciones MercadoLibre cada 5 min."""
    if _sync_locks["meli"]["running"]:
        print("[Scheduler MELI] Ya hay un sync corriendo, salto esta vez")
        return
    _sync_locks["meli"]["running"] = True
    try:
        print("[Scheduler MELI] Iniciando sync automático...")
        from mercadolibre import obtener_ordenes_meli
        from inventario import descontar_venta_inteligente, detectar_fulfillment_meli
        from datetime import datetime
        nuevas = 0
        canceladas = 0
        errores = []

        # Solo últimas 6 horas para mantener carga baja
        try:
            ordenes = obtener_ordenes_meli(limit=50, offset=0)
        except Exception as e:
            print(f"[Scheduler MELI] Error obteniendo órdenes: {e}")
            return

        for o in ordenes:
            try:
                order_id = str(o.get("id", ""))
                estado = o.get("status", "")
                meli_key = f"MELI-{order_id}"
                cancel_key = f"MELI-CANCEL-{order_id}"

                # ── Órdenes pagadas ──
                if estado in ("paid", "confirmed"):
                    if orden_ya_procesada_texto(meli_key):
                        continue

                    # Parsear fecha de compra
                    fecha_compra = None
                    try:
                        ds = (o.get("date_created", "") or "").replace("Z", "+00:00")
                        if ds:
                            fecha_compra = datetime.fromisoformat(ds)
                    except: pass

                    es_full = detectar_fulfillment_meli(o)
                    items_descontados = []

                    for item in o.get("order_items", []):
                        item_data = item.get("item", {})
                        item_id = item_data.get("id", "")
                        sku_seller = (
                            (item_data.get("seller_sku") or "").strip()
                            or (item_data.get("seller_custom_field") or "").strip()
                        )
                        cantidad = int(item.get("quantity", 1))
                        if not sku_seller:
                            continue

                        # Traducir SKU canal MELI a SKU Lusync (ej: ODJ3N001 → ODJ3NB001)
                        try:
                            from inventario import obtener_sku_lusync_por_canal
                            sku_lusync = obtener_sku_lusync_por_canal("mercadolibre", sku_canal=sku_seller, item_id_canal=item_id) or sku_seller
                        except Exception:
                            sku_lusync = sku_seller

                        # Descuento inteligente (Full vs Central)
                        try:
                            descontar_venta_inteligente(
                                sku=sku_lusync,
                                cantidad=cantidad,
                                canal="mercadolibre",
                                fulfillment=es_full,
                                orden_id=order_id,
                                fecha_compra_marketplace=fecha_compra
                            )
                            items_descontados.append((sku_lusync, cantidad))

                            # Sync a otros canales SOLO si fue Seller (Full no afecta otras bodegas)
                            if not es_full:
                                p_actual = next((pp for pp in cargar_productos() if pp["sku"] == sku_lusync), None)
                                if p_actual:
                                    sincronizar_stock_marketplaces(
                                        sku_lusync, p_actual["stock"],
                                        contexto="meli_orden_bg"
                                    )
                        except Exception as e:
                            errores.append(f"MELI {order_id}/{sku_seller}→{sku_lusync}: {e}")

                    if items_descontados:
                        marcar_orden_procesada_texto(meli_key)
                        nuevas += 1

                # ── Órdenes canceladas ──
                elif estado in ("cancelled", "canceled"):
                    if orden_ya_procesada_texto(cancel_key):
                        continue
                    if not orden_ya_procesada_texto(meli_key):
                        # Nunca se procesó la venta, no hay nada que reintegrar
                        marcar_orden_procesada_texto(cancel_key)
                        continue

                    items_reintegrados = []
                    for item in o.get("order_items", []):
                        item_data = item.get("item", {})
                        item_id_canc = item_data.get("id", "")
                        sku_seller = (
                            (item_data.get("seller_sku") or "").strip()
                            or (item_data.get("seller_custom_field") or "").strip()
                        )
                        cantidad = int(item.get("quantity", 1))
                        if not sku_seller: continue

                        # Traducir SKU canal a Lusync
                        try:
                            from inventario import obtener_sku_lusync_por_canal
                            sku_lusync = obtener_sku_lusync_por_canal("mercadolibre", sku_canal=sku_seller, item_id_canal=item_id_canc) or sku_seller
                        except Exception:
                            sku_lusync = sku_seller

                        productos = cargar_productos()
                        for p in productos:
                            if p["sku"] == sku_lusync:
                                p["stock"] += cantidad
                                guardar_producto(p)
                                registrar_movimiento(
                                    "entrada", p["sku"], p["nombre"], cantidad,
                                    f"Cancelación MELI orden {order_id}",
                                    usuario="Sistema", canal="MercadoLibre", orden_id=order_id
                                )
                                sincronizar_stock_marketplaces(
                                    p["sku"], p["stock"],
                                    contexto="meli_cancelacion_bg"
                                )
                                items_reintegrados.append(f"{p['nombre']} (SKU: {sku_seller}→{sku_lusync}) x{cantidad}")
                                break

                    if items_reintegrados:
                        try:
                            crear_alerta(
                                tipo="cancelacion",
                                titulo=f"Orden cancelada en MercadoLibre: {order_id}",
                                mensaje="Stock reintegrado:<br>" + "<br>".join(f"• {it}" for it in items_reintegrados),
                                sku=sku_lusync
                            )
                        except: pass
                        canceladas += 1

                    marcar_orden_procesada_texto(cancel_key)
            except Exception as e:
                errores.append(f"MELI orden: {e}")

        print(f"[Scheduler MELI] Sync OK — nuevas:{nuevas} canceladas:{canceladas} errores:{len(errores)}")
    except Exception as e:
        print(f"[Scheduler MELI] Error general: {e}")
    finally:
        _sync_locks["meli"]["running"] = False


def _sync_falabella_automatico():
    """Sync órdenes + cancelaciones Falabella cada 10 min."""
    if _sync_locks["falabella"]["running"]:
        print("[Scheduler Falabella] Ya hay un sync corriendo, salto")
        return
    _sync_locks["falabella"]["running"] = True
    try:
        print("[Scheduler Falabella] Iniciando sync automático...")
        # FIX: importar también obtener_items_orden_falabella — los items NO vienen
        # dentro del objeto orden, hay que pedirlos por separado con el order_id
        from falabella import obtener_ordenes_falabella, obtener_items_orden_falabella
        from inventario import obtener_sku_lusync_por_canal
        from bodegas_logic import descontar_venta, detectar_fulfillment_falabella
        nuevas = 0
        canceladas = 0
        errores = []

        # FIX: buscar en todos los estados relevantes en pasadas separadas para
        # no mezclar lógica. Estado=None devuelve todos (sin filtro).
        # Falabella SellerCenter soporta: pending, ready_to_ship, shipped, delivered, canceled
        estados_activos   = ["pending", "ready_to_ship", "shipped", "delivered"]
        estados_cancelados = ["canceled"]

        todas_ordenes = []
        for estado in estados_activos + estados_cancelados:
            try:
                lote = obtener_ordenes_falabella(estado=estado, dias=7, limit=50, offset=0)
                if isinstance(lote, list):
                    todas_ordenes.extend(lote)
            except Exception as e:
                errores.append(f"FA get estado {estado}: {e}")

        print(f"[Scheduler Falabella] Total órdenes obtenidas: {len(todas_ordenes)}")

        for o in todas_ordenes:
            if not isinstance(o, dict):
                continue
            # obtener_ordenes_falabella ya desenvuelve {"Order":{}} — o es la orden directa
            try:
                order_id = str(o.get("OrderId") or o.get("orderId") or o.get("id") or "")
                if not order_id:
                    continue
                # Estado puede venir en Statuses[0].Status o en Status directo
                statuses = o.get("Statuses") or []
                if isinstance(statuses, list) and statuses:
                    estado_orden = (statuses[0].get("Status") or "").lower()
                else:
                    estado_orden = (o.get("Status") or o.get("status") or "").lower()
                fa_key = f"FALABELLA-{order_id}"
                cancel_key = f"FALABELLA-CANCEL-{order_id}"

                # ── Órdenes nuevas (estados que descuentan stock) ──
                if estado_orden in ("ready_to_ship", "shipped", "delivered", "pending"):
                    if orden_ya_procesada_texto(fa_key):
                        continue

                    # FIX CRÍTICO: los items NO vienen en el objeto orden.
                    # Hay que pedirlos explícitamente por order_id
                    try:
                        items_orden = obtener_items_orden_falabella(order_id) or []
                    except Exception as e:
                        errores.append(f"FA items orden {order_id}: {e}")
                        items_orden = []

                    if not items_orden:
                        print(f"[Scheduler Falabella] Orden {order_id}: sin items, saltando")
                        continue

                    es_fbf = detectar_fulfillment_falabella(o)
                    tipo_str = "FBF" if es_fbf else "FBS"
                    items_descontados = []

                    # Parsear fecha real de compra (Falabella: CreatedAt = "2026-05-05 23:57:15" sin tz)
                    fecha_compra_fa = None
                    try:
                        import pytz as _pytz_fa
                        date_str = (o.get("CreatedAt") or o.get("created_at") or "")
                        if date_str:
                            try:
                                # Formato con T y/o timezone
                                fecha_compra_fa = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                            except ValueError:
                                try:
                                    # Formato "YYYY-MM-DD HH:MM:SS" sin timezone → asumir UTC
                                    fecha_naive = datetime.strptime(date_str.strip(), "%Y-%m-%d %H:%M:%S")
                                    fecha_compra_fa = _pytz_fa.utc.localize(fecha_naive)
                                except Exception:
                                    fecha_compra_fa = None
                    except Exception:
                        fecha_compra_fa = None

                    for item in items_orden:
                        # FIX: campo real de Falabella es "Sku" (no "SellerSku")
                        # obtener_items_orden_falabella ya lo normaliza en SellerSku=Sku
                        # y agrupa cantidad real (cada OrderItem = 1 unidad)
                        seller_sku = (item.get("SellerSku") or item.get("Sku") or
                                      item.get("sellerSku") or item.get("sku") or "").strip()
                        cantidad = int(item.get("Quantity") or item.get("quantity") or 1)
                        if not seller_sku:
                            continue

                        sku_lusync = obtener_sku_lusync_por_canal("falabella", seller_sku) or seller_sku

                        productos = cargar_productos()
                        prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                        if not prod:
                            print(f"[Scheduler Falabella] SKU '{sku_lusync}' no encontrado")
                            continue

                        resultado = descontar_venta(
                            sku=sku_lusync,
                            cantidad=cantidad,
                            canal="Falabella",
                            fulfillment=es_fbf,
                            orden_id=order_id,
                            motivo=f"Venta Falabella {tipo_str}",
                            usuario="Sistema",
                            fecha_compra_marketplace=fecha_compra_fa,
                            origen_registro="scheduler"
                        )
                        sincronizar_stock_marketplaces(
                            sku_lusync, resultado.get("stock_despues", 0),
                            contexto="falabella_orden_bg"
                        )
                        items_descontados.append(f"{seller_sku} x{cantidad}")
                        print(f"[Scheduler Falabella] {order_id} {tipo_str}: {sku_lusync} -{cantidad} desde {resultado.get('bodega','?')}")

                    if items_descontados:
                        marcar_orden_procesada_texto(fa_key)
                        nuevas += 1

                # ── Órdenes canceladas ──
                elif estado_orden in ("canceled", "cancelled"):
                    if orden_ya_procesada_texto(cancel_key):
                        continue
                    if not orden_ya_procesada_texto(fa_key):
                        # nunca se descontó, solo marcar cancelación para no volver
                        marcar_orden_procesada_texto(cancel_key)
                        continue

                    try:
                        items_orden = obtener_items_orden_falabella(order_id) or []
                    except Exception as e:
                        errores.append(f"FA items cancelada {order_id}: {e}")
                        items_orden = []

                    items_reintegrados = []
                    ultimo_sku = None
                    for item in items_orden:
                        seller_sku = (item.get("SellerSku") or item.get("Sku") or
                                      item.get("sellerSku") or item.get("sku") or "").strip()
                        cantidad = int(item.get("Quantity") or item.get("quantity") or 1)
                        if not seller_sku:
                            continue
                        ultimo_sku = seller_sku
                        sku_lusync = obtener_sku_lusync_por_canal("falabella", seller_sku) or seller_sku

                        productos = cargar_productos()
                        prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                        if not prod:
                            continue

                        prod["stock"] += cantidad
                        guardar_producto(prod)
                        registrar_movimiento(
                            "entrada", prod["sku"], prod["nombre"], cantidad,
                            f"Cancelación Falabella orden {order_id}",
                            usuario="Sistema", canal="Falabella", orden_id=order_id
                        )
                        sincronizar_stock_marketplaces(
                            prod["sku"], prod["stock"], contexto="falabella_cancelacion_bg"
                        )
                        items_reintegrados.append(f"{prod['nombre']} (SKU: {seller_sku}) x{cantidad}")

                    if items_reintegrados:
                        try:
                            crear_alerta(
                                tipo="cancelacion",
                                titulo=f"Orden cancelada en Falabella: {order_id}",
                                mensaje="Stock reintegrado:<br>" + "<br>".join(f"• {it}" for it in items_reintegrados),
                                sku=ultimo_sku
                            )
                        except:
                            pass
                        canceladas += 1

                    marcar_orden_procesada_texto(cancel_key)

            except Exception as e:
                errores.append(f"FALABELLA orden: {e}")

        import gc; gc.collect()
        print(f"[Scheduler Falabella] Sync OK — nuevas:{nuevas} canceladas:{canceladas} errores:{len(errores)}")
        if errores:
            print(f"[Scheduler Falabella] Errores: {errores[:3]}")
    except Exception as e:
        print(f"[Scheduler Falabella] Error general: {e}")
    finally:
        _sync_locks["falabella"]["running"] = False


def _sync_paris_automatico():
    """Sync órdenes + cancelaciones París cada 10 min."""
    if _sync_locks["paris"]["running"]:
        print("[Scheduler Paris] Ya hay un sync corriendo, salto")
        return
    _sync_locks["paris"]["running"] = True
    try:
        print("[Scheduler Paris] Iniciando sync automático...")
        from paris import obtener_ordenes_paris_todas
        from inventario import obtener_sku_lusync_por_canal
        from bodegas_logic import descontar_venta, detectar_fulfillment_paris
        nuevas = 0
        canceladas = 0
        errores = []

        # Traer sin filtro de estado (Paris no filtra bien por estado en la API)
        # dias=7 para cubrir órdenes recientes con margen
        try:
            ordenes = obtener_ordenes_paris_todas(dias=7)
        except Exception as e:
            print(f"[Scheduler Paris] Error obteniendo órdenes: {e}")
            return

        print(f"[Scheduler Paris] Órdenes obtenidas: {len(ordenes)}")

        for o in ordenes:
            try:
                sub_order = str(o.get("subOrderNumber") or o.get("subOrder") or o.get("orderNumber") or "")
                if not sub_order:
                    continue

                # FIX: usar PARIS- como prefijo (consistente con scheduler manual en paris.py)
                pa_key  = f"PARIS-{sub_order}"
                cancel_key = f"PA-CANCEL-{sub_order}"

                # FIX: el estado en París viene en itemStatus dentro de shipments[].items[],
                # NO a nivel de la orden. Tomar el estado general de la orden si existe,
                # pero NO descartar órdenes por estado — procesar todas las no marcadas.
                estado_orden = (
                    o.get("status") or o.get("itemStatus") or o.get("orderStatus") or ""
                ).lower()

                # Estados que indican cancelación explícita
                es_cancelada = estado_orden in ("canceled", "cancelled", "rejected", "failure")

                # ── Órdenes canceladas ──
                if es_cancelada:
                    if orden_ya_procesada_texto(cancel_key):
                        continue
                    if not orden_ya_procesada_texto(pa_key):
                        marcar_orden_procesada_texto(cancel_key)
                        continue
                    # Reintegrar: recorrer shipments → items (estructura real de París)
                    items_reintegrados = []
                    ultimo_sku = None
                    for ship in (o.get("shipments") or []):
                        for item in (ship.get("items") or []):
                            seller_sku = (item.get("seller_sku") or item.get("sellerSku") or "").strip()
                            cantidad = int(item.get("quantity") or 1)
                            if not seller_sku:
                                continue
                            ultimo_sku = seller_sku
                            sku_lusync = obtener_sku_lusync_por_canal("paris", seller_sku) or seller_sku
                            productos = cargar_productos()
                            prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                            if not prod:
                                continue
                            prod["stock"] += cantidad
                            guardar_producto(prod)
                            registrar_movimiento(
                                "entrada", prod["sku"], prod["nombre"], cantidad,
                                f"Cancelación París orden {sub_order}",
                                usuario="Sistema", canal="París", orden_id=sub_order
                            )
                            sincronizar_stock_marketplaces(prod["sku"], prod["stock"], contexto="paris_cancelacion_bg")
                            items_reintegrados.append(f"{prod['nombre']} (SKU: {seller_sku}) x{cantidad}")
                    if items_reintegrados:
                        try:
                            crear_alerta(
                                tipo="cancelacion",
                                titulo=f"Orden cancelada en París: {sub_order}",
                                mensaje="Stock reintegrado:<br>" + "<br>".join(f"• {it}" for it in items_reintegrados),
                                sku=ultimo_sku
                            )
                        except:
                            pass
                        canceladas += 1
                    marcar_orden_procesada_texto(cancel_key)
                    continue

                # ── Órdenes nuevas: procesar todas las no marcadas ──
                if orden_ya_procesada_texto(pa_key):
                    continue

                # FIX CRÍTICO: los items están en shipments[].items[], NO en o.get("items")
                es_cd = detectar_fulfillment_paris(o)
                tipo_str = "Fulfillment" if es_cd else "Seller"
                items_descontados = []

                # Parsear fecha real de compra (Paris API: createdAt en ISO UTC)
                fecha_compra_pa = None
                try:
                    date_str = (o.get("createdAt") or o.get("created_at") or "")
                    if date_str:
                        fecha_compra_pa = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                except Exception:
                    fecha_compra_pa = None

                for ship in (o.get("shipments") or []):
                    for item in (ship.get("items") or []):
                        seller_sku = (item.get("seller_sku") or item.get("sellerSku") or "").strip()
                        cantidad = int(item.get("quantity") or 1)
                        if not seller_sku:
                            continue

                        sku_lusync = obtener_sku_lusync_por_canal("paris", seller_sku) or seller_sku
                        productos = cargar_productos()
                        prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                        if not prod:
                            print(f"[Scheduler Paris] SKU '{sku_lusync}' no encontrado")
                            continue

                        resultado = descontar_venta(
                            sku=sku_lusync,
                            cantidad=cantidad,
                            canal="Paris",
                            fulfillment=es_cd,
                            orden_id=sub_order,
                            motivo=f"Venta Paris {tipo_str}",
                            usuario="Sistema",
                            fecha_compra_marketplace=fecha_compra_pa,
                            origen_registro="scheduler"
                        )
                        sincronizar_stock_marketplaces(
                            sku_lusync, resultado.get("stock_despues", 0),
                            contexto="paris_orden_bg"
                        )
                        items_descontados.append(f"{seller_sku} x{cantidad}")
                        print(f"[Scheduler Paris] {sub_order} {tipo_str}: {sku_lusync} -{cantidad} desde {resultado.get('bodega','?')}")

                if items_descontados:
                    marcar_orden_procesada_texto(pa_key)
                    nuevas += 1
                elif not items_descontados and o.get("shipments"):
                    # Tiene shipments pero todos los items fallaron (SKU no mapeado, etc.)
                    # Marcar igual para no reintentar infinitamente
                    marcar_orden_procesada_texto(pa_key)

            except Exception as e:
                errores.append(f"PA orden: {e}")

        import gc; gc.collect()
        print(f"[Scheduler Paris] Sync OK — nuevas:{nuevas} canceladas:{canceladas} errores:{len(errores)}")
        if errores:
            print(f"[Scheduler Paris] Errores: {errores[:3]}")
    except Exception as e:
        print(f"[Scheduler Paris] Error general: {e}")
    finally:
        _sync_locks["paris"]["running"] = False


def _sync_ripley_automatico():
    """Sync órdenes + cancelaciones Ripley cada 10 min."""
    if _sync_locks["ripley"]["running"]:
        print("[Scheduler Ripley] Ya hay un sync corriendo, salto")
        return
    _sync_locks["ripley"]["running"] = True
    try:
        print("[Scheduler Ripley] Iniciando sync automático...")
        from ripley import obtener_ordenes_ripley
        from inventario import obtener_sku_lusync_por_canal
        from bodegas_logic import descontar_venta, detectar_fulfillment_ripley
        nuevas = 0
        canceladas = 0
        errores = []

        # FIX: Mirakl/Ripley requiere filtrar por estado explícito — si se pasa estado=None
        # la API puede devolver lista vacía o error. Iterar por cada estado activo
        # igual que hace el scheduler manual en ripley.py.
        # FIX: los estados en Mirakl son: WAITING_ACCEPTANCE, WAITING_DEBIT, SHIPPING,
        # SHIPPED, RECEIVED, REFUSED, CANCELED — en MAYÚSCULAS.
        estados_activos    = ["WAITING_ACCEPTANCE", "WAITING_DEBIT", "SHIPPING", "SHIPPED", "RECEIVED"]
        estados_cancelados = ["REFUSED", "CANCELED"]

        todas_ordenes = []
        for estado in estados_activos + estados_cancelados:
            try:
                lote = obtener_ordenes_ripley(estado=estado, dias=7, max_resultados=50)
                if isinstance(lote, list):
                    todas_ordenes.extend(lote)
            except Exception as e:
                errores.append(f"RP get estado {estado}: {e}")

        print(f"[Scheduler Ripley] Total órdenes obtenidas: {len(todas_ordenes)}")

        for o in todas_ordenes:
            try:
                # FIX: Mirakl devuelve el ID en "order_id" y el estado en "order_state"
                order_id = str(
                    o.get("order_id") or o.get("commercial_id") or
                    o.get("id") or ""
                )
                if not order_id:
                    continue

                # FIX: campo correcto es order_state (no status ni state)
                estado = (
                    o.get("order_state") or o.get("status") or o.get("state") or ""
                ).upper()

                rp_key     = f"RIPLEY-{order_id}"
                cancel_key = f"RP-CANCEL-{order_id}"

                # FIX: los items en Mirakl vienen en order_lines[].order_line_items[]
                # o directamente en order_lines[] según versión. Intentar ambos.
                items_orden = o.get("order_lines") or o.get("items") or o.get("lines") or []

                # ── Órdenes canceladas ──
                if estado in ("REFUSED", "CANCELED", "CANCELLED"):
                    if orden_ya_procesada_texto(cancel_key):
                        continue
                    if not orden_ya_procesada_texto(rp_key):
                        marcar_orden_procesada_texto(cancel_key)
                        continue
                    items_reintegrados = []
                    ultimo_sku = None
                    for item in items_orden:
                        shop_sku = (item.get("offer_sku") or item.get("shop_sku") or
                                    item.get("sku") or item.get("seller_sku") or "").strip()
                        cantidad = int(item.get("quantity") or 1)
                        if not shop_sku:
                            continue
                        ultimo_sku = shop_sku
                        sku_lusync = obtener_sku_lusync_por_canal("ripley", shop_sku) or shop_sku
                        productos = cargar_productos()
                        prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                        if not prod:
                            continue
                        prod["stock"] += cantidad
                        guardar_producto(prod)
                        registrar_movimiento(
                            "entrada", prod["sku"], prod["nombre"], cantidad,
                            f"Cancelación Ripley orden {order_id}",
                            usuario="Sistema", canal="Ripley", orden_id=order_id
                        )
                        sincronizar_stock_marketplaces(prod["sku"], prod["stock"], contexto="ripley_cancelacion_bg")
                        items_reintegrados.append(f"{prod['nombre']} (SKU: {shop_sku}) x{cantidad}")
                    if items_reintegrados:
                        try:
                            crear_alerta(
                                tipo="cancelacion",
                                titulo=f"Orden cancelada en Ripley: {order_id}",
                                mensaje="Stock reintegrado:<br>" + "<br>".join(f"• {it}" for it in items_reintegrados),
                                sku=ultimo_sku
                            )
                        except:
                            pass
                        canceladas += 1
                    marcar_orden_procesada_texto(cancel_key)
                    continue

                # ── Órdenes nuevas activas ──
                if estado not in ("WAITING_ACCEPTANCE", "WAITING_DEBIT", "SHIPPING", "SHIPPED", "RECEIVED"):
                    continue  # estado desconocido, saltar
                if orden_ya_procesada_texto(rp_key):
                    continue

                es_fbr = detectar_fulfillment_ripley(o)
                tipo_str = "FBR" if es_fbr else "Seller"
                items_descontados = []

                # Parsear fecha real de compra (Mirakl/Ripley: created_date en ISO UTC)
                fecha_compra_rp = None
                try:
                    date_str = (o.get("created_date") or o.get("createdDate") or "")
                    if date_str:
                        fecha_compra_rp = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                except Exception:
                    fecha_compra_rp = None

                for item in items_orden:
                    shop_sku = (item.get("offer_sku") or item.get("shop_sku") or
                                item.get("sku") or item.get("seller_sku") or "").strip()
                    cantidad = int(item.get("quantity") or 1)
                    if not shop_sku:
                        continue

                    sku_lusync = obtener_sku_lusync_por_canal("ripley", shop_sku) or shop_sku
                    productos = cargar_productos()
                    prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                    if not prod:
                        print(f"[Scheduler Ripley] SKU '{sku_lusync}' no encontrado")
                        continue

                    resultado = descontar_venta(
                        sku=sku_lusync,
                        cantidad=cantidad,
                        canal="Ripley",
                        fulfillment=es_fbr,
                        orden_id=order_id,
                        motivo=f"Venta Ripley {tipo_str}",
                        usuario="Sistema",
                        fecha_compra_marketplace=fecha_compra_rp,
                        origen_registro="scheduler"
                    )
                    sincronizar_stock_marketplaces(
                        sku_lusync, resultado.get("stock_despues", 0),
                        contexto="ripley_orden_bg"
                    )
                    items_descontados.append(f"{shop_sku} x{cantidad}")
                    print(f"[Scheduler Ripley] {order_id} {tipo_str}: {sku_lusync} -{cantidad} desde {resultado.get('bodega','?')}")

                if items_descontados:
                    marcar_orden_procesada_texto(rp_key)
                    nuevas += 1

            except Exception as e:
                errores.append(f"RP orden: {e}")

        import gc; gc.collect()
        print(f"[Scheduler Ripley] Sync OK — nuevas:{nuevas} canceladas:{canceladas} errores:{len(errores)}")
        if errores:
            print(f"[Scheduler Ripley] Errores: {errores[:3]}")
    except Exception as e:
        print(f"[Scheduler Ripley] Error general: {e}")
    finally:
        _sync_locks["ripley"]["running"] = False


def _sync_woo_automatico():
    """Sync órdenes + cancelaciones WooCommerce cada 10 min.
    
    Si el sitio Woo (babymine.cl) tiene problemas SSL/timeout, sale rápido sin afectar
    los syncs de otros marketplaces.
    """
    if _sync_locks["woo"]["running"]:
        print("[Scheduler Woo] Ya hay un sync corriendo, salto")
        return
    _sync_locks["woo"]["running"] = True
    try:
        print("[Scheduler Woo] Iniciando sync automático...")
        nuevas = 0
        canceladas = 0
        errores = []

        # Fecha UTC moderna (datetime.utcnow() está deprecated en Python 3.12+)
        from datetime import timezone as _tz
        fecha_corte = (datetime.now(_tz.utc) - timedelta(days=2)).isoformat()

        # Órdenes nuevas (processing/completed)
        ordenes_nuevas = []
        try:
            res = requests.get(
                "https://www.babymine.cl/wp-json/wc/v3/orders",
                params={
                    "consumer_key": WC_KEY, "consumer_secret": WC_SECRET,
                    "status": "processing,completed", "per_page": 50,
                    "after": fecha_corte
                },
                timeout=10  # Timeout corto: si Woo está lento o caído, abortamos rápido
            )
            if res.status_code == 200:
                ordenes_nuevas = res.json() or []
            else:
                print(f"[Scheduler Woo] HTTP {res.status_code} consultando órdenes nuevas")
        except requests.exceptions.SSLError as e:
            print(f"[Scheduler Woo] SSL error en babymine.cl (sitio Woo no responde): {str(e)[:120]}")
            print(f"[Scheduler Woo] Saltando sync Woo esta vez. Verificar certificado SSL del sitio.")
            return  # Salir limpio, otros marketplaces siguen
        except requests.exceptions.Timeout:
            print(f"[Scheduler Woo] Timeout consultando Woo (>10s) — sitio lento. Saltando.")
            return
        except Exception as e:
            print(f"[Scheduler Woo] Error obteniendo nuevas: {e}")
            return

        for o in ordenes_nuevas:
            try:
                order_id = str(o.get("id", ""))
                woo_key = f"WOO-{order_id}"
                if orden_ya_procesada_texto(woo_key):
                    continue
                items_descontados = []
                # Parsear fecha real de compra (WooCommerce: date_created en ISO)
                fecha_compra_woo = None
                try:
                    date_str = (o.get("date_created") or o.get("date_created_gmt") or "")
                    if date_str:
                        fecha_compra_woo = datetime.fromisoformat(
                            str(date_str).replace("Z", "+00:00")
                        )
                except Exception:
                    fecha_compra_woo = None

                for line in o.get("line_items", []):
                    sku = (line.get("sku") or "").strip()
                    cantidad = int(line.get("quantity") or 1)
                    if not sku: continue
                    productos = cargar_productos()
                    for p in productos:
                        if p["sku"] == sku:
                            p["stock"] = max(0, p["stock"] - cantidad)
                            guardar_producto(p)
                            registrar_movimiento(
                                "salida", p["sku"], p["nombre"], cantidad,
                                f"Venta Web (Woo) orden {order_id}",
                                usuario="Sistema", canal="Web", orden_id=order_id,
                                fecha_override=fecha_compra_woo
                            )
                            sincronizar_stock_marketplaces(
                                p["sku"], p["stock"], contexto="woo_orden_bg"
                            )
                            items_descontados.append(sku)
                            break
                if items_descontados:
                    marcar_orden_procesada_texto(woo_key)
                    nuevas += 1
            except Exception as e:
                errores.append(f"Woo orden: {e}")

        # Órdenes canceladas (mismo manejo de errores robusto)
        ordenes_canc = []
        try:
            res_c = requests.get(
                "https://www.babymine.cl/wp-json/wc/v3/orders",
                params={
                    "consumer_key": WC_KEY, "consumer_secret": WC_SECRET,
                    "status": "cancelled,refunded,failed", "per_page": 50,
                    "after": fecha_corte
                },
                timeout=10
            )
            if res_c.status_code == 200:
                ordenes_canc = res_c.json() or []
            else:
                print(f"[Scheduler Woo] HTTP {res_c.status_code} consultando canceladas")
        except (requests.exceptions.SSLError, requests.exceptions.Timeout) as e:
            print(f"[Scheduler Woo] Saltando canceladas por error red: {str(e)[:80]}")
            ordenes_canc = []
        except Exception as e:
            print(f"[Scheduler Woo] Error obteniendo canceladas: {e}")
            ordenes_canc = []

        for o in ordenes_canc:
            try:
                order_id = str(o.get("id", ""))
                woo_key = f"WOO-{order_id}"
                cancel_key = f"WOO-CANCEL-{order_id}"
                if orden_ya_procesada_texto(cancel_key): continue
                if not orden_ya_procesada_texto(woo_key):
                    marcar_orden_procesada_texto(cancel_key)
                    continue
                items_reintegrados = []
                ultimo_sku = None
                for line in o.get("line_items", []):
                    sku = (line.get("sku") or "").strip()
                    cantidad = int(line.get("quantity") or 1)
                    if not sku: continue
                    ultimo_sku = sku
                    productos = cargar_productos()
                    for p in productos:
                        if p["sku"] == sku:
                            p["stock"] += cantidad
                            guardar_producto(p)
                            registrar_movimiento(
                                "entrada", p["sku"], p["nombre"], cantidad,
                                f"Cancelación Web orden {order_id}",
                                usuario="Sistema", canal="Web", orden_id=order_id
                            )
                            sincronizar_stock_marketplaces(
                                p["sku"], p["stock"], contexto="woo_cancelacion_bg"
                            )
                            items_reintegrados.append(f"{p['nombre']} (SKU: {sku}) x{cantidad}")
                            break
                if items_reintegrados:
                    try:
                        crear_alerta(
                            tipo="cancelacion",
                            titulo=f"Orden cancelada en Web: {order_id}",
                            mensaje="Stock reintegrado:<br>" + "<br>".join(f"• {it}" for it in items_reintegrados),
                            sku=ultimo_sku
                        )
                    except: pass
                    canceladas += 1
                marcar_orden_procesada_texto(cancel_key)
            except Exception as e:
                errores.append(f"Woo cancel: {e}")

        print(f"[Scheduler Woo] Sync OK — nuevas:{nuevas} canceladas:{canceladas} errores:{len(errores)}")
    except Exception as e:
        print(f"[Scheduler Woo] Error general: {e}")
    finally:
        _sync_locks["woo"]["running"] = False


# ── Registrar todos los schedulers (escalonados) ──
# Walmart cada 5 min (existente)
scheduler.add_job(_sync_walmart_automatico, "interval", minutes=5, id="walmart_sync")
# MELI cada 5 min (más vendido = más reactivo)
scheduler.add_job(_sync_meli_automatico, "interval", minutes=5, id="meli_sync",
                  next_run_time=(datetime.now() + timedelta(seconds=120)))
# Falabella cada 10 min (2do en volumen)
scheduler.add_job(_sync_falabella_automatico, "interval", minutes=10, id="falabella_sync",
                  next_run_time=(datetime.now() + timedelta(seconds=180)))
# París cada 10 min
scheduler.add_job(_sync_paris_automatico, "interval", minutes=10, id="paris_sync",
                  next_run_time=(datetime.now() + timedelta(seconds=360)))
# Ripley cada 10 min
scheduler.add_job(_sync_ripley_automatico, "interval", minutes=10, id="ripley_sync",
                  next_run_time=(datetime.now() + timedelta(seconds=480)))
# Woo cada 10 min (último por menor volumen)
scheduler.add_job(_sync_woo_automatico, "interval", minutes=10, id="woo_sync",
                  next_run_time=(datetime.now() + timedelta(seconds=600)))


# ════════════════════════════════════════════════════════════════════════════
# SCHEDULER DIARIO DE RESPALDO — VERIFICAR STOCK MELI FULL vs API
# ════════════════════════════════════════════════════════════════════════════
# Cada 24h consulta API MELI y compara con Lusync.
# Si hay desfase (webhook FBM perdido, error de red, etc.), Lusync se ajusta.
# Esta es la "red de seguridad" del sistema FBM automático.
# ════════════════════════════════════════════════════════════════════════════

def _sync_full_meli_diario():
    """Verifica stock MELI Full real vs Lusync. Ajusta si hay diferencias."""
    if _sync_locks.get("full_meli", {}).get("running"):
        print("[Scheduler Full MELI] Ya hay un sync corriendo, salto")
        return
    if "full_meli" not in _sync_locks:
        _sync_locks["full_meli"] = {"running": False}
    _sync_locks["full_meli"]["running"] = True
    try:
        print("[Scheduler Full MELI] Iniciando verificación diaria contra API...")
        from mercadolibre import obtener_stock_full_real_meli
        from inventario import (get_stock_bodega, ajustar_stock_bodega,
                                cargar_productos as _cp, crear_alerta)
        
        stock_real_meli = obtener_stock_full_real_meli()
        if stock_real_meli is None:
            print("[Scheduler Full MELI] No se pudo obtener stock real, salto")
            return
        
        if not stock_real_meli:
            print("[Scheduler Full MELI] Sin publicaciones Full activas")
            return
        
        ajustes = []
        productos_dict = {p["sku"]: p for p in _cp()}
        
        for sku_lusync, datos_meli in stock_real_meli.items():
            if sku_lusync not in productos_dict:
                continue
            
            stock_meli_available = datos_meli.get("available", 0)
            stock_meli_transit = datos_meli.get("in_transit", 0)
            
            # Comparar con Lusync
            stock_lusync_full = get_stock_bodega(sku_lusync, "MELI_FULL") or 0
            stock_lusync_transit = get_stock_bodega(sku_lusync, "MELI_FULL_TRANSITO") or 0
            
            diff_full = stock_meli_available - stock_lusync_full
            diff_transit = stock_meli_transit - stock_lusync_transit
            
            # Si hay diferencia significativa, ajustar
            if diff_full != 0:
                try:
                    ajustar_stock_bodega(sku_lusync, "MELI_FULL", diff_full)
                    ajustes.append(f"{sku_lusync}: FULL {stock_lusync_full}→{stock_meli_available} (diff {diff_full:+d})")
                except Exception as e:
                    print(f"[Scheduler Full MELI] Error ajustando {sku_lusync} FULL: {e}")
            
            if diff_transit != 0:
                try:
                    ajustar_stock_bodega(sku_lusync, "MELI_FULL_TRANSITO", diff_transit)
                    ajustes.append(f"{sku_lusync}: TRANSITO {stock_lusync_transit}→{stock_meli_transit} (diff {diff_transit:+d})")
                except Exception as e:
                    print(f"[Scheduler Full MELI] Error ajustando {sku_lusync} TRANSITO: {e}")
        
        # ── Reporte ──
        if ajustes:
            print(f"[Scheduler Full MELI] {len(ajustes)} ajustes aplicados:")
            for a in ajustes[:20]:
                print(f"  • {a}")
            
            # Crear alerta si hay muchos ajustes (puede indicar webhook perdido)
            if len(ajustes) >= 3:
                try:
                    crear_alerta(
                        tipo="full_resync",
                        titulo=f"⚙️ Sync diario Full MELI: {len(ajustes)} ajustes",
                        mensaje=f"Se detectaron diferencias entre stock MELI Full real y Lusync.<br>"
                               f"Sistema ajustó automáticamente.<br><br>"
                               f"Primeros ajustes:<br>" + "<br>".join(f"• {a}" for a in ajustes[:5]),
                        canal="mercadolibre"
                    )
                except: pass
        else:
            print(f"[Scheduler Full MELI] OK — Stock Lusync coincide con MELI ({len(stock_real_meli)} SKUs verificados)")
    except Exception as e:
        import traceback
        print(f"[Scheduler Full MELI] Error general: {e}")
        print(traceback.format_exc())
    finally:
        _sync_locks["full_meli"]["running"] = False


# Registrar scheduler diario (cada 24 horas, primera ejecución a las 6h después del arranque)
# Tiempo elegido: 6h después de arrancar para no saturar el deploy inicial
scheduler.add_job(_sync_full_meli_diario, "interval", hours=24, id="full_meli_diario",
                  next_run_time=(datetime.now() + timedelta(hours=6)))

scheduler.start()
atexit.register(lambda: scheduler.shutdown(wait=False))

# ── SYNC DE RECUPERACIÓN AL ARRANCAR ──
# Busca órdenes perdidas durante caídas del servidor
def _sync_recuperacion():
    try:
        print("[Recuperación] Buscando órdenes no procesadas...")
        productos = cargar_productos()
        recuperadas = 0
        for estado in ["Created", "Acknowledged", "Shipped", "Delivered"]:
            ordenes = obtener_ordenes_walmart(estado)
            for o in ordenes:
                order_id = o.get("purchaseOrderId")
                if not order_id:
                    continue
                customer_order_id = str(o.get("customerOrderId", order_id))
                if orden_ya_procesada_texto(customer_order_id):
                    continue

                # Marcar ANTES de procesar para evitar dobles descuentos
                marcar_orden_procesada_texto(customer_order_id)

                lineas = o.get("orderLines", {}).get("orderLine", [])
                if isinstance(lineas, dict):
                    lineas = [lineas]

                for linea in lineas:
                    try:
                        sku = linea.get("item", {}).get("sku")
                        if not sku:
                            continue
                        cantidad = 1
                        qty = linea.get("orderLineQuantity", {})
                        if qty and qty.get("amount"):
                            cantidad = int(float(qty.get("amount", 1)))
                        if cantidad == 1:
                            status_qty = linea.get("statusQuantity", {})
                            if status_qty and status_qty.get("amount"):
                                cantidad = int(float(status_qty.get("amount", 1)))

                        for p in productos:
                            if p["sku"] == sku:
                                p["stock"] = max(0, p["stock"] - cantidad)
                                guardar_producto(p)
                                registrar_movimiento("salida", p["sku"], p["nombre"],
                                                    cantidad, "Venta Walmart (recuperada)",
                                                    usuario="Sistema", canal="Walmart",
                                                    orden_id=customer_order_id)
                                sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
                                try:
                                    from inventario import sincronizar_stock_a_bodega_central
                                    sincronizar_stock_a_bodega_central(p["sku"])
                                except: pass
                                print(f"[Recuperación] SKU:{sku} Cant:{cantidad} OC:{customer_order_id}")
                    except Exception as e:
                        print(f"[Recuperación] Error linea: {e}")

                marcar_orden_procesada_texto(customer_order_id)
                recuperadas += 1

        # También recuperar cancelaciones
        try:
            canceladas = obtener_ordenes_walmart("Cancelled")
            for o in canceladas:
                order_id = o.get("purchaseOrderId")
                if not order_id:
                    continue
                customer_order_id = str(o.get("customerOrderId", order_id))
                cancel_key = f"CANCEL-{customer_order_id}"
                if not orden_ya_procesada_texto(customer_order_id):
                    continue
                if orden_ya_procesada_texto(cancel_key):
                    continue
                lineas = o.get("orderLines", {}).get("orderLine", [])
                if isinstance(lineas, dict):
                    lineas = [lineas]
                items_cancel_rec = []
                for linea in lineas:
                    sku = linea.get("item", {}).get("sku")
                    if not sku:
                        continue
                    cantidad = 1
                    qty = linea.get("orderLineQuantity", {})
                    if qty and qty.get("amount"):
                        cantidad = int(float(qty.get("amount", 1)))
                    for p in productos:
                        if p["sku"] == sku:
                            p["stock"] += cantidad
                            guardar_producto(p)
                            registrar_movimiento("entrada", p["sku"], p["nombre"],
                                                cantidad, "Cancelación Walmart (recuperada)",
                                                usuario="Sistema", canal="Walmart",
                                                orden_id=customer_order_id)
                            sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
                            items_cancel_rec.append(f"{p['nombre']} (SKU: {sku}) x{cantidad}")
                if items_cancel_rec:
                    try:
                        crear_alerta(
                            tipo="cancelacion",
                            canal="Walmart",
                            titulo=f"Orden cancelada en Walmart: {customer_order_id}",
                            mensaje="El cliente canceló la orden. Stock reintegrado automáticamente:<br><br>" +
                                    "<br>".join(f"• {it}" for it in items_cancel_rec),
                            orden_id=customer_order_id
                        )
                    except Exception as e:
                        print(f"[Recuperación] Error creando alerta: {e}")
                marcar_orden_procesada_texto(cancel_key)
        except Exception as e:
            print(f"[Recuperación] Error cancelaciones: {e}")

        print(f"[Recuperación] Completado — {recuperadas} órdenes recuperadas")
    except Exception as e:
        print(f"[Recuperación] Error general: {e}")

# Ejecutar recuperación 10 segundos después del arranque
scheduler.add_job(_sync_recuperacion, "date", 
                  run_date=__import__("datetime").datetime.now() + __import__("datetime").timedelta(seconds=10),
                  id="recovery_sync")

@app.route("/agregar", methods=["POST"])
def agregar():
    data = request.json
    p = {
        "sku": data["sku"],
        "nombre": data["nombre"],
        "stock": int(data["stock"]),
        "precio_normal": float(data.get("precio_normal", 0)),
        "precio_oferta": float(data.get("precio_oferta", 0))
    }
    guardar_producto(p)
    return {"ok": True}

@app.route("/importar_woo")
def importar():
    """Importa productos desde WooCommerce respetando el modelo padre/variante.

    Reglas:
    - Tipo 'simple'   → importa el producto directamente (es la unidad de stock)
    - Tipo 'variable' → importa SOLO las variantes (el padre NO lleva stock en Woo)
    - Tipo 'grouped'  → omite (es un grupo, no lleva stock propio)
    - Tipo 'external' → omite (link externo, sin stock)

    Para cada variante, prefiere el SKU manual configurado en Woo. Si una
    variante no tiene SKU, se omite y se reporta en el log (para que el usuario
    lo configure en Woo antes de re-intentar).
    """
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "importar_woo", entidad="productos",
                    detalle="Importación desde WooCommerce")

    productos = cargar_productos()
    skus_existentes = {p["sku"] for p in productos}

    # Contadores
    nuevos = 0
    actualizados = 0
    omitidos_padre_variable = 0
    omitidos_sin_sku = 0
    omitidos_tipo = 0
    ya_existian = 0
    errores = 0
    detalle = []  # log producto por producto

    # ── Paginación: Woo limita a 100 por página, hay que pedir varias ──
    page = 1
    todos_productos = []
    while True:
        try:
            res = requests.get(
                "https://www.babymine.cl/wp-json/wc/v3/products",
                params={
                    "consumer_key": WC_KEY,
                    "consumer_secret": WC_SECRET,
                    "per_page": 100,
                    "page": page,
                    "status": "publish"  # solo productos publicados
                },
                timeout=30
            )
            if res.status_code != 200:
                detalle.append(f"❌ Error HTTP {res.status_code} en página {page}: {res.text[:200]}")
                break
            lote = res.json()
            if not lote:
                break
            todos_productos.extend(lote)
            detalle.append(f"📥 Página {page}: {len(lote)} productos obtenidos")
            if len(lote) < 100:
                break
            page += 1
            if page > 50:  # safety: max 5000 productos
                detalle.append("⚠ Límite de 50 páginas alcanzado, detengo paginación")
                break
        except Exception as e:
            detalle.append(f"❌ Excepción en página {page}: {e}")
            errores += 1
            break

    detalle.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    detalle.append(f"Total productos Woo a procesar: {len(todos_productos)}")
    detalle.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    # ── Helper: formatear nombre con atributos de variante ──
    def _nombre_variante(producto_padre, variante):
        """Construye nombre 'Coche Reversible Pro - Color: Azul, Talla: M' a partir
        de los atributos de la variante."""
        nombre_base = (producto_padre.get("name") or "").strip()
        atributos = variante.get("attributes", []) or []
        # attributes en variations: [{"id": 1, "name": "Color", "option": "Azul"}, ...]
        partes = []
        for a in atributos:
            nombre_attr = a.get("name") or a.get("option_name") or ""
            valor_attr = a.get("option") or a.get("value") or ""
            if valor_attr:
                if nombre_attr:
                    partes.append(f"{nombre_attr}: {valor_attr}")
                else:
                    partes.append(valor_attr)
        if partes:
            return f"{nombre_base} - {', '.join(partes)}"
        return nombre_base

    # ── Procesar cada producto Woo ──
    for p in todos_productos:
        tipo = p.get("type", "")
        woo_id = p.get("id", "?")
        nombre_padre = (p.get("name") or "").strip()

        # Caso 1: producto SIMPLE (sin variantes) → importar directo
        if tipo == "simple":
            sku = (p.get("sku") or "").strip()
            if not sku:
                omitidos_sin_sku += 1
                detalle.append(f"⏭ [{woo_id}] '{nombre_padre[:50]}': SIMPLE sin SKU manual, omitido")
                continue

            stock = p.get("stock_quantity") or 0
            pn = p.get("regular_price") or "0"
            po = p.get("sale_price") or "0"
            try:
                precio_normal = float(pn) if pn else 0
                precio_oferta = float(po) if po else 0
            except: precio_normal, precio_oferta = 0, 0

            try:
                if sku in skus_existentes:
                    ya_existian += 1
                    detalle.append(f"= [{woo_id}] {sku} '{nombre_padre[:50]}': SIMPLE ya existe en Lusync, no se sobreescribe")
                else:
                    guardar_producto({
                        "sku": sku,
                        "nombre": nombre_padre,
                        "stock": stock,
                        "precio_normal": precio_normal,
                        "precio_oferta": precio_oferta
                    })
                    skus_existentes.add(sku)
                    nuevos += 1
                    detalle.append(f"+ [{woo_id}] {sku} '{nombre_padre[:50]}': SIMPLE importado (stock={stock})")
            except Exception as e:
                errores += 1
                detalle.append(f"❌ [{woo_id}] {sku} error: {e}")

        # Caso 2: producto VARIABLE → SOLO importar las variantes (NO el padre)
        elif tipo == "variable":
            omitidos_padre_variable += 1
            detalle.append(f"📦 [{woo_id}] '{nombre_padre[:50]}': VARIABLE (padre NO se importa, busco variantes...)")

            try:
                res_var = requests.get(
                    f"https://www.babymine.cl/wp-json/wc/v3/products/{p['id']}/variations",
                    params={
                        "consumer_key": WC_KEY,
                        "consumer_secret": WC_SECRET,
                        "per_page": 100
                    },
                    timeout=30
                )
                if res_var.status_code != 200:
                    errores += 1
                    detalle.append(f"  ❌ Error consultando variantes de {woo_id}: HTTP {res_var.status_code}")
                    continue
                variantes = res_var.json() or []
                detalle.append(f"  ↳ {len(variantes)} variante{'s' if len(variantes)!=1 else ''} encontrada{'s' if len(variantes)!=1 else ''}")
            except Exception as e:
                errores += 1
                detalle.append(f"  ❌ Excepción consultando variantes de {woo_id}: {e}")
                continue

            for v in variantes:
                v_id = v.get("id", "?")
                sku_v = (v.get("sku") or "").strip()
                nombre_v = _nombre_variante(p, v)

                # SKU manual obligatorio para variantes
                if not sku_v:
                    omitidos_sin_sku += 1
                    detalle.append(f"  ⏭ Variante [{v_id}] '{nombre_v[:50]}': SIN SKU manual en Woo, omitida")
                    continue

                stock_v = v.get("stock_quantity") or 0
                pn_v = v.get("regular_price") or "0"
                po_v = v.get("sale_price") or "0"
                try:
                    precio_normal_v = float(pn_v) if pn_v else 0
                    precio_oferta_v = float(po_v) if po_v else 0
                except: precio_normal_v, precio_oferta_v = 0, 0

                try:
                    if sku_v in skus_existentes:
                        ya_existian += 1
                        detalle.append(f"  = [{v_id}] {sku_v} '{nombre_v[:50]}': variante ya existe en Lusync")
                    else:
                        guardar_producto({
                            "sku": sku_v,
                            "nombre": nombre_v,
                            "stock": stock_v,
                            "precio_normal": precio_normal_v,
                            "precio_oferta": precio_oferta_v
                        })
                        skus_existentes.add(sku_v)
                        nuevos += 1
                        detalle.append(f"  + [{v_id}] {sku_v} '{nombre_v[:50]}': variante importada (stock={stock_v})")
                except Exception as e:
                    errores += 1
                    detalle.append(f"  ❌ Variante [{v_id}] {sku_v} error: {e}")

        # Caso 3: tipos no soportados (grouped, external, etc.)
        else:
            omitidos_tipo += 1
            detalle.append(f"⏭ [{woo_id}] '{nombre_padre[:50]}': tipo '{tipo}' no soportado, omitido")

    # ── Resumen ──
    total_woo = len(todos_productos)
    detalle.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    detalle.append(f"RESUMEN:")
    detalle.append(f"  Productos en Woo:        {total_woo}")
    detalle.append(f"  Importados nuevos:       {nuevos}")
    detalle.append(f"  Ya existían en Lusync:   {ya_existian}")
    detalle.append(f"  Padres 'variable' (no se importan, sólo variantes): {omitidos_padre_variable}")
    detalle.append(f"  Omitidos sin SKU:        {omitidos_sin_sku}")
    detalle.append(f"  Omitidos por tipo no soportado: {omitidos_tipo}")
    detalle.append(f"  Errores:                 {errores}")

    return jsonify({
        "ok": True,
        "mensaje": f"{nuevos} productos importados ({ya_existian} ya existían)",
        "total_woo": total_woo,
        "nuevos": nuevos,
        "actualizados": actualizados,
        "ya_existian": ya_existian,
        "omitidos_padre_variable": omitidos_padre_variable,
        "omitidos_sin_sku": omitidos_sin_sku,
        "omitidos_tipo": omitidos_tipo,
        "errores": errores,
        "log": detalle
    })

@app.route("/sincronizar_precios_woo")
def sincronizar_precios_woo():
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr, "sincronizar_precios", entidad="productos", detalle="Sincronización de precios WooCommerce")
    actualizados = 0
    res = requests.get(
        "https://www.babymine.cl/wp-json/wc/v3/products",
        params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET, "per_page": 100}
    )
    if res.status_code != 200:
        return {"error": "Woo error"}

    for p in res.json():
        if p["type"] == "simple":
            sku = p.get("sku") or str(p.get("id"))
            pn = p.get("regular_price") or "0"
            po = p.get("sale_price") or "0"
            actualizar_precios(sku,
                float(pn) if pn else 0,
                float(po) if po else 0)
            actualizados += 1

        if p["type"] == "variable":
            res_var = requests.get(
                f"https://www.babymine.cl/wp-json/wc/v3/products/{p['id']}/variations",
                params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET, "per_page": 100}
            )
            if res_var.status_code != 200:
                continue
            for v in res_var.json():
                sku = v.get("sku") or str(v.get("id"))
                vn = v.get("regular_price") or "0"
                vo = v.get("sale_price") or "0"
                actualizar_precios(sku,
                    float(vn) if vn else 0,
                    float(vo) if vo else 0)
                actualizados += 1

    return {"mensaje": f"{actualizados} precios sincronizados"}

@app.route("/actualizar_precios", methods=["POST"])
def actualizar_precios_route():
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr, "actualizar_precios", entidad="productos", detalle="Actualización manual de precios")
    data = request.json
    sku = data.get("sku")
    precio_normal = float(data.get("precio_normal", 0))
    precio_oferta = float(data.get("precio_oferta", 0))

    # Guardar en BD
    actualizar_precios(sku, precio_normal, precio_oferta)

    # Buscar el producto en WooCommerce por SKU
    try:
        res = requests.get(
            "https://www.babymine.cl/wp-json/wc/v3/products",
            params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET, "sku": sku}
        )
        if res.status_code == 200 and res.json():
            producto = res.json()[0]
            payload = {
                "regular_price": str(precio_normal),
                "sale_price": str(precio_oferta) if precio_oferta > 0 else ""
            }
            if producto["type"] == "simple":
                requests.put(
                    f"https://www.babymine.cl/wp-json/wc/v3/products/{producto['id']}",
                    params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET},
                    json=payload
                )
            elif producto["type"] == "variation":
                requests.put(
                    f"https://www.babymine.cl/wp-json/wc/v3/products/{producto['parent_id']}/variations/{producto['id']}",
                    params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET},
                    json=payload
                )
    except:
        pass

    return {"ok": True}

@app.route("/entrada", methods=["POST"])
def entrada():
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr, "entrada_manual", entidad="productos", detalle="Entrada manual de stock")
    data = request.json
    productos = cargar_productos()
    for p in productos:
        if p["sku"] == data["sku"]:
            p["stock"] += int(data["cantidad"])
            guardar_producto(p)
            registrar_movimiento("entrada", p["sku"], p["nombre"], int(data["cantidad"]), data.get("motivo"), usuario="Luis Padilla", canal="Manual")
            # Sync automático a los 6 marketplaces (resiliente: si uno falla, los demás siguen)
            syncs = sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="entrada_manual")
            return {"ok": True, "syncs": syncs}
    return {"error": "no encontrado"}

@app.route("/salida", methods=["POST"])
def salida():
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr, "salida_manual", entidad="productos", detalle="Salida manual de stock")
    data = request.json
    productos = cargar_productos()
    for p in productos:
        if p["sku"] == data["sku"]:
            if p["stock"] < int(data["cantidad"]):
                return {"error": "Stock insuficiente"}
            p["stock"] -= int(data["cantidad"])
            guardar_producto(p)
            registrar_movimiento("salida", p["sku"], p["nombre"], int(data["cantidad"]), data.get("motivo"), usuario="Luis Padilla", canal="Manual")
            # Sync automático a los 6 marketplaces
            syncs = sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="salida_manual")
            return {"ok": True, "syncs": syncs}
    return {"error": "no encontrado"}

@app.route("/sync_ordenes")
def sync_ordenes():
    try:
        res = requests.get(
            "https://www.babymine.cl/wp-json/wc/v3/orders",
            params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET, "status": "processing"},
            timeout=15
        )
    except requests.exceptions.Timeout:
        print("[WooCommerce] Timeout en sync_ordenes")
        return {"ok": True, "nuevas_ordenes": 0, "warn": "timeout"}
    except Exception as e:
        print(f"[WooCommerce] Error en sync_ordenes: {e}")
        return {"ok": True, "nuevas_ordenes": 0, "warn": str(e)}
    if res.status_code != 200:
        return {"error": "Woo error", "status": res.status_code}

    productos = cargar_productos()
    nuevas = 0

    for o in res.json():
        if orden_ya_procesada(o["id"]):
            continue

        # WooCommerce ya guarda en hora Chile — usar directamente sin convertir
        from datetime import datetime
        try:
            fecha_real = datetime.strptime(o.get("date_created",""), "%Y-%m-%dT%H:%M:%S")
        except:
            fecha_real = None

        for item in o["line_items"]:
            sku = item.get("sku")
            cantidad = item.get("quantity")
            for p in productos:
                if p["sku"] == sku:
                    p["stock"] -= cantidad
                    guardar_producto(p)
                    registrar_movimiento("salida", p["sku"], p["nombre"], cantidad, "Venta Web",
                                        usuario="Sistema", canal="WooCommerce",
                                        orden_id=str(o["id"]), fecha_override=fecha_real)
                    sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
        marcar_orden_procesada(o["id"])
        nuevas += 1

    return {"ok": True, "nuevas_ordenes": nuevas}

@app.route("/movimientos_hoy")
def movimientos_hoy():
    return {"ventas": cargar_movimientos_hoy()}

@app.route("/productos")
def ver_productos():
    if not session.get("logged"):
        return {"productos": [], "error": "no autorizado"}, 401
    try:
        return {"productos": cargar_productos()}
    except Exception as e:
        print(f"[/productos] Error: {e}")
        return {"productos": [], "error": str(e)}, 500

@app.route("/movimientos")
def ver_movimientos():
    if not session.get("logged"):
        return {"movimientos": [], "error": "no autorizado"}, 401
    try:
        limite = int(request.args.get("limite", 20))
        return {"movimientos": cargar_movimientos(limite)}
    except Exception as e:
        print(f"[/movimientos] Error: {e}")
        return {"movimientos": [], "error": str(e)}, 500

# ── WALMART ──

@app.route("/walmart/test")
def walmart_test():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    try:
        from walmart import get_token, WALMART_CLIENT_ID
        token = get_token()
        return {"conectado": True, "client_id": WALMART_CLIENT_ID[:8]+"..."}
    except Exception as e:
        return {"conectado": False, "error": str(e)}

@app.route("/walmart/diagnostico")
def walmart_diagnostico():
    """Diagnóstico completo de Walmart — items, SKUs y test de inventory"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    import requests as req
    from walmart import walmart_headers, WALMART_BASE_URL

    resultado = {}

    # 1. Traer items publicados en Walmart
    try:
        res = req.get(
            f"{WALMART_BASE_URL}/v3/items",
            headers=walmart_headers(),
            params={"limit": 5}
        )
        resultado["items_status"] = res.status_code
        resultado["items_respuesta"] = res.text[:800]
    except Exception as e:
        resultado["items_error"] = str(e)

    # 2. Buscar el producto por SKU específico
    try:
        res2 = req.get(
            f"{WALMART_BASE_URL}/v3/items/CBSNCPB001",
            headers=walmart_headers(),
            params={"productIdType": "SKU"}
        )
        resultado["busqueda_sku_status"] = res2.status_code
        resultado["busqueda_sku_respuesta"] = res2.text[:500]
    except Exception as e:
        resultado["busqueda_sku_error"] = str(e)

    # 3. Probar inventory con cantidad fija
    try:
        headers = walmart_headers()
        headers["Content-Type"] = "application/json"
        payload = {"quantity": {"unit": "EACH", "amount": 10}}
        res3 = req.put(
            f"{WALMART_BASE_URL}/v3/inventory",
            headers=headers,
            json=payload,
            params={"sku": "CBSNCPB001"}
        )
        resultado["inventory_sin_param_status"] = res3.status_code
        resultado["inventory_sin_param_respuesta"] = res3.text[:500]
    except Exception as e:
        resultado["inventory_error"] = str(e)

    return resultado

@app.route("/walmart/test_stock_one")
def walmart_test_stock_one():
    """Prueba actualizar stock de UN solo producto para debug"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    productos = cargar_productos()
    if not productos:
        return {"error": "sin productos"}
    p = productos[0]
    try:
        import requests as req
        from walmart import get_token, WALMART_BASE_URL, walmart_headers
        headers = walmart_headers()
        headers["Content-Type"] = "application/json"
        payload = {
            "sku": p["sku"],
            "quantity": {"unit": "EACH", "amount": int(p["stock"])}
        }
        res = req.put(
            f"{WALMART_BASE_URL}/v3/inventory",
            headers=headers,
            json=payload
        )
        return {
            "sku": p["sku"],
            "stock": p["stock"],
            "status": res.status_code,
            "respuesta": res.text[:500]
        }
    except Exception as e:
        return {"error": str(e)}

@app.route("/walmart/sync_stock", methods=["POST"])
def walmart_sync_stock():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    productos = cargar_productos()
    ok = 0
    error = 0
    errores_detalle = []
    print(f"[Sync Stock] Iniciando sync masivo — {len(productos)} productos a 6 marketplaces")
    for p in productos:
        if p.get("sku"):
            # Sync a TODOS los marketplaces (el helper es resiliente)
            syncs = sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="sync_masivo_manual")
            # Cuenta como "ok" si al menos Walmart respondió bien (manteniendo lógica original)
            if syncs.get("walmart") == "ok":
                ok += 1
            else:
                error += 1
                errores_detalle.append(p["sku"])
    print(f"[Sync Stock] Completado — OK:{ok} Error:{error}")
    return {"ok": ok, "error": error, "total": len(productos), "errores": errores_detalle[:5]}

@app.route("/walmart/sync_precios", methods=["POST"])
def walmart_sync_precios():
    """Envía precios actuales (precio_normal/oferta) tal cual a Walmart, sin transformaciones.
    Las comisiones, márgenes y redondeos se manejarán en el módulo Motor de Precios."""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    productos = cargar_productos()
    enviados = 0
    fallidos = 0
    log = []
    for p in productos:
        if not p.get("sku") or not p.get("precio_normal", 0) > 0:
            continue
        # Enviar precio tal cual: si hay oferta, usar oferta; si no, precio normal
        precio = p["precio_oferta"] if p.get("precio_oferta", 0) > 0 else p["precio_normal"]
        if actualizar_precio_walmart(p["sku"], precio):
            enviados += 1
            log.append(f"✓ {p['sku']} → ${precio}")
        else:
            fallidos += 1
            log.append(f"× {p['sku']} falló")
    return {"ok": True, "enviados": enviados, "fallidos": fallidos, "log": log[:30]}

# El endpoint /walmart/sync_ordenes se movió a walmart.py (Blueprint)
# Ver: walmart_bp en walmart.py

@app.route("/walmart/ver_ordenes")
def walmart_ver_ordenes():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    import requests as req
    from walmart import walmart_headers, WALMART_BASE_URL
    resultado = {}

    from datetime import datetime, timedelta
    fecha_inicio = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%dT00:00:00.000Z")

    # Probar sin filtro de estado pero con fecha
    try:
        h = walmart_headers()
        res = req.get(
            f"{WALMART_BASE_URL}/v3/orders",
            headers=h,
            params={"createdStartDate": fecha_inicio, "limit": 5}
        )
        resultado["sin_filtro_status"] = res.status_code
        resultado["sin_filtro_resp"] = res.text[:600]
    except Exception as e:
        resultado["sin_filtro_error"] = str(e)

    # Probar con cada estado
    for estado in ["Created", "Acknowledged", "Shipped", "Delivered"]:
        try:
            h2 = walmart_headers()
            res2 = req.get(
                f"{WALMART_BASE_URL}/v3/orders",
                headers=h2,
                params={"createdStartDate": fecha_inicio, "status": estado, "limit": 5}
            )
            resultado[estado+"_status"] = res2.status_code
            resultado[estado+"_resp"] = res2.text[:300]
        except Exception as e:
            resultado[estado+"_error"] = str(e)

    return resultado

@app.route("/fix_woo_limpiar_duplicados")
def fix_woo_limpiar_duplicados():
    """Limpia duplicados de WooCommerce y deja solo 1 movimiento por orden+SKU con fecha real"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    from datetime import datetime
    import pytz
    conn = get_conn()
    cur = conn.cursor()

    # 1. Borrar TODOS los movimientos de WooCommerce para empezar limpio
    cur.execute("DELETE FROM movimientos WHERE canal = 'WooCommerce'")
    borrados = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()

    # 2. Volver a registrar desde WooCommerce con fecha real de compra
    res = requests.get(
        "https://www.babymine.cl/wp-json/wc/v3/orders",
        params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET,
                "status": "processing", "per_page": 100}
    )
    if res.status_code != 200:
        return {"error": "Woo error", "borrados": borrados}

    productos = cargar_productos()
    registrados = 0
    chile_tz = pytz.timezone('America/Santiago')  # pytz maneja UTC-3/UTC-4 automáticamente

    for o in res.json():
        try:
            # WooCommerce ya guarda en hora Chile — sin conversión
            fecha_real = datetime.strptime(o.get("date_created",""), "%Y-%m-%dT%H:%M:%S")
        except:
            fecha_real = None

        for item in o.get("line_items", []):
            sku = item.get("sku")
            cantidad = item.get("quantity", 1)
            for p in productos:
                if p["sku"] == sku:
                    registrar_movimiento(
                        "salida", p["sku"], p["nombre"],
                        cantidad, "Venta Web",
                        usuario="Sistema", canal="WooCommerce",
                        orden_id=str(o["id"]),
                        fecha_override=fecha_real
                    )
                    registrados += 1

    return {"ok": True, "borrados": borrados, "registrados": registrados}

@app.route("/fix_woo_fechas")
def fix_woo_fechas():
    """Corrige la fecha de movimientos WooCommerce guardados con hora UTC incorrecta"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    # Restar 3 horas a movimientos de WooCommerce del 27/04 que son del 26/04 en Chile
    cur.execute("""
        UPDATE movimientos
        SET fecha = fecha - INTERVAL '3 hours'
        WHERE canal = 'WooCommerce'
        AND DATE(fecha) = '2026-04-27'
        AND EXTRACT(HOUR FROM fecha) < 7
    """)
    corregidos = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True, "corregidos": corregidos}

@app.route("/fix_woo_movimientos")
def fix_woo_movimientos():
    """Registra movimientos faltantes de órdenes WooCommerce ya procesadas"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401

    res = requests.get(
        "https://www.babymine.cl/wp-json/wc/v3/orders",
        params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET,
                "status": "processing", "per_page": 50}
    )
    if res.status_code != 200:
        return {"error": "Woo error"}

    productos = cargar_productos()
    registrados = 0

    for o in res.json():
        # Solo procesar las ya marcadas (que no tienen movimiento)
        if not orden_ya_procesada(o["id"]):
            continue

        # Verificar si ya tiene movimiento registrado
        from inventario import get_conn
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM movimientos WHERE orden_id = %s AND canal = 'WooCommerce'",
            (str(o["id"]),)
        )
        ya_tiene_movimiento = cur.fetchone() is not None
        cur.close()
        conn.close()

        if ya_tiene_movimiento:
            continue

        # Registrar el movimiento con la fecha REAL de la orden de WooCommerce
        from datetime import datetime
        import pytz
        chile_tz = pytz.timezone('America/Santiago')  # pytz maneja UTC-3/UTC-4 automáticamente
        fecha_orden_str = o.get("date_created", "")
        try:
            # WooCommerce devuelve fecha en UTC — convertir a Chile
            fecha_utc = datetime.strptime(fecha_orden_str, "%Y-%m-%dT%H:%M:%S")
            fecha_utc = pytz.utc.localize(fecha_utc)
            fecha_chile = fecha_utc.astimezone(chile_tz)
        except:
            fecha_chile = None

        for item in o.get("line_items", []):
            sku = item.get("sku")
            cantidad = item.get("quantity", 1)
            for p in productos:
                if p["sku"] == sku:
                    registrar_movimiento(
                        "salida", p["sku"], p["nombre"],
                        cantidad, "Venta Web",
                        usuario="Sistema", canal="WooCommerce",
                        orden_id=str(o["id"]),
                        fecha_override=fecha_chile
                    )
                    registrados += 1

    return {"ok": True, "movimientos_registrados": registrados}

@app.route("/debug_woo_ordenes")
def debug_woo_ordenes():
    """Ver órdenes de WooCommerce en estado processing"""
    res = requests.get(
        "https://www.babymine.cl/wp-json/wc/v3/orders",
        params={"consumer_key": WC_KEY, "consumer_secret": WC_SECRET, "status": "processing", "per_page": 10}
    )
    if res.status_code != 200:
        return {"error": res.status_code, "detalle": res.text[:200]}
    ordenes = res.json()
    resultado = []
    for o in ordenes:
        ya = orden_ya_procesada(o["id"])
        resultado.append({
            "id": o["id"],
            "fecha": o.get("date_created"),
            "ya_procesada": ya,
            "items": [{"sku": i.get("sku"), "cantidad": i.get("quantity")} for i in o.get("line_items", [])]
        })
    return {"total": len(ordenes), "ordenes": resultado}

@app.route("/hora_servidor")
def hora_servidor():
    from datetime import datetime
    import pytz
    utc_now = datetime.utcnow()
    chile_tz = pytz.timezone('America/Santiago')  # pytz maneja UTC-3/UTC-4 automáticamente
    chile_now = datetime.now(chile_tz)
    return {
        "utc": utc_now.strftime("%d/%m/%Y %H:%M:%S"),
        "chile_pytz": chile_now.strftime("%d/%m/%Y %H:%M:%S"),
        "chile_offset": str(chile_now.utcoffset()),
        "postgres_now": None
    }

@app.route("/fix_db")
def fix_db():
    """Crea columnas faltantes en la BD"""
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS orden_id TEXT DEFAULT NULL")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS usuario TEXT DEFAULT 'Sistema'")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS canal TEXT DEFAULT 'Sistema'")
        cur.execute("ALTER TABLE ordenes_procesadas ADD COLUMN IF NOT EXISTS order_id_texto TEXT")
        conn.commit()
        cur.close()
        conn.close()
        return {"ok": True, "mensaje": "Columnas creadas correctamente"}
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        return {"error": str(e)}

@app.route("/walmart/reset_y_limpiar")
def walmart_reset_y_limpiar():
    """Borra movimientos de Walmart y limpia órdenes procesadas para resincronizar limpio"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()

    # 1. Borrar movimientos de Walmart únicamente
    cur.execute("DELETE FROM movimientos WHERE canal = 'Walmart' AND motivo = 'Venta Walmart'")
    movimientos_borrados = cur.rowcount

    # 2. Limpiar SOLO órdenes de Walmart (las que tienen order_id_texto con formato P...)
    cur.execute("""
        DELETE FROM ordenes_procesadas
        WHERE order_id_texto IS NOT NULL
    """)
    ordenes_borradas = cur.rowcount

    # 3. Crear columna order_id_texto si no existe
    cur.execute("ALTER TABLE ordenes_procesadas ADD COLUMN IF NOT EXISTS order_id_texto TEXT")

    conn.commit()
    cur.close()
    conn.close()
    return {
        "ok": True,
        "movimientos_borrados": movimientos_borrados,
        "ordenes_borradas": ordenes_borradas,
        "mensaje": "Listo. Ahora sincroniza órdenes de Walmart desde el panel."
    }

@app.route("/walmart/ver_fechas")
def walmart_ver_fechas():
    """Ver fechas exactas de movimientos de Walmart"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            TO_CHAR(fecha, 'DD/MM/YYYY HH24:MI') as utc,
            TO_CHAR(fecha AT TIME ZONE 'America/Santiago', 'DD/MM/YYYY HH24:MI') as santiago,
            TO_CHAR(fecha AT TIME ZONE 'America/Santiago', 'DD/MM/YYYY') as fecha_santiago,
            motivo, canal
        FROM movimientos
        WHERE canal = 'Walmart'
        ORDER BY fecha DESC
        LIMIT 5
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {"movimientos": [
        {"utc":r[0],"santiago":r[1],"fecha_santiago":r[2],"motivo":r[3],"canal":r[4]}
        for r in rows
    ]}

@app.route("/walmart/ver_movimientos_db")
def walmart_ver_movimientos_db():
    """Ver movimientos de hoy en la BD para diagnóstico"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT tipo, sku, nombre, cantidad, motivo, canal, usuario,
               TO_CHAR(fecha AT TIME ZONE 'America/Santiago', 'HH24:MI') as hora
        FROM movimientos
        WHERE DATE(fecha AT TIME ZONE 'America/Santiago') = CURRENT_DATE
        ORDER BY fecha DESC
        LIMIT 20
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {"movimientos": [
        {"tipo":r[0],"sku":r[1],"nombre":r[2][:30],"cantidad":r[3],
         "motivo":r[4],"canal":r[5],"usuario":r[6],"hora":r[7]}
        for r in rows
    ]}

@app.route("/walmart/fix_canales")
def walmart_fix_canales():
    """Corrige hora UTC de movimientos de Walmart procesados antes del fix de timezone"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    # Restar 1 hora adicional a movimientos de Walmart del 27/04 (ya se restaron 4, falta 1 más → total 3h)
    cur.execute("""
        UPDATE movimientos
        SET fecha = fecha - INTERVAL '1 hour'
        WHERE canal = 'Walmart'
        AND motivo = 'Venta Walmart'
        AND DATE(fecha AT TIME ZONE 'UTC') = '2026-04-27'
        AND EXTRACT(HOUR FROM fecha AT TIME ZONE 'UTC') = 0
    """)
    actualizados = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True, "movimientos_corregidos": actualizados}

@app.route("/walmart/sync_debug")
def walmart_sync_debug():
    """Ejecuta el sync completo y retorna resultado detallado"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401

    productos = cargar_productos()
    log = []
    nuevas = 0

    for estado in ["Created", "Acknowledged", "Shipped", "Delivered"]:
        ordenes = obtener_ordenes_walmart(estado)
        log.append(f"Estado {estado}: {len(ordenes)} ordenes")

        for o in ordenes:
            order_id = o.get("purchaseOrderId")
            if not order_id:
                log.append("Sin order_id, saltando")
                continue

            # Bug ③ fix: usar customerOrderId consistente con el resto del sistema
            customer_order_id = str(o.get("customerOrderId", order_id))
            ya = orden_ya_procesada_texto(customer_order_id)
            log.append(f"Orden {order_id} customerOrderId:{customer_order_id} ya_procesada:{ya}")

            if ya:
                continue

            lineas = o.get("orderLines", {}).get("orderLine", [])
            if isinstance(lineas, dict):
                lineas = [lineas]

            log.append(f"  Lineas: {len(lineas)}")

            for linea in lineas:
                sku = linea.get("item", {}).get("sku")
                cantidad = 1
                qty = linea.get("orderLineQuantity", {})
                if qty and qty.get("amount"):
                    cantidad = int(float(qty["amount"]))
                if cantidad == 1:
                    status_qty = linea.get("statusQuantity", {})
                    if status_qty and status_qty.get("amount"):
                        cantidad = int(float(status_qty.get("amount", 1)))

                log.append(f"  SKU:{sku} Cantidad:{cantidad}")

                encontrado = False
                for p in productos:
                    if p["sku"] == sku:
                        encontrado = True
                        stock_antes = p["stock"]
                        p["stock"] = max(0, p["stock"] - cantidad)
                        guardar_producto(p)
                        registrar_movimiento("salida", p["sku"], p["nombre"],
                                            cantidad, "Venta Walmart",
                                            usuario="Sistema", canal="Walmart",
                                            orden_id=customer_order_id)
                        sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
                        try:
                            from inventario import sincronizar_stock_a_bodega_central
                            sincronizar_stock_a_bodega_central(p["sku"])
                        except: pass
                        log.append(f"  OK {p['nombre']} stock:{stock_antes}->{p['stock']}")

                if not encontrado:
                    log.append(f"  SKU {sku} no encontrado en Lusync")

            marcar_orden_procesada_texto(customer_order_id)
            nuevas += 1

    # ── CANCELACIONES en sync manual
    try:
        canceladas = obtener_ordenes_walmart("Cancelled")
        for o in canceladas:
            order_id = o.get("purchaseOrderId")
            if not order_id:
                continue
            customer_order_id = str(o.get("customerOrderId", order_id))
            cancel_key = f"CANCEL-{customer_order_id}"
            if not orden_ya_procesada_texto(customer_order_id):
                continue
            if orden_ya_procesada_texto(cancel_key):
                continue
            lineas = o.get("orderLines", {}).get("orderLine", [])
            if isinstance(lineas, dict):
                lineas = [lineas]
            items_cancel_man = []
            for linea in lineas:
                sku = linea.get("item", {}).get("sku")
                if not sku:
                    continue
                cantidad = 1
                qty = linea.get("orderLineQuantity", {})
                if qty and qty.get("amount"):
                    cantidad = int(float(qty.get("amount", 1)))
                for p in productos:
                    if p["sku"] == sku:
                        p["stock"] += cantidad
                        guardar_producto(p)
                        registrar_movimiento("entrada", p["sku"], p["nombre"],
                                            cantidad, "Cancelación Walmart",
                                            usuario="Sistema", canal="Walmart",
                                            orden_id=customer_order_id)
                        sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
                        items_cancel_man.append(f"{p['nombre']} (SKU: {sku}) x{cantidad}")
                        log.append(f"CANCELACION SKU:{sku} +{cantidad} Stock:{p['stock']}")
            if items_cancel_man:
                try:
                    crear_alerta(
                        tipo="cancelacion",
                        canal="Walmart",
                        titulo=f"Orden cancelada en Walmart: {customer_order_id}",
                        mensaje="El cliente canceló la orden. Stock reintegrado automáticamente:<br><br>" +
                                "<br>".join(f"• {it}" for it in items_cancel_man),
                        orden_id=customer_order_id
                    )
                except Exception as e:
                    log.append(f"Error creando alerta: {e}")
            marcar_orden_procesada_texto(cancel_key)
    except Exception as e:
        log.append(f"Error cancelaciones: {e}")

    return {"nuevas_ordenes": nuevas, "log": log}

@app.route("/walmart/debug_ordenes")
def walmart_debug_ordenes():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from datetime import datetime, timedelta
    import requests as req
    from walmart import walmart_headers, WALMART_BASE_URL

    fecha_inicio = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%dT00:00:00.000Z")
    h = walmart_headers()
    res = req.get(
        f"{WALMART_BASE_URL}/v3/orders",
        headers=h,
        params={"createdStartDate": fecha_inicio, "status": "Acknowledged", "limit": 2}
    )
    if res.status_code != 200:
        return {"error": res.text}

    data = res.json()
    ordenes = data.get("list", {}).get("elements", {}).get("order", [])
    if isinstance(ordenes, dict):
        ordenes = [ordenes]

    # Mostrar estructura completa de la primera orden
    if ordenes:
        o = ordenes[0]
        return {
            "purchaseOrderId": o.get("purchaseOrderId"),
            "orderLines_raw": str(o.get("orderLines", {}))[:1000],
            "keys_orden": list(o.keys()),
            "orden_completa": str(o)[:1500]
        }
    return {"mensaje": "sin ordenes"}

@app.route("/eliminar_producto", methods=["POST"])
def eliminar_producto_route():
    data_in = request.json or {}
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr, "eliminar_producto",
                    entidad="productos", entidad_id=data_in.get("sku","?"),
                    detalle=f"Eliminación producto SKU:{data_in.get('sku','?')}")
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = request.json
    sku = data.get("sku")
    if not sku:
        return {"error": "SKU requerido"}
    eliminar_producto(sku)
    return {"ok": True}

@app.route("/configuracion", methods=["GET","POST"])
def configuracion():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    if request.method == "POST":
        data = request.json
        set_configuracion(data)
        return {"ok": True}
    return {"config": get_configuracion()}

@app.route("/lead_time", methods=["POST"])
def lead_time():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = request.json
    set_lead_time(data.get("sku"), data.get("lead_time", 45))
    return {"ok": True}

# ── DEVOLUCIONES ──

@app.route("/devoluciones")
def devoluciones_list():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    estado = request.args.get("estado", "todas")
    return {"devoluciones": listar_devoluciones(estado)}

@app.route("/devoluciones/nueva", methods=["POST"])
def devoluciones_nueva():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = request.json
    dev_id = crear_devolucion(data)
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "crear_devolucion", entidad="devoluciones", entidad_id=str(dev_id),
                    detalle=f"Nueva DEV: OC={data.get('oc_origen')} SKU={data.get('sku')}")
    return {"ok": True, "id": dev_id}

@app.route("/devoluciones/<int:dev_id>")
def devoluciones_get(dev_id):
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    dev = get_devolucion(dev_id=dev_id)
    if not dev:
        return {"error": "no encontrada"}, 404
    return {"devolucion": dev}

@app.route("/devoluciones/buscar")
def devoluciones_buscar_codigo():
    """Lookup por código DEV para pistoleo"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    codigo = request.args.get("codigo", "").strip()
    dev = get_devolucion(codigo=codigo)
    if not dev:
        return {"error": "no encontrada"}, 404
    return {"devolucion": dev}

@app.route("/devoluciones/lookup_oc")
def devoluciones_lookup_oc():
    """Busca productos asociados a una OC en movimientos"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    oc = request.args.get("oc", "").strip()
    if not oc:
        return {"error": "OC requerida"}, 400
    conn = __import__('psycopg2').connect(__import__('os').environ.get("DATABASE_URL"))
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT m.sku, m.nombre, m.canal,
               ABS(m.cantidad) as cantidad,
               TO_CHAR(
                 CASE WHEN COALESCE(m.canal,'') IN ('Walmart','WooCommerce')
                      THEN m.fecha - INTERVAL '4 hours'
                      ELSE m.fecha
                 END, 'DD/MM/YYYY HH24:MI') as fecha
        FROM movimientos m
        WHERE m.orden_id = %s AND m.tipo = 'salida'
        ORDER BY m.sku
    """, (oc,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    if not rows:
        return {"error": "OC no encontrada en movimientos"}, 404
    items = [{"sku": r[0], "nombre": r[1], "canal": r[2],
              "cantidad": r[3], "fecha": r[4]} for r in rows]
    return {"items": items, "oc": oc}


@app.route("/devoluciones/buscar_orden", methods=["GET", "POST"])
def devoluciones_buscar_orden():
    """Busca una orden de manera inteligente:
    1) Primero en BD local (movimientos) — si la encuentra, ya sabe el canal
    2) Si no está local, intenta detectar marketplace por formato del N°
    3) Devuelve items + canal detectado para que el usuario seleccione qué devolver
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    if request.method == "POST":
        numero = (request.json or {}).get("numero", "").strip()
    else:
        numero = request.args.get("numero", "").strip()

    if not numero:
        return jsonify({"error": "Número de orden requerido"}), 400

    # Limpiar el número (quitar espacios, prefijos opcionales que el usuario haya puesto)
    numero_limpio = numero.upper().strip()
    # Si pegó algo como "ML-2000012757", quitar el prefijo
    for prefix in ("ML-", "FALABELLA-", "WM-", "PA-", "RP-", "WC-"):
        if numero_limpio.startswith(prefix):
            numero_limpio = numero_limpio[len(prefix):]
            break

    # ── Paso 1: Buscar en BD local ─────────────────────────────────
    conn = __import__('psycopg2').connect(__import__('os').environ.get("DATABASE_URL"))
    cur = conn.cursor()
    cur.execute("""
        SELECT m.sku, m.nombre, m.canal, ABS(m.cantidad) as cantidad,
               TO_CHAR(m.fecha, 'DD/MM/YYYY HH24:MI') as fecha,
               m.bodega_codigo, m.orden_id
        FROM movimientos m
        WHERE m.orden_id = %s AND m.tipo = 'salida'
        ORDER BY m.sku
    """, (numero_limpio,))
    rows = cur.fetchall()
    cur.close(); conn.close()

    if rows:
        # ¡Encontrada en BD! Devolvemos toda la info al toque
        canal_detectado = rows[0][2] or "Desconocido"
        items = []
        for r in rows:
            items.append({
                "sku": r[0],
                "nombre": r[1],
                "canal": r[2],
                "cantidad": int(r[3] or 1),
                "fecha": r[4],
                "bodega_origen": r[5] or "CENTRAL",
                "orden_id": r[6]
            })
        return jsonify({
            "encontrada": True,
            "fuente": "bd_local",
            "marketplace": canal_detectado,
            "marketplace_codigo": _normalizar_marketplace(canal_detectado),
            "numero_orden": numero_limpio,
            "items": items,
            "total_items": len(items)
        })

    # ── Paso 2: No encontrada local — detectar por formato e intentar API ──
    marketplace_sugerido = _detectar_marketplace_por_formato(numero_limpio)

    return jsonify({
        "encontrada": False,
        "fuente": "no_encontrada",
        "marketplace_sugerido": marketplace_sugerido,
        "numero_orden": numero_limpio,
        "mensaje": (
            f"Orden no encontrada en BD local. "
            f"Posible marketplace: {marketplace_sugerido or 'desconocido'}. "
            f"Puedes registrar la devolución manualmente seleccionando el SKU."
        )
    })


def _detectar_marketplace_por_formato(numero):
    """Detecta el marketplace por el formato del número de orden."""
    n = numero.strip().upper()
    # Reglas en orden de especificidad
    if n.endswith("-A") or n.endswith("-B"):
        return "Ripley"
    if n.startswith("CLP") or n.startswith("PAR"):
        return "Paris"
    if n.startswith("PO") or n.startswith("WMT"):
        return "Walmart"
    # Por longitud numérica
    if n.isdigit():
        if len(n) >= 14:
            return "MercadoLibre"  # 16 dígitos típicos
        if 7 <= len(n) <= 9:
            return "Falabella"
        if len(n) <= 6:
            return "WooCommerce"
    return None


def _normalizar_marketplace(canal_raw):
    """Convierte el nombre del canal en código estándar (lowercase sin espacios)"""
    if not canal_raw: return ""
    c = canal_raw.lower().strip()
    if "mercadolibre" in c or "meli" in c: return "mercadolibre"
    if "paris" in c or "parís" in c: return "paris"
    if "walmart" in c or "wfs" in c: return "walmart"
    if "woo" in c: return "woocommerce"
    if "ripley" in c: return "ripley"
    if "falabella" in c: return "falabella"
    if "hites" in c: return "hites"
    return c


# ════════════════════════════════════════════════════════════════════════════
# DEVOLUCIONES — SISTEMA AVANZADO con tipificación, deadline 72h hábiles y etiqueta
# ════════════════════════════════════════════════════════════════════════════

@app.route("/devoluciones/registrar_avanzado", methods=["POST"])
def devoluciones_registrar_avanzado():
    """Registra una devolución con el flujo completo:
    - OC asociada (manual o detectada)
    - Producto seleccionado de la OC
    - Tipificación (buen_estado / reparable / dado_de_baja / reembolsado / reenviado)
    - Motivo/notas
    - Calcula deadline = ahora + 72h hábiles
    - Genera código DEV-YYYY-NNNN

    Body JSON:
    {
      "oc_origen": "2000012654022175",
      "marketplace": "MercadoLibre",
      "sku": "SDCMM001",
      "nombre": "Silla de comer Menta",
      "cantidad": 1,
      "tipificacion": "dado_de_baja",
      "motivo_texto": "Carcasa fracturada lateral derecho...",
      "responsable": "Luis"
    }
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    try:
        from inventario import (crear_devolucion, generar_codigo_dev,
                                ajustar_stock_dev, registrar_audit, get_conn)
        from feriados import calcular_deadline_habil
        from datetime import datetime
        import json

        data = request.json or {}
        oc = (data.get("oc_origen") or "").strip()
        sku = (data.get("sku") or "").strip()
        tipificacion = (data.get("tipificacion") or "").strip()
        motivo_texto = data.get("motivo_texto", "").strip()
        cantidad = int(data.get("cantidad", 1))
        nombre = data.get("nombre", "")
        marketplace = data.get("marketplace", "")
        responsable = session.get("usuario", "Sistema")

        # Validaciones
        if not oc:
            return jsonify({"ok": False, "error": "OC origen requerida"}), 400
        if not sku:
            return jsonify({"ok": False, "error": "SKU requerido"}), 400
        tipificaciones_validas = {
            "buen_estado", "reenviado", "reparable",
            "dado_de_baja", "reembolsado"
        }
        if tipificacion not in tipificaciones_validas:
            return jsonify({"ok": False, "error": f"Tipificación inválida. Válidas: {tipificaciones_validas}"}), 400
        # Para casos críticos exigir motivo
        if tipificacion in ("reparable", "dado_de_baja") and not motivo_texto:
            return jsonify({"ok": False, "error": "Motivo obligatorio para reparable/dado_de_baja"}), 400

        ahora = datetime.now()
        deadline = calcular_deadline_habil(ahora, dias_habiles=3)
        codigo = generar_codigo_dev()

        # Snapshot de la orden (por si después se necesita)
        orden_snapshot = json.dumps({
            "oc_origen": oc,
            "marketplace": marketplace,
            "fecha_registro": ahora.isoformat()
        })

        # Crear devolución (usa función existente)
        dev_data = {
            "oc_origen": oc,
            "canal": marketplace,
            "sku": sku,
            "nombre": nombre,
            "cantidad": cantidad,
            "motivo_cliente": motivo_texto[:500],
            "estado_producto": tipificacion,
            "responsable": responsable,
            "estado": _estado_segun_tipificacion(tipificacion)
        }
        dev_id = crear_devolucion(dev_data)
        if not dev_id:
            return jsonify({"ok": False, "error": "No se pudo crear devolución en BD"}), 500

        # Actualizar campos avanzados
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            UPDATE devoluciones SET
                codigo = COALESCE(codigo, %s),
                tipificacion = %s,
                motivo_texto = %s,
                usuario_revisor = %s,
                fecha_deadline = %s,
                fecha_recepcion = %s,
                origen_datos = 'manual',
                orden_data_json = %s
            WHERE id = %s
        """, (codigo, tipificacion, motivo_texto, responsable,
              deadline, ahora, orden_snapshot, dev_id))
        conn.commit()
        cur.close(); conn.close()

        # Aplicar impacto en stock según tipificación
        impacto = _aplicar_impacto_devolucion(tipificacion, sku, cantidad, dev_id)

        registrar_audit(responsable, request.remote_addr,
                        "devolucion_avanzada",
                        entidad="devoluciones", entidad_id=str(dev_id),
                        detalle=f"OC {oc} · {sku} · {tipificacion} · {impacto}")

        return jsonify({
            "ok": True,
            "id": dev_id,
            "codigo": codigo,
            "tipificacion": tipificacion,
            "deadline_iso": deadline.isoformat(),
            "deadline_legible": deadline.strftime("%d/%m/%Y %H:%M"),
            "impacto_stock": impacto,
            "estado": _estado_segun_tipificacion(tipificacion)
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


def _estado_segun_tipificacion(tipif):
    """Mapea tipificación a estado interno"""
    return {
        "buen_estado": "aceptada_reintegrada",
        "reenviado": "aceptada_reenviada",
        "reparable": "en_reparacion",
        "dado_de_baja": "dada_de_baja",
        "reembolsado": "reembolsada"
    }.get(tipif, "pendiente")


def _aplicar_impacto_devolucion(tipificacion, sku, cantidad, dev_id):
    """Aplica el impacto en stock según la tipificación."""
    try:
        from inventario import get_conn, ajustar_stock_dev
        if tipificacion == "buen_estado":
            # Reintegra al stock CENTRAL
            ajustar_stock_dev(sku, cantidad, dev_id, "reintegro_buen_estado")
            return f"Reintegrado +{cantidad} a CENTRAL"
        elif tipificacion in ("reenviado", "reembolsado", "dado_de_baja", "reparable"):
            # No reintegra
            return "Sin impacto en stock (no reintegrable)"
        return "Sin impacto"
    except Exception as e:
        return f"Error aplicando impacto: {e}"


@app.route("/devoluciones/<int:dev_id>/etiqueta_pdf")
def devoluciones_etiqueta_pdf(dev_id):
    """Genera un PDF con código de barras para identificar el producto físico.

    Solo aplica para tipificaciones: reparable, dado_de_baja, reembolsado
    """
    if not session.get("logged"):
        return "No autorizado", 401
    try:
        from inventario import get_devolucion
        from io import BytesIO
        from flask import send_file
        from reportlab.lib.pagesizes import A6, landscape
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, black, white
        import barcode
        from barcode.writer import ImageWriter

        dev = get_devolucion(dev_id=dev_id)
        if not dev:
            return "Devolución no encontrada", 404

        tipif = dev.get("tipificacion", "")
        codigo = dev.get("codigo", f"DEV-{dev_id:06d}")
        sku = dev.get("sku", "")
        nombre = dev.get("nombre", "") or sku
        oc = dev.get("oc_origen", "")
        canal = dev.get("canal", "")
        motivo = dev.get("motivo_texto", "") or dev.get("motivo_cliente", "")
        fecha_recepcion = dev.get("fecha_recepcion") or dev.get("fecha_solicitud")

        # Configurar título y color según tipificación
        titulo_color, titulo_text = {
            "dado_de_baja": (HexColor("#7f1d1d"), "DAR DE BAJA"),
            "reparable": (HexColor("#92400e"), "EN REPARACION"),
            "reembolsado": (HexColor("#1e3a8a"), "REEMBOLSADO"),
            "buen_estado": (HexColor("#065f46"), "REINTEGRADO"),
            "reenviado": (HexColor("#854F0B"), "REENVIADO")
        }.get(tipif, (black, "DEVOLUCION"))

        # Generar código de barras como imagen en memoria
        EAN = barcode.get_barcode_class('code128')
        barcode_io = BytesIO()
        EAN(codigo, writer=ImageWriter()).write(barcode_io, options={
            "module_width": 0.4,
            "module_height": 12.0,
            "font_size": 10,
            "text_distance": 4.0,
            "quiet_zone": 2.0
        })
        barcode_io.seek(0)
        # Guardar a temporal para reportlab
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            tmp.write(barcode_io.read())
            barcode_path = tmp.name

        # Crear PDF
        pdf_buf = BytesIO()
        # Etiqueta tamaño A6 horizontal (10x15cm aprox)
        c = canvas.Canvas(pdf_buf, pagesize=landscape(A6))
        ancho, alto = landscape(A6)

        # Header
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(8*mm, alto - 8*mm, "DEVOLUCIÓN")
        c.setFillColor(titulo_color)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(8*mm, alto - 14*mm, titulo_text)

        # Línea divisoria
        c.setStrokeColor(black)
        c.setLineWidth(1)
        c.line(8*mm, alto - 17*mm, ancho - 8*mm, alto - 17*mm)

        # Datos lado izquierdo
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 8)
        y = alto - 22*mm
        info_lines = [
            ("SKU:", sku),
            ("Producto:", nombre[:35]),
            ("Orden:", f"{oc} ({canal})"),
            ("Recepción:", fecha_recepcion.strftime("%d/%m/%Y %H:%M") if fecha_recepcion and hasattr(fecha_recepcion, 'strftime') else "—")
        ]
        for label, value in info_lines:
            c.setFont("Helvetica-Bold", 7)
            c.drawString(8*mm, y, label)
            c.setFont("Helvetica", 7)
            c.drawString(28*mm, y, str(value))
            y -= 4*mm

        # Motivo (en caja con fondo)
        if motivo:
            y_motivo_top = y - 2*mm
            c.setFillColor(HexColor("#fef2f2") if tipif == "dado_de_baja" else HexColor("#fef3c7"))
            c.rect(8*mm, y_motivo_top - 18*mm, ancho - 16*mm, 18*mm, fill=1, stroke=0)
            c.setFillColor(titulo_color)
            c.setFont("Helvetica-Bold", 7)
            c.drawString(10*mm, y_motivo_top - 4*mm, "MOTIVO:")
            c.setFillColor(black)
            c.setFont("Helvetica", 7)
            # Wrap del texto
            words = motivo.split()
            linea = ""
            y_text = y_motivo_top - 8*mm
            for w in words:
                test = (linea + " " + w).strip()
                if len(test) > 60:
                    c.drawString(10*mm, y_text, linea)
                    y_text -= 3.5*mm
                    linea = w
                    if y_text < y_motivo_top - 16*mm: break
                else:
                    linea = test
            if linea:
                c.drawString(10*mm, y_text, linea)

        # Código de barras en la parte inferior
        from reportlab.lib.utils import ImageReader
        try:
            barcode_img = ImageReader(barcode_path)
            barcode_w = (ancho - 16*mm)
            c.drawImage(barcode_img, 8*mm, 5*mm, width=barcode_w, height=18*mm,
                        preserveAspectRatio=True, anchor='c')
        except Exception as e:
            c.setFont("Helvetica", 6)
            c.drawString(8*mm, 8*mm, f"Error generando barcode: {e}")

        # Footer
        c.setFont("Helvetica", 5)
        c.setFillColor(HexColor("#666666"))
        c.drawString(8*mm, 2*mm, "Lusync ERP · Babymine")

        c.showPage()
        c.save()
        pdf_buf.seek(0)

        # Limpiar tempfile
        try:
            import os
            os.unlink(barcode_path)
        except: pass

        # Marcar etiqueta como generada
        try:
            from inventario import get_conn
            conn = get_conn(); cur = conn.cursor()
            cur.execute("UPDATE devoluciones SET etiqueta_generada=TRUE WHERE id=%s", (dev_id,))
            conn.commit()
            cur.close(); conn.close()
        except: pass

        return send_file(pdf_buf, as_attachment=True,
                         download_name=f"etiqueta_{codigo}.pdf",
                         mimetype="application/pdf")
    except Exception as e:
        import traceback
        return f"Error generando PDF: {e}\n\n{traceback.format_exc()}", 500


@app.route("/devoluciones/pendientes_revision")
def devoluciones_pendientes_revision():
    """Lista devoluciones que están pendientes de revisión, con info de deadline."""
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        from feriados import descripcion_tiempo_restante, color_urgencia
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT id, codigo, oc_origen, canal, sku, nombre, cantidad,
                   tipificacion, motivo_texto, fecha_recepcion, fecha_deadline,
                   etiqueta_generada, estado
            FROM devoluciones
            WHERE estado IN ('pendiente', 'en_reparacion')
              AND fecha_deadline IS NOT NULL
            ORDER BY fecha_deadline ASC
            LIMIT 100
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()

        items = []
        for r in rows:
            deadline = r[10]
            items.append({
                "id": r[0], "codigo": r[1], "oc_origen": r[2], "canal": r[3],
                "sku": r[4], "nombre": r[5], "cantidad": r[6],
                "tipificacion": r[7], "motivo_texto": r[8],
                "fecha_recepcion": r[9].isoformat() if r[9] else None,
                "fecha_deadline": deadline.isoformat() if deadline else None,
                "etiqueta_generada": bool(r[11]),
                "estado": r[12],
                "tiempo_restante": descripcion_tiempo_restante(deadline) if deadline else "Sin deadline",
                "urgencia": color_urgencia(deadline) if deadline else "normal"
            })

        return jsonify({
            "total": len(items),
            "items": items
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/devoluciones/<int:dev_id>/actualizar", methods=["POST"])
def devoluciones_actualizar(dev_id):
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = request.json
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "actualizar_devolucion", entidad="devoluciones", entidad_id=str(dev_id),
                    detalle=f"Estado: {data.get('estado','?')} · Resolución: {data.get('resolucion','?')}")
    dev = get_devolucion(dev_id=dev_id)
    if not dev:
        return {"error": "no encontrada"}, 404
    actualizar_devolucion(dev_id, data)
    # Si se reingresa al stock, registrar movimiento
    if data.get("estado") == "reingresada" and dev.get("sku") and not dev.get("impacto_stock_reingresado"):
        productos = cargar_productos()
        for p in productos:
            if p["sku"] == dev["sku"]:
                p["stock"] += int(dev.get("cantidad", 1))
                guardar_producto(p)
                registrar_movimiento("entrada", p["sku"], p["nombre"],
                                     int(dev.get("cantidad", 1)), "Devolución reingresada",
                                     usuario=session.get("usuario", "Sistema"),
                                     canal="Manual", orden_id=dev.get("oc_origen"))
                sincronizar_stock_marketplaces(p["sku"], p["stock"], contexto="auto_sync")
                break
    return {"ok": True}

@app.route("/devoluciones/<int:dev_id>/eliminar", methods=["POST"])
def devoluciones_eliminar(dev_id):
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = request.json
    clave = data.get("clave", "")
    clave_admin = __import__('os').environ.get("PASSWORD", "")
    if clave != clave_admin:
        registrar_audit(session.get("usuario","?"), request.remote_addr,
                        "intento_eliminar_devolucion", entidad="devoluciones", entidad_id=str(dev_id),
                        resultado="fallido", detalle="Clave admin incorrecta")
        return {"error": "Clave incorrecta"}, 403
    conn = __import__('psycopg2').connect(__import__('os').environ.get("DATABASE_URL"))
    cur = conn.cursor()
    cur.execute("SELECT codigo, oc_origen, nombre FROM devoluciones WHERE id = %s", (dev_id,))
    row = cur.fetchone()
    detalle_dev = str(row) if row else str(dev_id)
    cur.execute("DELETE FROM devoluciones WHERE id = %s", (dev_id,))
    conn.commit()
    cur.close(); conn.close()
    registrar_audit(session.get("usuario","admin"), request.remote_addr,
                    "eliminar_devolucion", entidad="devoluciones", entidad_id=str(dev_id),
                    detalle=f"Devolución eliminada: {detalle_dev}", dato_antes=detalle_dev)
    return {"ok": True}

@app.route("/devoluciones/<int:dev_id>/generar_codigo", methods=["POST"])
def devoluciones_generar_codigo(dev_id):
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    dev = get_devolucion(dev_id=dev_id)
    if not dev:
        return {"error": "no encontrada"}, 404
    if dev.get("codigo"):
        return {"ok": True, "codigo": dev["codigo"]}
    codigo = generar_codigo_dev()
    asignar_codigo_dev(dev_id, codigo)
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "generar_codigo_dev", entidad="devoluciones", entidad_id=str(dev_id),
                    detalle=f"Código generado: {codigo}")
    return {"ok": True, "codigo": codigo}

# ── PARIS ──

@app.route("/paris/test")
def paris_test():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    result = verificar_conexion_paris()
    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "paris_test", detalle=f"Test conexión Paris: {result}")
    return result

@app.route("/paris/ordenes")
def paris_ordenes():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    dias = int(request.args.get("dias", 30))
    estado = request.args.get("estado") or None
    ordenes = obtener_ordenes_paris_todas(dias=dias, estado=estado)
    return {"ordenes": ordenes, "total": len(ordenes)}

@app.route("/paris/stock")
def paris_stock():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    data = obtener_stock_paris()
    return data or {"error": "sin datos"}

# /paris/sync_ordenes movido a paris.py (Blueprint)


# ── AUDIT LOG ──

@app.route("/audit")
def audit_view():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    # Asegurar tabla existe (por si el deploy no la creó)
    init_audit()
    # Registrar que el admin consultó el log
    registrar_audit(
        session.get("usuario", "admin"),
        request.remote_addr,
        "consultar_audit",
        detalle="Vista del Audit Log"
    )
    limite = int(request.args.get("limite", 200))
    filtro_accion    = request.args.get("accion") or None
    filtro_usuario   = request.args.get("usuario") or None
    filtro_resultado = request.args.get("resultado") or None
    logs = listar_audit(limite, filtro_accion, filtro_usuario, filtro_resultado)
    return {"logs": logs, "total": len(logs)}

@app.route("/audit/test", methods=["POST"])
def audit_test():
    """Endpoint para verificar que el audit funciona — solo admin"""
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    init_audit()
    registrar_audit(
        session.get("usuario", "admin"),
        request.remote_addr,
        "test_audit",
        detalle="Test manual del sistema de audit"
    )
    return {"ok": True, "mensaje": "Registro de prueba creado"}

# ── LOGIN / PANEL ──

@app.route("/")
def home():
    if session.get("logged"):
        return redirect("/panel")
    return render_template("login.html")

@app.route("/login_check", methods=["POST"])
def login_check():
    data = request.json
    if data.get("user") == USUARIO and data.get("password") == PASSWORD:
        session["logged"] = True
        session["usuario"] = data.get("user")
        registrar_audit(data.get("user"), request.remote_addr, "login", detalle="Inicio de sesión exitoso")
        return {"ok": True}
    registrar_audit(data.get("user","?"), request.remote_addr, "login", resultado="fallido", detalle="Clave incorrecta")
    return {"ok": False}

@app.route("/logout")
def logout():
    registrar_audit(session.get("usuario","?"), request.remote_addr, "logout", detalle="Cierre de sesión")
    session.clear()
    return redirect("/")

@app.route("/panel")
def panel():
    if not session.get("logged"):
        return redirect("/")
    return render_template("panel.html")

@app.route("/debug/estado_bd")
def debug_estado_bd():
    if not session.get("logged"):
        return {"error": "no autorizado"}, 401
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
          (SELECT COUNT(*) FROM ordenes_procesadas) as total_op,
          (SELECT COUNT(*) FROM ordenes_procesadas WHERE order_id_texto IS NOT NULL) as con_texto,
          (SELECT COUNT(DISTINCT order_id_texto) FROM ordenes_procesadas
           WHERE order_id_texto IS NOT NULL) as unicos,
          (SELECT COUNT(*) FROM movimientos
           WHERE canal='Walmart') as mov_walmart_total,
          (SELECT COUNT(*) FROM movimientos
           WHERE canal='Walmart' AND orden_id IN (
             SELECT orden_id FROM movimientos
             WHERE canal='Walmart' AND orden_id IS NOT NULL AND orden_id != ''
             GROUP BY orden_id HAVING COUNT(*) > 1
           )) as mov_con_orden_duplicada
    """)
    r = cur.fetchone()

    cur.execute("""
        SELECT orden_id, sku, COUNT(*) as veces,
               MIN(TO_CHAR(fecha, 'DD/MM HH24:MI')) as primera,
               MAX(TO_CHAR(fecha, 'DD/MM HH24:MI')) as ultima
        FROM movimientos
        WHERE canal='Walmart' AND orden_id IS NOT NULL AND orden_id != ''
        GROUP BY orden_id, sku
        HAVING COUNT(*) > 1
        ORDER BY veces DESC
        LIMIT 20
    """)
    dupes = [{"orden_id": x[0], "sku": x[1], "veces": x[2],
              "primera": x[3], "ultima": x[4]} for x in cur.fetchall()]

    cur.execute("""
        SELECT orden_id, order_id_texto,
               TO_CHAR(fecha, 'DD/MM HH24:MI') as fecha
        FROM ordenes_procesadas
        ORDER BY fecha DESC LIMIT 10
    """)
    ultimas_op = [{"orden_id": x[0], "texto": x[1], "fecha": x[2]}
                  for x in cur.fetchall()]

    cur.close(); conn.close()
    return {
        "ordenes_procesadas_total": r[0],
        "con_order_id_texto": r[1],
        "unicos": r[2],
        "movimientos_walmart_total": r[3],
        "movimientos_con_orden_duplicada": r[4],
        "duplicados_detalle": dupes,
        "ultimas_ordenes_procesadas": ultimas_op
    }


@app.route("/debug/paris_skus")
def debug_paris_skus():
    """Trae los SKUs reales de París para tu seller."""
    if not session.get("logged"): return {"error": "no autorizado"}, 401
    import requests as req
    from paris import paris_headers, PARIS_BASE_URL, obtener_stock_paris, obtener_productos_paris

    # Opción 1: stock real
    stock_data = obtener_stock_paris(limite=100, offset=0)

    # Opción 2: productos publicados
    prod_data = obtener_productos_paris(limite=25, offset=0)

    # Opción 3: llamada directa a v2/stock para ver estructura
    try:
        res = req.get(f"{PARIS_BASE_URL}/v2/stock",
                      headers=paris_headers(),
                      params={"limit": 50, "offset": 0},
                      timeout=15)
        stock_raw = {"status": res.status_code, "body": res.json() if res.status_code == 200 else res.text[:500]}
    except Exception as e:
        stock_raw = {"error": str(e)}

    return {
        "stock_v2": stock_raw,
        "productos_search": prod_data,
        "stock_data": stock_data
    }


# ── MAPEO SKUs ──────────────────────────────────────────────────────────────

@app.route("/sku_mapeo")
def ruta_sku_mapeo():
    if not session.get("logged"): return redirect("/")
    return jsonify(listar_sku_mapeo())

@app.route("/sku_mapeo/historial")
def ruta_sku_mapeo_historial():
    if not session.get("logged"): return redirect("/")
    return jsonify(listar_historial_mapeo())

@app.route("/sku_mapeo/guardar", methods=["POST"])
def ruta_sku_mapeo_guardar():
    if not session.get("logged"): return jsonify({"ok": False}), 401
    data = request.json or {}
    try:
        guardar_sku_mapeo_fila(
            data.get("sku_lusync", "").strip(),
            {
                "web":         data.get("sku_web", ""),
                "walmart":     data.get("sku_walmart", ""),
                "paris":       data.get("sku_paris", ""),
                "falabella":   data.get("sku_falabella", ""),
                "ripley":      data.get("sku_ripley", ""),
                "mercadolibre":data.get("sku_mercadolibre", ""),
                "hites":       data.get("sku_hites", "")
            }
        )
        return jsonify({"ok": True})
    except Exception as e:
        import traceback
        print(f"[guardar_sku_mapeo] ERROR: {e}")
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/sku_mapeo/plataforma_web", methods=["GET", "POST"])
def ruta_plataforma_web():
    if not session.get("logged"): return jsonify({}), 401
    if request.method == "POST":
        data = request.json or {}
        set_plataforma_web(data.get("plataforma", "woocommerce"))
        return jsonify({"ok": True})
    return jsonify({"plataforma": get_plataforma_web()})

@app.route("/sku_mapeo/exportar_excel")
def ruta_exportar_excel():
    if not session.get("logged"): return redirect("/")
    try:
        import io, openpyxl
        filas = listar_sku_mapeo()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mapeo SKUs"
        ws.append(["SKU Lusync","Producto","SKU Web","SKU Walmart","SKU Paris",
                   "SKU Falabella","SKU Ripley","SKU MercadoLibre","SKU Hites"])
        for f in filas:
            ws.append([f.get("sku_lusync",""), f.get("nombre",""),
                       f.get("sku_web",""), f.get("sku_walmart",""),
                       f.get("sku_paris",""), f.get("sku_falabella",""),
                       f.get("sku_ripley",""), f.get("sku_mercadolibre",""),
                       f.get("sku_hites","")])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, download_name="mapeo_skus.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/sku_mapeo/importar_excel", methods=["POST"])
def ruta_importar_excel():
    """Importa mapeo SKU → marketplaces desde Excel.

    Estructura esperada (columnas):
      A: SKU Lusync
      B: Producto (nombre, ignorado)
      C: SKU Web
      D: SKU Walmart
      E: SKU Paris
      F: SKU Falabella
      G: SKU Ripley
      H: SKU MercadoLibre
      I: SKU Hites

    Pobla DOS tablas:
      1. sku_mapeo (legacy 1:1) — para compatibilidad
      2. sku_mapeo_canal (nuevo multi-publicación) — para que UI muestre badges
    """
    if not session.get("logged"): return jsonify({"ok": False, "error": "no autorizado"}), 401
    try:
        import io, openpyxl
        from inventario import agregar_publicacion

        archivo = request.files.get("archivo")
        if not archivo:
            return jsonify({"ok": False, "error": "No se recibio archivo"})
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"ok": False, "error": "Archivo vacio o sin datos"})

        importados = 0
        publicaciones_agregadas = 0
        publicaciones_fallidas = 0
        errores = []
        log = []

        # Mapeo de columna → canal (índice de columna en el Excel)
        # Col A=0 sku_lusync, B=1 producto, C=2 web, D=3 walmart, E=4 paris,
        # F=5 falabella, G=6 ripley, H=7 mercadolibre, I=8 hites
        canales_columnas = [
            (2, "web"),
            (3, "walmart"),
            (4, "paris"),
            (5, "falabella"),
            (6, "ripley"),
            (7, "mercadolibre"),
            (8, "hites"),
        ]

        for i, row in enumerate(rows[1:], start=2):
            try:
                sku_lusync = str(row[0]).strip() if row[0] else ""
                if not sku_lusync or sku_lusync == "None":
                    continue

                # 1. Tabla legacy sku_mapeo (1:1)
                skus = {
                    "web":          str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    "walmart":      str(row[3]).strip() if len(row) > 3 and row[3] else "",
                    "paris":        str(row[4]).strip() if len(row) > 4 and row[4] else "",
                    "falabella":    str(row[5]).strip() if len(row) > 5 and row[5] else "",
                    "ripley":       str(row[6]).strip() if len(row) > 6 and row[6] else "",
                    "mercadolibre": str(row[7]).strip() if len(row) > 7 and row[7] else "",
                    "hites":        str(row[8]).strip() if len(row) > 8 and row[8] else "",
                }
                guardar_sku_mapeo_fila(sku_lusync, skus)
                importados += 1

                # 2. Tabla nueva sku_mapeo_canal (multi-publicación)
                # Para cada canal con SKU, agregamos UNA publicación (idempotente)
                for col_idx, canal in canales_columnas:
                    if col_idx < len(row) and row[col_idx]:
                        sku_canal_val = str(row[col_idx]).strip()
                        if sku_canal_val and sku_canal_val.lower() not in ("none", "nan", "null", ""):
                            # Para MELI: si el SKU empieza con MLC, es item_id; sino es seller_sku
                            item_id_canal = None
                            sku_para_guardar = sku_canal_val
                            if canal == "mercadolibre" and sku_canal_val.upper().startswith("MLC"):
                                item_id_canal = sku_canal_val
                                sku_para_guardar = sku_canal_val

                            # ── BLINDAJE: verificar si ya existe ANTES de insertar ──
                            # Esto evita los duplicados que se crean cuando se re-importa
                            # el Excel sobre un dataset que ya fue procesado por auto_mapeo_v2
                            try:
                                from inventario import get_conn
                                conn_check = get_conn(); cur_check = conn_check.cursor()
                                cur_check.execute("""
                                    SELECT COUNT(*) FROM sku_mapeo_canal
                                    WHERE canal = %s
                                      AND sku_lusync = %s
                                      AND sku_canal = %s
                                      AND activo = TRUE
                                """, (canal, sku_lusync, sku_para_guardar))
                                ya_existe = cur_check.fetchone()[0] > 0
                                cur_check.close(); conn_check.close()
                            except Exception as e_check:
                                ya_existe = False
                                log.append(f"Fila {i} {sku_lusync}/{canal}: error chequeando existencia: {e_check}")

                            if ya_existe:
                                log.append(f"Fila {i} {sku_lusync} → {canal}:{sku_canal_val}: ya existe, skip (blindaje anti-duplicado)")
                                continue

                            try:
                                mapeo_id = agregar_publicacion(
                                    sku_lusync=sku_lusync,
                                    canal=canal,
                                    sku_canal=sku_para_guardar,
                                    item_id_canal=item_id_canal,
                                    es_catalogo=False,
                                    notas="import_excel"
                                )
                                if mapeo_id:
                                    publicaciones_agregadas += 1
                                else:
                                    publicaciones_fallidas += 1
                                    log.append(f"Fila {i} {sku_lusync} → {canal}:{sku_canal_val}: agregar_publicacion devolvió None")
                            except Exception as e_pub:
                                publicaciones_fallidas += 1
                                log.append(f"Fila {i} {sku_lusync} → {canal}:{sku_canal_val}: {e_pub}")

            except Exception as e:
                errores.append(f"Fila {i}: {str(e)}")

        registrar_importacion_mapeo(
            session.get("usuario", "Sistema"),
            archivo.filename,
            importados,
            [{"fila": i, "error": e} for i, e in enumerate(errores)]
        )
        registrar_audit(
            session.get("usuario", "Sistema"), request.remote_addr,
            "importar_mapeo_excel",
            detalle=f"importados={importados} pubs={publicaciones_agregadas} pubs_fallidas={publicaciones_fallidas}"
        )

        return jsonify({
            "ok": True,
            "importados": importados,
            "publicaciones_agregadas": publicaciones_agregadas,
            "publicaciones_fallidas": publicaciones_fallidas,
            "errores": errores,
            "log_publicaciones": log[:50]  # primeros 50 errores de pub
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/debug/paris_stock")
def debug_paris_stock():
    """Diagnóstico: prueba envío de stock a Paris para todos los productos mapeados."""
    if not session.get("logged"): return redirect("/")
    try:
        from inventario import listar_sku_mapeo, cargar_productos
        from paris import actualizar_stock_paris, verificar_conexion_paris

        conexion = verificar_conexion_paris()
        productos = {p["sku"]: p for p in cargar_productos()}
        mapeo = listar_sku_mapeo()

        resultados = []
        for fila in mapeo:
            sku_lusync = fila.get("sku_lusync", "")
            sku_paris  = fila.get("sku_paris", "")
            if not sku_paris:
                continue
            prod = productos.get(sku_lusync)
            stock_actual = prod.get("stock", 0) if prod else 0

            ok = actualizar_stock_paris(sku_lusync, stock_actual)
            resultados.append({
                "sku_lusync": sku_lusync,
                "sku_paris":  sku_paris,
                "nombre":     fila.get("nombre", ""),
                "stock":      stock_actual,
                "ok":         ok
            })

        return jsonify({
            "conexion": conexion,
            "total_mapeados": len(resultados),
            "exitosos": sum(1 for r in resultados if r["ok"]),
            "fallidos": sum(1 for r in resultados if not r["ok"]),
            "detalle":  resultados
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/debug/paris_stock_raw")
def debug_paris_stock_raw():
    """Envía stock a Paris uno por uno y muestra status + body crudo de cada respuesta."""
    if not session.get("logged"): return redirect("/")
    try:
        import requests as req
        from inventario import listar_sku_mapeo, cargar_productos
        from paris import paris_headers, PARIS_BASE_URL

        productos = {p["sku"]: p for p in cargar_productos()}
        mapeo = listar_sku_mapeo()

        resultados = []
        for fila in mapeo:
            sku_lusync = fila.get("sku_lusync", "")
            sku_paris  = (fila.get("sku_paris", "") or "").strip()
            if not sku_paris:
                continue
            prod = productos.get(sku_lusync)
            stock_actual = int(prod.get("stock", 0)) if prod else 0

            payload = {"skus": [{"skuSeller": sku_paris, "quantity": stock_actual}]}
            try:
                res = req.post(
                    f"{PARIS_BASE_URL}/v1/stock/sku-seller",
                    headers=paris_headers(),
                    json=payload,
                    timeout=15
                )
                resultados.append({
                    "sku_lusync": sku_lusync,
                    "sku_paris": sku_paris,
                    "stock": stock_actual,
                    "status": res.status_code,
                    "request_body": payload,
                    "response_body": res.text[:500]
                })
            except Exception as e:
                resultados.append({
                    "sku_lusync": sku_lusync,
                    "sku_paris": sku_paris,
                    "stock": stock_actual,
                    "error": str(e),
                    "request_body": payload
                })

        return jsonify({"resultados": resultados})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/paris_stock_v2")
def debug_paris_stock_v2():
    """Prueba 3 variantes de payload para identificar el formato correcto."""
    if not session.get("logged"): return redirect("/")
    try:
        import requests as req
        from paris import paris_headers, PARIS_BASE_URL

        # Probar con un SKU conocido del catálogo
        sku_test = "CDPPASN001"
        stock = 99

        variantes = [
            {"nombre": "skuSeller (camelCase, actual)",
             "payload": {"skus": [{"skuSeller": sku_test, "quantity": stock}]}},
            {"nombre": "sku_seller (snake_case)",
             "payload": {"skus": [{"sku_seller": sku_test, "quantity": stock}]}},
            {"nombre": "sku (campo simple)",
             "payload": {"skus": [{"sku": sku_test, "quantity": stock}]}},
            {"nombre": "objeto plano sin skus[]",
             "payload": {"skuSeller": sku_test, "quantity": stock}},
            {"nombre": "snake_case plano",
             "payload": {"sku_seller": sku_test, "quantity": stock}},
        ]

        resultados = []
        for v in variantes:
            try:
                res = req.post(
                    f"{PARIS_BASE_URL}/v1/stock/sku-seller",
                    headers=paris_headers(),
                    json=v["payload"],
                    timeout=15
                )
                resultados.append({
                    "variante": v["nombre"],
                    "payload_enviado": v["payload"],
                    "status": res.status_code,
                    "response": res.text[:400]
                })
            except Exception as e:
                resultados.append({
                    "variante": v["nombre"],
                    "payload_enviado": v["payload"],
                    "error": str(e)
                })

        return jsonify({"sku_probado": sku_test, "stock": stock, "resultados": resultados})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/paris_stock_consultar")
def debug_paris_stock_consultar():
    """Consulta el stock actual desde la API de Paris para los SKUs mapeados."""
    if not session.get("logged"): return redirect("/")
    try:
        import requests as req
        from inventario import listar_sku_mapeo
        from paris import paris_headers, PARIS_BASE_URL

        mapeo = listar_sku_mapeo()
        resultados = []

        for fila in mapeo:
            sku_paris = (fila.get("sku_paris", "") or "").strip()
            if not sku_paris:
                continue

            # Endpoint v1: stock por sku_seller
            try:
                res = req.get(
                    f"{PARIS_BASE_URL}/v1/stock/sku-seller/{sku_paris}",
                    headers=paris_headers(),
                    timeout=15
                )
                resultados.append({
                    "sku_paris": sku_paris,
                    "sku_lusync": fila.get("sku_lusync"),
                    "endpoint": "v1/stock/sku-seller/{sku}",
                    "status": res.status_code,
                    "response": res.text[:600]
                })
            except Exception as e:
                resultados.append({
                    "sku_paris": sku_paris,
                    "error": str(e)
                })

        return jsonify({"resultados": resultados})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/paris_stock_listar")
def debug_paris_stock_listar():
    """Lista todo el stock actual del seller usando el endpoint v2/stock."""
    if not session.get("logged"): return redirect("/")
    try:
        import requests as req
        from paris import paris_headers, PARIS_BASE_URL

        # Probar varios endpoints para ver cuál funciona
        endpoints = [
            "/v2/stock",
            "/v1/stock",
            "/v2/stock/sku-seller",
            "/v1/stock/sku-seller",
        ]
        resultados = []
        for ep in endpoints:
            try:
                res = req.get(
                    f"{PARIS_BASE_URL}{ep}",
                    headers=paris_headers(),
                    params={"limit": 200, "offset": 0},
                    timeout=20
                )
                resultados.append({
                    "endpoint": ep,
                    "status": res.status_code,
                    "response": res.text[:1500]
                })
            except Exception as e:
                resultados.append({"endpoint": ep, "error": str(e)})
        return jsonify({"resultados": resultados})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/paris_stock_warehouse")
def debug_paris_stock_warehouse():
    """Probar enviando stock con campo warehouse explícito en distintos formatos."""
    if not session.get("logged"): return redirect("/")
    try:
        import requests as req
        from paris import paris_headers, PARIS_BASE_URL

        sku_test = "RHR2022-1"  # tiene stock real 25, voy a enviar 99 y ver si actualiza
        stock = 99

        variantes = [
            {"nombre": "sin warehouse",
             "payload": {"skus": [{"sku_seller": sku_test, "quantity": stock}]}},
            {"nombre": "warehouse=Dropshipping (nombre real)",
             "payload": {"skus": [{"sku_seller": sku_test, "quantity": stock, "warehouse": "Dropshipping"}]}},
            {"nombre": "warehouse=dropship (lo que retornó la API)",
             "payload": {"skus": [{"sku_seller": sku_test, "quantity": stock, "warehouse": "dropship"}]}},
            {"nombre": "warehouseName=Dropshipping",
             "payload": {"skus": [{"sku_seller": sku_test, "quantity": stock, "warehouseName": "Dropshipping"}]}},
        ]

        resultados = []
        for v in variantes:
            try:
                res = req.post(
                    f"{PARIS_BASE_URL}/v1/stock/sku-seller",
                    headers=paris_headers(),
                    json=v["payload"],
                    timeout=15
                )
                resultados.append({
                    "variante": v["nombre"],
                    "request": v["payload"],
                    "status": res.status_code,
                    "response": res.text[:600]
                })
            except Exception as e:
                resultados.append({"variante": v["nombre"], "error": str(e)})

        # Esperar 2 segundos y consultar el listado para ver si se reflejó
        import time
        time.sleep(2)
        listado = req.get(f"{PARIS_BASE_URL}/v2/stock", headers=paris_headers(),
                          params={"limit": 200}, timeout=20)

        # Buscar el SKU de prueba en el listado
        sku_actual = None
        if listado.status_code == 200:
            data = listado.json()
            for s in data.get("skus", []):
                if s.get("sku_seller") == sku_test:
                    sku_actual = {
                        "sku_seller": s.get("sku_seller"),
                        "quantity": s.get("quantity"),
                        "availableStock": s.get("availableStock"),
                        "warehouseName": s.get("warehouseName"),
                        "updatedAt": s.get("updatedAt")
                    }
                    break

        return jsonify({
            "sku_probado": sku_test,
            "stock_enviado": stock,
            "tests": resultados,
            "estado_actual_en_paris": sku_actual or "no encontrado en listado"
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# /paris/sync_estado movido a paris.py (Blueprint)


@app.route("/paris/forzar_sync_sku", methods=["POST"])
def ruta_paris_forzar_sync():
    """Re-envía el stock actual de un SKU específico a Paris."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from inventario import cargar_productos
        from paris import actualizar_stock_paris
        data = request.json or {}
        sku = data.get("sku_lusync", "").strip()
        if not sku:
            return jsonify({"ok": False, "error": "SKU no proporcionado"})
        prod = next((p for p in cargar_productos() if p["sku"] == sku), None)
        if not prod:
            return jsonify({"ok": False, "error": f"SKU {sku} no existe en inventario"})
        ok = actualizar_stock_paris(sku, prod["stock"])
        return jsonify({"ok": ok, "sku": sku, "stock_enviado": prod["stock"]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# /paris/forzar_sync_todos movido a paris.py (Blueprint)


# ════════════════════════════════════════════════════════════════════════════
# SYNC_ESTADO PARA MELI Y WALMART (NUEVO - patrón unificado)
# ════════════════════════════════════════════════════════════════════════════

@app.route("/mercadolibre/sync_estado")
def ruta_meli_sync_estado():
    """Compara stock en Lusync (CENTRAL) vs stock real en MercadoLibre.
    Usa el mismo formato que /paris/sync_estado para reutilizar UI."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from mercadolibre import obtener_publicaciones_meli, verificar_conexion_meli
        from inventario import listar_sku_mapeo, get_stock_bodega

        conexion = verificar_conexion_meli()

        # Obtener publicaciones del seller (las que están publicadas en MELI)
        publicaciones = obtener_publicaciones_meli(limite=50, offset=0)
        if publicaciones is None:
            return jsonify({"error": "No se pudo conectar con MercadoLibre", "conexion": conexion}), 500

        # Indexar por SKU MELI (item_id) y por seller_custom_field
        meli_dict_por_item = {}
        meli_dict_por_seller_sku = {}
        for it in publicaciones.get("items", []):
            item_id = it.get("item_id", "")
            seller_sku = (it.get("sku_seller", "") or "").strip()
            data_meli = {
                "stock": it.get("stock", 0),
                "title": it.get("title", ""),
                "price": it.get("price", 0),
                "status": it.get("status", "")
            }
            if item_id: meli_dict_por_item[item_id] = data_meli
            if seller_sku: meli_dict_por_seller_sku[seller_sku] = data_meli

        # Comparar stock CENTRAL vs MELI
        mapeo = listar_sku_mapeo()
        resultados = []
        for fila in mapeo:
            sku_meli = (fila.get("sku_mercadolibre", "") or "").strip()
            if not sku_meli:
                continue
            sku_lusync = fila.get("sku_lusync", "")
            stock_central = get_stock_bodega(sku_lusync, "CENTRAL")

            # Buscar primero por item_id, después por seller_custom_field
            meli_data = meli_dict_por_item.get(sku_meli) or meli_dict_por_seller_sku.get(sku_meli, {})
            stock_meli = meli_data.get("stock", None) if meli_data else None

            if stock_meli is None:
                estado = "no_encontrado"
            elif stock_meli == stock_central:
                estado = "sincronizado"
            else:
                estado = "desincronizado"

            resultados.append({
                "sku_lusync": sku_lusync,
                "sku_paris": sku_meli,  # nombre genérico para template
                "sku_meli": sku_meli,
                "nombre": fila.get("nombre", "") or meli_data.get("title", ""),
                "stock_lusync": stock_central,
                "stock_paris": stock_meli,  # nombre genérico
                "stock_meli": stock_meli,
                "diferencia": (stock_meli - stock_central) if stock_meli is not None else None,
                "ultima_actualizacion_paris": "",
                "status_meli": meli_data.get("status", ""),
                "estado": estado
            })

        return jsonify({
            "conexion": conexion,
            "total": len(resultados),
            "sincronizados": sum(1 for r in resultados if r["estado"] == "sincronizado"),
            "desincronizados": sum(1 for r in resultados if r["estado"] == "desincronizado"),
            "no_encontrados": sum(1 for r in resultados if r["estado"] == "no_encontrado"),
            "resultados": resultados
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/mercadolibre/forzar_sync_sku", methods=["POST"])
def ruta_meli_forzar_sync_sku():
    """Re-envía el stock de un SKU específico a MercadoLibre."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from mercadolibre import actualizar_stock_meli
        from inventario import listar_sku_mapeo, get_stock_bodega
        data = request.json or {}
        sku_lusync = data.get("sku_lusync", "").strip()
        if not sku_lusync:
            return jsonify({"ok": False, "error": "sku_lusync requerido"}), 400

        sku_meli = None
        # Obtener publicaciones MELI para el SKU (multi-publicación)
        from inventario import obtener_publicaciones_canal
        publicaciones = obtener_publicaciones_canal(sku_lusync, "mercadolibre")

        # Fallback legacy: si no hay mapeos en sku_mapeo_canal, usar tabla vieja
        if not publicaciones:
            sku_meli = ""
            for fila in listar_sku_mapeo():
                if fila.get("sku_lusync") == sku_lusync:
                    sku_meli = (fila.get("sku_mercadolibre", "") or "").strip()
                    break
            if not sku_meli:
                return jsonify({"ok": False, "error": f"SKU {sku_lusync} no tiene mapeo MELI"}), 400

        stock = get_stock_bodega(sku_lusync, "CENTRAL")
        # Usar wrapper multi-publicación
        from mercadolibre import actualizar_stock_meli
        resultado = actualizar_stock_meli(sku_lusync, stock)
        return jsonify({
            "ok": resultado.get("ok", False),
            "sku": sku_lusync,
            "stock_enviado": stock,
            "publicaciones_actualizadas": resultado.get("exitosas", 0),
            "publicaciones_fallidas": resultado.get("fallidas", 0),
            "total_publicaciones": resultado.get("total_publicaciones", 0),
            "log": resultado.get("log", [])
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/walmart/sync_estado")
def ruta_walmart_sync_estado():
    """Compara stock en Lusync (CENTRAL) vs stock real en Walmart.
    Devuelve el mismo formato que /paris/sync_estado para reutilizar UI."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from walmart import walmart_headers, WALMART_BASE_URL
        from inventario import listar_sku_mapeo, get_stock_bodega
        import requests as req

        # Obtener stock actual de Walmart
        # Walmart paginar items - usar endpoint /v3/items
        walmart_dict = {}
        try:
            # Buscar items del seller
            res = req.get(
                f"{WALMART_BASE_URL}/v3/items",
                headers=walmart_headers(),
                params={"limit": 200, "offset": 0},
                timeout=20
            )
            if res.status_code == 200:
                data = res.json()
                items = data.get("ItemResponse", []) or data.get("itemResponse", []) or []
                for it in items:
                    sku = it.get("sku") or it.get("itemSku", "")
                    if sku:
                        walmart_dict[sku] = {
                            "stock": int(it.get("availableQuantity") or it.get("totalAvailableQty") or 0),
                            "title": it.get("productName", "") or it.get("title", ""),
                            "price": float(it.get("price", {}).get("amount", 0)) if isinstance(it.get("price"), dict) else 0,
                            "status": it.get("status", ""),
                            "wfs": it.get("wfsEnabled", False)
                        }
        except Exception as e:
            return jsonify({"error_walmart": str(e)}), 500

        # Comparar
        mapeo = listar_sku_mapeo()
        resultados = []
        for fila in mapeo:
            sku_walmart = (fila.get("sku_walmart", "") or "").strip()
            if not sku_walmart:
                continue
            sku_lusync = fila.get("sku_lusync", "")
            stock_central = get_stock_bodega(sku_lusync, "CENTRAL")
            wm_data = walmart_dict.get(sku_walmart, {})
            stock_walmart = wm_data.get("stock", None) if wm_data else None

            if stock_walmart is None:
                estado = "no_encontrado"
            elif stock_walmart == stock_central:
                estado = "sincronizado"
            else:
                estado = "desincronizado"

            resultados.append({
                "sku_lusync": sku_lusync,
                "sku_paris": sku_walmart,  # nombre genérico
                "sku_walmart": sku_walmart,
                "nombre": fila.get("nombre", "") or wm_data.get("title", ""),
                "stock_lusync": stock_central,
                "stock_paris": stock_walmart,  # nombre genérico
                "stock_walmart": stock_walmart,
                "diferencia": (stock_walmart - stock_central) if stock_walmart is not None else None,
                "ultima_actualizacion_paris": "",
                "wfs": wm_data.get("wfs", False),
                "estado": estado
            })

        return jsonify({
            "total": len(resultados),
            "sincronizados": sum(1 for r in resultados if r["estado"] == "sincronizado"),
            "desincronizados": sum(1 for r in resultados if r["estado"] == "desincronizado"),
            "no_encontrados": sum(1 for r in resultados if r["estado"] == "no_encontrado"),
            "resultados": resultados
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/walmart/forzar_sync_sku", methods=["POST"])
def ruta_walmart_forzar_sync_sku():
    """Re-envía el stock de un SKU específico a Walmart."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from walmart import actualizar_stock_walmart
        from inventario import listar_sku_mapeo, get_stock_bodega
        data = request.json or {}
        sku_lusync = data.get("sku_lusync", "").strip()
        if not sku_lusync:
            return jsonify({"ok": False, "error": "sku_lusync requerido"}), 400

        sku_walmart = None
        for fila in listar_sku_mapeo():
            if fila.get("sku_lusync") == sku_lusync:
                sku_walmart = (fila.get("sku_walmart", "") or "").strip()
                break
        if not sku_walmart:
            return jsonify({"ok": False, "error": f"SKU {sku_lusync} no tiene mapeo Walmart"}), 400

        stock = get_stock_bodega(sku_lusync, "CENTRAL")
        ok = actualizar_stock_walmart(sku_walmart, stock)
        return jsonify({"ok": ok, "sku": sku_lusync, "sku_walmart": sku_walmart, "stock_enviado": stock})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/walmart/forzar_sync_todos", methods=["POST"])
def ruta_walmart_forzar_sync_todos():
    """Re-envía el stock de todos los SKUs mapeados a Walmart."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from walmart import actualizar_stock_walmart
        from inventario import listar_sku_mapeo, get_stock_bodega
        enviados = 0
        fallidos = 0
        for fila in listar_sku_mapeo():
            sku_walmart = (fila.get("sku_walmart", "") or "").strip()
            sku_lusync = fila.get("sku_lusync", "")
            if not sku_walmart or not sku_lusync:
                continue
            stock = get_stock_bodega(sku_lusync, "CENTRAL")
            if actualizar_stock_walmart(sku_walmart, stock):
                enviados += 1
            else:
                fallidos += 1
        return jsonify({"ok": True, "enviados": enviados, "fallidos": fallidos})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500






@app.route("/debug/orden_paris_raw/<sub_order_number>")
def debug_orden_paris_raw(sub_order_number):
    """Devuelve el JSON crudo de la orden de París para identificar el campo de estado correcto."""
    if not session.get("logged"): return redirect("/")
    try:
        from paris import obtener_ordenes_paris_todas
        # Buscar en TODOS los estados posibles que conocemos
        for estado in ["awaiting_fullfillment", "ready_to_ship", "shipped", "delivered", "cancelled"]:
            try:
                todas = obtener_ordenes_paris_todas(dias=60, estado=estado)
                for so in todas:
                    if str(so.get("subOrderNumber", "")) == str(sub_order_number):
                        # Devolver TODA la orden para ver qué campos tiene
                        return jsonify({
                            "encontrada_buscando_estado": estado,
                            "campos_top_nivel": list(so.keys()),
                            "orden_completa": so
                        })
            except Exception as e:
                print(f"[debug] Error buscando en {estado}: {e}")
                continue
        return jsonify({"error": f"Orden {sub_order_number} no encontrada en ningún estado"}), 404
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/buscar_orden_paris/<sub_order_number>")
def debug_buscar_orden_paris(sub_order_number):
    """Busca una orden Paris SIN filtro de estado para encontrarla."""
    if not session.get("logged"): return redirect("/")
    try:
        from paris import paris_headers, PARIS_BASE_URL, get_seller_id
        import requests as req
        from datetime import datetime, timedelta

        # Probar varias estrategias
        resultados = {"intentos": []}

        # Intento 1: buscar SIN filtro de estado (todas las órdenes 90 días)
        for dias in [30, 60, 90, 180]:
            fecha_desde = (datetime.utcnow() - timedelta(days=dias)).strftime("%Y-%m-%d")
            params = {
                "gteCreatedAt": fecha_desde,
                "limit": 100,
                "offset": 0
            }
            seller_id = get_seller_id()
            if seller_id:
                params["sellerId"] = seller_id

            try:
                res = req.get(f"{PARIS_BASE_URL}/v2/sub-orders",
                              headers=paris_headers(),
                              params=params, timeout=20)

                if res.status_code != 200:
                    resultados["intentos"].append({
                        "dias": dias, "status_code": res.status_code,
                        "error": res.text[:200]
                    })
                    continue

                data = res.json()
                ordenes = data.get("data", [])
                total = data.get("count", 0)

                # Buscar la sub-orden
                encontrada = None
                for o in ordenes:
                    if str(o.get("subOrderNumber", "")) == str(sub_order_number):
                        encontrada = o
                        break

                resultados["intentos"].append({
                    "dias": dias,
                    "total_ordenes_traidas": len(ordenes),
                    "total_ordenes_paris": total,
                    "encontrada": encontrada is not None,
                    "campos_si_encontrada": list(encontrada.keys()) if encontrada else None,
                    "orden_completa": encontrada
                })
                if encontrada:
                    break  # ya la encontró, no seguir
            except Exception as e:
                resultados["intentos"].append({"dias": dias, "error": str(e)})

        # Intento 2: buscar directamente por orderNumber/subOrderNumber en endpoint específico
        try:
            res = req.get(f"{PARIS_BASE_URL}/v2/sub-orders/{sub_order_number}",
                          headers=paris_headers(), timeout=15)
            resultados["endpoint_directo"] = {
                "url": f"/v2/sub-orders/{sub_order_number}",
                "status_code": res.status_code,
                "respuesta": res.json() if res.status_code == 200 else res.text[:300]
            }
        except Exception as e:
            resultados["endpoint_directo"] = {"error": str(e)}

        return jsonify(resultados)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/orden_meli_raw/<order_id>")
def debug_orden_meli_raw(order_id):
    """Devuelve datos crudos de MELI sobre una orden, probando varios endpoints."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import meli_headers, MELI_API_URL
        from inventario import get_meli_auth
        import requests as req

        auth = get_meli_auth()
        if not auth or not auth.get("access_token"):
            return jsonify({"error": "MELI no conectado"}), 400

        resultados = {
            "user_id_seller_lusync": auth.get("user_id"),
            "intentos": []
        }

        # Intento 1: /orders/{id} directo
        try:
            res = req.get(f"{MELI_API_URL}/orders/{order_id}",
                          headers=meli_headers(), timeout=15)
            resultados["intentos"].append({
                "endpoint": f"/orders/{order_id}",
                "status_code": res.status_code,
                "respuesta": res.json() if res.status_code == 200 else res.text[:500]
            })
        except Exception as e:
            resultados["intentos"].append({"endpoint": f"/orders/{order_id}", "error": str(e)})

        # Intento 2: buscar la orden en /orders/search?seller=... con q=order_id
        try:
            res = req.get(f"{MELI_API_URL}/orders/search",
                          headers=meli_headers(),
                          params={"seller": auth.get("user_id"), "q": order_id, "limit": 10},
                          timeout=15)
            resultados["intentos"].append({
                "endpoint": "/orders/search?q=...",
                "status_code": res.status_code,
                "total_resultados": res.json().get("paging", {}).get("total", 0) if res.status_code == 200 else None,
                "primeras_3_ordenes": res.json().get("results", [])[:3] if res.status_code == 200 else res.text[:500]
            })
        except Exception as e:
            resultados["intentos"].append({"endpoint": "/orders/search", "error": str(e)})

        # Intento 3: ¿es un PACK? Probar /packs/{id}
        try:
            res = req.get(f"{MELI_API_URL}/packs/{order_id}",
                          headers=meli_headers(), timeout=15)
            resultados["intentos"].append({
                "endpoint": f"/packs/{order_id}",
                "status_code": res.status_code,
                "respuesta": res.json() if res.status_code == 200 else res.text[:300]
            })
        except Exception as e:
            resultados["intentos"].append({"endpoint": f"/packs/{order_id}", "error": str(e)})

        return jsonify(resultados)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/orden_meli/<order_id>")
def debug_orden_meli(order_id):
    """Diagnóstico exhaustivo de una orden MELI específica.
    Te dice por qué NO se descontó el stock."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import obtener_orden_meli
        from inventario import (orden_ya_procesada_texto, listar_sku_mapeo,
                                cargar_productos, get_stock_bodega,
                                determinar_bodega_para_canal)
        from bodegas_logic import detectar_fulfillment_meli

        # 1. Traer la orden de MELI
        orden = obtener_orden_meli(order_id)
        if not orden:
            return jsonify({
                "error": f"Orden {order_id} no encontrada en MELI API",
                "posibles_causas": [
                    "El order_id es incorrecto",
                    "El token OAuth expiró",
                    "Esta orden no pertenece a tu cuenta MELI"
                ]
            }), 404

        # 2. Estado de la orden
        estado = orden.get("status", "")
        order_items = orden.get("order_items", [])

        # 3. ¿Ya está marcada como procesada?
        meli_key = f"MELI-{order_id}"
        ya_procesada = orden_ya_procesada_texto(meli_key)

        # 4. Detectar Full vs Seller
        es_full = detectar_fulfillment_meli(orden)
        bodega_correcta = determinar_bodega_para_canal("MercadoLibre", fulfillment=es_full)

        # 5. Para cada item: buscar mapeo y verificar
        productos_dict = {p["sku"]: p for p in cargar_productos()}
        mapeo = listar_sku_mapeo()
        items_diagnostico = []
        for item in order_items:
            sku_seller = (item.get("item", {}).get("seller_custom_field", "") or "").strip()
            item_id = item.get("item", {}).get("id", "")
            qty = int(item.get("quantity", 1) or 1)

            # Buscar SKU Lusync
            sku_lusync_por_mapeo = None
            for fila in mapeo:
                if (fila.get("sku_mercadolibre") == item_id or
                    fila.get("sku_mercadolibre") == sku_seller):
                    sku_lusync_por_mapeo = fila.get("sku_lusync")
                    break

            sku_lusync_final = sku_lusync_por_mapeo or sku_seller
            existe_en_inventario = sku_lusync_final in productos_dict

            stock_actual_central = get_stock_bodega(sku_lusync_final, "CENTRAL") if existe_en_inventario else None
            stock_actual_bodega_correcta = get_stock_bodega(sku_lusync_final, bodega_correcta) if existe_en_inventario else None

            items_diagnostico.append({
                "item_id_meli": item_id,
                "sku_seller_custom_field": sku_seller,
                "cantidad": qty,
                "sku_encontrado_en_mapeo": sku_lusync_por_mapeo,
                "sku_lusync_que_se_usaria": sku_lusync_final,
                "existe_en_inventario_lusync": existe_en_inventario,
                "stock_actual_bodega_central": stock_actual_central,
                "stock_actual_bodega_destino": stock_actual_bodega_correcta,
                "bodega_donde_se_descontaria": bodega_correcta
            })

        # 6. Diagnóstico final: ¿por qué falló?
        razones = []
        if estado not in ("paid", "confirmed", "payment_required"):
            razones.append(f"Estado de la orden es '{estado}', solo se procesan: paid/confirmed/payment_required")
        if ya_procesada:
            razones.append("La orden YA estaba marcada como procesada (se ignora en sync)")
        for it in items_diagnostico:
            if not it["existe_en_inventario_lusync"]:
                razones.append(f"SKU '{it['sku_lusync_que_se_usaria']}' NO existe en inventario Lusync")

        return jsonify({
            "order_id": order_id,
            "estado_orden_meli": estado,
            "es_fulfillment_full": es_full,
            "bodega_donde_descontaria": bodega_correcta,
            "ya_marcada_como_procesada": ya_procesada,
            "items": items_diagnostico,
            "razones_no_descontado": razones if razones else ["✓ Debería descontar correctamente"],
            "fecha_orden": orden.get("date_created", ""),
            "comprador": orden.get("buyer", {}).get("nickname", "")
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/orden_paris/<sub_order_number>")
def debug_orden_paris(sub_order_number):
    """Diagnóstico exhaustivo de una orden París específica."""
    if not session.get("logged"): return redirect("/")
    try:
        from paris import obtener_orden_paris, obtener_ordenes_paris_todas
        from inventario import (orden_ya_procesada_texto, listar_sku_mapeo,
                                cargar_productos, get_stock_bodega,
                                determinar_bodega_para_canal)
        from bodegas_logic import detectar_fulfillment_paris

        # 1. Buscar la orden — primero intentar directo, si falla buscar en lista
        orden = None
        try:
            orden = obtener_orden_paris(sub_order_number)
        except: pass

        if not orden:
            # Buscar en todas las órdenes recientes
            for estado in ["awaiting_fullfillment", "ready_to_ship", "shipped", "delivered", "cancelled"]:
                try:
                    todas = obtener_ordenes_paris_todas(dias=60, estado=estado)
                    for so in todas:
                        if str(so.get("subOrderNumber", "")) == str(sub_order_number):
                            orden = so
                            break
                    if orden: break
                except: continue

        if not orden:
            return jsonify({
                "error": f"Orden {sub_order_number} no encontrada en París API (buscado 60 días, todos los estados)",
                "posibles_causas": [
                    "El subOrderNumber es incorrecto",
                    "La orden tiene más de 60 días",
                    "El estado no está en la lista buscada"
                ]
            }), 404

        # 2. Datos de la orden
        estado_paris = orden.get("statusName") or orden.get("status") or ""
        paris_key = f"PARIS-{sub_order_number}"
        ya_procesada = orden_ya_procesada_texto(paris_key)

        # 3. Detectar CD vs Seller
        es_cd = detectar_fulfillment_paris(orden)
        bodega_correcta = determinar_bodega_para_canal("Paris", fulfillment=es_cd)

        # 4. Diagnóstico de cada item
        productos_dict = {p["sku"]: p for p in cargar_productos()}
        mapeo = listar_sku_mapeo()
        items_diagnostico = []
        shipments = orden.get("shipments", [])
        for ship in shipments:
            for item in ship.get("items", []):
                sku_paris = item.get("seller_sku") or item.get("sellerSku") or ""
                qty = int(item.get("quantity", 1) or 1)

                # Buscar mapeo
                sku_lusync_por_mapeo = None
                for fila in mapeo:
                    if fila.get("sku_paris") == sku_paris:
                        sku_lusync_por_mapeo = fila.get("sku_lusync")
                        break

                sku_lusync_final = sku_lusync_por_mapeo or sku_paris
                existe = sku_lusync_final in productos_dict

                stock_central = get_stock_bodega(sku_lusync_final, "CENTRAL") if existe else None
                stock_destino = get_stock_bodega(sku_lusync_final, bodega_correcta) if existe else None

                items_diagnostico.append({
                    "sku_paris_recibido": sku_paris,
                    "cantidad": qty,
                    "sku_encontrado_en_mapeo": sku_lusync_por_mapeo,
                    "sku_lusync_que_se_usaria": sku_lusync_final,
                    "existe_en_inventario_lusync": existe,
                    "stock_actual_bodega_central": stock_central,
                    "stock_actual_bodega_destino": stock_destino,
                    "bodega_donde_se_descontaria": bodega_correcta
                })

        # 5. Razones de no descuento
        razones = []
        estados_validos = ["awaiting_fullfillment", "ready_to_ship", "shipped", "delivered"]
        if estado_paris not in estados_validos:
            razones.append(f"Estado '{estado_paris}' no está en la lista que el sync procesa: {estados_validos}")
        if ya_procesada:
            razones.append("La orden YA estaba marcada como procesada")
        for it in items_diagnostico:
            if not it["existe_en_inventario_lusync"]:
                razones.append(f"SKU '{it['sku_lusync_que_se_usaria']}' NO existe en inventario Lusync")

        return jsonify({
            "sub_order_number": sub_order_number,
            "estado_paris": estado_paris,
            "es_fulfillment_cd": es_cd,
            "bodega_donde_descontaria": bodega_correcta,
            "ya_marcada_como_procesada": ya_procesada,
            "items": items_diagnostico,
            "razones_no_descontado": razones if razones else ["✓ Debería descontar correctamente"],
            "datos_orden_completos": {
                "fecha": orden.get("createdAt") or orden.get("date_created", ""),
                "shipments_count": len(shipments)
            }
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/orden_meli_shipment/<order_id>")
def debug_orden_meli_shipment(order_id):
    """Investiga el shipment de una orden MELI para identificar Full vs Seller."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import obtener_orden_meli, meli_headers, MELI_API_URL
        import requests as req

        orden = obtener_orden_meli(order_id)
        if not orden:
            return jsonify({"error": "Orden no encontrada"}), 404

        # Obtener shipping_id
        shipping = orden.get("shipping", {}) or {}
        shipping_id = shipping.get("id")

        resultado = {
            "order_id": order_id,
            "shipping_object_en_orden": shipping,
            "shipping_id": shipping_id,
            "tags_orden": orden.get("tags", []),
            "fulfilled_orden": orden.get("fulfilled"),
        }

        # Si hay shipping_id, consultar /shipments/{id}
        if shipping_id:
            try:
                res = req.get(f"{MELI_API_URL}/shipments/{shipping_id}",
                              headers=meli_headers(), timeout=15)
                if res.status_code == 200:
                    ship_detail = res.json()
                    resultado["shipment_detalle_completo"] = ship_detail
                    resultado["CAMPOS_CLAVE_PARA_DETECTAR_FULL"] = {
                        "logistic_type": ship_detail.get("logistic_type"),
                        "logistic.type": (ship_detail.get("logistic") or {}).get("type"),
                        "logistic.mode": (ship_detail.get("logistic") or {}).get("mode"),
                        "shipping_mode": ship_detail.get("shipping_mode"),
                        "shipping_option_name": (ship_detail.get("shipping_option") or {}).get("name"),
                        "service_id": ship_detail.get("service_id"),
                        "tags": ship_detail.get("tags", [])
                    }
                else:
                    resultado["shipment_error"] = f"{res.status_code}: {res.text[:300]}"
            except Exception as e:
                resultado["shipment_error"] = str(e)

        return jsonify(resultado)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/orden_meli_completa/<order_id>")
def debug_orden_meli_completa(order_id):
    """Vuelca la orden + el item asociado para encontrar dónde está el SKU."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import obtener_orden_meli, meli_headers, MELI_API_URL
        import requests as req

        # 1. Traer la orden completa
        orden = obtener_orden_meli(order_id)
        if not orden:
            return jsonify({"error": "Orden no encontrada"}), 404

        # 2. Para cada item, traer también el detalle del item
        items_completos = []
        for item in orden.get("order_items", []):
            item_data = item.get("item", {})
            item_id_meli = item_data.get("id", "")

            # Traer el ítem completo de la API
            item_detalle = None
            try:
                res = req.get(f"{MELI_API_URL}/items/{item_id_meli}",
                              headers=meli_headers(), timeout=15)
                if res.status_code == 200:
                    item_detalle = res.json()
            except Exception as e:
                item_detalle = {"error": str(e)}

            items_completos.append({
                "item_en_orden": item_data,
                "campos_top_nivel_item": list(item_data.keys()),
                "item_detalle_completo": item_detalle,
                "campos_top_nivel_detalle": list(item_detalle.keys()) if isinstance(item_detalle, dict) else None,
                # Buscar SKU en posibles ubicaciones del detalle
                "POSIBLES_SKU_DETECTADOS": {
                    "item.seller_custom_field": item_data.get("seller_custom_field"),
                    "item.seller_sku": item_data.get("seller_sku"),
                    "detalle.seller_custom_field": item_detalle.get("seller_custom_field") if isinstance(item_detalle, dict) else None,
                    "detalle.seller_sku": item_detalle.get("seller_sku") if isinstance(item_detalle, dict) else None,
                    "detalle.attributes": [
                        {"id": a.get("id"), "name": a.get("name"), "value": a.get("value_name")}
                        for a in (item_detalle.get("attributes", []) if isinstance(item_detalle, dict) else [])
                        if "SKU" in (a.get("id", "") + a.get("name", "")).upper()
                    ]
                }
            })

        return jsonify({
            "order_id": order_id,
            "campos_top_nivel_orden": list(orden.keys()),
            "items": items_completos
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/mapeo_completo")
def debug_mapeo_completo():
    """Lista TODA la tabla sku_mapeo con todos los campos para detectar dónde está E10E11E12."""
    if not session.get("logged"): return redirect("/")
    try:
        from inventario import listar_sku_mapeo
        mapeo = listar_sku_mapeo()

        # Filtrar el SKU específico
        E10 = [m for m in mapeo if m.get("sku_lusync") == "E10E11E12"]

        # Buscar por 'MLC1584290001' en cualquier campo
        contiene_MLC = []
        for m in mapeo:
            for campo, valor in m.items():
                if isinstance(valor, str) and "MLC1584290001" in valor:
                    contiene_MLC.append({"sku_lusync": m.get("sku_lusync"), "campo": campo, "valor": valor})

        return jsonify({
            "total_filas_en_tabla": len(mapeo),
            "todos_los_campos_disponibles": list(mapeo[0].keys()) if mapeo else [],
            "sku_E10E11E12": E10,
            "filas_con_MLC1584290001": contiene_MLC,
            "primeras_5_filas_completas": mapeo[:5]
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/mapeo_meli")
def debug_mapeo_meli():
    """Lista todos los mapeos de SKU MercadoLibre para detectar duplicados o errores."""
    if not session.get("logged"): return redirect("/")
    try:
        from inventario import listar_sku_mapeo
        mapeo = listar_sku_mapeo()

        # Filtrar solo los que tienen sku_mercadolibre
        con_meli = [m for m in mapeo if m.get("sku_mercadolibre")]

        # Detectar duplicados (mismo MLC apunta a distintos sku_lusync)
        meli_to_lusync = {}
        for m in con_meli:
            sku_meli = m.get("sku_mercadolibre", "").strip()
            sku_lusync = m.get("sku_lusync", "")
            if sku_meli not in meli_to_lusync:
                meli_to_lusync[sku_meli] = []
            meli_to_lusync[sku_meli].append(sku_lusync)

        duplicados = {k: v for k, v in meli_to_lusync.items() if len(v) > 1}

        return jsonify({
            "total_mapeos_con_meli": len(con_meli),
            "lista_completa": [
                {"sku_lusync": m.get("sku_lusync"),
                 "sku_mercadolibre": m.get("sku_mercadolibre"),
                 "nombre": m.get("nombre", "")}
                for m in con_meli
            ],
            "DUPLICADOS_DETECTADOS": duplicados,
            "buscar_MLC1584290001": [
                m for m in con_meli
                if m.get("sku_mercadolibre", "").strip() == "MLC1584290001"
            ]
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/precios_productos")
def debug_precios_productos():
    """Diagnóstico: muestra los precios cargados en BD para los productos vendidos."""
    if not session.get("logged"): return redirect("/")
    try:
        from datetime import datetime, timedelta
        productos = cargar_productos()
        skus_vendidos_top = stats_top_productos_vendidos(
            (datetime.now().date() - timedelta(days=30)),
            datetime.now().date(),
            limite=20
        )
        skus_set = {p["sku"] for p in skus_vendidos_top}
        relevantes = [p for p in productos if p["sku"] in skus_set]
        return jsonify({
            "total_productos_bd": len(productos),
            "productos_con_precio_normal": sum(1 for p in productos if (p.get("precio_normal") or 0) > 0),
            "productos_con_precio_oferta": sum(1 for p in productos if (p.get("precio_oferta") or 0) > 0),
            "productos_sin_precio": sum(1 for p in productos if (p.get("precio_normal") or 0) <= 0 and (p.get("precio_oferta") or 0) <= 0),
            "skus_vendidos_recientemente_y_sus_precios": [
                {
                    "sku": p["sku"],
                    "nombre": p["nombre"],
                    "precio_normal": p.get("precio_normal", 0),
                    "precio_oferta": p.get("precio_oferta", 0),
                    "stock": p.get("stock", 0)
                }
                for p in relevantes
            ],
            "muestra_5_productos": [
                {
                    "sku": p["sku"],
                    "nombre": p["nombre"][:40],
                    "precio_normal": p.get("precio_normal", 0),
                    "precio_oferta": p.get("precio_oferta", 0)
                }
                for p in productos[:5]
            ]
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500



# ── BODEGAS ─────────────────────────────────────────────────────────────────

@app.route("/bodegas")
def ruta_bodegas_listar():
    if not session.get("logged"): return jsonify({}), 401
    return jsonify(listar_bodegas())

@app.route("/bodegas/stock")
def ruta_bodegas_stock():
    """Tabla completa producto × bodega para administración."""
    if not session.get("logged"): return jsonify([]), 401
    return jsonify(listar_stock_completo())

@app.route("/bodegas/totales")
def ruta_bodegas_totales():
    """Totales por bodega para dashboard."""
    if not session.get("logged"): return jsonify({}), 401
    return jsonify(stock_total_por_bodega())

@app.route("/bodegas/stock_sku")
def ruta_bodegas_stock_sku():
    """Stock detallado de un SKU específico en todas las bodegas."""
    if not session.get("logged"): return jsonify({}), 401
    sku = request.args.get("sku", "").strip()
    if not sku: return jsonify({"error": "sku requerido"}), 400
    return jsonify(stock_por_bodega(sku))

@app.route("/bodegas/set_stock", methods=["POST"])
def ruta_bodegas_set_stock():
    """Establece stock de un SKU en una bodega (override)."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    data = request.json or {}
    sku = (data.get("sku") or "").strip()
    bodega = (data.get("bodega_codigo") or "").strip()
    cantidad = int(data.get("cantidad", 0))
    if not sku or not bodega:
        return jsonify({"ok": False, "error": "sku y bodega_codigo requeridos"})
    try:
        set_stock_bodega(sku, bodega, cantidad)
        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "set_stock_bodega",
                        detalle=f"SKU {sku} en bodega {bodega} = {cantidad}")
        return jsonify({"ok": True, "nuevo_total": get_stock_bodega(sku, bodega)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/bodegas/actualizar_nombres", methods=["GET", "POST"])
def ruta_bodegas_actualizar_nombres():
    """Sobrescribe nombres de bodegas con los valores actuales del código.
    Útil cuando renombramos una bodega y la BD tiene el nombre viejo."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        actualizadas = actualizar_nombres_bodegas()
        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "actualizar_nombres_bodegas",
                        detalle=f"{len(actualizadas)} bodegas refrescadas")
        return jsonify({"ok": True, "actualizadas": actualizadas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/bodegas/matriz")
def ruta_bodegas_matriz():
    """Devuelve la matriz completa SKU × Bodega para la UI.
    Returns: {bodegas: [...], skus: [{sku, nombre, stocks: {bodega: cantidad}, total}]}
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn, listar_bodegas
        bodegas = listar_bodegas()
        bodega_codigos = [b["codigo"] for b in bodegas]

        conn = get_conn(); cur = conn.cursor()
        # Obtener todos los productos
        cur.execute("SELECT sku, nombre FROM productos ORDER BY nombre")
        productos = cur.fetchall()

        # Obtener todo el stock_bodega en una sola query
        cur.execute("SELECT sku, bodega_codigo, cantidad FROM stock_bodega")
        stock_dict = {}
        for sku, bod, cant in cur.fetchall():
            stock_dict.setdefault(sku, {})[bod] = int(cant or 0)
        cur.close(); conn.close()

        # Construir matriz
        skus = []
        for sku, nombre in productos:
            stocks = {b: stock_dict.get(sku, {}).get(b, 0) for b in bodega_codigos}
            total = sum(stocks.values())
            skus.append({
                "sku": sku,
                "nombre": nombre or sku,
                "stocks": stocks,
                "total": total
            })

        return jsonify({
            "bodegas": bodegas,
            "skus": skus,
            "total_skus": len(skus)
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/bodegas/descargar_plantilla")
def ruta_bodegas_descargar_plantilla():
    """Genera Excel con todos los SKUs y columnas por bodega para llenar."""
    if not session.get("logged"): return redirect("/")
    try:
        from inventario import get_conn, listar_bodegas
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        from datetime import datetime

        bodegas = listar_bodegas()

        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT sku, nombre FROM productos ORDER BY nombre")
        productos = cur.fetchall()
        cur.execute("SELECT sku, bodega_codigo, cantidad FROM stock_bodega")
        stock_dict = {}
        for sku, bod, cant in cur.fetchall():
            stock_dict.setdefault(sku, {})[bod] = int(cant or 0)
        cur.close(); conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock por Bodega"

        # Encabezados: sku, nombre, + una columna por bodega
        headers = ["sku_lusync", "nombre"] + [b["codigo"] for b in bodegas]
        ws.append(headers)

        # Estilizar encabezado
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EEEDFE")
            cell.alignment = Alignment(horizontal="center")

        # Filas con datos
        for sku, nombre in productos:
            row = [sku, nombre or sku]
            for b in bodegas:
                row.append(stock_dict.get(sku, {}).get(b["codigo"], 0))
            ws.append(row)

        # Hoja con instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        ws2.append(["Plantilla de Stock por Bodega — Lusync"])
        ws2.append([])
        ws2.append(["Cómo usar:"])
        ws2.append(["1. Edita las cantidades en cada celda según tu inventario real"])
        ws2.append(["2. NO modifiques los nombres de columnas (encabezados)"])
        ws2.append(["3. NO modifiques la columna 'sku_lusync' (es el identificador)"])
        ws2.append(["4. Guarda el archivo y súbelo desde el botón 'Importar Excel'"])
        ws2.append([])
        ws2.append(["Bodegas disponibles:"])
        for b in bodegas:
            ws2.append([b["codigo"], "→", b["nombre"]])

        # Generar archivo en memoria
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre_archivo = f"plantilla_stock_bodegas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=nombre_archivo,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/bodegas/importar_excel", methods=["POST"])
def ruta_bodegas_importar_excel():
    """Recibe un Excel con stock por bodega y lo aplica.
    Procesa en lotes y registra el resultado en bodegas_imports."""
    if not session.get("logged"): return jsonify({"ok": False, "error": "no autorizado"}), 401
    try:
        from openpyxl import load_workbook
        from inventario import (set_stock_bodega, listar_bodegas,
                                crear_import_log, actualizar_import_log)

        if "archivo" not in request.files:
            return jsonify({"ok": False, "error": "Sin archivo"}), 400

        archivo = request.files["archivo"]
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        # Leer encabezados (fila 1)
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if "sku_lusync" not in headers:
            return jsonify({
                "ok": False,
                "error": "Columna 'sku_lusync' no encontrada en el Excel"
            }), 400

        idx_sku = headers.index("sku_lusync")
        idx_nombre = headers.index("nombre") if "nombre" in headers else None

        # Identificar columnas de bodegas (códigos válidos)
        bodegas_validas = {b["codigo"] for b in listar_bodegas()}
        cols_bodega = {}  # {col_idx: codigo_bodega}
        for i, h in enumerate(headers):
            if h in bodegas_validas:
                cols_bodega[i] = h

        if not cols_bodega:
            return jsonify({
                "ok": False,
                "error": f"No se encontraron columnas de bodegas válidas. Columnas válidas: {sorted(bodegas_validas)}"
            }), 400

        # Contar filas (descontando encabezado)
        total_filas = ws.max_row - 1

        # Crear log de import
        usuario = session.get("usuario", "Sistema")
        nombre_archivo = archivo.filename or "stock.xlsx"
        import_id = crear_import_log(nombre_archivo, usuario, total_filas)

        procesados = 0
        advertencias = 0
        errores = 0
        log_lines = []

        # Procesar fila por fila
        for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                sku = str(row[idx_sku]).strip() if row[idx_sku] else ""
                if not sku:
                    errores += 1
                    log_lines.append(f"× Fila {fila_idx}: SKU vacío")
                    continue

                # Aplicar cada bodega
                cambios = []
                for col_idx, bod_codigo in cols_bodega.items():
                    val = row[col_idx]
                    if val is None or val == "":
                        continue
                    try:
                        cantidad = int(float(val))
                        if cantidad < 0:
                            advertencias += 1
                            log_lines.append(f"! Fila {fila_idx} {sku}: {bod_codigo} negativo, ajustado a 0")
                            cantidad = 0
                        set_stock_bodega(sku, bod_codigo, cantidad)
                        cambios.append(f"{bod_codigo}={cantidad}")
                    except (ValueError, TypeError):
                        advertencias += 1
                        log_lines.append(f"! Fila {fila_idx} {sku}: {bod_codigo} valor inválido '{val}'")

                if cambios:
                    procesados += 1
                    log_lines.append(f"✓ Fila {fila_idx} {sku}: {', '.join(cambios)}")

                # Actualizar progreso cada 25 filas
                if fila_idx % 25 == 0:
                    actualizar_import_log(import_id, procesados=procesados,
                                          advertencias=advertencias, errores=errores)
            except Exception as e:
                errores += 1
                log_lines.append(f"× Fila {fila_idx}: {str(e)[:100]}")

        # Estado final
        if errores > 0 and procesados == 0:
            estado_final = "error"
        elif advertencias > 0 or errores > 0:
            estado_final = "advertencias"
        else:
            estado_final = "ok"

        actualizar_import_log(import_id,
                              procesados=procesados,
                              advertencias=advertencias,
                              errores=errores,
                              estado=estado_final,
                              log="\n".join(log_lines[-200:]))  # últimas 200 líneas

        registrar_audit(usuario, request.remote_addr, "importar_stock_bodegas",
                        entidad="stock_bodega",
                        detalle=f"{procesados} OK, {advertencias} adv, {errores} err")

        return jsonify({
            "ok": True,
            "import_id": import_id,
            "procesados": procesados,
            "advertencias": advertencias,
            "errores": errores,
            "estado": estado_final,
            "ultimas_lineas": log_lines[-20:]
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/bodegas/imports")
def ruta_bodegas_imports():
    """Lista las últimas importaciones."""
    if not session.get("logged"): return jsonify({"imports": []}), 401
    try:
        from inventario import listar_imports_recientes
        return jsonify({"imports": listar_imports_recientes(limit=20)})
    except Exception as e:
        return jsonify({"imports": [], "error": str(e)}), 500


@app.route("/bodegas/imports/<int:import_id>")
def ruta_bodegas_import_detalle(import_id):
    """Detalle de una importación específica con log completo."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import obtener_import_log
        det = obtener_import_log(import_id)
        if not det:
            return jsonify({"error": "Import no encontrado"}), 404
        return jsonify(det)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/bodegas/guardar_lote", methods=["POST"])
def ruta_bodegas_guardar_lote():
    """Guarda múltiples cambios de stock en una sola llamada (edición inline en UI)."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from inventario import set_stock_bodega
        data = request.json or {}
        cambios = data.get("cambios", [])  # [{sku, bodega_codigo, cantidad}, ...]

        guardados = 0
        errores = []
        for c in cambios:
            try:
                set_stock_bodega(c["sku"], c["bodega_codigo"], int(c["cantidad"]))
                guardados += 1
            except Exception as e:
                errores.append(f"{c.get('sku')}/{c.get('bodega_codigo')}: {e}")

        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "editar_stock_bodegas", entidad="stock_bodega",
                        detalle=f"{guardados} celdas actualizadas")

        return jsonify({"ok": True, "guardados": guardados, "errores": errores})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/bodegas/transferir", methods=["POST"])
def ruta_bodegas_transferir():
    """Mueve stock de una bodega a otra para un SKU."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    data = request.json or {}
    sku = (data.get("sku") or "").strip()
    desde = (data.get("desde") or "").strip()
    hasta = (data.get("hasta") or "").strip()
    cantidad = int(data.get("cantidad", 0))
    if not sku or not desde or not hasta or cantidad <= 0:
        return jsonify({"ok": False, "error": "Faltan parámetros"})

    stock_origen = get_stock_bodega(sku, desde)
    if stock_origen < cantidad:
        return jsonify({"ok": False, "error": f"Stock insuficiente en {desde}: {stock_origen}"})

    try:
        ajustar_stock_bodega(sku, desde, -cantidad)
        ajustar_stock_bodega(sku, hasta, cantidad)
        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "transferir_stock",
                        detalle=f"SKU {sku}: {cantidad}u {desde} → {hasta}")
        return jsonify({"ok": True,
                        "stock_desde": get_stock_bodega(sku, desde),
                        "stock_hasta": get_stock_bodega(sku, hasta)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── DASHBOARD STATS ─────────────────────────────────────────────────────────

def _parse_rango_fechas():
    """Lee desde/hasta de query params, default últimos 7 días."""
    from datetime import datetime, timedelta
    hoy = datetime.now().date()
    desde_str = request.args.get("desde", "").strip()
    hasta_str = request.args.get("hasta", "").strip()
    try:
        hasta = datetime.strptime(hasta_str, "%Y-%m-%d").date() if hasta_str else hoy
    except: hasta = hoy
    try:
        desde = datetime.strptime(desde_str, "%Y-%m-%d").date() if desde_str else (hoy - timedelta(days=6))
    except: desde = hoy - timedelta(days=6)
    return desde, hasta


@app.route("/stats/kpis")
def ruta_stats_kpis():
    if not session.get("logged"): return jsonify({}), 401
    desde, hasta = _parse_rango_fechas()
    return jsonify(stats_kpis_dashboard(desde, hasta))


@app.route("/stats/ventas_por_canal")
def ruta_stats_ventas_canal():
    if not session.get("logged"): return jsonify([]), 401
    desde, hasta = _parse_rango_fechas()
    return jsonify(stats_ventas_por_canal_dia(desde, hasta))


@app.route("/stats/top_productos")
def ruta_stats_top():
    if not session.get("logged"): return jsonify([]), 401
    desde, hasta = _parse_rango_fechas()
    limite = int(request.args.get("limite", 10))
    return jsonify(stats_top_productos_vendidos(desde, hasta, limite))


@app.route("/stats/movimientos_dia")
def ruta_stats_movs():
    if not session.get("logged"): return jsonify([]), 401
    desde, hasta = _parse_rango_fechas()
    return jsonify(stats_movimientos_dia(desde, hasta))


@app.route("/stats/distribucion_stock")
def ruta_stats_distribucion():
    if not session.get("logged"): return jsonify([]), 401
    return jsonify(stats_distribucion_stock_canal())


@app.route("/stats/propia_vs_fulfillment")
def ruta_stats_propia_vs_fulfillment():
    """Desglose de ventas por bodega: propia (CENTRAL) vs fulfillment (otras).
    Devuelve montos, unidades y % por cada bodega individual."""
    if not session.get("logged"): return jsonify({}), 401
    desde, hasta = _parse_rango_fechas()
    try:
        from inventario import get_conn, listar_bodegas
        conn = get_conn(); cur = conn.cursor()

        # Query: agrupar movimientos de salida por bodega y calcular monto + cantidad
        # Para el monto: cantidad × precio del producto
        cur.execute("""
            SELECT
              COALESCE(m.bodega_codigo, 'CENTRAL') AS bodega,
              COUNT(DISTINCT m.orden_id) AS ventas,
              COALESCE(SUM(m.cantidad), 0) AS unidades,
              COALESCE(SUM(m.cantidad * COALESCE(p.precio_oferta, p.precio_normal, 0)), 0) AS monto
            FROM movimientos m
            LEFT JOIN productos p ON p.sku = m.sku
            WHERE m.tipo = 'salida'
              AND m.canal NOT IN ('Manual', 'Sistema')
              AND m.fecha::date BETWEEN %s AND %s
            GROUP BY COALESCE(m.bodega_codigo, 'CENTRAL')
        """, (desde, hasta))
        rows = cur.fetchall()
        cur.close(); conn.close()

        # Mapear por bodega y enriquecer con metadata
        bodegas_meta = {b["codigo"]: b for b in listar_bodegas()}
        por_bodega = {}
        for bod_codigo, ventas, unidades, monto in rows:
            por_bodega[bod_codigo] = {
                "codigo": bod_codigo,
                "nombre": bodegas_meta.get(bod_codigo, {}).get("nombre", bod_codigo),
                "tipo": bodegas_meta.get(bod_codigo, {}).get("tipo", "propia"),
                "ventas": int(ventas or 0),
                "unidades": int(unidades or 0),
                "monto": float(monto or 0)
            }

        # Asegurar que todas las bodegas activas aparezcan (aunque sea con 0)
        for b in listar_bodegas():
            if b["codigo"] not in por_bodega:
                por_bodega[b["codigo"]] = {
                    "codigo": b["codigo"],
                    "nombre": b["nombre"],
                    "tipo": b["tipo"],
                    "ventas": 0, "unidades": 0, "monto": 0
                }

        # Totales agregados: propia vs fulfillment
        total_propia_monto = sum(b["monto"] for b in por_bodega.values() if b["tipo"] == "propia")
        total_propia_ventas = sum(b["ventas"] for b in por_bodega.values() if b["tipo"] == "propia")
        total_propia_unidades = sum(b["unidades"] for b in por_bodega.values() if b["tipo"] == "propia")
        total_full_monto = sum(b["monto"] for b in por_bodega.values() if b["tipo"] != "propia")
        total_full_ventas = sum(b["ventas"] for b in por_bodega.values() if b["tipo"] != "propia")
        total_full_unidades = sum(b["unidades"] for b in por_bodega.values() if b["tipo"] != "propia")
        total_general = total_propia_monto + total_full_monto

        return jsonify({
            "desde": str(desde),
            "hasta": str(hasta),
            "propia": {
                "monto": total_propia_monto,
                "ventas": total_propia_ventas,
                "unidades": total_propia_unidades,
                "porcentaje": round((total_propia_monto / total_general * 100) if total_general > 0 else 0, 1)
            },
            "fulfillment": {
                "monto": total_full_monto,
                "ventas": total_full_ventas,
                "unidades": total_full_unidades,
                "porcentaje": round((total_full_monto / total_general * 100) if total_general > 0 else 0, 1)
            },
            "total": total_general,
            "por_bodega": sorted(por_bodega.values(), key=lambda x: -x["monto"])
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── MERCADOLIBRE ─────────────────────────────────────────────────────────────

@app.route("/mercadolibre/conectar")
def ruta_meli_conectar():
    """Inicia el flujo OAuth2 redirigiendo al usuario al login de MercadoLibre."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import construir_url_autorizacion
        url = construir_url_autorizacion(state=session.get("usuario", "lusync"))
        return redirect(url)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/mercadolibre/callback")
def ruta_meli_callback():
    """Recibe el code de MercadoLibre tras autorización del usuario."""
    code = request.args.get("code")
    error = request.args.get("error")
    if error:
        return f"<h2>Error de MercadoLibre</h2><p>{error}: {request.args.get('error_description','')}</p>", 400
    if not code:
        return "<h2>Falta el parámetro code</h2>", 400

    try:
        from mercadolibre import intercambiar_codigo_por_token
        import time
        data = intercambiar_codigo_por_token(code)
        set_meli_auth({
            "access_token":  data["access_token"],
            "refresh_token": data.get("refresh_token", ""),
            "user_id":       data.get("user_id"),
            "expires_at":    int(time.time()) + int(data.get("expires_in", 21600))
        })
        # Redirigir al panel con confirmación
        return redirect("/panel?meli=ok")
    except Exception as e:
        import traceback
        return f"<h2>Error conectando MercadoLibre</h2><pre>{str(e)}\n\n{traceback.format_exc()}</pre>", 500


@app.route("/mercadolibre/desconectar", methods=["POST"])
def ruta_meli_desconectar():
    """Borra el token guardado."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    borrar_meli_auth()
    return jsonify({"ok": True})


@app.route("/mercadolibre/estado")
def ruta_meli_estado():
    """Devuelve el estado de la conexión (para mostrar en el panel)."""
    if not session.get("logged"): return jsonify({}), 401
    try:
        from mercadolibre import verificar_conexion_meli
        return jsonify(verificar_conexion_meli())
    except Exception as e:
        return jsonify({"conectado": False, "error": str(e)})


@app.route("/mercadolibre/publicaciones")
def ruta_meli_publicaciones():
    """Lista las publicaciones del seller (para mapeo de SKUs)."""
    if not session.get("logged"): return jsonify({}), 401
    try:
        from mercadolibre import obtener_publicaciones_meli
        offset = int(request.args.get("offset", 0))
        return jsonify(obtener_publicaciones_meli(limite=50, offset=offset) or {"items":[],"total":0})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/mercadolibre/sync_precios", methods=["POST"])
def ruta_meli_sync_precios():
    """Envía precios actuales tal cual a MercadoLibre, sin transformaciones.
    Las comisiones, márgenes y redondeos se manejarán en el módulo Motor de Precios."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from mercadolibre import actualizar_precio_meli
        from inventario import listar_sku_mapeo

        productos = cargar_productos()
        productos_dict = {p["sku"]: p for p in productos}
        mapeos = listar_sku_mapeo()

        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "sync_meli_precios", detalle="Sync masivo precios MercadoLibre")

        enviados = 0
        fallidos = 0
        log = []
        for fila in mapeos:
            sku_lusync = fila.get("sku_lusync", "")
            sku_meli = (fila.get("sku_mercadolibre", "") or "").strip()
            if not sku_meli or not sku_lusync:
                continue
            p = productos_dict.get(sku_lusync, {})
            precio_normal = p.get("precio_normal") or 0
            precio_oferta = p.get("precio_oferta") or 0
            # Enviar precio tal cual: oferta si existe, si no normal
            precio = precio_oferta if precio_oferta > 0 else precio_normal
            if precio <= 0:
                log.append(f"⚠ {sku_lusync} sin precio")
                continue
            if actualizar_precio_meli(sku_meli, precio):
                enviados += 1
                log.append(f"✓ {sku_meli} → ${precio}")
            else:
                fallidos += 1
                log.append(f"× {sku_meli} falló")
        return jsonify({"ok": True, "enviados": enviados, "fallidos": fallidos, "log": log[:30]})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/paris/sync_precios", methods=["POST"])
def ruta_paris_sync_precios():
    """Envía precios actuales tal cual a París, sin transformaciones.
    Las comisiones, márgenes y redondeos se manejarán en el módulo Motor de Precios."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from paris import actualizar_precio_paris
        from inventario import listar_sku_mapeo

        productos = cargar_productos()
        productos_dict = {p["sku"]: p for p in productos}
        mapeos = listar_sku_mapeo()

        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "sync_paris_precios", detalle="Sync masivo precios Paris")

        enviados = 0
        fallidos = 0
        log = []
        for fila in mapeos:
            sku_lusync = fila.get("sku_lusync", "")
            sku_paris = (fila.get("sku_paris", "") or "").strip()
            if not sku_paris or not sku_lusync:
                continue
            p = productos_dict.get(sku_lusync, {})
            precio_normal = p.get("precio_normal") or 0
            precio_oferta = p.get("precio_oferta") or 0
            if precio_normal <= 0:
                log.append(f"⚠ {sku_lusync} sin precio_normal")
                continue
            # Enviar precios tal cual: precio_normal y precio_oferta si existe
            precio_oferta_final = precio_oferta if (precio_oferta > 0 and precio_oferta < precio_normal) else None
            ok = actualizar_precio_paris(sku_paris, precio_normal, precio_oferta_final)
            if ok:
                enviados += 1
                txt = f"✓ {sku_paris} → ${precio_normal}"
                if precio_oferta_final: txt += f" (oferta ${precio_oferta_final})"
                log.append(txt)
            else:
                fallidos += 1
                log.append(f"× {sku_paris} falló")
        return jsonify({"ok": True, "enviados": enviados, "fallidos": fallidos, "log": log[:30]})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/mercadolibre/sync_ordenes")
def ruta_meli_sync_ordenes():
    """Descarga órdenes históricas de MercadoLibre y descuenta stock de bodega correcta.
    Optimizado para no agotar RAM en Render."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from mercadolibre import obtener_ordenes_meli, obtener_sku_de_item_meli
        from inventario import (descontar_venta_inteligente, detectar_fulfillment_meli,
                                listar_sku_mapeo)
        from datetime import datetime
        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "sync_meli", detalle="Sync manual órdenes MercadoLibre")

        productos_dict = {p["sku"]: p for p in cargar_productos()}
        nuevas = 0
        errores = []
        log = []

        # Cache de SKUs ya resueltos via /items/{id} para no re-consultar la misma publicación
        cache_sku_por_item = {}

        # ⚠️ PROTECCIÓN RAM: traer máx 2 páginas de 50 = 100 órdenes por request
        # Para cargar más históricos, llamar el endpoint varias veces o usar ?paginas=4
        max_paginas = int(request.args.get("paginas", 2))
        log.append(f"Configuración: máx {max_paginas} páginas × 50 = {max_paginas*50} órdenes")

        for pagina in range(max_paginas):
            offset = pagina * 50
            try:
                ordenes = obtener_ordenes_meli(limit=50, offset=offset)
                log.append(f"Página {pagina+1} (offset {offset}): {len(ordenes)} órdenes")
                if not ordenes:
                    break
                for o in ordenes:
                    order_id = str(o.get("id", ""))
                    estado = o.get("status", "")
                    if estado not in ("paid", "confirmed"):
                        continue
                    meli_key = f"MELI-{order_id}"
                    if orden_ya_procesada_texto(meli_key):
                        continue
                    marcar_orden_procesada_texto(meli_key)

                    # ── Extraer fecha real de compra del marketplace ────────
                    # MELI devuelve date_created en ISO con timezone (ej: 2026-05-03T18:32:15.000-04:00)
                    # Lo parseamos y dejamos que descontar_venta_inteligente lo convierta a Chile
                    fecha_compra_meli = None
                    try:
                        import pytz as _pytz
                        date_str = o.get("date_created", "") or ""
                        if date_str:
                            # Manejar ambos formatos: con .000 milisegundos o sin
                            if "." in date_str:
                                # Formato: 2026-05-03T18:32:15.000-04:00
                                date_str_clean = date_str.replace("Z", "+00:00")
                                fecha_compra_meli = datetime.fromisoformat(date_str_clean)
                            else:
                                # Formato: 2026-05-03T18:32:15-04:00 o con Z
                                date_str_clean = date_str.replace("Z", "+00:00")
                                fecha_compra_meli = datetime.fromisoformat(date_str_clean)
                            # DEBUG temporal: loggear lo que se parseó
                            log.append(f"  Orden {order_id}: date_created='{date_str}' → parsed={fecha_compra_meli}")
                        else:
                            log.append(f"  Orden {order_id}: date_created VACÍO en payload")
                    except Exception as e:
                        log.append(f"  Orden {order_id}: NO se pudo parsear date_created '{o.get('date_created','')}': {e}")
                        fecha_compra_meli = None

                    # Detectar Full vs Seller
                    es_full = detectar_fulfillment_meli(o)
                    tipo_str = "FULL" if es_full else "Seller"

                    for item in o.get("order_items", []):
                        item_data = item.get("item", {})
                        item_id = item_data.get("id", "")
                        # Leer SKU: seller_sku primero, luego seller_custom_field
                        sku_seller = (
                            (item_data.get("seller_sku") or "").strip()
                            or (item_data.get("seller_custom_field") or "").strip()
                        )
                        # Si vacío, consultar detalle (con cache)
                        if not sku_seller and item_id:
                            if item_id in cache_sku_por_item:
                                sku_seller = cache_sku_por_item[item_id]
                            else:
                                try:
                                    sku_resuelto = obtener_sku_de_item_meli(item_id)
                                    if sku_resuelto:
                                        sku_seller = sku_resuelto
                                        cache_sku_por_item[item_id] = sku_resuelto
                                        log.append(f"  SKU resuelto desde item detail: {sku_seller}")
                                    else:
                                        cache_sku_por_item[item_id] = None
                                except Exception as e:
                                    log.append(f"  Error consultando item: {e}")
                        qty = int(item.get("quantity", 1) or 1)

                        # Buscar SKU Lusync vía sku_mapeo_canal (PRIORIDAD: por item_id, luego por sku_seller)
                        sku_lusync = None
                        try:
                            from inventario import obtener_sku_lusync_por_canal
                            sku_lusync = obtener_sku_lusync_por_canal(
                                "mercadolibre",
                                item_id_canal=item_id,
                                sku_canal=sku_seller
                            )
                        except Exception as e:
                            log.append(f"  Error consultando sku_mapeo_canal: {e}")

                        # Fallback legacy: tabla sku_mapeo vieja
                        if not sku_lusync:
                            try:
                                for fila in listar_sku_mapeo():
                                    sku_mapped = (fila.get("sku_mercadolibre") or "").strip()
                                    if sku_mapped and (sku_mapped == item_id or sku_mapped == sku_seller):
                                        sku_lusync = fila.get("sku_lusync")
                                        break
                            except: pass

                        # Último fallback: usar el SKU del seller tal cual
                        if not sku_lusync and sku_seller:
                            sku_lusync = sku_seller

                        if not sku_lusync or sku_lusync not in productos_dict:
                            log.append(f"Orden {order_id}: SKU '{sku_lusync or item_id}' no encontrado")
                            continue

                        resultado = descontar_venta_inteligente(
                            sku=sku_lusync,
                            cantidad=qty,
                            canal="MercadoLibre",
                            fulfillment=es_full,
                            orden_id=order_id,
                            motivo=f"Venta MercadoLibre {tipo_str}",
                            usuario="Sistema",
                            fecha_compra_marketplace=fecha_compra_meli,
                            origen_registro="sync_manual"
                        )
                        log.append(f"{order_id} {tipo_str}: {sku_lusync} -{qty} desde {resultado['bodega']}")

                        # Sync a otros canales SOLO si fue Seller (afectó Central)
                        if not es_full:
                            try:
                                from bodegas_logic import sincronizar_stock_a_marketplaces
                                sincronizar_stock_a_marketplaces(sku_lusync, excepto=["mercadolibre"])
                            except Exception as e:
                                log.append(f"  Sync cruzado falló: {e}")
                    nuevas += 1

                # Liberar memoria entre páginas
                del ordenes
                import gc
                gc.collect()
            except Exception as e:
                errores.append(f"página {pagina+1}: {str(e)}")
                log.append(f"Página {pagina+1}: ERROR {str(e)}")
                break

        log.append(f"Cache SKU items consultados: {len(cache_sku_por_item)} ítems")
        return jsonify({
            "ok": True,
            "nuevas_ordenes": nuevas,
            "errores": errores,
            "log": log
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/mercadolibre/reclasificar_bodegas", methods=["GET", "POST"])
def ruta_meli_reclasificar_bodegas():
    """Re-clasifica las ventas históricas de MercadoLibre verificando si fueron Full o Seller.
    Para cada movimiento MELI ya registrado, consulta el shipment y mueve el stock
    de bodega CENTRAL a MELI_FULL si era Full (y viceversa).

    Útil para corregir descuentos hechos con el detector viejo que no consultaba shipments."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from mercadolibre import obtener_orden_meli
        from inventario import (get_conn, ajustar_stock_bodega, get_stock_bodega,
                                determinar_bodega_para_canal)
        from bodegas_logic import detectar_fulfillment_meli
        registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                        "reclasificar_bodegas_meli",
                        detalle="Re-clasificar ventas MELI por bodega correcta")

        max_ordenes = int(request.args.get("max", 100))
        dry_run = request.args.get("dry_run", "0") == "1"
        log = []
        movidas = 0
        ya_correctas = 0
        sin_orden = 0
        errores = []

        # Buscar movimientos MELI únicos por (orden_id, sku) tomando el más reciente
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""SELECT orden_id, sku, cantidad, bodega_codigo
                       FROM (
                           SELECT orden_id, sku, cantidad, bodega_codigo, id,
                                  ROW_NUMBER() OVER (PARTITION BY orden_id, sku ORDER BY id DESC) AS rn
                           FROM movimientos
                           WHERE canal = 'MercadoLibre'
                             AND tipo = 'salida'
                             AND orden_id IS NOT NULL
                             AND orden_id != ''
                       ) sub
                       WHERE rn = 1
                       ORDER BY id DESC
                       LIMIT %s""", (max_ordenes,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        log.append(f"Revisando {len(rows)} movimientos MELI")

        for orden_id, sku, cantidad, bodega_actual in rows:
            try:
                orden = obtener_orden_meli(orden_id)
                if not orden:
                    sin_orden += 1
                    continue

                es_full = detectar_fulfillment_meli(orden)
                bodega_correcta = determinar_bodega_para_canal("MercadoLibre", fulfillment=es_full)

                if bodega_actual == bodega_correcta:
                    ya_correctas += 1
                    continue

                # Hay que mover: reintegrar a la bodega vieja, descontar de la nueva
                if dry_run:
                    log.append(f"[DRY-RUN] {orden_id} {sku} ({cantidad}u): {bodega_actual} → {bodega_correcta} ({'Full' if es_full else 'Seller'})")
                else:
                    # Reintegrar a bodega actual (deshacer descuento original)
                    ajustar_stock_bodega(sku, bodega_actual, cantidad)
                    # Descontar de bodega correcta
                    ajustar_stock_bodega(sku, bodega_correcta, -cantidad)

                    # Actualizar el bodega_codigo en el movimiento
                    conn2 = get_conn(); cur2 = conn2.cursor()
                    cur2.execute("""UPDATE movimientos SET bodega_codigo=%s
                                   WHERE canal='MercadoLibre' AND tipo='salida'
                                     AND orden_id=%s AND sku=%s""",
                                (bodega_correcta, orden_id, sku))
                    conn2.commit()
                    cur2.close(); conn2.close()
                    log.append(f"✓ {orden_id} {sku} ({cantidad}u): {bodega_actual} → {bodega_correcta}")
                movidas += 1
            except Exception as e:
                errores.append(f"{orden_id}: {str(e)}")

        return jsonify({
            "ok": True,
            "total_revisadas": len(rows),
            "ya_correctas": ya_correctas,
            "movidas": movidas,
            "sin_orden_en_meli": sin_orden,
            "dry_run": dry_run,
            "errores": errores[:10],
            "log": log[:60]
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/mercadolibre/forzar_sync_todos", methods=["POST"])
def ruta_meli_forzar_todos():
    """Re-envía el stock de todos los SKUs mapeados a MercadoLibre."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        from mercadolibre import actualizar_stock_meli
        productos = {p["sku"]: p for p in cargar_productos()}
        mapeo = listar_sku_mapeo()
        enviados = 0
        fallidos = 0
        for fila in mapeo:
            sku_lusync = fila.get("sku_lusync", "")
            sku_meli = (fila.get("sku_mercadolibre", "") or "").strip()
            if not sku_meli or sku_lusync not in productos:
                continue
            stock = productos[sku_lusync]["stock"]
            if actualizar_stock_meli(sku_lusync, stock):
                enviados += 1
            else:
                fallidos += 1
        return jsonify({"ok": True, "enviados": enviados, "fallidos": fallidos})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/mercadolibre/webhook", methods=["GET", "POST"])
def ruta_meli_webhook():
    """Endpoint receptor de TODAS las notificaciones de MercadoLibre.
    
    MELI envía todas las notificaciones (orders_v2, fbm_stock_operations, items, etc.)
    a esta URL única configurada en DevCenter. Aquí ruteamos según el topic.
    """
    # GET solo para que MELI valide que la URL responde
    if request.method == "GET":
        return jsonify({"status": "ok"}), 200
    
    try:
        payload = request.json or {}
        topic = (payload.get("topic") or "").lower()
        
        # ── ROUTING POR TOPIC ──
        if topic in ("orders_v2", "orders"):
            # Órdenes de venta y cancelaciones (lo que ya teníamos)
            from mercadolibre import procesar_webhook_meli
            ok = procesar_webhook_meli(payload)
        elif topic in ("fbm_stock_operations", "marketplace_fbm_stock"):
            # Cambios de stock en MELI Full (inbound, damaged, lost, returns)
            try:
                ok = procesar_webhook_fbm(payload)
            except Exception as e:
                import traceback
                print(f"[FBM Webhook] Error: {e}")
                print(traceback.format_exc())
                ok = False
        else:
            # Otros topics (items, questions, claims, etc.) → loguear pero no procesar
            print(f"[MELI Webhook] Topic '{topic}' recibido, no procesado (resource: {payload.get('resource', '?')})")
            ok = True
        
        # MELI espera 200 rápido; si tarda mucho reintenta
        return jsonify({"ok": ok}), 200
    except Exception as e:
        print(f"[MELI Webhook] Error general: {e}")
        # Importante: devolver 200 igual para que MELI no reintente infinito
        return jsonify({"ok": False, "error": str(e)}), 200


# ════════════════════════════════════════════════════════════════════════════
# WEBHOOK FBM (FULFILLMENT BY MERCADOLIBRE) - OPERACIONES DE STOCK FULL
# ════════════════════════════════════════════════════════════════════════════
# MELI notifica cambios en el stock Full vía topic "fbm_stock_operations".
# Tipos de operación que procesa este webhook:
#
#   📦 inbound_reception  → MELI recibió tu envío. Mover MELI_FULL_TRANSITO → MELI_FULL
#   🔄 inbound_returns    → Devolución del comprador llegó a MELI. Sumar a MELI_FULL
#   ⚠️ damaged            → Mercadería dañada en bodega MELI. Restar de MELI_FULL + alerta
#   ⚠️ lost               → Mercadería extraviada. Restar de MELI_FULL + alerta
#
# NO procesa sale_confirmation (eso ya lo maneja /mercadolibre/webhook con orders_v2)
#
# Idempotencia: cada operación tiene operation_id único.
# Si MELI manda 2 veces la misma, solo procesa la primera.
# ════════════════════════════════════════════════════════════════════════════

@app.route("/mercadolibre/webhook_fbm", methods=["GET", "POST"])
def ruta_meli_webhook_fbm():
    """Webhook receptor de notificaciones FBM (stock Full)."""
    # GET solo para que MELI valide que la URL responde
    if request.method == "GET":
        return jsonify({"status": "ok", "endpoint": "fbm_stock_operations"}), 200
    
    try:
        payload = request.json or {}
        # Procesar de forma resiliente: si algo falla, devolver 200 igual
        # (MELI reintenta cada 5min hasta 7 días si recibe error)
        try:
            ok = procesar_webhook_fbm(payload)
        except Exception as e:
            import traceback
            print(f"[FBM Webhook] Error procesando: {e}")
            print(traceback.format_exc())
            ok = False
        
        return jsonify({"ok": ok}), 200
    except Exception as e:
        print(f"[FBM Webhook] Error general: {e}")
        return jsonify({"ok": False, "error": str(e)}), 200


def procesar_webhook_fbm(payload):
    """Procesa una notificación FBM de cambio de stock Full.
    
    Estructura típica del payload (según docs MELI):
    {
      "topic": "fbm_stock_operations",
      "resource": "/inventories/{INVENTORY_ID}/stock/fulfillment/operations/{OPERATION_ID}",
      "user_id": 123456789,
      "application_id": 12345,
      "sent": "2026-05-05T10:00:00Z",
      "received": "2026-05-05T10:00:00Z"
    }
    
    Procesa:
    - inbound_reception: TRANSITO → FULL
    - damaged/lost: resta FULL + alerta
    - inbound_returns: suma FULL
    """
    from inventario import (orden_ya_procesada_texto, marcar_orden_procesada_texto,
                            ajustar_stock_bodega, get_stock_bodega, crear_alerta,
                            cargar_productos as _cp)
    from mercadolibre import get_meli_token
    import requests as _req
    
    topic = payload.get("topic", "")
    resource = payload.get("resource", "")
    
    print(f"[FBM Webhook] Recibido: topic={topic} resource={resource}")
    
    if topic != "fbm_stock_operations":
        print(f"[FBM Webhook] Topic no relevante: {topic} - ignorando")
        return True
    
    # Extraer operation_id del resource: /inventories/{INV}/stock/fulfillment/operations/{OP_ID}
    operation_id = ""
    inventory_id = ""
    try:
        parts = resource.strip("/").split("/")
        # parts: ["inventories", "{INV}", "stock", "fulfillment", "operations", "{OP_ID}"]
        if len(parts) >= 6:
            inventory_id = parts[1]
            operation_id = parts[5]
    except Exception as e:
        print(f"[FBM Webhook] No se pudo parsear resource: {e}")
        return False
    
    if not operation_id:
        print(f"[FBM Webhook] operation_id vacío, ignorando")
        return False
    
    # ── Idempotencia ATÓMICA: si otra request ya procesó esta operación, salir ──
    fbm_key = f"FBM-OP-{operation_id}"
    from inventario import intentar_marcar_orden_atomic
    if not intentar_marcar_orden_atomic(fbm_key):
        print(f"[FBM Webhook] Operación {operation_id} ya procesada (atomic)")
        return True
    
    # ── Consultar el detalle de la operación a MELI ──
    # Necesitamos saber: tipo, SKU, cantidad
    try:
        token = get_meli_token()
        url = f"https://api.mercadolibre.com{resource}"
        r = _req.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
        if r.status_code != 200:
            print(f"[FBM Webhook] HTTP {r.status_code} consultando operación {operation_id}")
            print(f"[FBM Webhook] Body: {r.text[:200]}")
            return False
        operacion = r.json()
    except Exception as e:
        print(f"[FBM Webhook] Error consultando detalle de operación: {e}")
        return False
    
    # ── Extraer datos de la operación ──
    # Estructura típica del response:
    # {
    #   "id": "operation_id",
    #   "type": "inbound_reception" | "damaged" | "lost" | "sale_confirmation" | ...
    #   "date_created": "...",
    #   "items": [
    #     {"sku": "MLBxxx", "quantity": 5, "seller_sku": "ODJM001", ...}
    #   ]
    # }
    
    op_type = (operacion.get("type") or operacion.get("operation_type") or "").lower()
    items = operacion.get("items", []) or []
    
    print(f"[FBM Webhook] Tipo: {op_type}, Items: {len(items)}")
    
    # ── Filtrar tipos que NO procesa este webhook ──
    # sale_confirmation lo maneja /mercadolibre/webhook con orders_v2
    if op_type in ("sale_confirmation", "sale", "outbound"):
        print(f"[FBM Webhook] Tipo {op_type} es manejado por webhook orders_v2, ignorando")
        marcar_orden_procesada_texto(fbm_key)  # Marcar para no re-evaluarla
        return True
    
    # ── Procesar según tipo ──
    items_procesados = []
    
    for item in items:
        # MELI puede usar varios nombres para el SKU del seller
        seller_sku = (
            (item.get("seller_sku") or "").strip()
            or (item.get("seller_custom_field") or "").strip()
            or (item.get("sku") or "").strip()
        )
        cantidad = int(item.get("quantity") or item.get("qty") or 0)
        
        if not seller_sku or cantidad <= 0:
            continue
        
        # Buscar SKU Lusync vía mapeo (por si MELI usa item_id en vez de SKU)
        sku_lusync = seller_sku
        try:
            from inventario import listar_sku_mapeo
            for fila in listar_sku_mapeo():
                sku_meli_mapped = (fila.get("sku_mercadolibre") or "").strip()
                if sku_meli_mapped == seller_sku:
                    sku_lusync = fila.get("sku_lusync") or seller_sku
                    break
        except: pass
        
        # Verificar que el SKU exista en productos
        productos = _cp()
        producto = next((p for p in productos if p["sku"] == sku_lusync), None)
        if not producto:
            print(f"[FBM Webhook] SKU '{sku_lusync}' no existe en Lusync, saltando")
            continue
        
        # ╔══════════════════════════════════════════════════════════╗
        # ║ INBOUND_RECEPTION → MELI recibió tu envío                ║
        # ║ Mover de MELI_FULL_TRANSITO a MELI_FULL                  ║
        # ╚══════════════════════════════════════════════════════════╝
        if op_type == "inbound_reception":
            try:
                # Obtener stock actual de TRANSITO
                stock_transito_actual = get_stock_bodega(sku_lusync, "MELI_FULL_TRANSITO") or 0
                # Cantidad a mover: mínimo entre lo recibido y lo que está en tránsito
                # Esto evita números negativos si hay desfase con el reporte Excel
                cantidad_a_mover = min(cantidad, stock_transito_actual) if stock_transito_actual > 0 else cantidad
                
                # Restar de TRANSITO (si había)
                if stock_transito_actual > 0:
                    ajustar_stock_bodega(sku_lusync, "MELI_FULL_TRANSITO", -cantidad_a_mover)
                
                # Sumar a FULL
                ajustar_stock_bodega(sku_lusync, "MELI_FULL", cantidad)
                
                # Registrar movimiento
                registrar_movimiento(
                    "entrada", sku_lusync, producto["nombre"], cantidad,
                    f"Ingreso a MELI Full confirmado (op {operation_id})",
                    usuario="Sistema (Webhook FBM)", canal="MercadoLibre",
                    orden_id=operation_id
                )
                
                # Crear alerta visible en panel
                try:
                    crear_alerta(
                        tipo="full_ingreso",
                        titulo=f"📦 {sku_lusync} ingresó a MELI Full",
                        mensaje=f"<strong>{cantidad}</strong> unidades de <strong>{producto['nombre']}</strong> ya están aptas para vender en MELI Full. Operación: {operation_id}",
                        sku=sku_lusync, canal="mercadolibre"
                    )
                except: pass
                
                items_procesados.append(f"{sku_lusync} +{cantidad} a Full (de tránsito)")
                print(f"[FBM Webhook] inbound_reception: {sku_lusync} +{cantidad} → MELI_FULL")
            except Exception as e:
                print(f"[FBM Webhook] Error inbound_reception {sku_lusync}: {e}")
        
        # ╔══════════════════════════════════════════════════════════╗
        # ║ DAMAGED / LOST → Mercadería perdida o dañada en MELI    ║
        # ║ Restar de MELI_FULL + alerta de pérdida                  ║
        # ╚══════════════════════════════════════════════════════════╝
        elif op_type in ("damaged", "lost", "loss", "destruction"):
            try:
                ajustar_stock_bodega(sku_lusync, "MELI_FULL", -cantidad)
                
                registrar_movimiento(
                    "salida", sku_lusync, producto["nombre"], cantidad,
                    f"Mercadería {op_type} en MELI Full (op {operation_id})",
                    usuario="Sistema (Webhook FBM)", canal="MercadoLibre",
                    orden_id=operation_id
                )
                
                try:
                    crear_alerta(
                        tipo="full_perdida",
                        titulo=f"⚠️ Pérdida en MELI Full: {sku_lusync}",
                        mensaje=f"MELI reportó <strong>{cantidad}</strong> unidades de <strong>{producto['nombre']}</strong> como <strong>{op_type}</strong>. Operación: {operation_id}",
                        sku=sku_lusync, canal="mercadolibre"
                    )
                except: pass
                
                items_procesados.append(f"{sku_lusync} -{cantidad} ({op_type})")
                print(f"[FBM Webhook] {op_type}: {sku_lusync} -{cantidad} de MELI_FULL")
            except Exception as e:
                print(f"[FBM Webhook] Error {op_type} {sku_lusync}: {e}")
        
        # ╔══════════════════════════════════════════════════════════╗
        # ║ INBOUND_RETURNS → Devolución del comprador               ║
        # ║ Sumar a MELI_FULL (vuelve a estar disponible)            ║
        # ╚══════════════════════════════════════════════════════════╝
        elif op_type in ("inbound_returns", "return", "customer_return"):
            try:
                ajustar_stock_bodega(sku_lusync, "MELI_FULL", cantidad)
                
                registrar_movimiento(
                    "entrada", sku_lusync, producto["nombre"], cantidad,
                    f"Devolución cliente MELI Full (op {operation_id})",
                    usuario="Sistema (Webhook FBM)", canal="MercadoLibre",
                    orden_id=operation_id
                )
                
                try:
                    crear_alerta(
                        tipo="full_devolucion",
                        titulo=f"🔄 Devolución MELI Full: {sku_lusync}",
                        mensaje=f"<strong>{cantidad}</strong> unidades de <strong>{producto['nombre']}</strong> volvieron a Full por devolución. Operación: {operation_id}",
                        sku=sku_lusync, canal="mercadolibre"
                    )
                except: pass
                
                items_procesados.append(f"{sku_lusync} +{cantidad} (devolución)")
                print(f"[FBM Webhook] inbound_returns: {sku_lusync} +{cantidad} → MELI_FULL")
            except Exception as e:
                print(f"[FBM Webhook] Error inbound_returns {sku_lusync}: {e}")
        
        # ╔══════════════════════════════════════════════════════════╗
        # ║ Otros tipos: solo loguear (no procesar)                  ║
        # ╚══════════════════════════════════════════════════════════╝
        else:
            print(f"[FBM Webhook] Tipo {op_type} no procesado (no implementado)")
    
    # Marcar operación como procesada (idempotencia)
    marcar_orden_procesada_texto(fbm_key)
    
    print(f"[FBM Webhook] OK — Operación {operation_id} ({op_type}): {len(items_procesados)} items procesados")
    return True


@app.route("/debug/meli_test_conexion")
def debug_meli_test():
    """Diagnóstico de conexión con MercadoLibre."""
    if not session.get("logged"): return redirect("/")
    try:
        from mercadolibre import verificar_conexion_meli
        import os
        return jsonify({
            "app_id_configurado": bool(os.environ.get("MERCADOLIBRE_APP_ID")),
            "secret_configurado": bool(os.environ.get("MERCADOLIBRE_CLIENT_SECRET")),
            "redirect_uri": os.environ.get("MERCADOLIBRE_REDIRECT_URI", "(default)"),
            "estado": verificar_conexion_meli()
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── ALERTAS ─────────────────────────────────────────────────────────────────

@app.route("/alertas")
def ruta_alertas():
    if not session.get("logged"): return redirect("/")
    solo_no = request.args.get("solo_no_leidas", "false").lower() == "true"
    return jsonify(listar_alertas(limite=100, solo_no_leidas=solo_no))

@app.route("/alertas/contador")
def ruta_alertas_contador():
    if not session.get("logged"): return jsonify({"count": 0})
    return jsonify({"count": contar_alertas_no_leidas()})

@app.route("/alertas/leer/<int:alerta_id>", methods=["POST"])
def ruta_alerta_leer(alerta_id):
    if not session.get("logged"): return jsonify({"ok": False}), 401
    marcar_alerta_leida(alerta_id)
    return jsonify({"ok": True})

@app.route("/alertas/leer_todas", methods=["POST"])
def ruta_alertas_leer_todas():
    if not session.get("logged"): return jsonify({"ok": False}), 401
    marcar_todas_leidas()
    return jsonify({"ok": True})

@app.route("/alertas/config", methods=["GET", "POST"])
def ruta_alertas_config():
    if not session.get("logged"): return jsonify({}), 401
    if request.method == "POST":
        data = request.json or {}
        # Filtrar solo claves válidas
        permitidas = {"smtp_host","smtp_port","smtp_user","smtp_password",
                      "smtp_from","destinatarios","notif_cancelaciones","notif_errores_api"}
        filtered = {k: v for k, v in data.items() if k in permitidas}
        set_alertas_config(filtered)
        return jsonify({"ok": True})
    cfg = get_alertas_config()
    # No exponer la contraseña SMTP en GET (solo si está configurada)
    if cfg.get("smtp_password"):
        cfg["smtp_password"] = "********"
    return jsonify(cfg)

@app.route("/alertas/test_email", methods=["POST"])
def ruta_alertas_test():
    """Envía un email de prueba con la configuración actual."""
    if not session.get("logged"): return jsonify({"ok": False}), 401
    try:
        crear_alerta(
            tipo="test",
            canal="Sistema",
            titulo="Email de prueba Lusync",
            mensaje="Si recibes este correo, las notificaciones por email están funcionando correctamente."
        )
        return jsonify({"ok": True, "mensaje": "Alerta creada y email enviado (revisa configuración SMTP en logs si no llega)"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO DE TRAZABILIDAD DE FECHAS (temporal/debug)
# ════════════════════════════════════════════════════════════════════════════

@app.route("/debug/movimientos_trazabilidad")
def debug_movimientos_trazabilidad():
    """Muestra los últimos 30 movimientos con TODOS los campos de trazabilidad.
    Sirve para verificar si fecha_compra_marketplace se está guardando correctamente.
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn(); cur = conn.cursor()

        # Asegurar columnas
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS fecha_compra_marketplace TIMESTAMP")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen_registro TEXT DEFAULT 'sistema'")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_antes INTEGER")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_despues INTEGER")
        conn.commit()

        # Resumen agregado
        cur.execute("""
            SELECT COALESCE(canal, '?') as canal,
                   COALESCE(bodega_codigo, 'CENTRAL') as bodega,
                   COALESCE(origen_registro, 'sistema') as origen,
                   COUNT(*) as total,
                   COUNT(fecha_compra_marketplace) as con_fecha_real,
                   COUNT(fecha_importacion) as con_fecha_import,
                   COUNT(stock_antes) as con_stock_antes,
                   TO_CHAR(MIN(fecha), 'DD/MM HH24:MI') as primer_mov,
                   TO_CHAR(MAX(fecha), 'DD/MM HH24:MI') as ultimo_mov
            FROM movimientos
            WHERE fecha > NOW() - INTERVAL '7 days'
            GROUP BY canal, bodega_codigo, origen_registro
            ORDER BY total DESC
        """)
        resumen = []
        for r in cur.fetchall():
            resumen.append({
                "canal": r[0], "bodega": r[1], "origen": r[2],
                "total": r[3], "con_fecha_real": r[4],
                "con_fecha_import": r[5], "con_stock_antes": r[6],
                "primer": r[7], "ultimo": r[8],
                "porcentaje_con_fecha_real": round((r[4]/r[3])*100, 1) if r[3] > 0 else 0
            })

        # Detalle de los últimos 30 movimientos
        cur.execute("""
            SELECT id, tipo, sku, canal, bodega_codigo, orden_id,
                   TO_CHAR(fecha, 'DD/MM/YYYY HH24:MI') as fecha_mov,
                   TO_CHAR(fecha_importacion, 'DD/MM/YYYY HH24:MI') as fecha_imp,
                   TO_CHAR(fecha_compra_marketplace, 'DD/MM/YYYY HH24:MI') as fecha_compra,
                   origen_registro, stock_antes, stock_despues, cantidad
            FROM movimientos
            ORDER BY fecha DESC
            LIMIT 30
        """)
        detalles = []
        for r in cur.fetchall():
            detalles.append({
                "id": r[0], "tipo": r[1], "sku": r[2], "canal": r[3],
                "bodega": r[4], "orden_id": r[5],
                "fecha_mov": r[6], "fecha_import": r[7],
                "fecha_compra_marketplace": r[8],
                "origen": r[9],
                "stock_antes": r[10], "stock_despues": r[11],
                "cantidad": r[12]
            })

        cur.close(); conn.close()

        # Diagnóstico automático
        diagnostico = []
        total_recientes = sum(r["total"] for r in resumen)
        total_con_fecha = sum(r["con_fecha_real"] for r in resumen)

        if total_recientes == 0:
            diagnostico.append("⚠ No hay movimientos en los últimos 7 días")
        elif total_con_fecha == 0:
            diagnostico.append("✗ NINGÚN movimiento tiene fecha_compra_marketplace. El parsing/guardado está fallando.")
        elif total_con_fecha < total_recientes:
            faltantes = total_recientes - total_con_fecha
            diagnostico.append(f"⚠ {faltantes}/{total_recientes} movimientos sin fecha_compra_marketplace")
        else:
            diagnostico.append(f"✓ Todos los {total_recientes} movimientos recientes tienen fecha de compra real")

        # Detectar si todavía hay movimientos sin bodega_codigo
        sin_bodega = [r for r in resumen if not r["bodega"] or r["bodega"] == "CENTRAL" and "FULL" not in str(r)]

        return jsonify({
            "diagnostico": diagnostico,
            "resumen_por_canal": resumen,
            "ultimos_30_movimientos": detalles
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/payload_orden_meli/<order_id>")
def debug_payload_orden_meli(order_id):
    """Muestra el payload crudo de una orden MELI específica para verificar
    qué campos vienen y poder diagnosticar el parsing de date_created."""
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from mercadolibre import obtener_orden_meli
        from datetime import datetime as _dt
        orden = obtener_orden_meli(order_id)
        if not orden:
            return jsonify({"error": f"Orden {order_id} no encontrada"}), 404

        # Intentar parsear date_created como lo hace el sync
        date_str = orden.get("date_created", "") or ""
        parseo_result = {
            "date_created_raw": date_str,
            "tipo": str(type(date_str).__name__),
            "es_string_vacio": date_str == "",
            "longitud": len(str(date_str))
        }
        try:
            if date_str:
                date_str_clean = date_str.replace("Z", "+00:00")
                fecha = _dt.fromisoformat(date_str_clean)
                parseo_result["parseado_ok"] = True
                parseo_result["datetime_parseado"] = str(fecha)
                parseo_result["tiene_tzinfo"] = fecha.tzinfo is not None
        except Exception as e:
            parseo_result["parseado_ok"] = False
            parseo_result["error"] = str(e)

        return jsonify({
            "order_id": order_id,
            "campos_fecha_disponibles": {
                "date_created": orden.get("date_created"),
                "date_closed": orden.get("date_closed"),
                "last_updated": orden.get("last_updated")
            },
            "diagnostico_parseo": parseo_result,
            "shipping_id": (orden.get("shipping") or {}).get("id"),
            "status": orden.get("status"),
            "total_items": len(orden.get("order_items", []))
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/debug/test_parseo_fechas")
def debug_test_parseo_fechas():
    """Test directo del código de parseo con strings de ejemplo de cada marketplace."""
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    from datetime import datetime as _dt
    import pytz as _pytz
    TZ = _pytz.timezone('America/Santiago')

    casos = [
        ("MELI con tz Chile", "2026-05-03T18:32:15.000-04:00"),
        ("MELI con tz UTC", "2026-05-03T22:32:15.000Z"),
        ("Paris UTC con Z", "2026-05-03T05:02:00.000Z"),
        ("Ripley UTC con Z", "2026-05-03T18:32:15Z"),
        ("Falabella sin tz", "2026-05-03 14:32:15"),
        ("Falabella +0000", "2026-05-03T14:32:15+0000")
    ]
    resultados = []
    for nombre, date_str in casos:
        item = {"nombre": nombre, "input": date_str}
        try:
            try:
                fecha = _dt.fromisoformat(date_str.replace("Z", "+00:00"))
                item["parseado"] = str(fecha)
                if fecha.tzinfo:
                    fecha_chile = fecha.astimezone(TZ).replace(tzinfo=None)
                    item["en_chile"] = str(fecha_chile)
                else:
                    item["en_chile"] = "sin tz, no se convierte"
            except ValueError:
                # Fallback Falabella sin T
                fecha_naive = _dt.strptime(date_str.strip(), "%Y-%m-%d %H:%M:%S")
                fecha = _pytz.utc.localize(fecha_naive)
                fecha_chile = fecha.astimezone(TZ).replace(tzinfo=None)
                item["parseado"] = str(fecha) + " (UTC asumido)"
                item["en_chile"] = str(fecha_chile)
            item["ok"] = True
        except Exception as e:
            item["ok"] = False
            item["error"] = str(e)
        resultados.append(item)
    return jsonify({"resultados": resultados})


# ════════════════════════════════════════════════════════════════════════════
# RECONSTRUCCIÓN DE FECHAS DE COMPRA HISTÓRICAS
# ════════════════════════════════════════════════════════════════════════════
# Endpoint admin para rellenar la columna fecha_compra_marketplace de los
# movimientos antiguos consultando los APIs de cada marketplace.
#
# Estrategia: traer todas las órdenes de los últimos N días una sola vez por
# canal (mucho más eficiente que consultar de a una), construir un diccionario
# {orden_id: fecha_compra} en memoria, y luego hacer UN UPDATE por movimiento.
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/reconstruir_fechas_compra", methods=["GET", "POST"])
def admin_reconstruir_fechas_compra():
    """Reconstruye fecha_compra_marketplace para movimientos antiguos.

    Query string:
      ?canales=mercadolibre,paris,walmart,falabella   (default: los 4)
      ?dias=30                                         (default: 30)
      ?dry_run=1                                       (default: 0; si 1 no escribe)

    Usa los APIs de cada marketplace para obtener la fecha real de compra,
    luego hace UPDATE en bloque por orden_id.
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    from datetime import datetime as _dt
    import pytz as _pytz
    TZ_CHILE = _pytz.timezone('America/Santiago')

    canales_str = request.args.get("canales", "mercadolibre,paris,walmart,falabella")
    canales_pedidos = set(c.strip().lower() for c in canales_str.split(",") if c.strip())
    dias = int(request.args.get("dias", 30))
    dry_run = request.args.get("dry_run", "0") == "1"

    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "reconstruir_fechas_compra",
                    detalle=f"canales={canales_pedidos} dias={dias} dry_run={dry_run}")

    log = []
    log.append(f"Configuración: canales={list(canales_pedidos)}, dias={dias}, dry_run={dry_run}")

    # ── 1) Cargar movimientos pendientes de la BD ──
    from inventario import get_conn
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT id, canal, orden_id
        FROM movimientos
        WHERE fecha_compra_marketplace IS NULL
          AND orden_id IS NOT NULL AND orden_id != ''
          AND fecha > NOW() - (%s || ' days')::INTERVAL
        ORDER BY fecha DESC
    """, (dias,))
    pendientes = cur.fetchall()
    cur.close(); conn.close()

    log.append(f"Movimientos pendientes en BD: {len(pendientes)}")

    # Agrupar pendientes por canal normalizado
    pendientes_por_canal = {}
    for mov_id, canal, orden_id in pendientes:
        canal_norm = (canal or "").lower()
        # Normalizar nombres de canales
        if "mercadolibre" in canal_norm or "meli" in canal_norm:
            canal_norm = "mercadolibre"
        elif "paris" in canal_norm or "parís" in canal_norm:
            canal_norm = "paris"
        elif "walmart" in canal_norm or "wfs" in canal_norm:
            canal_norm = "walmart"
        elif "falabella" in canal_norm:
            canal_norm = "falabella"
        else:
            continue  # Manual, WooCommerce, Sistema, etc → no se reconstruye desde API
        if canal_norm not in canales_pedidos:
            continue
        pendientes_por_canal.setdefault(canal_norm, []).append((mov_id, str(orden_id)))

    for c, items in pendientes_por_canal.items():
        log.append(f"  {c}: {len(items)} movimientos pendientes")

    # ── 2) Para cada canal, traer las órdenes y armar mapa orden_id → fecha ──
    fecha_por_orden = {}  # {(canal_norm, orden_id): datetime con tz}

    # MELI: consultar orden por orden (no hay endpoint que filtre por order_id en bloque útil)
    if "mercadolibre" in pendientes_por_canal:
        ids_meli = pendientes_por_canal["mercadolibre"]
        log.append(f"[MELI] Consultando {len(set(o for _, o in ids_meli))} órdenes únicas...")
        try:
            from mercadolibre import obtener_orden_meli
            ya_consultados = set()
            for mov_id, orden_id in ids_meli:
                if orden_id in ya_consultados:
                    continue
                ya_consultados.add(orden_id)
                try:
                    orden = obtener_orden_meli(orden_id)
                    if orden:
                        date_str = orden.get("date_created", "") or ""
                        if date_str:
                            try:
                                fecha = _dt.fromisoformat(date_str.replace("Z", "+00:00"))
                                fecha_por_orden[("mercadolibre", orden_id)] = fecha
                            except Exception as e:
                                log.append(f"  MELI {orden_id}: error parseando '{date_str}': {e}")
                except Exception as e:
                    log.append(f"  MELI {orden_id}: error consultando API: {e}")
            log.append(f"[MELI] Fechas obtenidas: {len([k for k in fecha_por_orden if k[0]=='mercadolibre'])}")
        except Exception as e:
            log.append(f"[MELI] ERROR general: {e}")

    # PARÍS: traer todas las órdenes recientes y filtrar
    if "paris" in pendientes_por_canal:
        log.append(f"[Paris] Trayendo órdenes últimos {dias} días...")
        try:
            from paris import obtener_ordenes_paris_todas
            ordenes = obtener_ordenes_paris_todas(dias=dias, estado=None)
            log.append(f"[Paris] {len(ordenes)} órdenes obtenidas")
            for so in ordenes:
                sub_order_num = str(so.get("subOrderNumber", "") or "")
                if not sub_order_num:
                    continue
                date_str = (so.get("createdAt") or so.get("created_at") or "")
                if date_str:
                    try:
                        fecha = _dt.fromisoformat(date_str.replace("Z", "+00:00"))
                        fecha_por_orden[("paris", sub_order_num)] = fecha
                    except Exception as e:
                        pass
            log.append(f"[Paris] Fechas obtenidas: {len([k for k in fecha_por_orden if k[0]=='paris'])}")
        except Exception as e:
            log.append(f"[Paris] ERROR: {e}")
            import gc; gc.collect()

    # WALMART: traer órdenes recientes (todos los estados que conocemos)
    if "walmart" in pendientes_por_canal:
        log.append(f"[Walmart] Trayendo órdenes últimos {dias} días...")
        try:
            from walmart import obtener_ordenes_walmart
            ya_consultados = set()
            # Walmart requiere consultar por estado. Iteramos los estados activos.
            for estado in ["Created", "Acknowledged", "Shipped", "Delivered"]:
                try:
                    ords_walmart = obtener_ordenes_walmart(estado=estado, max_paginas=2, limit=50, dias=dias)
                    log.append(f"[Walmart] {estado}: {len(ords_walmart)} órdenes")
                    for o in ords_walmart:
                        po = str(o.get("purchaseOrderId", "") or "")
                        co = str(o.get("customerOrderId", "") or po)
                        # En BD usamos customerOrderId como orden_id
                        if co in ya_consultados:
                            continue
                        ya_consultados.add(co)
                        # Walmart trae orderDate como timestamp en ms (epoch)
                        order_date_raw = o.get("orderDate")
                        fecha = None
                        if order_date_raw:
                            try:
                                # Si es int (epoch ms), convertir
                                if isinstance(order_date_raw, (int, float)):
                                    fecha = _dt.fromtimestamp(order_date_raw / 1000, tz=_pytz.utc)
                                elif isinstance(order_date_raw, str):
                                    if order_date_raw.isdigit():
                                        fecha = _dt.fromtimestamp(int(order_date_raw) / 1000, tz=_pytz.utc)
                                    else:
                                        fecha = _dt.fromisoformat(order_date_raw.replace("Z", "+00:00"))
                            except Exception as e:
                                log.append(f"  Walmart {co}: error parseando orderDate '{order_date_raw}': {e}")
                        if fecha:
                            fecha_por_orden[("walmart", co)] = fecha
                            # Walmart en BD a veces usa purchaseOrderId, también guardamos por po
                            if po and po != co:
                                fecha_por_orden[("walmart", po)] = fecha
                except Exception as e:
                    log.append(f"[Walmart] estado {estado}: error {e}")
            log.append(f"[Walmart] Fechas obtenidas: {len([k for k in fecha_por_orden if k[0]=='walmart'])}")
            import gc; gc.collect()
        except Exception as e:
            log.append(f"[Walmart] ERROR general: {e}")

    # FALABELLA: traer órdenes recientes (varios estados)
    if "falabella" in pendientes_por_canal:
        log.append(f"[Falabella] Trayendo órdenes últimos {dias} días...")
        try:
            from falabella import obtener_ordenes_falabella
            ya_consultados = set()
            for estado in [None, "pending", "ready_to_ship", "shipped", "delivered"]:
                try:
                    ords_fal = obtener_ordenes_falabella(estado=estado, dias=dias, limit=100)
                    log.append(f"[Falabella] {estado or 'todos'}: {len(ords_fal)} órdenes")
                    for o in ords_fal:
                        oid = str(o.get("OrderId") or o.get("OrderNumber") or "")
                        if not oid or oid in ya_consultados:
                            continue
                        ya_consultados.add(oid)
                        date_str = (o.get("CreatedAt") or o.get("created_at") or "")
                        if date_str:
                            try:
                                fecha = _dt.fromisoformat(date_str.replace("Z", "+00:00"))
                            except ValueError:
                                try:
                                    fecha_naive = _dt.strptime(date_str.strip(), "%Y-%m-%d %H:%M:%S")
                                    fecha = _pytz.utc.localize(fecha_naive)
                                except Exception as e:
                                    log.append(f"  Falabella {oid}: error parseando '{date_str}': {e}")
                                    continue
                            fecha_por_orden[("falabella", oid)] = fecha
                except Exception as e:
                    log.append(f"[Falabella] estado {estado}: error {e}")
            log.append(f"[Falabella] Fechas obtenidas: {len([k for k in fecha_por_orden if k[0]=='falabella'])}")
            import gc; gc.collect()
        except Exception as e:
            log.append(f"[Falabella] ERROR general: {e}")

    # ── 3) Hacer UPDATE en BD por cada movimiento que tenga match ──
    actualizados = 0
    sin_match = 0
    errores_db = 0
    detalle_no_encontrados = []

    if dry_run:
        log.append("DRY RUN: no se escribirá nada en BD")

    conn = get_conn(); cur = conn.cursor()
    for canal_norm, lista in pendientes_por_canal.items():
        for mov_id, orden_id in lista:
            fecha = fecha_por_orden.get((canal_norm, orden_id))
            if fecha is None:
                sin_match += 1
                if len(detalle_no_encontrados) < 30:
                    detalle_no_encontrados.append(f"{canal_norm}:{orden_id} (mov_id={mov_id})")
                continue
            # Convertir a Chile sin tz para guardar
            try:
                if hasattr(fecha, 'tzinfo') and fecha.tzinfo:
                    fecha_chile = fecha.astimezone(TZ_CHILE).replace(tzinfo=None)
                else:
                    fecha_chile = fecha
                if not dry_run:
                    cur.execute("""UPDATE movimientos
                                   SET fecha_compra_marketplace = %s
                                   WHERE id = %s""", (fecha_chile, mov_id))
                actualizados += 1
            except Exception as e:
                errores_db += 1
                log.append(f"  ERR DB mov_id={mov_id}: {e}")

    if not dry_run:
        try:
            conn.commit()
        except Exception as e:
            conn.rollback()
            log.append(f"COMMIT ERROR: {e}")
            errores_db += 1
    cur.close(); conn.close()

    # ── 4) Verificar cuántos quedan pendientes después ──
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM movimientos
        WHERE fecha_compra_marketplace IS NULL
          AND orden_id IS NOT NULL AND orden_id != ''
          AND fecha > NOW() - (%s || ' days')::INTERVAL
    """, (dias,))
    quedan_pendientes = cur.fetchone()[0]
    cur.close(); conn.close()

    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "movimientos_pendientes_iniciales": len(pendientes),
        "actualizados": actualizados,
        "sin_match_en_api": sin_match,
        "errores_db": errores_db,
        "movimientos_pendientes_restantes": quedan_pendientes,
        "ejemplos_no_encontrados": detalle_no_encontrados,
        "log": log
    })


@app.route("/admin/estado_reconstruccion")
def admin_estado_reconstruccion():
    """Muestra cuántos movimientos tienen/no tienen fecha_compra_marketplace,
    agrupado por canal. Útil para ver el progreso antes/después de ejecutar
    /admin/reconstruir_fechas_compra."""
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(canal, '?') as canal,
                   COUNT(*) as total,
                   COUNT(fecha_compra_marketplace) as con_fecha,
                   COUNT(*) - COUNT(fecha_compra_marketplace) as sin_fecha,
                   TO_CHAR(MIN(fecha), 'DD/MM/YYYY') as desde,
                   TO_CHAR(MAX(fecha), 'DD/MM/YYYY') as hasta
            FROM movimientos
            WHERE orden_id IS NOT NULL AND orden_id != ''
            GROUP BY canal
            ORDER BY total DESC
        """)
        resumen = []
        for r in cur.fetchall():
            pct = round((r[2]/r[1])*100, 1) if r[1] > 0 else 0
            resumen.append({
                "canal": r[0],
                "total": r[1],
                "con_fecha_real": r[2],
                "sin_fecha_real": r[3],
                "porcentaje_completo": pct,
                "desde": r[4],
                "hasta": r[5]
            })
        cur.close(); conn.close()

        total_general = sum(r["total"] for r in resumen)
        total_con = sum(r["con_fecha_real"] for r in resumen)
        return jsonify({
            "total_movimientos_con_orden": total_general,
            "con_fecha_real": total_con,
            "sin_fecha_real": total_general - total_con,
            "porcentaje_completo": round((total_con/total_general)*100, 1) if total_general > 0 else 0,
            "por_canal": resumen
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ════════════════════════════════════════════════════════════════════════════
# PLANTILLAS DE STOCK Y RESET CONTROLADO DE MOVIMIENTOS
# ════════════════════════════════════════════════════════════════════════════
# Endpoints para el flujo de "empezar limpio":
#   1) /admin/plantilla_stock_central   → descarga Excel solo con CENTRAL
#   2) /admin/plantilla_stock_fulfillment → descarga Excel con todas las
#      bodegas fulfillment (MELI_FULL, PARIS_CD, etc.)
#   3) /admin/reset_movimientos         → borra movimientos + ordenes_procesadas
#      con confirmación obligatoria y backup automático
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/plantilla_stock_central")
def admin_plantilla_stock_central():
    """Descarga plantilla Excel con todos los SKUs de productos
    + columna CENTRAL para que el usuario complete el stock.
    Si el SKU ya tenía stock en CENTRAL, lo prellena para que sirva
    también como exportación del estado actual."""
    if not session.get("logged"): return redirect("/")
    try:
        import io, openpyxl
        from inventario import cargar_productos, get_stock_bodega

        productos = cargar_productos()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Central"

        # Encabezados
        headers = ["sku_lusync", "nombre", "CENTRAL"]
        ws.append(headers)

        # Estilo encabezado
        from openpyxl.styles import Font, PatternFill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")

        # Datos: sku, nombre, stock actual en CENTRAL
        for p in productos:
            sku = p.get("sku", "")
            stock_actual = 0
            try:
                stock_actual = get_stock_bodega(sku, "CENTRAL") or 0
            except:
                stock_actual = 0
            ws.append([sku, p.get("nombre", ""), stock_actual])

        # Hoja de instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["INSTRUCCIONES — Plantilla de stock CENTRAL"],
            [""],
            ["1. Esta plantilla contiene TODOS los productos de tu sistema."],
            ["2. La columna CENTRAL contiene el stock ACTUAL en bodega central."],
            ["3. Edita los valores que necesites cambiar y guarda el archivo."],
            ["4. Sube este archivo desde la sección Bodegas → Importar Excel."],
            [""],
            ["IMPORTANTE:"],
            ["- NO borres ni cambies los nombres de las columnas (sku_lusync, nombre, CENTRAL)."],
            ["- NO cambies los valores de la columna sku_lusync."],
            ["- Si dejas una celda CENTRAL vacía, el sistema NO modifica ese SKU."],
            ["- Si pones 0, el sistema dejará ese SKU con stock cero."],
            [""],
            ["Nombres de bodega válidos para esta plantilla: CENTRAL"]
        ]
        for fila in instrucciones:
            ws2.append(fila)
        ws2.column_dimensions['A'].width = 80

        # Ajustar ancho columnas hoja principal
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, download_name="plantilla_stock_central.xlsx",
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/plantilla_stock_fulfillment")
def admin_plantilla_stock_fulfillment():
    """Descarga plantilla Excel con todos los SKUs + columnas para cada
    bodega de fulfillment (MELI_FULL, PARIS_CD, WALMART_FBM, FALABELLA_FBM,
    RIPLEY_FBM, HITES_FBM, WOO_DROP). Prellena con el stock actual."""
    if not session.get("logged"): return redirect("/")
    try:
        import io, openpyxl
        from inventario import cargar_productos, get_stock_bodega, listar_bodegas

        productos = cargar_productos()

        # Solo bodegas tipo fulfillment o dropship (excluye CENTRAL)
        bodegas = [b for b in listar_bodegas() if b.get("tipo") in ("fulfillment", "dropship")]
        bodegas_codigos = [b["codigo"] for b in bodegas]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Fulfillment"

        # Encabezados: sku, nombre, + cada bodega fulfillment
        headers = ["sku_lusync", "nombre"] + bodegas_codigos
        ws.append(headers)

        # Estilo encabezado
        from openpyxl.styles import Font, PatternFill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7C4A03", end_color="7C4A03", fill_type="solid")

        # Datos: sku, nombre, stock actual por cada fulfillment
        for p in productos:
            sku = p.get("sku", "")
            row = [sku, p.get("nombre", "")]
            for cod in bodegas_codigos:
                try:
                    row.append(get_stock_bodega(sku, cod) or 0)
                except:
                    row.append(0)
            ws.append(row)

        # Hoja de referencia con nombres legibles de cada bodega
        ws_ref = wb.create_sheet("Referencia bodegas")
        ws_ref.append(["Código", "Nombre legible", "Tipo", "Canal asociado"])
        from openpyxl.styles import Font as _F
        ws_ref.cell(row=1, column=1).font = _F(bold=True)
        ws_ref.cell(row=1, column=2).font = _F(bold=True)
        ws_ref.cell(row=1, column=3).font = _F(bold=True)
        ws_ref.cell(row=1, column=4).font = _F(bold=True)
        for b in bodegas:
            ws_ref.append([b["codigo"], b.get("nombre",""),
                           b.get("tipo",""), b.get("canal","") or "-"])
        ws_ref.column_dimensions['A'].width = 18
        ws_ref.column_dimensions['B'].width = 30
        ws_ref.column_dimensions['C'].width = 15
        ws_ref.column_dimensions['D'].width = 18

        # Hoja de instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        ws2.append(["INSTRUCCIONES — Plantilla de stock FULFILLMENT"])
        ws2.append([""])
        ws2.append(["Esta plantilla contiene una columna por cada bodega de fulfillment."])
        ws2.append(["Las columnas pre-completadas tienen el stock ACTUAL en cada bodega."])
        ws2.append([""])
        ws2.append(["IMPORTANTE:"])
        ws2.append(["- Solo edita las columnas de bodegas que necesites actualizar."])
        ws2.append(["- Si una celda está vacía, el sistema NO toca ese SKU/bodega."])
        ws2.append(["- Si pones 0, el sistema dejará ese SKU/bodega con stock cero."])
        ws2.append(["- NO cambies los nombres de las columnas (sku_lusync, MELI_FULL, etc)."])
        ws2.append([""])
        ws2.append(["Para descargar plantilla solo de Bodega Central usa:"])
        ws2.append(["/admin/plantilla_stock_central"])
        ws2.column_dimensions['A'].width = 80

        # Ajustar columnas principal
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        for i, _ in enumerate(bodegas_codigos):
            col_letter = openpyxl.utils.get_column_letter(3 + i)
            ws.column_dimensions[col_letter].width = 16

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, download_name="plantilla_stock_fulfillment.xlsx",
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/reset_movimientos", methods=["POST", "GET"])
def admin_reset_movimientos():
    """Borra TODOS los movimientos + las marcas de órdenes procesadas, para
    permitir un sync limpio desde cero. Hace backup automático de movimientos
    a una tabla temporal por si algo sale mal.

    REQUIERE confirmación EXPLÍCITA: ?confirmar=SI_BORRAR_TODO

    NO TOCA: productos, sku_mapeo, bodegas, stock_bodega, usuarios, audit_log,
             devoluciones, alertas_config.

    Después de ejecutar, los syncs de los marketplaces re-procesarán todas
    las órdenes que todavía estén en sus listados (últimos 30 días típicamente)."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    confirmar = request.args.get("confirmar", "")
    if confirmar != "SI_BORRAR_TODO":
        # Mostrar info de qué se va a borrar antes de ejecutar
        try:
            from inventario import get_conn
            conn = get_conn(); cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM movimientos")
            total_mov = cur.fetchone()[0]
            try:
                cur.execute("SELECT COUNT(*) FROM ordenes_procesadas")
                total_ord = cur.fetchone()[0]
            except:
                total_ord = 0
            cur.close(); conn.close()
            return jsonify({
                "ok": False,
                "modo": "preview",
                "mensaje": "Para confirmar la operación, agrega ?confirmar=SI_BORRAR_TODO a la URL",
                "se_borrarian": {
                    "movimientos": total_mov,
                    "ordenes_procesadas": total_ord
                },
                "se_mantienen": [
                    "productos", "sku_mapeo", "bodegas", "stock_bodega",
                    "usuarios", "audit_log", "devoluciones", "alertas",
                    "configuracion", "import_logs"
                ],
                "url_para_confirmar": "/admin/reset_movimientos?confirmar=SI_BORRAR_TODO"
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # Ejecutar reset con backup
    try:
        from inventario import get_conn
        from datetime import datetime as _dt
        conn = get_conn(); cur = conn.cursor()

        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        backup_movimientos = f"movimientos_backup_{timestamp}"
        backup_ordenes = f"ordenes_procesadas_backup_{timestamp}"

        log = []

        # 1. Backup de movimientos a tabla nueva
        try:
            cur.execute(f'CREATE TABLE "{backup_movimientos}" AS SELECT * FROM movimientos')
            cur.execute(f'SELECT COUNT(*) FROM "{backup_movimientos}"')
            count_mov = cur.fetchone()[0]
            log.append(f"✓ Backup de movimientos creado: {backup_movimientos} ({count_mov} filas)")
        except Exception as e:
            log.append(f"⚠ Error creando backup movimientos: {e}")
            conn.rollback()
            return jsonify({"ok": False, "error": f"No se pudo crear backup: {e}", "log": log}), 500

        # 2. Backup de ordenes_procesadas (si existe)
        try:
            cur.execute(f'CREATE TABLE "{backup_ordenes}" AS SELECT * FROM ordenes_procesadas')
            cur.execute(f'SELECT COUNT(*) FROM "{backup_ordenes}"')
            count_ord = cur.fetchone()[0]
            log.append(f"✓ Backup de ordenes_procesadas creado: {backup_ordenes} ({count_ord} filas)")
        except Exception as e:
            log.append(f"  (ordenes_procesadas no existía o error: {e})")
            conn.rollback()

        # 3. Borrar movimientos
        try:
            cur.execute("DELETE FROM movimientos")
            log.append(f"✓ Movimientos borrados")
        except Exception as e:
            conn.rollback()
            return jsonify({"ok": False, "error": f"Error borrando movimientos: {e}", "log": log}), 500

        # 4. Borrar ordenes_procesadas (si existe)
        try:
            cur.execute("DELETE FROM ordenes_procesadas")
            log.append(f"✓ Ordenes_procesadas borradas")
        except Exception as e:
            log.append(f"  (no se pudo borrar ordenes_procesadas: {e})")
            conn.rollback()
            # Reintentar de nuevo el delete de movimientos
            cur.execute("DELETE FROM movimientos")

        conn.commit()

        # 5. Audit log
        try:
            registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                            "RESET_MOVIMIENTOS", entidad="movimientos",
                            detalle=f"backup={backup_movimientos}, ord_backup={backup_ordenes}")
            log.append("✓ Audit log registrado")
        except Exception as e:
            log.append(f"  No se pudo registrar audit: {e}")

        cur.close(); conn.close()

        return jsonify({
            "ok": True,
            "mensaje": "Reset completado correctamente",
            "tablas_backup": [backup_movimientos, backup_ordenes],
            "instrucciones_siguientes": [
                "1. Verifica que el panel de Movimientos esté vacío",
                "2. Carga el stock CENTRAL via plantilla Excel (Bodegas → Importar Excel)",
                "3. Carga el stock FULFILLMENT via segunda plantilla Excel",
                "4. Ejecuta los syncs de cada marketplace para traer las órdenes recientes",
                "5. Si algo sale mal, los datos están en las tablas backup mencionadas arriba"
            ],
            "para_recuperar_backup": (
                f"INSERT INTO movimientos SELECT * FROM \"{backup_movimientos}\"; "
                f"INSERT INTO ordenes_procesadas SELECT * FROM \"{backup_ordenes}\";"
            ),
            "log": log
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/listar_backups")
def admin_listar_backups():
    """Lista todas las tablas de backup creadas por reset_movimientos.
    Útil si necesitas recuperar datos o limpiar backups antiguos."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT tablename FROM pg_tables
            WHERE schemaname = 'public'
              AND (tablename LIKE 'movimientos_backup_%'
                OR tablename LIKE 'ordenes_procesadas_backup_%')
            ORDER BY tablename DESC
        """)
        tablas = []
        for (nombre,) in cur.fetchall():
            try:
                cur.execute(f'SELECT COUNT(*) FROM "{nombre}"')
                cnt = cur.fetchone()[0]
            except:
                cnt = -1
            tablas.append({"tabla": nombre, "filas": cnt})
        cur.close(); conn.close()
        return jsonify({"backups": tablas, "total": len(tablas)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════
# AUTO-MAPEO DE SKUs ENTRE LUSYNC Y MARKETPLACES
# ════════════════════════════════════════════════════════════════════════════
# Endpoints para traer productos de cada marketplace y matchearlos
# automáticamente contra los productos Lusync por SKU, nombre y precio.
# ════════════════════════════════════════════════════════════════════════════

def _normalizar_texto(s):
    """Normaliza un texto para comparación: lowercase, sin acentos, sin espacios extra."""
    if not s: return ""
    import unicodedata, re
    s = str(s).lower().strip()
    # Quitar acentos
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    # Quitar caracteres no alfanuméricos
    s = re.sub(r'[^a-z0-9\s]', ' ', s)
    # Colapsar espacios
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _ratio_similitud(s1, s2):
    """Calcula similitud entre dos strings (0.0 a 1.0) usando SequenceMatcher."""
    from difflib import SequenceMatcher
    if not s1 or not s2: return 0.0
    return SequenceMatcher(None, _normalizar_texto(s1), _normalizar_texto(s2)).ratio()


def _calcular_score_match(producto_lusync, item_marketplace, sku_field, title_field, price_field):
    """Calcula score 0-100 de qué tan probable es que sean el mismo producto.

    Prioridad:
      1. SKU exacto del marketplace = SKU Lusync → 100 (match seguro)
      2. SKU Lusync contenido en título o sku del marketplace → 85
      3. Nombre normalizado idéntico → 80
      4. Nombre similitud > 0.85 + precio cercano (±10%) → 70-90
      5. Solo nombre similitud > 0.85 → 50-65
      6. Nombre similitud > 0.70 → 30-50
    """
    sku_lusync = (producto_lusync.get("sku") or "").strip().upper()
    nombre_lusync = producto_lusync.get("nombre") or ""
    precio_lusync = producto_lusync.get("precio") or 0

    sku_mp = str(item_marketplace.get(sku_field, "") or "").strip().upper()
    title_mp = item_marketplace.get(title_field, "") or ""
    price_mp = item_marketplace.get(price_field, 0) or 0
    try: price_mp = float(price_mp)
    except: price_mp = 0

    motivos = []
    score = 0

    # 1. SKU exacto
    if sku_lusync and sku_mp and sku_lusync == sku_mp:
        return 100, ["sku_exacto"]

    # 2. SKU Lusync contenido en sku del marketplace o en título
    if sku_lusync and len(sku_lusync) >= 4:
        if sku_mp and sku_lusync in sku_mp:
            score = 85
            motivos.append("sku_contenido_en_sku_mp")
        elif sku_lusync in (title_mp or "").upper():
            score = 80
            motivos.append("sku_en_titulo")
        if score >= 80:
            return score, motivos

    # 3. Nombre exacto normalizado
    nl = _normalizar_texto(nombre_lusync)
    nm = _normalizar_texto(title_mp)
    if nl and nm and nl == nm:
        score = 80
        motivos.append("nombre_exacto")

    # 4. Similitud de nombre
    if not score:
        sim = _ratio_similitud(nombre_lusync, title_mp)
        if sim >= 0.85:
            score = int(50 + sim * 30)  # 75-80
            motivos.append(f"nombre_sim_{int(sim*100)}%")
        elif sim >= 0.70:
            score = int(30 + sim * 20)  # 44-44
            motivos.append(f"nombre_sim_{int(sim*100)}%")

    # 5. Bonus por precio cercano (si tenemos un score base)
    if score > 0 and precio_lusync and price_mp:
        try:
            diff = abs(float(precio_lusync) - price_mp) / float(precio_lusync)
            if diff < 0.05:  # ±5%
                score += 15
                motivos.append("precio_muy_cercano")
            elif diff < 0.10:  # ±10%
                score += 10
                motivos.append("precio_cercano")
            elif diff < 0.20:  # ±20%
                score += 5
                motivos.append("precio_aproximado")
        except: pass

    return min(score, 99), motivos  # 99 max sin SKU exacto


def _buscar_mejor_match(producto_lusync, items_mp, sku_field, title_field, price_field):
    """Para un producto Lusync, encuentra el item del marketplace con mejor score.
    Devuelve (mejor_match_dict, score, motivos) o (None, 0, [])"""
    if not items_mp:
        return None, 0, []
    mejor = None
    mejor_score = 0
    mejor_motivos = []
    for item in items_mp:
        score, motivos = _calcular_score_match(
            producto_lusync, item, sku_field, title_field, price_field
        )
        if score > mejor_score:
            mejor_score = score
            mejor = item
            mejor_motivos = motivos
            if score >= 100:  # match perfecto, no seguir
                break
    return mejor, mejor_score, mejor_motivos


@app.route("/admin/auto_mapeo_skus", methods=["GET", "POST"])
def admin_auto_mapeo_skus():
    """Trae productos publicados de los marketplaces conectados y propone
    matches contra los productos Lusync.

    Query params:
      ?canales=mercadolibre,paris,walmart,falabella,ripley   (default: todos)
      ?guardar=1   (si 1, guarda automáticamente los matches con score >= 90)
      ?formato=json|excel   (default: json)

    Returns:
      JSON con la propuesta de mapeo, o un Excel descargable.
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    canales_str = request.args.get("canales", "mercadolibre,paris,walmart,falabella,ripley")
    canales_pedidos = set(c.strip().lower() for c in canales_str.split(",") if c.strip())
    guardar_auto = request.args.get("guardar", "0") == "1"
    formato = request.args.get("formato", "json").lower()
    umbral_auto = int(request.args.get("umbral_auto", 90))  # score mínimo para auto-guardar

    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "auto_mapeo_skus",
                    detalle=f"canales={canales_pedidos} guardar={guardar_auto}")

    log = []

    # ── 1. Cargar productos Lusync ──
    from inventario import cargar_productos, listar_sku_mapeo
    productos_lusync = cargar_productos()
    log.append(f"Productos Lusync: {len(productos_lusync)}")

    mapeo_actual = {}
    try:
        for fila in listar_sku_mapeo():
            mapeo_actual[fila["sku_lusync"]] = fila
    except Exception as e:
        log.append(f"Error cargando mapeo actual: {e}")

    # ── 2. Traer items de cada marketplace ──
    items_por_canal = {}

    if "mercadolibre" in canales_pedidos:
        try:
            from mercadolibre import obtener_publicaciones_meli
            todos = []
            offset = 0
            for _ in range(20):  # max 20 páginas (1000 items)
                data = obtener_publicaciones_meli(limite=50, offset=offset)
                if not data or not data.get("items"):
                    break
                todos.extend(data["items"])
                if len(data["items"]) < 50:
                    break
                offset += 50
            items_por_canal["mercadolibre"] = todos
            log.append(f"[MELI] {len(todos)} publicaciones obtenidas")
        except Exception as e:
            log.append(f"[MELI] ERROR: {e}")
            items_por_canal["mercadolibre"] = []

    if "paris" in canales_pedidos:
        try:
            from paris import obtener_productos_paris
            todos = []
            offset = 0
            for _ in range(20):
                data = obtener_productos_paris(limite=25, offset=offset)
                if not data:
                    break
                productos = data.get("products", []) if isinstance(data, dict) else (data or [])
                if isinstance(productos, dict):
                    productos = [productos]
                if not productos:
                    break
                # París devuelve formato distinto, normalizamos
                for p in productos:
                    items_paris = {
                        "sellerSku":    p.get("sellerSku") or p.get("sku") or "",
                        "name":         p.get("name") or p.get("productName") or "",
                        "price":        (p.get("price") or {}).get("normal") if isinstance(p.get("price"), dict) else p.get("price"),
                        "stock":        p.get("stock", 0),
                        "status":       p.get("status", "")
                    }
                    todos.append(items_paris)
                if len(productos) < 25:
                    break
                offset += 25
            items_por_canal["paris"] = todos
            log.append(f"[Paris] {len(todos)} productos obtenidos")
        except Exception as e:
            log.append(f"[Paris] ERROR: {e}")
            items_por_canal["paris"] = []

    if "walmart" in canales_pedidos:
        try:
            from walmart import obtener_productos_walmart
            todos = obtener_productos_walmart(limit=200, max_paginas=5)
            items_por_canal["walmart"] = todos
            log.append(f"[Walmart] {len(todos)} productos obtenidos")
        except Exception as e:
            log.append(f"[Walmart] ERROR: {e}")
            items_por_canal["walmart"] = []

    if "falabella" in canales_pedidos:
        try:
            from falabella import obtener_productos_falabella
            todos = []
            offset = 0
            for _ in range(20):
                productos = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                if not productos:
                    break
                # Normalizar formato Falabella
                for p in productos:
                    item = {
                        "sellerSku":  p.get("SellerSku") or p.get("sellerSku") or "",
                        "name":       p.get("Name") or p.get("name") or "",
                        "price":      p.get("Price") or p.get("price"),
                        "quantity":   p.get("Quantity") or p.get("quantity") or 0,
                        "status":     p.get("Status") or p.get("status", "")
                    }
                    try: item["price"] = float(item["price"]) if item["price"] else 0
                    except: item["price"] = 0
                    todos.append(item)
                if len(productos) < 100:
                    break
                offset += 100
            items_por_canal["falabella"] = todos
            log.append(f"[Falabella] {len(todos)} productos obtenidos")
        except Exception as e:
            log.append(f"[Falabella] ERROR: {e}")
            items_por_canal["falabella"] = []

    if "ripley" in canales_pedidos:
        try:
            from ripley import obtener_productos_ripley
            todos = obtener_productos_ripley(max_paginas=15, page_size=100)
            items_por_canal["ripley"] = todos
            log.append(f"[Ripley] {len(todos)} productos obtenidos")
        except Exception as e:
            log.append(f"[Ripley] ERROR: {e}")
            items_por_canal["ripley"] = []

    # ── 3. Para cada producto Lusync, buscar mejor match en cada marketplace ──
    # Configuración de campos por canal
    canal_config = {
        "mercadolibre": {"sku_field": "sku_seller", "title": "title", "price": "price"},
        "paris":        {"sku_field": "sellerSku", "title": "name", "price": "price"},
        "walmart":      {"sku_field": "sku", "title": "productName", "price": "price"},
        "falabella":    {"sku_field": "sellerSku", "title": "name", "price": "price"},
        "ripley":       {"sku_field": "shop_sku", "title": "product_title", "price": "price"}
    }

    propuesta_mapeo = []
    auto_guardados = 0
    necesitan_revision = 0
    sin_match = 0

    for p in productos_lusync:
        fila_propuesta = {
            "sku_lusync": p.get("sku", ""),
            "nombre":     p.get("nombre", ""),
            "precio":     p.get("precio", 0),
            "matches":    {}
        }

        # Mapeo actual (para no sobreescribir si ya existe match manual)
        m_actual = mapeo_actual.get(p.get("sku", ""), {})

        for canal, items in items_por_canal.items():
            cfg = canal_config[canal]
            mejor, score, motivos = _buscar_mejor_match(
                p, items, cfg["sku_field"], cfg["title"], cfg["price"]
            )

            # Determinar el SKU del marketplace que se guardaría
            sku_propuesto = ""
            titulo_mp = ""
            precio_mp = 0
            if mejor:
                sku_propuesto = str(mejor.get(cfg["sku_field"], "") or "").strip()
                titulo_mp = mejor.get(cfg["title"], "") or ""
                precio_mp = mejor.get(cfg["price"], 0) or 0

            # Para MELI usamos item_id si no hay sku_seller
            if canal == "mercadolibre" and mejor and not sku_propuesto:
                sku_propuesto = str(mejor.get("item_id", "") or "")

            # SKU actualmente en BD para este canal
            campo_bd = {
                "mercadolibre": "sku_mercadolibre",
                "paris":        "sku_paris",
                "walmart":      "sku_walmart",
                "falabella":    "sku_falabella",
                "ripley":       "sku_ripley"
            }.get(canal, "")
            sku_actual = m_actual.get(campo_bd, "") if campo_bd else ""

            fila_propuesta["matches"][canal] = {
                "sku_propuesto":  sku_propuesto,
                "sku_actual_bd":  sku_actual,
                "score":          score,
                "motivos":        motivos,
                "titulo_mp":      titulo_mp,
                "precio_mp":      precio_mp,
                "estado": (
                    "auto" if score >= umbral_auto else
                    "revisar" if score >= 50 else
                    "sin_match"
                )
            }

        # Contar resúmenes
        algun_auto = any(m["estado"] == "auto" for m in fila_propuesta["matches"].values())
        algun_revisar = any(m["estado"] == "revisar" for m in fila_propuesta["matches"].values())
        if algun_auto:
            auto_guardados += 1
        elif algun_revisar:
            necesitan_revision += 1
        else:
            sin_match += 1

        propuesta_mapeo.append(fila_propuesta)

    # ── 4. Guardar automáticamente los matches con score >= umbral_auto ──
    if guardar_auto:
        from inventario import guardar_sku_mapeo_fila
        guardados_efectivos = 0
        for fila in propuesta_mapeo:
            sku_lus = fila["sku_lusync"]
            m_actual = mapeo_actual.get(sku_lus, {})
            skus_a_guardar = {
                "web":          m_actual.get("sku_web", ""),
                "walmart":      m_actual.get("sku_walmart", ""),
                "paris":        m_actual.get("sku_paris", ""),
                "falabella":    m_actual.get("sku_falabella", ""),
                "ripley":       m_actual.get("sku_ripley", ""),
                "mercadolibre": m_actual.get("sku_mercadolibre", ""),
                "hites":        m_actual.get("sku_hites", "")
            }
            algun_cambio = False
            for canal_name, match_data in fila["matches"].items():
                if match_data["estado"] == "auto" and match_data["sku_propuesto"]:
                    if not skus_a_guardar.get(canal_name):  # solo si no había nada antes
                        skus_a_guardar[canal_name] = match_data["sku_propuesto"]
                        algun_cambio = True
            if algun_cambio:
                try:
                    guardar_sku_mapeo_fila(sku_lus, skus_a_guardar)
                    guardados_efectivos += 1
                except Exception as e:
                    log.append(f"  ERROR guardando {sku_lus}: {e}")
        log.append(f"Auto-guardados: {guardados_efectivos} mapeos actualizados en BD")

    # ── 5. Devolver según formato ──
    if formato == "excel":
        return _generar_excel_propuesta(propuesta_mapeo, items_por_canal, log)

    return jsonify({
        "ok": True,
        "resumen": {
            "total_productos_lusync": len(productos_lusync),
            "con_matches_seguros": auto_guardados,
            "necesitan_revision": necesitan_revision,
            "sin_match": sin_match,
            "umbral_auto": umbral_auto
        },
        "items_por_canal": {c: len(v) for c, v in items_por_canal.items()},
        "log": log,
        "propuesta": propuesta_mapeo
    })


def _generar_excel_propuesta(propuesta_mapeo, items_por_canal, log):
    """Genera un Excel descargable con la propuesta de mapeo."""
    import io, openpyxl
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Hoja 1: Propuesta principal
    ws = wb.active
    ws.title = "Propuesta de Mapeo"
    headers = [
        "SKU Lusync", "Nombre Lusync", "Precio Lusync",
        "MELI: SKU", "MELI: score", "MELI: título encontrado",
        "PARIS: SKU", "PARIS: score", "PARIS: título encontrado",
        "WALMART: SKU", "WALMART: score", "WALMART: título encontrado",
        "FALABELLA: SKU", "FALABELLA: score", "FALABELLA: título encontrado",
        "RIPLEY: SKU", "RIPLEY: score", "RIPLEY: título encontrado",
    ]
    ws.append(headers)
    # Estilo header
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colores por estado
    fill_auto    = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # verde
    fill_revisar = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # amarillo
    fill_nada    = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # rojo

    for fila in propuesta_mapeo:
        row = [fila["sku_lusync"], fila["nombre"], fila["precio"]]
        for canal in ["mercadolibre", "paris", "walmart", "falabella", "ripley"]:
            m = fila["matches"].get(canal, {})
            row.extend([
                m.get("sku_propuesto", ""),
                m.get("score", 0),
                (m.get("titulo_mp", "") or "")[:80]
            ])
        ws.append(row)

        # Aplicar color a las celdas de cada canal según estado
        row_idx = ws.max_row
        for i, canal in enumerate(["mercadolibre", "paris", "walmart", "falabella", "ripley"]):
            m = fila["matches"].get(canal, {})
            estado = m.get("estado", "sin_match")
            fill = fill_auto if estado == "auto" else (fill_revisar if estado == "revisar" else fill_nada)
            for offset in range(3):  # las 3 columnas de cada canal
                col = 4 + (i * 3) + offset
                ws.cell(row=row_idx, column=col).fill = fill

    # Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    for letter in ['D', 'G', 'J', 'M', 'P']:  # SKU columns
        ws.column_dimensions[letter].width = 18
    for letter in ['E', 'H', 'K', 'N', 'Q']:  # score
        ws.column_dimensions[letter].width = 8
    for letter in ['F', 'I', 'L', 'O', 'R']:  # título
        ws.column_dimensions[letter].width = 35
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "D2"

    # Hoja 2: instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instr = [
        ["INSTRUCCIONES — Auto-mapeo de SKUs"],
        [""],
        ["LEYENDA DE COLORES:"],
        ["🟢 Verde:    Score ≥ 90 (match seguro, recomendado guardar)"],
        ["🟡 Amarillo: Score 50-89 (revisar manualmente, puede ser correcto)"],
        ["🔴 Rojo:     Score < 50 (probablemente no hay match, revisa el catálogo)"],
        [""],
        ["CÓMO USAR:"],
        ["1. Revisa esta hoja, especialmente las celdas amarillas y rojas."],
        ["2. Si un match es correcto, no hagas nada (ya está propuesto)."],
        ["3. Si un match es incorrecto, edita la columna 'XXX: SKU' con el SKU correcto."],
        ["4. Si no hay match para un canal, deja la celda vacía."],
        ["5. Cuando termines, guarda este Excel."],
        ["6. Súbelo desde el panel: sección 'Mapeo SKUs' → 'Importar Excel'."],
        [""],
        ["IMPORTANTE:"],
        ["- NO cambies el SKU Lusync (columna A) ni los nombres de columnas."],
        ["- El sistema solo considera la columna 'SKU' de cada canal al importar."],
        ["- Las columnas 'score' y 'título encontrado' son informativas."],
        [""],
        ["LOG DE EJECUCIÓN:"],
    ]
    for i in instr:
        ws2.append(i)
    for line in log:
        ws2.append([line])
    ws2.column_dimensions['A'].width = 90

    # Hoja 3: items crudos por canal (debug)
    for canal, items in items_por_canal.items():
        ws_c = wb.create_sheet(f"Items {canal[:8]}")
        if items:
            keys = list(items[0].keys())
            ws_c.append(keys)
            for it in items:
                ws_c.append([str(it.get(k, ""))[:200] for k in keys])
            for c in range(1, len(keys) + 1):
                ws_c.cell(row=1, column=c).font = Font(bold=True)
            ws_c.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(
        buf,
        download_name=f"propuesta_mapeo_skus_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO DE SKUs POR MARKETPLACE (ver qué devuelve cada API)
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/skus_marketplace/<canal>")
def admin_skus_marketplace(canal):
    """Devuelve la lista cruda de SKUs publicados en un marketplace específico.
    Útil para ver QUÉ SKU exacto tiene cada item en cada marketplace.

    canal: mercadolibre | paris | walmart | falabella | ripley
    Query params:
      ?formato=json (default) | excel
      ?limite=200 (cantidad máxima a traer)
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    canal_l = canal.lower().strip()
    formato = request.args.get("formato", "json").lower()
    limite_max = int(request.args.get("limite", 500))

    items = []
    log = []
    error = None

    try:
        if canal_l in ("mercadolibre", "meli"):
            from mercadolibre import obtener_publicaciones_meli
            offset = 0
            while len(items) < limite_max:
                data = obtener_publicaciones_meli(limite=50, offset=offset)
                if not data or not data.get("items"):
                    break
                for it in data["items"]:
                    items.append({
                        "item_id":         it.get("item_id"),
                        "titulo":          it.get("title"),
                        "sku_seller":      it.get("sku_seller", ""),
                        "sku_origen":      it.get("sku_origen", ""),
                        "variantes_skus":  it.get("variantes_skus", []),
                        "stock":           it.get("stock"),
                        "precio":          it.get("price"),
                        "status":          it.get("status")
                    })
                if len(data["items"]) < 50:
                    break
                offset += 50
            log.append(f"MELI: {len(items)} publicaciones obtenidas")

        elif canal_l == "paris":
            from paris import obtener_productos_paris
            offset = 0
            total_reportado = None
            while len(items) < limite_max:
                data = obtener_productos_paris(limite=25, offset=offset)
                if not data:
                    break
                # París devuelve estructura: {results: [...], total, offset, limit}
                if isinstance(data, dict):
                    productos = data.get("results") or data.get("products") or data.get("items") or []
                    if total_reportado is None:
                        total_reportado = data.get("total", 0)
                        log.append(f"  Paris: total reportado por API = {total_reportado}")
                elif isinstance(data, list):
                    productos = data
                else:
                    productos = []
                if not productos:
                    log.append(f"  Paris offset {offset}: respuesta vacía o estructura inesperada")
                    log.append(f"  data keys: {list(data.keys()) if isinstance(data, dict) else type(data).__name__}")
                    break
                for p in productos:
                    # París estructura típica: sellerSku, partnerSku, name, price, stock, status
                    items.append({
                        "sku_paris":       p.get("sellerSku") or p.get("partnerSku") or p.get("sku") or "",
                        "partner_sku":     p.get("partnerSku", ""),
                        "titulo":          p.get("name") or p.get("productName") or p.get("title") or "",
                        "stock":           p.get("stock") or (p.get("offer") or {}).get("stock") or 0,
                        "precio":          (p.get("price") or {}).get("normal") if isinstance(p.get("price"), dict) else (p.get("price") or (p.get("offer") or {}).get("price")),
                        "status":          p.get("status") or p.get("itemStatus") or "",
                        "raw_keys":        list(p.keys())[:15]  # primeras 15 keys del payload crudo (debug)
                    })
                # Salir si: ya trajimos todo o la página vino con menos del límite
                if len(productos) < 25:
                    log.append(f"  Paris: última página (vino {len(productos)} < 25)")
                    break
                if total_reportado and len(items) >= total_reportado:
                    log.append(f"  Paris: alcanzado total reportado ({total_reportado})")
                    break
                offset += 25
            log.append(f"Paris: {len(items)} productos obtenidos")

        elif canal_l == "walmart":
            from walmart import obtener_productos_walmart
            # Walmart Chile máximo 50 por página → 20 páginas = 1000 items max
            resultado_debug = obtener_productos_walmart(limit=50, max_paginas=20, debug=True)
            items_raw = resultado_debug.get("items", [])
            for line in resultado_debug.get("debug_log", []):
                log.append(f"  Walmart: {line}")
            for p in items_raw[:limite_max]:
                items.append({
                    "sku_walmart":   p.get("sku", ""),
                    "wpid":          p.get("wpid", ""),
                    "titulo":        p.get("productName", ""),
                    "stock":         p.get("availableInventory", 0),
                    "precio":        p.get("price"),
                    "status":        p.get("status", "")
                })
            log.append(f"Walmart: {len(items)} productos obtenidos")

        elif canal_l == "falabella":
            from falabella import obtener_productos_falabella
            offset = 0
            while len(items) < limite_max:
                productos = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                if not productos:
                    break
                for p in productos:
                    items.append({
                        "sku_falabella":    p.get("SellerSku") or p.get("sellerSku") or "",
                        "shopSku":          p.get("ShopSku") or "",
                        "titulo":           p.get("Name") or p.get("name") or "",
                        "stock":            p.get("Quantity") or p.get("quantity") or 0,
                        "precio":           p.get("Price") or p.get("price"),
                        "status":           p.get("Status") or p.get("status", "")
                    })
                if len(productos) < 100:
                    break
                offset += 100
            log.append(f"Falabella: {len(items)} productos obtenidos")

        elif canal_l == "ripley":
            from ripley import obtener_productos_ripley
            items_raw = obtener_productos_ripley(max_paginas=15, page_size=100)
            for p in items_raw[:limite_max]:
                items.append({
                    "shop_sku":      p.get("shop_sku", ""),
                    "product_sku":   p.get("product_sku", ""),
                    "titulo":        p.get("product_title", ""),
                    "stock":         p.get("quantity", 0),
                    "precio":        p.get("price"),
                    "state_code":    p.get("state_code", ""),
                    "active":        p.get("active", True)
                })
            log.append(f"Ripley: {len(items)} ofertas obtenidas")

        else:
            return jsonify({"error": f"Canal '{canal}' no válido. Usa: mercadolibre, paris, walmart, falabella, ripley"}), 400

    except Exception as e:
        import traceback
        error = str(e)
        log.append(f"ERROR: {error}")
        log.append(traceback.format_exc())

    if formato == "excel":
        try:
            import io, openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"SKUs {canal_l[:8]}"
            if items:
                headers = list(items[0].keys())
                ws.append(headers)
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=c)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")
                for it in items:
                    ws.append([str(it.get(h, "")) for h in headers])
                for c in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 22
                ws.freeze_panes = "A2"

            ws_log = wb.create_sheet("Log")
            for line in log:
                ws_log.append([line])
            ws_log.column_dimensions['A'].width = 100

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            return send_file(buf, download_name=f"skus_{canal_l}.xlsx",
                             as_attachment=True,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return jsonify({"error": f"Error generando Excel: {e}", "items": items, "log": log}), 500

    return jsonify({
        "ok": error is None,
        "canal": canal_l,
        "total": len(items),
        "log": log,
        "error": error,
        "items": items
    })


# ════════════════════════════════════════════════════════════════════════════
# AUTO-MAPEO MULTI-PUBLICACIÓN (v2) — POBLAR sku_mapeo_canal
# ════════════════════════════════════════════════════════════════════════════
# Trae todas las publicaciones de cada marketplace y las inserta en la nueva
# tabla sku_mapeo_canal, soportando MÚLTIPLES publicaciones por SKU Lusync.
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/auto_mapeo_v2", methods=["GET", "POST"])
def admin_auto_mapeo_v2():
    """Auto-mapeo de SKUs en modelo multi-publicación.

    Para cada marketplace, trae TODAS las publicaciones y las inserta en
    sku_mapeo_canal vinculadas al producto Lusync correspondiente.

    Si una publicación tiene SKU exacto a un producto Lusync → se vincula.
    Si una publicación tiene SKU vacío → se vincula al SKU Lusync por similitud
    de nombre (umbral configurable).

    Query params:
      ?canales=mercadolibre,paris,walmart,falabella,ripley   (default: todos)
      ?dry_run=1                                             (default 0)
      ?umbral_nombre=0.85                                    (similitud nombre)
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    canales_str = request.args.get("canales", "mercadolibre,paris,walmart,falabella,ripley")
    canales_pedidos = set(c.strip().lower() for c in canales_str.split(",") if c.strip())
    dry_run = request.args.get("dry_run", "0") == "1"
    umbral_nombre = float(request.args.get("umbral_nombre", "0.85"))

    registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                    "auto_mapeo_v2",
                    detalle=f"canales={canales_pedidos} dry_run={dry_run}")

    log = [f"Auto-mapeo v2: canales={list(canales_pedidos)}, dry_run={dry_run}, umbral_nombre={umbral_nombre}"]

    # 1. Cargar productos Lusync
    from inventario import (cargar_productos, agregar_publicacion,
                            obtener_publicaciones_canal)
    productos_lusync = cargar_productos()
    log.append(f"Productos Lusync: {len(productos_lusync)}")

    # Diccionario {sku_normalizado: sku_lusync_real} para match exacto rápido
    skus_norm = {}
    for p in productos_lusync:
        sku = (p.get("sku") or "").strip()
        if sku:
            skus_norm[sku.upper()] = sku

    # 2. Para cada marketplace, traer publicaciones e insertar
    resumen = {}
    publicaciones_creadas = []

    def _vincular_publicacion(canal, sku_canal, item_id_canal, titulo, precio, log_canal):
        """Encuentra el SKU Lusync para una publicación e inserta en sku_mapeo_canal.

        Estrategia:
          1. Match exacto por SKU (sku_canal == sku Lusync)
          2. Match por SKU contenido (ej: 'ODJM001-AZUL' contiene 'ODJM001')
          3. Match por similitud de nombre (umbral configurable)
        """
        sku_canal_clean = (sku_canal or "").strip()
        sku_lusync_match = None
        razon = ""

        # 1. Match exacto
        if sku_canal_clean and sku_canal_clean.upper() in skus_norm:
            sku_lusync_match = skus_norm[sku_canal_clean.upper()]
            razon = "sku_exacto"

        # 2. Match parcial: SKU Lusync contenido en sku_canal o título
        if not sku_lusync_match and sku_canal_clean:
            sku_canal_upper = sku_canal_clean.upper()
            titulo_upper = (titulo or "").upper()
            for sku_norm, sku_real in skus_norm.items():
                if len(sku_norm) >= 4 and (sku_norm in sku_canal_upper or sku_norm in titulo_upper):
                    sku_lusync_match = sku_real
                    razon = "sku_parcial" if sku_norm in sku_canal_upper else "sku_en_titulo"
                    break

        # 3. Match por nombre (similitud)
        if not sku_lusync_match and titulo:
            mejor_sim = 0
            mejor_sku = None
            for p in productos_lusync:
                nombre_p = p.get("nombre") or ""
                if not nombre_p: continue
                sim = _ratio_similitud(nombre_p, titulo)
                if sim >= umbral_nombre and sim > mejor_sim:
                    mejor_sim = sim
                    mejor_sku = p.get("sku")
            if mejor_sku:
                sku_lusync_match = mejor_sku
                razon = f"nombre_sim_{int(mejor_sim*100)}%"

        if not sku_lusync_match:
            log_canal.append(f"  [SIN MATCH] {sku_canal_clean or item_id_canal}: '{titulo[:50]}'")
            return None

        # Insertar en sku_mapeo_canal (idempotente — agregar_publicacion hace UPSERT)
        if not dry_run:
            try:
                mapeo_id = agregar_publicacion(
                    sku_lusync=sku_lusync_match,
                    canal=canal,
                    sku_canal=sku_canal_clean or sku_lusync_match,
                    item_id_canal=item_id_canal,
                    es_catalogo=False,
                    notas=f"auto_v2:{razon}"
                )
                if mapeo_id:
                    publicaciones_creadas.append({
                        "canal": canal, "sku_lusync": sku_lusync_match,
                        "sku_canal": sku_canal_clean, "item_id_canal": item_id_canal,
                        "razon": razon
                    })
                    log_canal.append(f"  [{razon}] {sku_canal_clean or item_id_canal} → {sku_lusync_match}")
                    return mapeo_id
            except Exception as e:
                log_canal.append(f"  [ERROR] {sku_canal_clean}: {e}")
        else:
            log_canal.append(f"  [DRY {razon}] {sku_canal_clean or item_id_canal} → {sku_lusync_match}")
        return None

    # ─── MERCADOLIBRE ───
    if "mercadolibre" in canales_pedidos:
        log_meli = []
        try:
            from mercadolibre import obtener_publicaciones_meli
            todos_meli = []
            offset = 0
            for _ in range(20):  # max 1000 items
                data = obtener_publicaciones_meli(limite=50, offset=offset)
                if not data or not data.get("items"): break
                todos_meli.extend(data["items"])
                if len(data["items"]) < 50: break
                offset += 50
            log_meli.append(f"MELI: {len(todos_meli)} publicaciones obtenidas")

            mapeados_ok = 0
            sin_match = 0
            for it in todos_meli:
                item_id = it.get("item_id")
                sku_seller = it.get("sku_seller", "")
                titulo = it.get("title", "")
                precio = it.get("price")
                # MELI: usar sku_seller si existe, sino el item_id como sku_canal
                sku_para_mapear = sku_seller if sku_seller else item_id
                resultado = _vincular_publicacion(
                    "mercadolibre", sku_para_mapear, item_id, titulo, precio, log_meli
                )
                if resultado: mapeados_ok += 1
                elif not dry_run: sin_match += 1
            resumen["mercadolibre"] = {
                "publicaciones": len(todos_meli),
                "mapeadas": mapeados_ok,
                "sin_match": sin_match
            }
        except Exception as e:
            log_meli.append(f"MELI ERROR: {e}")
            resumen["mercadolibre"] = {"error": str(e)}
        log.extend(log_meli[:50])  # primeras 50 líneas para no inundar

    # ─── PARIS ───
    if "paris" in canales_pedidos:
        log_paris = []
        try:
            from paris import obtener_productos_paris
            todos_paris = []
            offset = 0
            total_reportado = None
            for _ in range(20):
                data = obtener_productos_paris(limite=25, offset=offset)
                if not data: break
                if isinstance(data, dict):
                    productos = data.get("results") or data.get("products") or []
                    if total_reportado is None:
                        total_reportado = data.get("total", 0)
                else:
                    productos = data if isinstance(data, list) else []
                if not productos: break
                todos_paris.extend(productos)
                if len(productos) < 25: break
                if total_reportado and len(todos_paris) >= total_reportado: break
                offset += 25
            log_paris.append(f"Paris: {len(todos_paris)} productos obtenidos")

            mapeados_ok = 0
            sin_match = 0
            for p in todos_paris:
                sku_paris = p.get("sellerSku") or p.get("partnerSku") or p.get("sku") or ""
                titulo = p.get("name") or p.get("productName") or ""
                precio = (p.get("price") or {}).get("normal") if isinstance(p.get("price"), dict) else p.get("price")
                if not sku_paris and not titulo: continue
                resultado = _vincular_publicacion(
                    "paris", sku_paris, None, titulo, precio, log_paris
                )
                if resultado: mapeados_ok += 1
                elif not dry_run: sin_match += 1
            resumen["paris"] = {
                "publicaciones": len(todos_paris),
                "mapeadas": mapeados_ok,
                "sin_match": sin_match
            }
        except Exception as e:
            log_paris.append(f"Paris ERROR: {e}")
            resumen["paris"] = {"error": str(e)}
        log.extend(log_paris[:50])

    # ─── WALMART ───
    if "walmart" in canales_pedidos:
        log_wm = []
        try:
            from walmart import obtener_productos_walmart
            todos_wm = obtener_productos_walmart(limit=50, max_paginas=20)
            log_wm.append(f"Walmart: {len(todos_wm)} productos obtenidos")

            mapeados_ok = 0
            sin_match = 0
            for p in todos_wm:
                sku_wm = p.get("sku", "")
                titulo = p.get("productName", "")
                precio = p.get("price")
                wpid = p.get("wpid")
                if not sku_wm and not titulo: continue
                resultado = _vincular_publicacion(
                    "walmart", sku_wm, wpid, titulo, precio, log_wm
                )
                if resultado: mapeados_ok += 1
                elif not dry_run: sin_match += 1
            resumen["walmart"] = {
                "publicaciones": len(todos_wm),
                "mapeadas": mapeados_ok,
                "sin_match": sin_match
            }
        except Exception as e:
            log_wm.append(f"Walmart ERROR: {e}")
            resumen["walmart"] = {"error": str(e)}
        log.extend(log_wm[:50])

    # ─── FALABELLA ───
    if "falabella" in canales_pedidos:
        log_fa = []
        try:
            from falabella import obtener_productos_falabella
            todos_fa = []
            offset = 0
            for _ in range(20):
                productos = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                if not productos: break
                todos_fa.extend(productos)
                if len(productos) < 100: break
                offset += 100
            log_fa.append(f"Falabella: {len(todos_fa)} productos obtenidos")

            mapeados_ok = 0
            sin_match = 0
            for p in todos_fa:
                sku_fa = p.get("SellerSku") or p.get("sellerSku") or ""
                titulo = p.get("Name") or p.get("name") or ""
                shop_sku = p.get("ShopSku") or ""
                precio = p.get("Price") or p.get("price")
                if not sku_fa and not titulo: continue
                resultado = _vincular_publicacion(
                    "falabella", sku_fa, shop_sku, titulo, precio, log_fa
                )
                if resultado: mapeados_ok += 1
                elif not dry_run: sin_match += 1
            resumen["falabella"] = {
                "publicaciones": len(todos_fa),
                "mapeadas": mapeados_ok,
                "sin_match": sin_match
            }
        except Exception as e:
            log_fa.append(f"Falabella ERROR: {e}")
            resumen["falabella"] = {"error": str(e)}
        log.extend(log_fa[:50])

    # ─── RIPLEY ───
    if "ripley" in canales_pedidos:
        log_rp = []
        try:
            from ripley import obtener_productos_ripley
            todos_rp = obtener_productos_ripley(max_paginas=15, page_size=100)
            log_rp.append(f"Ripley: {len(todos_rp)} ofertas obtenidas")

            mapeados_ok = 0
            sin_match = 0
            for p in todos_rp:
                sku_rp = p.get("shop_sku", "")
                titulo = p.get("product_title", "")
                product_sku = p.get("product_sku", "")
                precio = p.get("price")
                if not sku_rp and not titulo: continue
                resultado = _vincular_publicacion(
                    "ripley", sku_rp, product_sku, titulo, precio, log_rp
                )
                if resultado: mapeados_ok += 1
                elif not dry_run: sin_match += 1
            resumen["ripley"] = {
                "publicaciones": len(todos_rp),
                "mapeadas": mapeados_ok,
                "sin_match": sin_match
            }
        except Exception as e:
            log_rp.append(f"Ripley ERROR: {e}")
            resumen["ripley"] = {"error": str(e)}
        log.extend(log_rp[:50])

    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "resumen": resumen,
        "publicaciones_creadas_total": len(publicaciones_creadas),
        "ejemplos_creadas": publicaciones_creadas[:30],
        "log": log
    })


# ────────────────────────────────────────────────────────────────────────────
# CRUD de sku_mapeo_canal (multi-publicación)
# ────────────────────────────────────────────────────────────────────────────

@app.route("/sku_mapeo_canal/listar")
def ruta_sku_mapeo_canal_listar():
    """Lista todos los mapeos en sku_mapeo_canal, agrupados por SKU Lusync.
    Devuelve formato listo para UI multi-fila."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import listar_mapeos_canal, cargar_productos
        canal = request.args.get("canal")  # opcional
        sku_lusync = request.args.get("sku_lusync")  # opcional
        mapeos = listar_mapeos_canal(canal=canal, sku_lusync=sku_lusync)

        # Agrupar por sku_lusync para que UI los muestre en bloques
        productos = {p["sku"]: p for p in cargar_productos()}
        agrupado = {}
        for m in mapeos:
            sku = m["sku_lusync"]
            if sku not in agrupado:
                p = productos.get(sku, {})
                agrupado[sku] = {
                    "sku_lusync": sku,
                    "nombre": p.get("nombre", ""),
                    "publicaciones": []
                }
            agrupado[sku]["publicaciones"].append(m)

        return jsonify({
            "ok": True,
            "total": len(mapeos),
            "agrupado": list(agrupado.values()),
            "mapeos": mapeos
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/sku_mapeo_canal/agregar", methods=["POST"])
def ruta_sku_mapeo_canal_agregar():
    """Agrega manualmente una publicación a sku_mapeo_canal.
    Body JSON: {sku_lusync, canal, sku_canal, item_id_canal?, es_catalogo?, notas?}"""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import agregar_publicacion
        data = request.get_json() or {}
        sku_lusync = (data.get("sku_lusync") or "").strip()
        canal = (data.get("canal") or "").strip().lower()
        sku_canal = (data.get("sku_canal") or "").strip()
        item_id_canal = data.get("item_id_canal")
        es_catalogo = bool(data.get("es_catalogo", False))
        notas = data.get("notas")

        if not (sku_lusync and canal and sku_canal):
            return jsonify({"ok": False, "error": "Faltan parámetros: sku_lusync, canal, sku_canal"}), 400

        mapeo_id = agregar_publicacion(sku_lusync, canal, sku_canal,
                                       item_id_canal, es_catalogo, notas)
        if mapeo_id:
            registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                            "sku_mapeo_canal_agregar",
                            detalle=f"{sku_lusync} → {canal}:{sku_canal} (id={mapeo_id})")
            return jsonify({"ok": True, "mapeo_id": mapeo_id})
        return jsonify({"ok": False, "error": "No se pudo guardar"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/sku_mapeo_canal/eliminar", methods=["POST"])
def ruta_sku_mapeo_canal_eliminar():
    """Elimina (soft delete) una publicación de sku_mapeo_canal.
    Body JSON: {mapeo_id}"""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import eliminar_publicacion
        data = request.get_json() or {}
        mapeo_id = data.get("mapeo_id")
        if not mapeo_id:
            return jsonify({"ok": False, "error": "Falta mapeo_id"}), 400
        ok = eliminar_publicacion(int(mapeo_id))
        if ok:
            registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                            "sku_mapeo_canal_eliminar", detalle=f"id={mapeo_id}")
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/sku_mapeo_canal/contar")
def ruta_sku_mapeo_canal_contar():
    """Devuelve {sku_lusync: {canal: cantidad_publicaciones}} para UI."""
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import contar_publicaciones_por_sku
        return jsonify({"ok": True, "datos": contar_publicaciones_por_sku()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════
# MIGRACIÓN: sku_mapeo (legacy) → sku_mapeo_canal (nuevo multi-publicación)
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/migrar_sku_mapeo_a_canal", methods=["GET", "POST"])
def admin_migrar_sku_mapeo_a_canal():
    """Lee la tabla legacy sku_mapeo (1:1) y crea filas en sku_mapeo_canal.

    Para cada fila de sku_mapeo, crea hasta 7 filas en sku_mapeo_canal
    (una por cada canal con SKU no vacío).

    Es idempotente: si una publicación ya existe en sku_mapeo_canal, la actualiza
    pero no la duplica (gracias al UNIQUE de la tabla).

    Query params:
      ?dry_run=1   simula sin escribir
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    dry_run = request.args.get("dry_run", "0") == "1"

    try:
        from inventario import listar_sku_mapeo, agregar_publicacion

        mapeos = listar_sku_mapeo()
        log = []
        publicaciones_creadas = 0
        publicaciones_fallidas = 0
        skus_procesados = 0

        # Mapeo campo BD → canal
        campos_canal = [
            ("sku_web",          "web"),
            ("sku_walmart",      "walmart"),
            ("sku_paris",        "paris"),
            ("sku_falabella",    "falabella"),
            ("sku_ripley",       "ripley"),
            ("sku_mercadolibre", "mercadolibre"),
            ("sku_hites",        "hites"),
        ]

        for fila in mapeos:
            sku_lusync = (fila.get("sku_lusync") or "").strip()
            if not sku_lusync:
                continue
            skus_procesados += 1

            for campo_bd, canal in campos_canal:
                sku_canal_val = (fila.get(campo_bd) or "").strip()
                if not sku_canal_val or sku_canal_val.lower() in ("none", "nan", "null"):
                    continue

                # Para MELI: detectar si es item_id (MLC...) o seller_sku
                item_id_canal = None
                if canal == "mercadolibre" and sku_canal_val.upper().startswith("MLC"):
                    item_id_canal = sku_canal_val

                # ── BLINDAJE: verificar si ya existe ANTES de insertar ──
                if not dry_run:
                    try:
                        from inventario import get_conn
                        conn_check = get_conn(); cur_check = conn_check.cursor()
                        cur_check.execute("""
                            SELECT COUNT(*) FROM sku_mapeo_canal
                            WHERE canal = %s
                              AND sku_lusync = %s
                              AND sku_canal = %s
                              AND activo = TRUE
                        """, (canal, sku_lusync, sku_canal_val))
                        ya_existe = cur_check.fetchone()[0] > 0
                        cur_check.close(); conn_check.close()
                    except:
                        ya_existe = False
                    if ya_existe:
                        log.append(f"⏭ {sku_lusync} → {canal}:{sku_canal_val}: ya existe, skip")
                        continue

                if dry_run:
                    log.append(f"[DRY] {sku_lusync} → {canal}:{sku_canal_val}" +
                               (f" (item_id={item_id_canal})" if item_id_canal else ""))
                    publicaciones_creadas += 1
                else:
                    try:
                        mapeo_id = agregar_publicacion(
                            sku_lusync=sku_lusync,
                            canal=canal,
                            sku_canal=sku_canal_val,
                            item_id_canal=item_id_canal,
                            es_catalogo=False,
                            notas="migracion_legacy"
                        )
                        if mapeo_id:
                            publicaciones_creadas += 1
                            log.append(f"✓ {sku_lusync} → {canal}:{sku_canal_val} (id={mapeo_id})")
                        else:
                            publicaciones_fallidas += 1
                            log.append(f"✗ {sku_lusync} → {canal}:{sku_canal_val}: agregar devolvió None")
                    except Exception as e:
                        publicaciones_fallidas += 1
                        log.append(f"✗ {sku_lusync} → {canal}:{sku_canal_val}: {e}")

        if not dry_run:
            registrar_audit(
                session.get("usuario", "Sistema"), request.remote_addr,
                "migrar_sku_mapeo_a_canal",
                detalle=f"skus={skus_procesados} pubs_creadas={publicaciones_creadas} fallidas={publicaciones_fallidas}"
            )

        return jsonify({
            "ok": True,
            "dry_run": dry_run,
            "skus_procesados": skus_procesados,
            "publicaciones_creadas": publicaciones_creadas,
            "publicaciones_fallidas": publicaciones_fallidas,
            "log": log[:200]  # primeras 200 líneas
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/estado_tablas")
def admin_estado_tablas():
    """Devuelve el conteo de filas de las tablas principales para diagnóstico rápido.

    Útil para saber si los productos están cargados, si el mapeo está poblado,
    si las publicaciones existen, etc.
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn(); cur = conn.cursor()

        info = {}

        # Productos
        try:
            cur.execute("SELECT COUNT(*) FROM productos")
            info["productos"] = {"total": cur.fetchone()[0]}
            cur.execute("SELECT sku, nombre FROM productos ORDER BY sku LIMIT 5")
            info["productos"]["primeros_5"] = [{"sku": r[0], "nombre": r[1]} for r in cur.fetchall()]
        except Exception as e:
            info["productos"] = {"error": str(e)}

        # sku_mapeo (legacy)
        try:
            cur.execute("SELECT COUNT(*) FROM sku_mapeo")
            info["sku_mapeo_legacy"] = {"total": cur.fetchone()[0]}
            cur.execute("SELECT sku_lusync, sku_paris, sku_mercadolibre FROM sku_mapeo ORDER BY sku_lusync LIMIT 5")
            info["sku_mapeo_legacy"]["primeros_5"] = [
                {"sku_lusync": r[0], "sku_paris": r[1], "sku_mercadolibre": r[2]}
                for r in cur.fetchall()
            ]
        except Exception as e:
            info["sku_mapeo_legacy"] = {"error": str(e)}

        # sku_mapeo_canal (nuevo)
        try:
            cur.execute("SELECT COUNT(*) FROM sku_mapeo_canal WHERE activo = TRUE")
            info["sku_mapeo_canal"] = {"total": cur.fetchone()[0]}
            cur.execute("SELECT canal, COUNT(*) FROM sku_mapeo_canal WHERE activo = TRUE GROUP BY canal")
            info["sku_mapeo_canal"]["por_canal"] = {r[0]: r[1] for r in cur.fetchall()}
        except Exception as e:
            info["sku_mapeo_canal"] = {"error": str(e)}

        # Bodegas y stock
        try:
            cur.execute("SELECT COUNT(*) FROM bodegas")
            info["bodegas"] = {"total": cur.fetchone()[0]}
            cur.execute("SELECT COUNT(*) FROM stock_bodega WHERE cantidad > 0")
            info["stock_bodega_con_stock"] = cur.fetchone()[0]
        except Exception as e:
            info["bodegas"] = {"error": str(e)}

        # Movimientos
        try:
            cur.execute("SELECT COUNT(*) FROM movimientos")
            info["movimientos"] = {"total": cur.fetchone()[0]}
        except Exception as e:
            info["movimientos"] = {"error": str(e)}

        # Diagnóstico cruzado: ¿cuántos sku_mapeo NO tienen producto correspondiente?
        try:
            cur.execute("""
                SELECT COUNT(DISTINCT m.sku_lusync)
                FROM sku_mapeo m
                LEFT JOIN productos p ON p.sku = m.sku_lusync
                WHERE p.sku IS NULL
            """)
            huerfanos = cur.fetchone()[0]
            info["mapeos_huerfanos"] = {
                "cantidad": huerfanos,
                "explicacion": "SKUs en sku_mapeo que NO tienen producto en tabla 'productos'"
            }
            if huerfanos > 0:
                cur.execute("""
                    SELECT m.sku_lusync FROM sku_mapeo m
                    LEFT JOIN productos p ON p.sku = m.sku_lusync
                    WHERE p.sku IS NULL LIMIT 10
                """)
                info["mapeos_huerfanos"]["ejemplos"] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            info["mapeos_huerfanos"] = {"error": str(e)}

        # Tablas de backup (si has hecho reset)
        try:
            cur.execute("""
                SELECT tablename FROM pg_tables
                WHERE schemaname = 'public'
                  AND (tablename LIKE 'productos_backup_%'
                    OR tablename LIKE 'movimientos_backup_%')
                ORDER BY tablename DESC
                LIMIT 10
            """)
            info["backups"] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            info["backups"] = {"error": str(e)}

        cur.close(); conn.close()

        # Diagnóstico automático
        diagnostico = []
        prod_total = info.get("productos", {}).get("total", 0)
        mapeo_total = info.get("sku_mapeo_legacy", {}).get("total", 0)
        canal_total = info.get("sku_mapeo_canal", {}).get("total", 0)
        huerfanos = info.get("mapeos_huerfanos", {}).get("cantidad", 0)

        if prod_total == 0:
            diagnostico.append("⚠ Tabla 'productos' está VACÍA. Ejecuta /importar_woo para poblar.")
        if mapeo_total > 0 and prod_total == 0:
            diagnostico.append(f"⚠ Tienes {mapeo_total} mapeos en sku_mapeo pero 0 productos. Por eso la UI muestra vacío.")
        if huerfanos > 0:
            diagnostico.append(f"⚠ {huerfanos} mapeos están huérfanos (sin producto en BD).")
        if canal_total > 0 and prod_total == 0:
            diagnostico.append(f"ℹ {canal_total} publicaciones en sku_mapeo_canal listas. Solo falta importar productos.")
        if not diagnostico:
            diagnostico.append("✓ Todo en orden")

        return jsonify({
            "ok": True,
            "tablas": info,
            "diagnostico": diagnostico
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ════════════════════════════════════════════════════════════════════════════
# LIMPIEZA DE DUPLICADOS EN sku_mapeo_canal
# ════════════════════════════════════════════════════════════════════════════
# Detecta y elimina (soft delete) duplicados generados cuando se ejecutó
# auto_mapeo_v2 + migración Excel sobre el mismo dataset.
#
# Reglas:
#   1. Por cada (sku_lusync, canal, sku_canal): si hay 2+ filas, mantener la(s)
#      que tienen item_id_canal y borrar las que no lo tienen.
#   2. Si ninguna tiene item_id, mantener la primera y borrar las demás.
#   3. Por separado: borrar TODAS las entradas del canal 'web' (son redundantes
#      porque el SKU Lusync ya es el SKU Web en este sistema basado en Woo).
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/limpiar_duplicados_mapeo", methods=["GET", "POST"])
def admin_limpiar_duplicados_mapeo():
    """Limpia duplicados en sku_mapeo_canal usando soft delete (activo=FALSE).

    Query params:
      ?dry_run=1     simula sin borrar
      ?incluir_web=0 NO borrar entradas del canal web (default: 1, sí borra)
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    dry_run = request.args.get("dry_run", "0") == "1"
    incluir_web = request.args.get("incluir_web", "1") == "1"

    try:
        from inventario import get_conn

        conn = get_conn(); cur = conn.cursor()

        log = []
        a_borrar_duplicados = []  # ids de filas a marcar como inactivas (duplicados)
        a_borrar_web = []          # ids de filas del canal web a marcar como inactivas
        mantener = []              # ids de filas que sobreviven

        # ── 1. Detectar duplicados por (sku_lusync, canal, sku_canal) ──
        cur.execute("""
            SELECT sku_lusync, canal, sku_canal,
                   ARRAY_AGG(id ORDER BY id) AS ids,
                   ARRAY_AGG(item_id_canal ORDER BY id) AS item_ids,
                   COUNT(*) AS cantidad
            FROM sku_mapeo_canal
            WHERE activo = TRUE
            GROUP BY sku_lusync, canal, sku_canal
            HAVING COUNT(*) > 1
            ORDER BY sku_lusync, canal
        """)
        grupos_duplicados = cur.fetchall()
        log.append(f"Grupos de duplicados encontrados (mismo sku_lusync+canal+sku_canal): {len(grupos_duplicados)}")

        for sku_lusync, canal, sku_canal, ids, item_ids, cantidad in grupos_duplicados:
            # Separar en "con item_id" y "sin item_id"
            con_item = []
            sin_item = []
            for i, mapeo_id in enumerate(ids):
                if item_ids[i]:
                    con_item.append((mapeo_id, item_ids[i]))
                else:
                    sin_item.append(mapeo_id)

            if con_item and sin_item:
                # Caso típico: 1+ con item_id (de auto-mapeo) + 1+ sin item_id (de Excel)
                # Mantener todas las con item_id, borrar las sin item_id
                for mapeo_id in sin_item:
                    a_borrar_duplicados.append({
                        "id": mapeo_id, "sku_lusync": sku_lusync, "canal": canal,
                        "sku_canal": sku_canal, "razon": "duplicado_sin_item_id"
                    })
                for mapeo_id, iid in con_item:
                    mantener.append({
                        "id": mapeo_id, "sku_lusync": sku_lusync, "canal": canal,
                        "sku_canal": sku_canal, "item_id": iid
                    })
                log.append(f"  {sku_lusync}/{canal}/{sku_canal}: mantengo {len(con_item)} con item_id, borro {len(sin_item)} sin item_id")
            elif con_item:
                # Solo con item_id, todos distintos → mantener todos (son publicaciones reales)
                for mapeo_id, iid in con_item:
                    mantener.append({
                        "id": mapeo_id, "sku_lusync": sku_lusync, "canal": canal,
                        "sku_canal": sku_canal, "item_id": iid
                    })
                log.append(f"  {sku_lusync}/{canal}/{sku_canal}: {len(con_item)} todas con item_id distintos, mantengo todas")
            elif sin_item:
                # Solo sin item_id: mantener primera, borrar resto
                mantener.append({
                    "id": sin_item[0], "sku_lusync": sku_lusync, "canal": canal,
                    "sku_canal": sku_canal, "item_id": None
                })
                for mapeo_id in sin_item[1:]:
                    a_borrar_duplicados.append({
                        "id": mapeo_id, "sku_lusync": sku_lusync, "canal": canal,
                        "sku_canal": sku_canal, "razon": "duplicado_sin_item_id_resto"
                    })
                log.append(f"  {sku_lusync}/{canal}/{sku_canal}: {len(sin_item)} sin item_id, mantengo 1 borro {len(sin_item)-1}")

        # ── 2. Borrar entradas del canal 'web' (todas son redundantes) ──
        if incluir_web:
            cur.execute("""
                SELECT id, sku_lusync, sku_canal
                FROM sku_mapeo_canal
                WHERE activo = TRUE AND canal = 'web'
            """)
            for r in cur.fetchall():
                a_borrar_web.append({
                    "id": r[0], "sku_lusync": r[1], "sku_canal": r[2],
                    "razon": "canal_web_redundante"
                })
            log.append(f"Entradas del canal 'web' a borrar: {len(a_borrar_web)} (redundantes con SKU Lusync)")
        else:
            log.append("Canal 'web' preservado (incluir_web=0)")

        total_a_borrar = len(a_borrar_duplicados) + len(a_borrar_web)

        # ── 3. Ejecutar el soft delete (si no es dry_run) ──
        if not dry_run and total_a_borrar > 0:
            ids_a_borrar = [d["id"] for d in a_borrar_duplicados] + [d["id"] for d in a_borrar_web]
            # Soft delete por lotes de 100 para no hacer query gigante
            for i in range(0, len(ids_a_borrar), 100):
                lote = ids_a_borrar[i:i+100]
                cur.execute(f"""
                    UPDATE sku_mapeo_canal
                    SET activo = FALSE, actualizado_at = NOW(),
                        notas = COALESCE(notas, '') || ' | limpieza_duplicados'
                    WHERE id = ANY(%s)
                """, (lote,))
            conn.commit()
            log.append(f"✓ Soft delete ejecutado para {total_a_borrar} filas")
        elif dry_run:
            log.append(f"DRY RUN: no se ejecutó delete (se borrarían {total_a_borrar} filas)")
        else:
            log.append("No hay nada que borrar")

        # ── 4. Estado final ──
        cur.execute("SELECT COUNT(*) FROM sku_mapeo_canal WHERE activo = TRUE")
        activos_despues = cur.fetchone()[0]

        cur.execute("SELECT canal, COUNT(*) FROM sku_mapeo_canal WHERE activo = TRUE GROUP BY canal")
        por_canal_despues = {r[0]: r[1] for r in cur.fetchall()}

        cur.close(); conn.close()

        if not dry_run:
            registrar_audit(
                session.get("usuario", "Sistema"), request.remote_addr,
                "limpiar_duplicados_mapeo",
                detalle=f"borrados={total_a_borrar} (dups={len(a_borrar_duplicados)} web={len(a_borrar_web)})"
            )

        return jsonify({
            "ok": True,
            "dry_run": dry_run,
            "incluir_web": incluir_web,
            "resumen": {
                "duplicados_a_borrar": len(a_borrar_duplicados),
                "web_a_borrar": len(a_borrar_web),
                "total_a_borrar": total_a_borrar,
                "publicaciones_que_quedan": activos_despues if not dry_run else "(dry_run no actualiza)",
                "por_canal_despues": por_canal_despues if not dry_run else None
            },
            "duplicados_a_borrar": a_borrar_duplicados[:30],
            "web_a_borrar_ejemplos": a_borrar_web[:30],
            "log": log[:80]
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


# ════════════════════════════════════════════════════════════════════════════
# RESET DE sku_mapeo_canal Y REPOBLAR DESDE EL EXCEL (sku_mapeo legacy)
# ════════════════════════════════════════════════════════════════════════════
# Borra TODAS las publicaciones de sku_mapeo_canal y las regenera SOLO desde
# lo que está en sku_mapeo (la tabla legacy poblada por el import de Excel).
#
# Esto garantiza que el mapeo refleje EXACTAMENTE lo que el usuario cargó
# en el Excel — sin contaminación del auto-mapeo por similitud de nombre.
#
# Reglas:
#   - HARD DELETE de sku_mapeo_canal (no soft delete, queremos limpieza total)
#   - Para cada fila de sku_mapeo, crear UNA fila por canal con SKU no vacío
#   - EXCLUIR canal 'web' (ya decidiste que es redundante)
#   - Para MELI: si el SKU empieza con MLC, lo guarda como item_id_canal
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# ENRIQUECIMIENTO DE item_id_canal (los IDs específicos de cada marketplace)
# ════════════════════════════════════════════════════════════════════════════
# Después del reset+repoblar desde Excel, todas las publicaciones tienen
# item_id_canal=NULL porque el Excel solo tiene SKUs.
#
# Este endpoint llena los item_id consultando cada marketplace y haciendo match
# por SKU EXACTO (sin similitud por nombre, evitando falsos positivos).
#
# Por canal:
#   - MELI: SELLER_SKU del attribute → item_id es el MLC...
#   - Falabella: shop_sku numérico (ej: 116363873) por SellerSku
#   - Ripley: product_id Mirakl por shop_sku
#   - Walmart: wpid o productId por sku del seller
#   - París: el sellerSku ES el SKU mismo, raramente hay item_id distinto
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# IMPORTAR STOCK MELI FULL DESDE EXCEL
# ════════════════════════════════════════════════════════════════════════════
# Procesa el reporte de stock que MELI exporta:
#  - Columna "SKU"                → identificador del producto
#  - Columna "Aptas para vender"  → stock REAL en MELI Full (vendible)
#  - Columna "En camino a Full"   → stock en tránsito (entrará pronto)
#
# UI amigable: /admin/cargar_full_meli_ui (formulario web)
# API directa: /admin/importar_stock_full_meli (POST con JSON o archivo)
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/cargar_full_meli_ui", methods=["GET"])
def admin_cargar_full_meli_ui():
    """UI HTML para cargar stock Full MELI fácilmente desde el navegador."""
    html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Cargar Stock MELI Full · Lusync</title>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui, sans-serif;
    margin: 0; padding: 20px; background: #f5f4ed; color: #2c2c2c; line-height: 1.5;
  }
  .container { max-width: 900px; margin: 0 auto; }
  h1 { font-size: 22px; font-weight: 600; margin-bottom: 6px; }
  .subtitle { color: #888; font-size: 13px; margin-bottom: 24px; }
  .card {
    background: #fff; border: 1px solid #e8e6dc; border-radius: 12px;
    padding: 20px; margin-bottom: 16px;
  }
  label { display: block; font-size: 12px; font-weight: 600; color: #555; margin-bottom: 6px; text-transform: uppercase; letter-spacing: 0.5px; }
  input[type="text"], textarea {
    width: 100%; padding: 10px 12px; font-size: 13px;
    border: 1px solid #ddd; border-radius: 6px; font-family: ui-monospace, monospace;
    background: #fafaf6;
  }
  textarea { min-height: 280px; resize: vertical; }
  input[type="text"]:focus, textarea:focus { outline: 2px solid #6366f1; }
  .btn {
    padding: 12px 20px; font-size: 14px; font-weight: 600; border: none;
    border-radius: 8px; cursor: pointer; margin-right: 8px; margin-top: 12px;
    transition: opacity 0.2s;
  }
  .btn:disabled { opacity: 0.5; cursor: not-allowed; }
  .btn-test { background: #fff3cd; color: #856404; border: 1px solid #ffc107; }
  .btn-real { background: #155724; color: #fff; }
  .btn:hover:not(:disabled) { opacity: 0.85; }
  .resultado {
    background: #fafaf6; border-left: 3px solid #6366f1; padding: 16px;
    border-radius: 6px; font-size: 13px; margin-top: 16px; display: none;
  }
  .resultado.ok { border-left-color: #155724; background: #d4edda; }
  .resultado.error { border-left-color: #dc3545; background: #f8d7da; }
  .resultado pre { background: rgba(0,0,0,0.05); padding: 10px; border-radius: 4px; overflow-x: auto; font-size: 11px; }
  .stat { display: inline-block; padding: 6px 12px; background: #fff; border-radius: 99px; margin: 4px; font-size: 12px; border: 1px solid #ddd; }
  .help { font-size: 12px; color: #888; margin-top: 8px; }
  .badge-warn { background: #fff3cd; color: #856404; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }
</style>
</head>
<body>
<div class="container">
  <h1>📦 Cargar Stock MELI Full</h1>
  <p class="subtitle">Importa el stock real de MercadoLibre Full a Lusync. Separa "Aptas para vender" y "En camino a Full" en bodegas distintas.</p>

  <div class="card">
    <label>🔑 Token de acceso</label>
    <input type="text" id="token" placeholder="Pega aquí tu token..." value="lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw">
    <p class="help">Token para acceder al endpoint sin login. Cambialo en Render como variable ADMIN_BYPASS_TOKEN si quieres.</p>
  </div>

  <div class="card">
    <label>📋 Datos a cargar (JSON con SKUs)</label>
    <textarea id="datos">{
  "items": [
    {"sku": "CDBUCVN001", "aptas": 0,  "transito": 2},
    {"sku": "ODJ3NR001",  "aptas": 9,  "transito": 5},
    {"sku": "CDBRWD001",  "aptas": 2,  "transito": 0},
    {"sku": "CCCN001",    "aptas": 10, "transito": 0},
    {"sku": "ODJ4N001",   "aptas": 9,  "transito": 10},
    {"sku": "ODJM001",    "aptas": 0,  "transito": 30},
    {"sku": "CTSECNSB001","aptas": 0,  "transito": 2},
    {"sku": "CDBRWN001",  "aptas": 8,  "transito": 0},
    {"sku": "ODJ3NA001",  "aptas": 8,  "transito": 5},
    {"sku": "ODJ3N001",   "aptas": 26, "transito": 20},
    {"sku": "CDBRWB001",  "aptas": 1,  "transito": 0},
    {"sku": "CBSNBPN001", "aptas": 3,  "transito": 0},
    {"sku": "ODJA001",    "aptas": 1,  "transito": 5},
    {"sku": "GPPLA001",   "aptas": 1,  "transito": 0},
    {"sku": "SDCEMR001",  "aptas": 0,  "transito": 1},
    {"sku": "SDCEG001",   "aptas": 2,  "transito": 0},
    {"sku": "SDCED001",   "aptas": 3,  "transito": 0},
    {"sku": "PBEAMR001",  "aptas": 7,  "transito": 3},
    {"sku": "GPPLR001",   "aptas": 1,  "transito": 2},
    {"sku": "CBRMSCN001", "aptas": 7,  "transito": 0},
    {"sku": "SDCER001",   "aptas": 2,  "transito": 0},
    {"sku": "EDLABA001",  "aptas": 4,  "transito": 0},
    {"sku": "EDLABR001",  "aptas": 2,  "transito": 0},
    {"sku": "MAD003",     "aptas": 8,  "transito": 0},
    {"sku": "MAD005",     "aptas": 10, "transito": 5},
    {"sku": "MAD006",     "aptas": 19, "transito": 0},
    {"sku": "MAD004",     "aptas": 12, "transito": 8},
    {"sku": "CBRMLRR001", "aptas": 6,  "transito": 0},
    {"sku": "PBEAMG001",  "aptas": 0,  "transito": 3},
    {"sku": "SDCR2021",   "aptas": 34, "transito": 0}
  ]
}</textarea>
    <p class="help">
      <span class="badge-warn">PRE-CARGADO</span> Estos son tus 30 SKUs Full. Puedes editarlos si necesitas.
      <strong>aptas</strong> = stock vendible HOY · <strong>transito</strong> = en camino a Full.
    </p>

    <button class="btn btn-test" onclick="ejecutar(true)">⚠️ Probar primero (sin escribir)</button>
    <button class="btn btn-real" onclick="ejecutar(false)" id="btnReal" disabled>✅ Cargar de verdad</button>
    <p class="help">Primero prueba (dry-run). Si todo se ve bien, se habilita el botón verde.</p>
  </div>

  <div class="resultado" id="resultado"></div>
</div>

<script>
function ejecutar(dryRun) {
  const token = document.getElementById('token').value.trim();
  const datosTxt = document.getElementById('datos').value.trim();
  const resultado = document.getElementById('resultado');
  const btnReal = document.getElementById('btnReal');

  if (!token) { alert('Falta el token'); return; }
  let datos;
  try { datos = JSON.parse(datosTxt); }
  catch(e) { alert('JSON inválido: ' + e.message); return; }

  resultado.style.display = 'block';
  resultado.className = 'resultado';
  resultado.innerHTML = '<strong>⏳ Procesando...</strong>';

  const url = '/admin/importar_stock_full_meli?token=' + encodeURIComponent(token) + (dryRun ? '&dry_run=1' : '');

  fetch(url, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: datosTxt
  })
  .then(r => r.json())
  .then(res => {
    if (res.ok) {
      resultado.className = 'resultado ok';
      let html = dryRun
        ? '<strong>✅ Prueba OK (sin escribir nada)</strong><br><br>'
        : '<strong>✅ ¡CARGA REAL EXITOSA!</strong><br><br>';
      const r = res.resumen;
      html += '<div>';
      html += '<span class="stat">📦 Items procesados: <strong>' + r.items_en_excel + '</strong></span>';
      html += '<span class="stat">✓ Matched: <strong>' + r.matched_con_lusync + '</strong></span>';
      html += '<span class="stat">⚠️ No matched: <strong>' + r.no_matched + '</strong></span>';
      html += '<span class="stat">📊 Stock Aptas: <strong>' + r.stock_aptas_total + '</strong> u.</span>';
      html += '<span class="stat">🚚 Stock Tránsito: <strong>' + r.stock_transito_total + '</strong> u.</span>';
      html += '</div>';
      if (res.no_matched && res.no_matched.length > 0) {
        html += '<br><strong>SKUs que no matchean (revisar):</strong><pre>' + JSON.stringify(res.no_matched, null, 2) + '</pre>';
      }
      if (res.errores && res.errores.length > 0) {
        html += '<br><strong>Errores:</strong><pre>' + JSON.stringify(res.errores, null, 2) + '</pre>';
      }
      if (dryRun) {
        html += '<br><strong>👉 Si todo se ve bien, presiona "✅ Cargar de verdad"</strong>';
        btnReal.disabled = false;
      } else {
        html += '<br><strong>🎉 Stock cargado correctamente. Ya puedes cerrar esta página.</strong>';
        html += '<br><br>Verifica en: <a href="/bodegas">Bodegas</a> · <a href="/productos">Productos</a>';
      }
      resultado.innerHTML = html;
    } else {
      resultado.className = 'resultado error';
      resultado.innerHTML = '<strong>❌ Error:</strong> ' + (res.error || 'desconocido') + '<br><pre>' + JSON.stringify(res, null, 2) + '</pre>';
    }
  })
  .catch(err => {
    resultado.className = 'resultado error';
    resultado.innerHTML = '<strong>❌ Error de conexión:</strong> ' + err.message;
  });
}
</script>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


# ════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO DE STOCK POR SKU
# ════════════════════════════════════════════════════════════════════════════
# Te dice TODO sobre el stock de un SKU:
#  - Stock legacy (campo productos.stock)
#  - Stock por bodega (CENTRAL, MELI_FULL, etc.)
#  - Suma de bodegas vs legacy (¿son consistentes?)
#  - Últimos 20 movimientos (con stock_antes, stock_despues, bodega)
#
# Uso: /admin/diagnostico_stock?sku=GPPLA001&token=TU_TOKEN
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# REPROCESAR ÓRDENES ESPECÍFICAS (recuperación manual)
# ════════════════════════════════════════════════════════════════════════════
# Permite forzar el procesamiento de órdenes que se perdieron por algún motivo
# (bug, caída de servidor, webhook perdido, etc.).
#
# Funciona así:
# 1. Recibe una lista de orden_ids + canal (meli, walmart, etc.)
# 2. Borra la marca "ya_procesada" de cada una
# 3. Llama al scheduler del canal correspondiente para que las re-procese
#
# Uso desde browser:
#   /admin/reprocesar_ordenes?canal=meli&order_ids=2000012841095175,2000012839784797&token=XXX
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/reprocesar_ordenes", methods=["GET", "POST"])
def admin_reprocesar_ordenes():
    """Borra marcas de órdenes específicas para que el scheduler las re-procese."""
    # Bypass por token
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return jsonify({"error": "no autorizado"}), 401
    
    # Acepta GET con query params o POST con JSON
    if request.method == "POST" and request.is_json:
        data = request.get_json()
        canal = (data.get("canal") or "").lower().strip()
        order_ids_raw = data.get("order_ids") or ""
    else:
        canal = (request.args.get("canal") or "").lower().strip()
        order_ids_raw = request.args.get("order_ids") or ""
    
    # Parsear orden_ids (separados por coma)
    if isinstance(order_ids_raw, list):
        order_ids = [str(x).strip() for x in order_ids_raw if x]
    else:
        order_ids = [s.strip() for s in str(order_ids_raw).split(",") if s.strip()]
    
    if not canal or not order_ids:
        return jsonify({
            "error": "Faltan parámetros",
            "uso": {
                "canal": "meli, walmart, falabella, paris, ripley, woo",
                "order_ids": "lista separada por comas (ej: 2000012841095175,2000012839784797)",
                "ejemplo_url": "/admin/reprocesar_ordenes?canal=meli&order_ids=12345,67890&token=XXX"
            }
        }), 400
    
    # Mapeo de canales a prefijos de keys
    prefijos = {
        "meli":      ["MELI-", "MELI-CANCEL-"],
        "walmart":   ["WM-", "WM-CANCEL-"],
        "falabella": ["FALABELLA-", "FALABELLA-CANCEL-"],
        "paris":     ["PA-", "PA-CANCEL-"],
        "ripley":    ["RP-", "RP-CANCEL-"],
        "woo":       ["WOO-", "WOO-CANCEL-"],
    }
    
    if canal not in prefijos:
        return jsonify({
            "error": f"Canal '{canal}' no reconocido",
            "validos": list(prefijos.keys())
        }), 400
    
    # ── Borrar marcas de cada orden ──
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    
    resultados = {
        "canal": canal,
        "order_ids_solicitadas": order_ids,
        "marcas_borradas": [],
        "errores": []
    }
    
    try:
        for order_id in order_ids:
            order_id_clean = str(order_id).strip()
            for prefijo in prefijos[canal]:
                key = f"{prefijo}{order_id_clean}"
                try:
                    cur.execute("""
                        DELETE FROM ordenes_procesadas 
                        WHERE order_id_texto = %s
                        RETURNING order_id_texto
                    """, (key,))
                    r = cur.fetchall()
                    if r:
                        resultados["marcas_borradas"].append(key)
                except Exception as e:
                    resultados["errores"].append(f"{key}: {str(e)[:80]}")
        conn.commit()
    except Exception as e:
        conn.rollback()
        cur.close(); conn.close()
        return jsonify({"error": f"Error general: {e}"}), 500
    
    cur.close(); conn.close()
    
    # ── Disparar el scheduler correspondiente ──
    schedulers = {
        "meli":      _sync_meli_automatico,
        "walmart":   _sync_walmart_automatico,
        "falabella": _sync_falabella_automatico,
        "paris":     _sync_paris_automatico,
        "ripley":    _sync_ripley_automatico,
        "woo":       _sync_woo_automatico,
    }
    
    scheduler_func = schedulers.get(canal)
    if scheduler_func:
        try:
            # Ejecutar en thread para no bloquear la respuesta
            import threading
            t = threading.Thread(target=scheduler_func, daemon=True)
            t.start()
            resultados["scheduler_ejecutado"] = True
            resultados["mensaje"] = f"Scheduler {canal} ejecutándose en background. Las órdenes serán procesadas en los próximos segundos."
        except Exception as e:
            resultados["errores"].append(f"Scheduler: {str(e)[:80]}")
    
    # Audit log
    try:
        registrar_audit(
            session.get("usuario", "Sistema (token)"), request.remote_addr,
            "reprocesar_ordenes",
            detalle=f"canal={canal} ordenes={','.join(order_ids[:10])} marcas_borradas={len(resultados['marcas_borradas'])}"
        )
    except: pass
    
    return jsonify(resultados)


# ════════════════════════════════════════════════════════════════════════════
# ENDPOINT: MOVIMIENTOS POR RANGO DE FECHAS
# ════════════════════════════════════════════════════════════════════════════
# Devuelve TODOS los movimientos en un rango de fechas (sin tope de 200/300).
# Útil para auditoría cruzada con Excels exportados de marketplaces.
#
# Uso desde browser:
#   /admin/movimientos_por_fecha?desde=2026-05-01&hasta=2026-05-06&token=XXX
#   /admin/movimientos_por_fecha?desde=2026-05-01&hasta=2026-05-06&canal=mercadolibre&token=XXX
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# ENDPOINT: FORZAR SYNC MELI EN RANGO DE FECHAS
# ════════════════════════════════════════════════════════════════════════════
# Trae TODAS las órdenes MELI en un rango de fechas y las procesa.
# Útil para recuperar órdenes históricas que no llegaron por webhook ni scheduler.
#
# Uso:
#   /admin/sync_meli_rango?desde=2026-05-01&hasta=2026-05-06&token=XXX
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# ENDPOINT: FORZAR SYNC DE CUALQUIER CANAL (genérico)
# ════════════════════════════════════════════════════════════════════════════
# Fuerza la ejecución del scheduler de un canal específico.
# Útil para recuperar órdenes después de bugs o caídas.
#
# Uso:
#   /admin/forzar_sync_canal?canal=walmart&token=XXX
#   /admin/forzar_sync_canal?canal=falabella&token=XXX
#   /admin/forzar_sync_canal?canal=paris&token=XXX
#   /admin/forzar_sync_canal?canal=ripley&token=XXX
#   /admin/forzar_sync_canal?canal=woo&token=XXX
# ════════════════════════════════════════════════════════════════════════════

@app.route("/admin/forzar_sync_canal", methods=["GET"])
def admin_forzar_sync_canal():
    """Fuerza la ejecución del scheduler de un canal específico (en background)."""
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return jsonify({"error": "no autorizado"}), 401
    
    canal = (request.args.get("canal") or "").strip().lower()
    
    schedulers_disponibles = {
        "walmart": _sync_walmart_automatico,
        "falabella": _sync_falabella_automatico,
        "paris": _sync_paris_automatico,
        "ripley": _sync_ripley_automatico,
        "woo": _sync_woo_automatico,
        "meli": _sync_meli_automatico,
    }
    
    if canal not in schedulers_disponibles:
        return jsonify({
            "error": f"Canal '{canal}' no válido",
            "canales_validos": list(schedulers_disponibles.keys()),
            "ejemplo": "/admin/forzar_sync_canal?canal=walmart&token=XXX"
        }), 400
    
    # Liberar lock si quedó pegado
    if canal in _sync_locks:
        _sync_locks[canal]["running"] = False
    
    # Ejecutar scheduler en background
    import threading
    sched_func = schedulers_disponibles[canal]
    t = threading.Thread(target=sched_func, daemon=True)
    t.start()
    
    return jsonify({
        "canal": canal,
        "scheduler_ejecutado": True,
        "mensaje": f"Scheduler {canal} corriendo en background. Las órdenes se procesarán en los próximos segundos.",
        "verificar_resultado": f"/admin/movimientos_por_fecha?desde=2026-05-01&hasta=2026-05-06&canal={canal}&token={bypass_token}"
    })


@app.route("/admin/sync_meli_rango", methods=["GET"])
def admin_sync_meli_rango():
    """Fuerza sync de órdenes MELI en rango de fechas específico (con paginación)."""
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return jsonify({"error": "no autorizado"}), 401
    
    desde = request.args.get("desde", "")  # YYYY-MM-DD
    hasta = request.args.get("hasta", "")  # YYYY-MM-DD
    
    if not desde or not hasta:
        return jsonify({
            "error": "Faltan desde/hasta (formato YYYY-MM-DD)",
            "ejemplo": "/admin/sync_meli_rango?desde=2026-05-01&hasta=2026-05-06&token=XXX"
        }), 400
    
    # Convertir a formato MELI (ISO con timezone Chile UTC-4)
    date_from = f"{desde}T00:00:00.000-04:00"
    date_to = f"{hasta}T23:59:59.999-04:00"
    
    try:
        from mercadolibre import obtener_todas_ordenes_meli_rango
        from inventario import descontar_venta_inteligente, detectar_fulfillment_meli, intentar_marcar_orden_atomic, orden_ya_procesada_texto, marcar_orden_procesada_texto, obtener_sku_lusync_por_canal
        from datetime import datetime
        
        # Traer todas las órdenes del rango (con paginación)
        print(f"[Sync MELI Rango] Trayendo órdenes desde {date_from} hasta {date_to}")
        ordenes = obtener_todas_ordenes_meli_rango(date_from, date_to, max_paginas=20)
        print(f"[Sync MELI Rango] Total órdenes obtenidas: {len(ordenes)}")
        
        nuevas = 0
        ya_procesadas = 0
        canceladas = 0
        errores = []
        ordenes_procesadas_ids = []
        
        for o in ordenes:
            try:
                order_id = str(o.get("id", ""))
                estado = o.get("status", "")
                meli_key = f"MELI-{order_id}"
                cancel_key = f"MELI-CANCEL-{order_id}"
                
                # ── Órdenes pagadas ──
                if estado in ("paid", "confirmed"):
                    if orden_ya_procesada_texto(meli_key):
                        ya_procesadas += 1
                        continue
                    
                    fecha_compra = None
                    try:
                        ds = (o.get("date_created", "") or "").replace("Z", "+00:00")
                        if ds:
                            fecha_compra = datetime.fromisoformat(ds)
                    except: pass
                    
                    es_full = detectar_fulfillment_meli(o)
                    items_descontados = []
                    
                    for item in o.get("order_items", []):
                        item_data = item.get("item", {})
                        item_id = item_data.get("id", "")
                        sku_seller = (
                            (item_data.get("seller_sku") or "").strip()
                            or (item_data.get("seller_custom_field") or "").strip()
                        )
                        cantidad = int(item.get("quantity", 1))
                        if not sku_seller:
                            continue
                        
                        # Traducir SKU canal a Lusync
                        try:
                            sku_lusync = obtener_sku_lusync_por_canal("mercadolibre", sku_canal=sku_seller, item_id_canal=item_id) or sku_seller
                        except Exception:
                            sku_lusync = sku_seller
                        
                        try:
                            descontar_venta_inteligente(
                                sku=sku_lusync,
                                cantidad=cantidad,
                                canal="mercadolibre",
                                fulfillment=es_full,
                                orden_id=order_id,
                                fecha_compra_marketplace=fecha_compra
                            )
                            items_descontados.append({"sku_canal": sku_seller, "sku_lusync": sku_lusync, "cantidad": cantidad})
                        except Exception as e:
                            errores.append(f"{order_id}/{sku_seller}→{sku_lusync}: {str(e)[:100]}")
                    
                    if items_descontados:
                        marcar_orden_procesada_texto(meli_key)
                        nuevas += 1
                        ordenes_procesadas_ids.append({"id": order_id, "items": items_descontados, "full": es_full})
                
                # ── Órdenes canceladas ──
                elif estado in ("cancelled", "canceled"):
                    if orden_ya_procesada_texto(cancel_key):
                        continue
                    if not orden_ya_procesada_texto(meli_key):
                        # Nunca se procesó la venta, solo marcar
                        marcar_orden_procesada_texto(cancel_key)
                        continue
                    canceladas += 1
                    marcar_orden_procesada_texto(cancel_key)
            except Exception as e:
                errores.append(f"Orden {order_id}: {str(e)[:100]}")
        
        return jsonify({
            "rango": {"desde": desde, "hasta": hasta},
            "total_ordenes_meli": len(ordenes),
            "procesadas_ahora": nuevas,
            "ya_procesadas_antes": ya_procesadas,
            "canceladas": canceladas,
            "errores_count": len(errores),
            "errores": errores[:30],  # solo primeros 30
            "ordenes_procesadas_ahora": ordenes_procesadas_ids[:50]  # solo primeras 50
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()[:500]}), 500


@app.route("/admin/movimientos_por_fecha", methods=["GET"])
def admin_movimientos_por_fecha():
    """Devuelve movimientos filtrados por rango de fechas (sin tope)."""
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return jsonify({"error": "no autorizado"}), 401
    
    desde = request.args.get("desde", "")  # YYYY-MM-DD
    hasta = request.args.get("hasta", "")  # YYYY-MM-DD
    canal_filtro = (request.args.get("canal") or "").strip().lower()
    
    if not desde or not hasta:
        return jsonify({
            "error": "Faltan parámetros desde/hasta",
            "uso": {
                "desde": "fecha desde formato YYYY-MM-DD (ej: 2026-05-01)",
                "hasta": "fecha hasta formato YYYY-MM-DD (ej: 2026-05-06)",
                "canal": "(opcional) filtrar por canal: mercadolibre, walmart, falabella, paris, ripley, web",
                "ejemplo": "/admin/movimientos_por_fecha?desde=2026-05-01&hasta=2026-05-06&token=XXX"
            }
        }), 400
    
    from inventario import get_conn
    conn = get_conn()
    cur = conn.cursor()
    
    try:
        # Asegurar columnas
        try:
            cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS bodega_codigo TEXT DEFAULT 'CENTRAL'")
            cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS fecha_compra_marketplace TIMESTAMP")
            cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen_registro TEXT DEFAULT 'sistema'")
            cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_antes INTEGER")
            cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_despues INTEGER")
            conn.commit()
        except: conn.rollback()
        
        # Query con filtro de fechas
        query = """
            SELECT tipo, sku, nombre, cantidad, motivo,
                   TO_CHAR(fecha, 'DD/MM/YYYY'),
                   TO_CHAR(fecha, 'HH24:MI'),
                   COALESCE(usuario, 'Sistema'),
                   COALESCE(canal, 'Sistema'),
                   COALESCE(orden_id, ''),
                   COALESCE(bodega_codigo, 'CENTRAL'),
                   TO_CHAR(fecha_compra_marketplace, 'DD/MM/YYYY HH24:MI'),
                   COALESCE(origen_registro, 'sistema'),
                   stock_antes,
                   stock_despues,
                   TO_CHAR(fecha, 'YYYY-MM-DD HH24:MI:SS')
            FROM movimientos
            WHERE fecha::date >= %s::date AND fecha::date <= %s::date
        """
        params = [desde, hasta]
        
        if canal_filtro:
            query += " AND LOWER(COALESCE(canal, '')) LIKE %s"
            params.append(f"%{canal_filtro}%")
        
        query += " ORDER BY fecha DESC"
        
        cur.execute(query, params)
        rows = cur.fetchall()
        
        movimientos = []
        for r in rows:
            movimientos.append({
                "tipo": r[0], "sku": r[1], "nombre": r[2], "cantidad": r[3],
                "motivo": r[4], "fecha": r[5], "hora": r[6], "usuario": r[7],
                "canal": r[8], "orden_id": r[9] or "", "bodega": r[10],
                "fecha_compra": r[11] or "", "origen": r[12],
                "stock_antes": r[13], "stock_despues": r[14], "fecha_iso": r[15]
            })
        
        # Estadísticas
        canales_count = {}
        ids_por_canal = {}
        for m in movimientos:
            c = m["canal"]
            canales_count[c] = canales_count.get(c, 0) + 1
            if c not in ids_por_canal: ids_por_canal[c] = set()
            if m["orden_id"]: ids_por_canal[c].add(m["orden_id"])
        
        ids_por_canal_lista = {k: sorted(list(v)) for k, v in ids_por_canal.items()}
        
        cur.close(); conn.close()
        
        return jsonify({
            "rango": {"desde": desde, "hasta": hasta, "canal": canal_filtro or "todos"},
            "total_movimientos": len(movimientos),
            "movimientos_por_canal": canales_count,
            "ordenes_unicas_por_canal": {k: len(v) for k, v in ids_por_canal.items()},
            "ids_por_canal": ids_por_canal_lista,
            "movimientos": movimientos
        })
    except Exception as e:
        try: cur.close(); conn.close()
        except: pass
        return jsonify({"error": str(e)}), 500


@app.route("/admin/diagnostico_stock", methods=["GET"])
def admin_diagnostico_stock():
    """Diagnóstico completo del stock de un SKU."""
    # Bypass por token (igual que importar)
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    if not autorizado:
        return jsonify({"error": "no autorizado"}), 401
    
    sku = (request.args.get("sku") or "").strip().upper()
    if not sku:
        return jsonify({"error": "Falta parámetro ?sku=XXX"}), 400
    
    try:
        from inventario import get_conn, listar_bodegas, get_stock_bodega
        
        conn = get_conn(); cur = conn.cursor()
        
        # ── 1. Datos del producto ──
        cur.execute("SELECT sku, nombre, stock FROM productos WHERE UPPER(sku)=%s LIMIT 1", (sku,))
        r = cur.fetchone()
        if not r:
            cur.close(); conn.close()
            return jsonify({"error": f"SKU '{sku}' no existe en productos"}), 404
        
        sku_real, nombre, stock_legacy = r
        
        # ── 2. Stock por bodega ──
        bodegas_info = {}
        suma_bodegas = 0
        try:
            for b in listar_bodegas(solo_activas=False):
                cant = get_stock_bodega(sku_real, b["codigo"]) or 0
                bodegas_info[b["codigo"]] = {
                    "nombre": b["nombre"],
                    "tipo": b.get("tipo", "?"),
                    "stock": cant
                }
                suma_bodegas += cant
        except Exception as e:
            bodegas_info["error"] = str(e)
        
        # ── 3. Últimos movimientos ──
        cur.execute("""
            SELECT fecha, tipo, cantidad, motivo, canal, bodega_codigo, 
                   stock_antes, stock_despues, orden_id, usuario, origen_registro
            FROM movimientos 
            WHERE sku=%s 
            ORDER BY fecha DESC 
            LIMIT 20
        """, (sku_real,))
        movimientos = []
        for m in cur.fetchall():
            movimientos.append({
                "fecha": m[0].strftime("%Y-%m-%d %H:%M:%S") if m[0] else None,
                "tipo": m[1],
                "cantidad": m[2],
                "motivo": m[3],
                "canal": m[4],
                "bodega": m[5],
                "stock_antes": m[6],
                "stock_despues": m[7],
                "orden_id": m[8],
                "usuario": m[9],
                "origen": m[10]
            })
        
        cur.close(); conn.close()
        
        # ── 4. Análisis de consistencia ──
        consistente = (suma_bodegas == stock_legacy)
        
        return jsonify({
            "sku": sku_real,
            "nombre": nombre,
            "stock_legacy_productos": stock_legacy,
            "stock_total_bodegas": suma_bodegas,
            "consistente": consistente,
            "diferencia": stock_legacy - suma_bodegas,
            "bodegas": bodegas_info,
            "ultimos_20_movimientos": movimientos,
            "ayuda": {
                "stock_legacy_productos": "Campo `stock` en tabla productos (sistema viejo)",
                "stock_total_bodegas": "Suma de stock_bodega para todas las bodegas",
                "consistente": "True si ambos suman lo mismo (deberían)",
                "stock_antes/stock_despues": "Snapshot de bodega ANTES y DESPUÉS del movimiento"
            }
        })
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "trace": traceback.format_exc()
        }), 500


@app.route("/admin/importar_stock_full_meli", methods=["GET", "POST"])
def admin_importar_stock_full_meli():
    """Importa stock MELI Full desde el Excel oficial de MELI.
    
    Acepta:
    - GET: muestra instrucciones / preview de SKUs en Lusync
    - POST con archivo: importa
      - file=<archivo.xlsx o .csv>
    - POST con JSON: importa desde lista
      - {"items": [{"sku": "ODJM001", "aptas": 0, "transito": 30}, ...]}
    
    Query params:
    - dry_run=1: simula sin escribir
    - token=XXX: bypass de login (para carga inicial via curl/script)
              Token configurable via env ADMIN_BYPASS_TOKEN
              Default: lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw
    """
    # Verificación: login normal O token válido
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token_recibido = request.args.get("token", "")
    autorizado = session.get("logged") or (token_recibido and token_recibido == bypass_token)
    
    if not autorizado:
        return jsonify({
            "error": "no autorizado",
            "como_autenticarse": "logueate en /login O agrega ?token=TU_TOKEN al URL"
        }), 401
    
    dry_run = request.args.get("dry_run", "0") == "1"
    
    # GET → instrucciones
    if request.method == "GET":
        productos = cargar_productos()
        return jsonify({
            "instrucciones": [
                "Sube el Excel exportado de MELI Full (botón 'Exportar' en la sección Stock Full)",
                "El Excel debe tener las columnas: SKU, Aptas para vender, En camino a Full",
                "POST este endpoint con archivo en form-data 'file' o JSON {items: [...]}"
            ],
            "skus_lusync_disponibles": len(productos),
            "skus_ejemplo": [p["sku"] for p in productos[:10]],
            "agregar_dry_run": "?dry_run=1 para simular sin escribir"
        })
    
    # POST → procesar
    items_a_procesar = []
    
    try:
        # ── Modo 1: archivo subido ──
        if "file" in request.files:
            file = request.files["file"]
            filename = (file.filename or "").lower()
            
            try:
                if filename.endswith((".xlsx", ".xls")):
                    import openpyxl
                    from io import BytesIO
                    wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
                    ws = wb.active
                    
                    # Encontrar columnas por nombre (case-insensitive, con/sin acentos)
                    headers_row = None
                    for row_idx in range(1, 6):  # buscar headers en primeras 5 filas
                        row = [str(c.value or "").strip() for c in ws[row_idx]]
                        if any("sku" in h.lower() for h in row) and any("apta" in h.lower() for h in row):
                            headers_row = row_idx
                            break
                    
                    if headers_row is None:
                        return jsonify({"ok": False, "error": "No se encontró fila de headers con 'SKU' y 'Aptas'"}), 400
                    
                    headers = [str(c.value or "").strip().lower() for c in ws[headers_row]]
                    
                    # Identificar índices de columnas
                    col_sku = next((i for i, h in enumerate(headers) if h == "sku"), None)
                    col_aptas = next((i for i, h in enumerate(headers) if "apta" in h and "vender" in h), None)
                    col_transito = next((i for i, h in enumerate(headers) if "camino" in h and "full" in h), None)
                    
                    if col_sku is None or col_aptas is None:
                        return jsonify({
                            "ok": False,
                            "error": "Columnas requeridas no encontradas",
                            "headers_detectados": headers,
                            "necesarias": ["SKU", "Aptas para vender (en Unidades en Full)"]
                        }), 400
                    
                    # Leer filas
                    for row in ws.iter_rows(min_row=headers_row+1, values_only=True):
                        if not row or not row[col_sku]: continue
                        sku = str(row[col_sku]).strip()
                        if not sku: continue
                        try:
                            aptas = int(row[col_aptas] or 0)
                        except (ValueError, TypeError):
                            aptas = 0
                        try:
                            transito = int(row[col_transito] or 0) if col_transito is not None else 0
                        except (ValueError, TypeError):
                            transito = 0
                        items_a_procesar.append({"sku": sku, "aptas": aptas, "transito": transito})
                
                elif filename.endswith(".csv") or filename.endswith(".tsv"):
                    import csv
                    from io import StringIO
                    sep = "\t" if filename.endswith(".tsv") else ","
                    text = file.read().decode("utf-8-sig")
                    reader = csv.DictReader(StringIO(text), delimiter=sep)
                    for row in reader:
                        # Normalizar keys
                        row_norm = {k.lower().strip(): v for k, v in row.items() if k}
                        sku = (row_norm.get("sku") or "").strip()
                        if not sku: continue
                        # Buscar columnas con nombre flexible
                        aptas_val = None
                        transito_val = None
                        for k, v in row_norm.items():
                            if "apta" in k and "vender" in k:
                                aptas_val = v
                            elif "camino" in k and "full" in k:
                                transito_val = v
                        try: aptas = int(aptas_val or 0)
                        except: aptas = 0
                        try: transito = int(transito_val or 0)
                        except: transito = 0
                        items_a_procesar.append({"sku": sku, "aptas": aptas, "transito": transito})
                else:
                    return jsonify({"ok": False, "error": "Formato no soportado. Usa .xlsx, .xls, .csv o .tsv"}), 400
            except Exception as e:
                import traceback
                return jsonify({
                    "ok": False, "error": f"Error parseando archivo: {e}",
                    "trace": traceback.format_exc()
                }), 400
        
        # ── Modo 2: JSON directo ──
        elif request.is_json:
            data = request.get_json()
            items_a_procesar = data.get("items", [])
        
        else:
            return jsonify({"ok": False, "error": "Sube file=<archivo> o JSON con items[]"}), 400
        
        if not items_a_procesar:
            return jsonify({"ok": False, "error": "No se encontraron items para procesar"}), 400
        
        # ── Verificar bodegas existen ──
        from inventario import (init_bodegas, set_stock_bodega, listar_bodegas,
                                cargar_productos as _cp)
        init_bodegas()  # Crea MELI_FULL y MELI_FULL_TRANSITO si faltan
        
        # ── Procesar items ──
        productos_lusync = {p["sku"].upper(): p for p in _cp()}
        
        resultados = {
            "matched": [],
            "no_matched": [],
            "actualizados_full": 0,
            "actualizados_transito": 0,
            "total_aptas": 0,
            "total_transito": 0,
            "errores": []
        }
        
        for item in items_a_procesar:
            sku = (item.get("sku") or "").strip()
            aptas = int(item.get("aptas", 0) or 0)
            transito = int(item.get("transito", 0) or 0)
            
            if not sku: continue
            
            # Match exacto (case-insensitive)
            sku_upper = sku.upper()
            if sku_upper not in productos_lusync:
                resultados["no_matched"].append({
                    "sku_excel": sku, "aptas": aptas, "transito": transito
                })
                continue
            
            sku_real = productos_lusync[sku_upper]["sku"]
            
            try:
                if not dry_run:
                    # Stock APTAS PARA VENDER → MELI_FULL
                    set_stock_bodega(sku_real, "MELI_FULL", aptas)
                    resultados["actualizados_full"] += 1
                    # Stock EN CAMINO → MELI_FULL_TRANSITO
                    set_stock_bodega(sku_real, "MELI_FULL_TRANSITO", transito)
                    resultados["actualizados_transito"] += 1
                
                resultados["matched"].append({
                    "sku": sku_real,
                    "aptas": aptas,
                    "transito": transito
                })
                resultados["total_aptas"] += aptas
                resultados["total_transito"] += transito
            except Exception as e:
                resultados["errores"].append(f"SKU {sku_real}: {str(e)[:80]}")
        
        # ── Audit log ──
        if not dry_run:
            registrar_audit(
                session.get("usuario", "Sistema"), request.remote_addr,
                "importar_stock_full_meli",
                detalle=f"matched={len(resultados['matched'])} no_match={len(resultados['no_matched'])} aptas={resultados['total_aptas']} transito={resultados['total_transito']}"
            )
        
        return jsonify({
            "ok": True,
            "dry_run": dry_run,
            "resumen": {
                "items_en_excel": len(items_a_procesar),
                "matched_con_lusync": len(resultados["matched"]),
                "no_matched": len(resultados["no_matched"]),
                "stock_aptas_total": resultados["total_aptas"],
                "stock_transito_total": resultados["total_transito"],
            },
            "matched": resultados["matched"][:50],  # Limitar para no saturar response
            "no_matched": resultados["no_matched"],
            "errores": resultados["errores"]
        })
    
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False, "error": str(e),
            "trace": traceback.format_exc()
        }), 500


@app.route("/admin/enriquecer_item_ids", methods=["GET", "POST"])
def admin_enriquecer_item_ids():
    """Enriquece sku_mapeo_canal con los item_ids específicos de cada marketplace.

    NO crea filas nuevas. Solo UPDATEa las que tienen item_id_canal=NULL.
    Match SOLO por SKU exacto (case-insensitive). Sin similitud de nombre.

    Query params:
      ?canales=mercadolibre,falabella,ripley,walmart,paris  (default: todos)
      ?dry_run=1   simula sin escribir
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    canales_str = request.args.get("canales", "mercadolibre,falabella,ripley,walmart,paris")
    canales_pedidos = [c.strip().lower() for c in canales_str.split(",") if c.strip()]
    dry_run = request.args.get("dry_run", "0") == "1"

    try:
        from inventario import get_conn
        from psycopg2.extras import execute_values
        import time

        t_start = time.time()
        log = []
        resumen_por_canal = {}

        # ── Helper: aplicar updates en bulk ──
        def _aplicar_updates(canal, updates):
            """updates = [(item_id, sku_lusync, sku_canal), ...]"""
            if not updates:
                return 0
            if dry_run:
                return len(updates)
            conn = get_conn(); cur = conn.cursor()
            actualizados = 0
            try:
                # Para cada update, hacer UPDATE puntual (porque los registros tienen sku distintos)
                # Usamos batch via execute_batch o simplemente loop
                for item_id, sku_lusync, sku_canal in updates:
                    cur.execute("""
                        UPDATE sku_mapeo_canal
                        SET item_id_canal = %s,
                            actualizado_at = NOW(),
                            notas = COALESCE(notas, '') || ' | enriquecido_' || %s
                        WHERE canal = %s
                          AND sku_lusync = %s
                          AND sku_canal = %s
                          AND item_id_canal IS NULL
                          AND activo = TRUE
                    """, (item_id, canal, canal, sku_lusync, sku_canal))
                    if cur.rowcount > 0:
                        actualizados += 1
                conn.commit()
            except Exception as e:
                conn.rollback()
                log.append(f"❌ Error aplicando updates de {canal}: {e}")
                actualizados = 0
            finally:
                cur.close(); conn.close()
            return actualizados

        # ── Helper: obtener mapeos pendientes (sin item_id) por canal ──
        def _obtener_mapeos_sin_item(canal):
            conn = get_conn(); cur = conn.cursor()
            cur.execute("""
                SELECT sku_lusync, sku_canal FROM sku_mapeo_canal
                WHERE canal = %s AND item_id_canal IS NULL AND activo = TRUE
            """, (canal,))
            rows = cur.fetchall()
            cur.close(); conn.close()
            # Diccionario: {sku_canal_upper: (sku_lusync, sku_canal_original)}
            return {r[1].upper(): (r[0], r[1]) for r in rows}

        # ╔════════════════════════════════════════════════════════════╗
        # ║ MERCADOLIBRE                                               ║
        # ╚════════════════════════════════════════════════════════════╝
        if "mercadolibre" in canales_pedidos:
            t_canal = time.time()
            try:
                from mercadolibre import get_meli_token
                from inventario import get_meli_auth
                import requests

                pendientes = _obtener_mapeos_sin_item("mercadolibre")
                log.append(f"[MELI] {len(pendientes)} pendientes sin item_id")

                if pendientes:
                    token = get_meli_token()
                    auth = get_meli_auth() or {}
                    seller_id = auth.get("user_id")
                    if not seller_id:
                        r_user = requests.get("https://api.mercadolibre.com/users/me",
                                              headers={"Authorization": f"Bearer {token}"}, timeout=10)
                        seller_id = r_user.json().get("id") if r_user.status_code == 200 else None

                    if not seller_id:
                        log.append(f"[MELI] ❌ No se pudo obtener seller_id")
                        resumen_por_canal["mercadolibre"] = {"error": "sin seller_id", "actualizados": 0}
                    else:
                        # Obtener item_ids del seller
                        item_ids = []
                        offset = 0
                        while True:
                            r_search = requests.get(
                                f"https://api.mercadolibre.com/users/{seller_id}/items/search",
                                params={"limit": 50, "offset": offset},
                                headers={"Authorization": f"Bearer {token}"},
                                timeout=15
                            )
                            if r_search.status_code != 200: break
                            d = r_search.json()
                            ids_lote = d.get("results", []) or []
                            if not ids_lote: break
                            item_ids.extend(ids_lote)
                            offset += len(ids_lote)
                            if offset >= d.get("paging", {}).get("total", 0) or len(item_ids) >= 5000:
                                break

                        log.append(f"[MELI] {len(item_ids)} publicaciones en MELI")

                        # Obtener detalle (con SELLER_SKU) en lotes de 20
                        updates = []
                        meli_multi_pubs = {}  # sku_canal_upper → [item_ids]
                        for i in range(0, len(item_ids), 20):
                            ids_param = ",".join(item_ids[i:i+20])
                            r_items = requests.get(
                                "https://api.mercadolibre.com/items",
                                params={"ids": ids_param, "attributes": "id,attributes,seller_custom_field"},
                                headers={"Authorization": f"Bearer {token}"},
                                timeout=20
                            )
                            if r_items.status_code != 200: continue
                            for item_resp in r_items.json():
                                if item_resp.get("code") != 200: continue
                                body = item_resp.get("body", {})
                                item_id_meli = body.get("id", "")
                                seller_sku = ""
                                for attr in body.get("attributes", []) or []:
                                    if attr.get("id") == "SELLER_SKU":
                                        seller_sku = (attr.get("value_name") or "").strip()
                                        break
                                if not seller_sku:
                                    seller_sku = (body.get("seller_custom_field") or "").strip()
                                if not seller_sku:
                                    continue
                                seller_sku_upper = seller_sku.upper()

                                if seller_sku_upper in pendientes:
                                    sku_lusync, sku_canal_orig = pendientes[seller_sku_upper]
                                    # Si ya hay otro item_id para este SKU, lo guardamos en meli_multi_pubs
                                    # para crearlo como fila nueva (multi-publicación)
                                    if seller_sku_upper not in meli_multi_pubs:
                                        meli_multi_pubs[seller_sku_upper] = []
                                    meli_multi_pubs[seller_sku_upper].append({
                                        "item_id": item_id_meli,
                                        "sku_lusync": sku_lusync,
                                        "sku_canal": sku_canal_orig
                                    })

                        # Para MELI: la primera publicación → UPDATE de la fila existente
                        # Las adicionales (2da, 3ra...) → INSERT nuevas filas (multi-pub)
                        actualizados = 0
                        nuevas_multipub = 0

                        for sku_upper, pubs in meli_multi_pubs.items():
                            if not pubs:
                                continue
                            primera = pubs[0]
                            updates.append((
                                primera["item_id"],
                                primera["sku_lusync"],
                                primera["sku_canal"]
                            ))
                            # Las demás se insertan como nuevas filas
                            if len(pubs) > 1 and not dry_run:
                                conn = get_conn(); cur = conn.cursor()
                                filas_nuevas = [
                                    (p["sku_lusync"], "mercadolibre", p["sku_canal"],
                                     p["item_id"], False, "enriquecimiento_multi_pub")
                                    for p in pubs[1:]
                                ]
                                try:
                                    execute_values(
                                        cur,
                                        """INSERT INTO sku_mapeo_canal
                                           (sku_lusync, canal, sku_canal, item_id_canal, es_catalogo, notas, activo, creado_at, actualizado_at)
                                           VALUES %s
                                           ON CONFLICT DO NOTHING""",
                                        filas_nuevas,
                                        template="(%s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())",
                                        page_size=200
                                    )
                                    nuevas_multipub += cur.rowcount
                                    conn.commit()
                                except Exception as e:
                                    conn.rollback()
                                    log.append(f"[MELI] Error insertando multi-pub: {e}")
                                finally:
                                    cur.close(); conn.close()
                            elif len(pubs) > 1 and dry_run:
                                nuevas_multipub += len(pubs) - 1

                        actualizados = _aplicar_updates("mercadolibre", updates)

                        resumen_por_canal["mercadolibre"] = {
                            "publicaciones_meli": len(item_ids),
                            "pendientes_sin_item_id": len(pendientes),
                            "match_exacto_skus": len(meli_multi_pubs),
                            "actualizados": actualizados,
                            "multi_pubs_creadas": nuevas_multipub,
                            "tiempo_seg": f"{time.time()-t_canal:.2f}"
                        }
                        log.append(f"[MELI] ✓ Actualizados: {actualizados} | Multi-pubs nuevas: {nuevas_multipub}")
            except Exception as e:
                log.append(f"[MELI] ❌ {e}")
                resumen_por_canal["mercadolibre"] = {"error": str(e)}

        # ╔════════════════════════════════════════════════════════════╗
        # ║ FALABELLA                                                  ║
        # ╚════════════════════════════════════════════════════════════╝
        if "falabella" in canales_pedidos:
            t_canal = time.time()
            try:
                from falabella import obtener_productos_falabella

                pendientes = _obtener_mapeos_sin_item("falabella")
                log.append(f"[Falabella] {len(pendientes)} pendientes sin item_id")

                if pendientes:
                    # Falabella tiene paginación: traer todos los productos
                    # obtener_productos_falabella devuelve LIST directa, no dict
                    productos_fa = []
                    offset = 0
                    while True:
                        try:
                            lote = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                            if not lote or not isinstance(lote, list):
                                break
                            productos_fa.extend(lote)
                            offset += len(lote)
                            if len(lote) < 100 or offset >= 5000: break
                        except Exception as e:
                            log.append(f"[Falabella] Error en offset {offset}: {e}")
                            break
                    log.append(f"[Falabella] {len(productos_fa)} productos en API")

                    updates = []
                    for prod in productos_fa:
                        seller_sku = (prod.get("SellerSku") or prod.get("sku") or "").strip()
                        shop_sku = (str(prod.get("ShopSku") or prod.get("shop_sku") or "")).strip()
                        if not seller_sku or not shop_sku:
                            continue
                        if seller_sku.upper() in pendientes:
                            sku_lusync, sku_canal_orig = pendientes[seller_sku.upper()]
                            updates.append((shop_sku, sku_lusync, sku_canal_orig))

                    actualizados = _aplicar_updates("falabella", updates)
                    resumen_por_canal["falabella"] = {
                        "productos_api": len(productos_fa),
                        "pendientes_sin_item_id": len(pendientes),
                        "match_exacto": len(updates),
                        "actualizados": actualizados,
                        "tiempo_seg": f"{time.time()-t_canal:.2f}"
                    }
                    log.append(f"[Falabella] ✓ Actualizados: {actualizados}")
            except Exception as e:
                log.append(f"[Falabella] ❌ {e}")
                resumen_por_canal["falabella"] = {"error": str(e)}

        # ╔════════════════════════════════════════════════════════════╗
        # ║ RIPLEY                                                     ║
        # ╚════════════════════════════════════════════════════════════╝
        if "ripley" in canales_pedidos:
            t_canal = time.time()
            try:
                from ripley import obtener_productos_ripley

                pendientes = _obtener_mapeos_sin_item("ripley")
                log.append(f"[Ripley] {len(pendientes)} pendientes sin item_id")

                if pendientes:
                    productos_rp = obtener_productos_ripley(max_paginas=20, page_size=100)
                    if not isinstance(productos_rp, list):
                        productos_rp = []
                    log.append(f"[Ripley] {len(productos_rp)} productos en API")

                    # En Ripley la función devuelve: shop_sku, product_sku, product_title, etc.
                    # En Mirakl, el SHOP_SKU es el id del seller (lo que el seller cargó)
                    # No hay un "item_id" separado distinto al shop_sku
                    # Por consistencia, usamos el product_sku (catálogo Mirakl) como item_id
                    # Si product_sku es vacío, usamos el shop_sku mismo (redundante pero válido)
                    updates = []
                    for prod in productos_rp:
                        shop_sku = (prod.get("shop_sku") or "").strip()
                        product_sku = (prod.get("product_sku") or "").strip()
                        if not shop_sku:
                            continue
                        # item_id en Ripley = product_sku Mirakl si distinto, sino shop_sku
                        item_id = product_sku if (product_sku and product_sku != shop_sku) else shop_sku
                        if shop_sku.upper() in pendientes:
                            sku_lusync, sku_canal_orig = pendientes[shop_sku.upper()]
                            # Solo actualizamos si el item_id es DIFERENTE al sku_canal
                            # (Ripley sin product_sku distinto = item_id == sku_canal, redundante)
                            if item_id and item_id != sku_canal_orig:
                                updates.append((item_id, sku_lusync, sku_canal_orig))

                    actualizados = _aplicar_updates("ripley", updates)
                    resumen_por_canal["ripley"] = {
                        "productos_api": len(productos_rp),
                        "pendientes_sin_item_id": len(pendientes),
                        "match_con_item_id_distinto": len(updates),
                        "actualizados": actualizados,
                        "nota": "Ripley: si product_sku == shop_sku, no se actualiza (es redundante)",
                        "tiempo_seg": f"{time.time()-t_canal:.2f}"
                    }
                    log.append(f"[Ripley] ✓ Actualizados: {actualizados}")
            except Exception as e:
                log.append(f"[Ripley] ❌ {e}")
                resumen_por_canal["ripley"] = {"error": str(e)}

        # ╔════════════════════════════════════════════════════════════╗
        # ║ WALMART                                                    ║
        # ╚════════════════════════════════════════════════════════════╝
        if "walmart" in canales_pedidos:
            t_canal = time.time()
            try:
                from walmart import obtener_productos_walmart

                pendientes = _obtener_mapeos_sin_item("walmart")
                log.append(f"[Walmart] {len(pendientes)} pendientes sin item_id")

                if pendientes:
                    # max_paginas=50 para asegurar que trae TODOS (50 items × 50 páginas = 2500)
                    productos_wm = obtener_productos_walmart(limit=50, max_paginas=50)
                    if not isinstance(productos_wm, list):
                        productos_wm = []
                    log.append(f"[Walmart] {len(productos_wm)} productos en API")

                    updates = []
                    sin_wpid = 0
                    for prod in productos_wm:
                        # Walmart: sku es seller_sku, item_id es wpid
                        seller_sku = (prod.get("sku") or "").strip()
                        wpid = (str(prod.get("wpid") or "")).strip()
                        if not seller_sku:
                            continue
                        if not wpid:
                            sin_wpid += 1
                            continue
                        if seller_sku.upper() in pendientes:
                            sku_lusync, sku_canal_orig = pendientes[seller_sku.upper()]
                            updates.append((wpid, sku_lusync, sku_canal_orig))

                    actualizados = _aplicar_updates("walmart", updates)
                    resumen_por_canal["walmart"] = {
                        "productos_api": len(productos_wm),
                        "pendientes_sin_item_id": len(pendientes),
                        "match_exacto": len(updates),
                        "sin_wpid_en_api": sin_wpid,
                        "actualizados": actualizados,
                        "tiempo_seg": f"{time.time()-t_canal:.2f}"
                    }
                    log.append(f"[Walmart] ✓ Actualizados: {actualizados} | Sin wpid: {sin_wpid}")
            except Exception as e:
                log.append(f"[Walmart] ❌ {e}")
                resumen_por_canal["walmart"] = {"error": str(e)}

        # ╔════════════════════════════════════════════════════════════╗
        # ║ PARIS                                                      ║
        # ╚════════════════════════════════════════════════════════════╝
        if "paris" in canales_pedidos:
            t_canal = time.time()
            try:
                # Usar obtener_productos_paris (no obtener_stock_paris) para listar publicaciones
                from paris import obtener_productos_paris

                pendientes = _obtener_mapeos_sin_item("paris")
                log.append(f"[Paris] {len(pendientes)} pendientes sin item_id")

                if pendientes:
                    # París: paginar via /v2/products/search
                    productos_pa = []
                    offset = 0
                    while True:
                        try:
                            data = obtener_productos_paris(limite=100, offset=offset)
                            if not data:
                                break
                            # Estructura usual: {"results": [...], "total": N} o array directo
                            if isinstance(data, dict):
                                lote = data.get("results") or data.get("products") or data.get("productos") or data.get("data") or []
                            elif isinstance(data, list):
                                lote = data
                            else:
                                lote = []
                            if not lote: break
                            productos_pa.extend(lote)
                            offset += len(lote)
                            if len(lote) < 100 or offset >= 5000: break
                        except Exception as e:
                            log.append(f"[Paris] Error en offset {offset}: {e}")
                            break
                    log.append(f"[Paris] {len(productos_pa)} productos en API")

                    # DIAGNÓSTICO: muestra las keys del primer producto para ver estructura real
                    if productos_pa:
                        sample_keys = list(productos_pa[0].keys()) if isinstance(productos_pa[0], dict) else []
                        log.append(f"[Paris DEBUG] Keys del 1er producto: {sample_keys[:15]}")
                        # Sample de los primeros 3 SKUs detectados
                        for i, p in enumerate(productos_pa[:3]):
                            sku_intentos = {
                                "sellerSku": p.get("sellerSku"),
                                "sku": p.get("sku"),
                                "partnerSku": p.get("partnerSku"),
                                "offerId": p.get("offerId"),
                                "id": p.get("id"),
                                "sellerProductId": p.get("sellerProductId"),
                                "productCode": p.get("productCode"),
                            }
                            sku_intentos = {k: v for k, v in sku_intentos.items() if v is not None}
                            log.append(f"[Paris DEBUG] Producto {i}: {sku_intentos}")

                    updates = []
                    no_match_ejemplos = []
                    for prod in productos_pa:
                        # París: probar MÚLTIPLES nombres de SKU posibles
                        seller_sku = (
                            prod.get("sellerSku") or
                            prod.get("sku") or
                            prod.get("partnerSku") or
                            prod.get("sellerProductId") or
                            prod.get("productCode") or
                            prod.get("id") or
                            ""
                        )
                        seller_sku = str(seller_sku).strip()

                        # item_id distinto (si lo hay): offerId, partnerSku, etc.
                        item_id_paris = (
                            str(prod.get("offerId") or
                                prod.get("partnerSku") or
                                prod.get("id") or
                                seller_sku)
                        ).strip()

                        if not seller_sku:
                            continue
                        if seller_sku.upper() in pendientes:
                            sku_lusync, sku_canal_orig = pendientes[seller_sku.upper()]
                            # Solo actualizamos si el item_id es DIFERENTE al sku_canal
                            if item_id_paris and item_id_paris != sku_canal_orig:
                                updates.append((item_id_paris, sku_lusync, sku_canal_orig))
                        else:
                            if len(no_match_ejemplos) < 5:
                                no_match_ejemplos.append({"sku": seller_sku, "item_id": item_id_paris})

                    actualizados = _aplicar_updates("paris", updates)
                    resumen_por_canal["paris"] = {
                        "productos_api": len(productos_pa),
                        "pendientes_sin_item_id": len(pendientes),
                        "match_con_item_id_distinto": len(updates),
                        "actualizados": actualizados,
                        "ejemplos_sin_match": no_match_ejemplos,
                        "nota": "Paris: si item_id == sku_canal, no se actualiza (es redundante)",
                        "tiempo_seg": f"{time.time()-t_canal:.2f}"
                    }
                    log.append(f"[Paris] ✓ Actualizados: {actualizados}")
            except Exception as e:
                log.append(f"[Paris] ❌ {e}")
                resumen_por_canal["paris"] = {"error": str(e)}

        # ── Resumen final ──
        t_total = time.time() - t_start
        total_actualizados = sum(
            v.get("actualizados", 0) for v in resumen_por_canal.values()
            if isinstance(v, dict)
        )

        if not dry_run:
            registrar_audit(
                session.get("usuario", "Sistema"), request.remote_addr,
                "enriquecer_item_ids",
                detalle=f"canales={canales_pedidos} actualizados={total_actualizados} t={t_total:.1f}s"
            )

        return jsonify({
            "ok": True,
            "dry_run": dry_run,
            "canales_procesados": canales_pedidos,
            "tiempo_total_seg": f"{t_total:.2f}",
            "total_actualizados": total_actualizados,
            "resumen_por_canal": resumen_por_canal,
            "log": log
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/auto_mapeo_meli_seguro", methods=["GET", "POST"])
def admin_auto_mapeo_meli_seguro():
    """Auto-mapeo SEGURO para MercadoLibre con match SOLO por SKU exacto.

    Diferencias con auto_mapeo_v2 (que es genérico):
      - Solo procesa MELI
      - Solo hace match por SKU EXACTO (sin similitud de nombre)
      - Captura publicaciones múltiples (varios MLC con mismo SELLER_SKU)
      - Tiene blindaje: no duplica si ya existe (canal, item_id_canal)
      - Optimizado con bulk insert

    Useful después de un reset desde Excel: agrega las publicaciones múltiples
    MELI sin contaminar con falsos positivos.

    Query params:
      ?dry_run=1   simula sin escribir
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    dry_run = request.args.get("dry_run", "0") == "1"

    try:
        from inventario import get_conn, cargar_productos, get_meli_auth
        from mercadolibre import get_meli_token
        import time

        t_start = time.time()
        log = []

        # ── 1. Cargar productos Lusync para hacer match exacto ──
        productos = cargar_productos()
        skus_lusync = set((p.get("sku") or "").strip().upper() for p in productos if p.get("sku"))
        log.append(f"Productos Lusync: {len(skus_lusync)}")

        # ── 2. Obtener token y publicaciones MELI ──
        try:
            token = get_meli_token()
        except Exception as e:
            return jsonify({"ok": False, "error": f"No hay token MELI válido: {e}"}), 401

        # Obtener seller_id desde la tabla auth (más confiable)
        auth = get_meli_auth()
        seller_id = auth.get("user_id") if auth else None

        # Si no está en auth, lo obtenemos via /users/me
        if not seller_id:
            import requests
            r_user = requests.get("https://api.mercadolibre.com/users/me",
                                  headers={"Authorization": f"Bearer {token}"}, timeout=10)
            if r_user.status_code != 200:
                return jsonify({"ok": False, "error": f"No se pudo obtener seller_id: {r_user.status_code}"}), 500
            seller_id = r_user.json().get("id")

        log.append(f"Seller ID MELI: {seller_id}")

        import requests

        # Obtener todas las publicaciones del seller
        item_ids = []
        offset = 0
        while True:
            r_search = requests.get(
                f"https://api.mercadolibre.com/users/{seller_id}/items/search",
                params={"limit": 50, "offset": offset},
                headers={"Authorization": f"Bearer {token}"},
                timeout=15
            )
            if r_search.status_code != 200:
                log.append(f"Error en search offset={offset}: {r_search.status_code}")
                break
            datos = r_search.json()
            ids_lote = datos.get("results", []) or []
            if not ids_lote:
                break
            item_ids.extend(ids_lote)
            total = datos.get("paging", {}).get("total", 0)
            log.append(f"  offset={offset}: {len(ids_lote)} ids (total: {total})")
            offset += len(ids_lote)
            if offset >= total or len(item_ids) >= 5000:  # safety
                break

        log.append(f"Total publicaciones MELI: {len(item_ids)}")

        # ── 3. Para cada item, obtener detalle (con SELLER_SKU) en lotes de 20 ──
        publicaciones_a_insertar = []  # (sku_lusync, sku_canal, item_id, titulo)
        sin_match = []
        sin_sku = []

        for i in range(0, len(item_ids), 20):
            lote = item_ids[i:i+20]
            ids_param = ",".join(lote)
            r_items = requests.get(
                "https://api.mercadolibre.com/items",
                params={"ids": ids_param, "attributes": "id,title,attributes,seller_custom_field"},
                headers={"Authorization": f"Bearer {token}"},
                timeout=20
            )
            if r_items.status_code != 200:
                log.append(f"Error obteniendo detalle lote {i}: {r_items.status_code}")
                continue

            for item_resp in r_items.json():
                if item_resp.get("code") != 200:
                    continue
                body = item_resp.get("body", {})
                item_id = body.get("id", "")
                titulo = body.get("title", "")

                # Buscar SELLER_SKU en attributes
                seller_sku = ""
                for attr in body.get("attributes", []) or []:
                    if attr.get("id") == "SELLER_SKU":
                        seller_sku = (attr.get("value_name") or "").strip()
                        break

                # Fallback: seller_custom_field
                if not seller_sku:
                    seller_sku = (body.get("seller_custom_field") or "").strip()

                if not seller_sku:
                    sin_sku.append({"item_id": item_id, "titulo": titulo[:50]})
                    continue

                # Match EXACTO contra productos Lusync
                if seller_sku.upper() in skus_lusync:
                    publicaciones_a_insertar.append({
                        "sku_lusync": seller_sku,  # ya está en mayúsculas exacto
                        "sku_canal": seller_sku,
                        "item_id_canal": item_id,
                        "titulo": titulo
                    })
                else:
                    sin_match.append({"item_id": item_id, "seller_sku": seller_sku, "titulo": titulo[:50]})

        log.append(f"Publicaciones con match exacto: {len(publicaciones_a_insertar)}")
        log.append(f"Publicaciones sin SELLER_SKU: {len(sin_sku)}")
        log.append(f"Publicaciones sin match: {len(sin_match)}")

        # ── 4. Blindaje: filtrar las que ya están en sku_mapeo_canal ──
        if not dry_run and publicaciones_a_insertar:
            conn = get_conn(); cur = conn.cursor()
            # Obtener todos los item_id ya registrados para MELI
            cur.execute("""
                SELECT item_id_canal FROM sku_mapeo_canal
                WHERE canal='mercadolibre' AND item_id_canal IS NOT NULL AND activo=TRUE
            """)
            item_ids_existentes = set(r[0] for r in cur.fetchall())
            cur.close(); conn.close()

            antes_filtro = len(publicaciones_a_insertar)
            publicaciones_a_insertar = [
                p for p in publicaciones_a_insertar
                if p["item_id_canal"] not in item_ids_existentes
            ]
            ya_existian = antes_filtro - len(publicaciones_a_insertar)
            log.append(f"Ya existían (saltadas): {ya_existian}")
            log.append(f"Nuevas a insertar: {len(publicaciones_a_insertar)}")

        # ── 5. Bulk insert ──
        publicaciones_creadas = 0
        if not dry_run and publicaciones_a_insertar:
            try:
                from psycopg2.extras import execute_values
                conn = get_conn(); cur = conn.cursor()
                filas = [
                    (p["sku_lusync"], "mercadolibre", p["sku_canal"], p["item_id_canal"],
                     False, "auto_mapeo_meli_seguro")
                    for p in publicaciones_a_insertar
                ]
                t_ins = time.time()
                execute_values(
                    cur,
                    """INSERT INTO sku_mapeo_canal
                       (sku_lusync, canal, sku_canal, item_id_canal, es_catalogo, notas, activo, creado_at, actualizado_at)
                       VALUES %s
                       ON CONFLICT DO NOTHING""",
                    filas,
                    template="(%s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())",
                    page_size=500
                )
                publicaciones_creadas = cur.rowcount
                conn.commit()
                cur.close(); conn.close()
                log.append(f"✓ BULK INSERT: {publicaciones_creadas} filas en {(time.time()-t_ins)*1000:.0f}ms")
            except Exception as e:
                log.append(f"❌ Error en bulk insert: {e}")
                return jsonify({"ok": False, "error": str(e), "log": log}), 500

        t_total = time.time() - t_start

        if not dry_run:
            registrar_audit(
                session.get("usuario", "Sistema"), request.remote_addr,
                "auto_mapeo_meli_seguro",
                detalle=f"creadas={publicaciones_creadas} t={t_total:.2f}s"
            )

        return jsonify({
            "ok": True,
            "dry_run": dry_run,
            "total_publicaciones_meli": len(item_ids),
            "con_match_exacto": len(publicaciones_a_insertar) if dry_run else publicaciones_creadas,
            "ya_existian": (antes_filtro - len(publicaciones_a_insertar)) if (not dry_run and 'antes_filtro' in dir()) else 0,
            "sin_seller_sku": len(sin_sku),
            "sin_match_exacto": len(sin_match),
            "tiempo_segundos": f"{t_total:.2f}",
            "log": log,
            "ejemplos_sin_match": sin_match[:20],
            "ejemplos_sin_seller_sku": sin_sku[:20],
            "ejemplos_publicaciones_creadas": [
                {"sku_lusync": p["sku_lusync"], "item_id": p["item_id_canal"], "titulo": p["titulo"][:60]}
                for p in publicaciones_a_insertar[:30]
            ]
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/reset_sku_mapeo_canal", methods=["GET", "POST"])
def admin_reset_sku_mapeo_canal():
    """Borra y repuebla sku_mapeo_canal desde sku_mapeo (Excel).

    OPTIMIZADO con bulk insert masivo: usa 1 sola query INSERT VALUES (...)
    en lugar de N queries individuales. Procesa miles de SKUs en <1 segundo.

    Query params:
      ?confirmar=SI         requerido para ejecutar (sin esto, solo preview)
      ?incluir_web=0        no incluir canal web (default: 0)
      ?dry_run=1            simular sin ejecutar
    """
    if not session.get("logged"): return jsonify({"error": "no autorizado"}), 401

    confirmar = request.args.get("confirmar", "")
    dry_run = request.args.get("dry_run", "0") == "1"
    incluir_web = request.args.get("incluir_web", "0") == "1"

    # Sin confirmación, solo preview
    if confirmar != "SI" and not dry_run:
        try:
            from inventario import get_conn
            conn = get_conn(); cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM sku_mapeo_canal WHERE activo = TRUE")
            actuales = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM sku_mapeo")
            mapeo_legacy = cur.fetchone()[0]
            cur.close(); conn.close()
            return jsonify({
                "ok": False,
                "modo": "preview",
                "mensaje": "Para confirmar, agrega ?confirmar=SI a la URL",
                "se_borrarian_de_sku_mapeo_canal": actuales,
                "se_recrearian_desde_sku_mapeo": mapeo_legacy,
                "incluir_web": incluir_web,
                "url_dry_run": "/admin/reset_sku_mapeo_canal?dry_run=1",
                "url_para_confirmar": f"/admin/reset_sku_mapeo_canal?confirmar=SI{'&incluir_web=1' if incluir_web else ''}"
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    try:
        from inventario import get_conn, listar_sku_mapeo
        from datetime import datetime as _dt
        import time

        log = []
        t_start = time.time()

        # ── 1. Leer datos ──
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM sku_mapeo_canal")
        total_antes = cur.fetchone()[0]
        log.append(f"Filas en sku_mapeo_canal antes: {total_antes}")

        mapeos = listar_sku_mapeo()
        log.append(f"Filas en sku_mapeo (Excel): {len(mapeos)}")

        # ── 2. Construir lista de filas a insertar (en memoria, rápido) ──
        campos_canal = [
            ("sku_walmart",      "walmart"),
            ("sku_paris",        "paris"),
            ("sku_falabella",    "falabella"),
            ("sku_ripley",       "ripley"),
            ("sku_mercadolibre", "mercadolibre"),
            ("sku_hites",        "hites"),
        ]
        if incluir_web:
            campos_canal.insert(0, ("sku_web", "web"))

        # filas_para_insertar = [(sku_lusync, canal, sku_canal, item_id_canal, es_catalogo, notas), ...]
        filas_para_insertar = []
        skus_procesados = 0
        log_creaciones_sample = []  # solo guardamos los primeros 50 ejemplos para el log

        for fila in mapeos:
            sku_lusync = (fila.get("sku_lusync") or "").strip()
            if not sku_lusync:
                continue
            skus_procesados += 1

            for campo_bd, canal in campos_canal:
                sku_canal_val = (fila.get(campo_bd) or "").strip()
                if not sku_canal_val or sku_canal_val.lower() in ("none", "nan", "null"):
                    continue

                # Para MELI: detectar si es item_id (MLC...) o seller_sku
                item_id_canal = None
                if canal == "mercadolibre" and sku_canal_val.upper().startswith("MLC"):
                    item_id_canal = sku_canal_val

                filas_para_insertar.append((
                    sku_lusync, canal, sku_canal_val,
                    item_id_canal, False, "reset_desde_excel"
                ))

                if len(log_creaciones_sample) < 50:
                    log_creaciones_sample.append(
                        f"{'[DRY] ' if dry_run else ''}{sku_lusync} → {canal}:{sku_canal_val}" +
                        (f" (item_id={item_id_canal})" if item_id_canal else "")
                    )

        t_build = time.time() - t_start
        log.append(f"Construidas {len(filas_para_insertar)} filas en {t_build:.2f}s")

        if dry_run:
            cur.close(); conn.close()
            log.append(f"[DRY RUN] No se ejecutó nada en BD")
            return jsonify({
                "ok": True,
                "dry_run": True,
                "incluir_web": incluir_web,
                "skus_procesados": skus_procesados,
                "publicaciones_a_crear": len(filas_para_insertar),
                "publicaciones_creadas": len(filas_para_insertar),
                "publicaciones_fallidas": 0,
                "filas_borradas": "(dry_run)",
                "filas_despues": None,
                "estado_final_por_canal": None,
                "tiempo_segundos": f"{t_build:.2f}",
                "log": log,
                "log_creaciones": log_creaciones_sample
            })

        # ── 3. EJECUCIÓN REAL: backup + delete + bulk insert ──

        # 3a. Backup
        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        backup_tabla = f"sku_mapeo_canal_backup_{timestamp}"
        try:
            cur.execute(f'CREATE TABLE "{backup_tabla}" AS SELECT * FROM sku_mapeo_canal')
            log.append(f"✓ Backup creado: {backup_tabla}")
        except Exception as e:
            conn.rollback()
            cur.close(); conn.close()
            return jsonify({"ok": False, "error": f"Backup falló: {e}", "log": log}), 500

        # 3b. HARD DELETE
        cur.execute("DELETE FROM sku_mapeo_canal")
        log.append(f"✓ DELETE ejecutado en sku_mapeo_canal")

        # 3c. BULK INSERT masivo (1 sola query con todos los VALUES)
        # Usamos psycopg2.extras.execute_values para optimización máxima
        publicaciones_creadas = 0
        publicaciones_fallidas = 0

        if filas_para_insertar:
            try:
                # Importar execute_values para bulk insert óptimo
                from psycopg2.extras import execute_values
                t_insert = time.time()
                execute_values(
                    cur,
                    """INSERT INTO sku_mapeo_canal
                       (sku_lusync, canal, sku_canal, item_id_canal, es_catalogo, notas, activo, creado_at, actualizado_at)
                       VALUES %s
                       ON CONFLICT DO NOTHING
                       RETURNING id""",
                    filas_para_insertar,
                    template="(%s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())",
                    page_size=500  # batch de 500 filas por iteración
                )
                ids_creados = [r[0] for r in cur.fetchall()] if cur.description else []
                publicaciones_creadas = len(ids_creados) if ids_creados else len(filas_para_insertar)
                t_ins_ms = (time.time() - t_insert) * 1000
                log.append(f"✓ BULK INSERT: {len(filas_para_insertar)} filas en {t_ins_ms:.0f}ms")
            except ImportError:
                # Fallback si no hay psycopg2.extras
                log.append("⚠ psycopg2.extras no disponible, usando INSERT individual (más lento)")
                for f in filas_para_insertar:
                    try:
                        cur.execute("""
                            INSERT INTO sku_mapeo_canal
                            (sku_lusync, canal, sku_canal, item_id_canal, es_catalogo, notas, activo, creado_at, actualizado_at)
                            VALUES (%s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())
                            ON CONFLICT DO NOTHING
                        """, f)
                        publicaciones_creadas += 1
                    except Exception as e_ins:
                        publicaciones_fallidas += 1
            except Exception as e:
                conn.rollback()
                cur.close(); conn.close()
                return jsonify({
                    "ok": False,
                    "error": f"Bulk insert falló: {e}",
                    "url_recuperar_backup": f"INSERT INTO sku_mapeo_canal SELECT * FROM \"{backup_tabla}\";",
                    "log": log
                }), 500

        conn.commit()
        log.append(f"✓ Commit ejecutado")

        # 3d. Estado final
        cur.execute("SELECT canal, COUNT(*) FROM sku_mapeo_canal GROUP BY canal ORDER BY canal")
        estado_final = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute("SELECT COUNT(*) FROM sku_mapeo_canal")
        total_despues = cur.fetchone()[0]
        log.append(f"Filas en sku_mapeo_canal después: {total_despues}")

        cur.close(); conn.close()

        t_total = time.time() - t_start

        registrar_audit(
            session.get("usuario", "Sistema"), request.remote_addr,
            "reset_sku_mapeo_canal",
            detalle=f"borradas={total_antes} creadas={publicaciones_creadas} backup={backup_tabla} t={t_total:.2f}s"
        )

        return jsonify({
            "ok": True,
            "dry_run": False,
            "incluir_web": incluir_web,
            "skus_procesados": skus_procesados,
            "publicaciones_creadas": publicaciones_creadas,
            "publicaciones_fallidas": publicaciones_fallidas,
            "filas_borradas": total_antes,
            "filas_despues": total_despues,
            "estado_final_por_canal": estado_final,
            "tiempo_segundos": f"{t_total:.2f}",
            "backup_tabla": backup_tabla,
            "url_recuperar_backup": f"INSERT INTO sku_mapeo_canal SELECT * FROM \"{backup_tabla}\";",
            "log": log,
            "log_creaciones": log_creaciones_sample
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/debug_falabella_ordenes")
def admin_debug_falabella_ordenes():
    """Diagnóstico profundo de por qué Falabella no registra órdenes.
    Uso: /admin/debug_falabella_ordenes?dias=7&token=XXX
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token = request.args.get("token", "")
    if token != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    dias = int(request.args.get("dias", 7))

    try:
        from falabella import obtener_ordenes_falabella, obtener_items_orden_falabella
        from inventario import obtener_sku_lusync_por_canal, orden_ya_procesada_texto

        reporte = {}

        # 1. Órdenes por estado (ya desenvueltas)
        for est in ["pending", "ready_to_ship", "shipped", "delivered", "canceled"]:
            try:
                ordenes = obtener_ordenes_falabella(estado=est, dias=dias, limit=5)
                reporte[est] = {"count": len(ordenes)}
                if ordenes:
                    o = ordenes[0]
                    order_id = str(o.get("OrderId") or "")
                    statuses = o.get("Statuses") or []
                    estado_leido = (statuses[0].get("Status") if statuses else o.get("Status") or "")
                    fa_key = f"FALABELLA-{order_id}"
                    ya_procesada = orden_ya_procesada_texto(fa_key)

                    # Items de esa orden
                    items = []
                    try:
                        items_raw = obtener_items_orden_falabella(order_id) or []
                        for item in items_raw:
                            sku_fa = (item.get("SellerSku") or item.get("sellerSku") or "")
                            sku_lusync = obtener_sku_lusync_por_canal("falabella", sku_fa) or "NO MAPEADO"
                            items.append({
                                "SellerSku": sku_fa,
                                "sku_lusync": sku_lusync,
                                "cantidad": item.get("Quantity") or item.get("quantity"),
                                "Status": item.get("Status") or item.get("status")
                            })
                    except Exception as e:
                        items = [f"ERROR items: {e}"]

                    reporte[est]["primera_orden"] = {
                        "OrderId": order_id,
                        "OrderNumber": o.get("OrderNumber"),
                        "CreatedAt": o.get("CreatedAt"),
                        "estado_leido_del_campo": estado_leido,
                        "fa_key": fa_key,
                        "ya_marcada_en_bd": ya_procesada,
                        "items": items,
                        "campos_disponibles": list(o.keys())
                    }
            except Exception as e:
                reporte[est] = {"error": str(e)}

        # 2. Simular exactamente lo que haría el scheduler
        simulacion = []
        try:
            todas = []
            for est in ["pending", "ready_to_ship", "shipped", "delivered"]:
                lote = obtener_ordenes_falabella(estado=est, dias=dias, limit=5) or []
                todas.extend(lote)

            for o in todas[:5]:
                order_id = str(o.get("OrderId") or o.get("orderId") or "")
                statuses = o.get("Statuses") or []
                estado_ord = (statuses[0].get("Status") if statuses else o.get("Status") or "").lower()
                fa_key = f"FALABELLA-{order_id}"
                ya_proc = orden_ya_procesada_texto(fa_key)
                items_raw = obtener_items_orden_falabella(order_id) or []

                sim = {
                    "order_id": order_id,
                    "estado_leido": estado_ord,
                    "estado_valido": estado_ord in ("ready_to_ship", "shipped", "delivered", "pending"),
                    "fa_key": fa_key,
                    "ya_procesada": ya_proc,
                    "items_count": len(items_raw),
                    "resultado": None
                }

                if ya_proc:
                    sim["resultado"] = "SKIP — ya marcada en BD"
                elif estado_ord not in ("ready_to_ship", "shipped", "delivered", "pending"):
                    sim["resultado"] = f"SKIP — estado '{estado_ord}' no reconocido"
                elif not items_raw:
                    sim["resultado"] = "SKIP — sin items"
                else:
                    sku_fa = (items_raw[0].get("SellerSku") or items_raw[0].get("sellerSku") or "")
                    sku_lusync = obtener_sku_lusync_por_canal("falabella", sku_fa) or sku_fa
                    productos = cargar_productos()
                    prod = next((p for p in productos if p["sku"] == sku_lusync), None)
                    if not prod:
                        sim["resultado"] = f"SKIP — SKU '{sku_lusync}' no en productos"
                    else:
                        sim["resultado"] = f"✅ PROCESARÍA — {sku_lusync} stock={prod['stock']}"

                simulacion.append(sim)
        except Exception as e:
            simulacion = [f"ERROR simulación: {e}"]

        return jsonify({
            "ok": True,
            "ordenes_por_estado": reporte,
            "simulacion_scheduler": simulacion
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/exportar_movimientos_excel")
def admin_exportar_movimientos_excel():
    """Exporta todos los movimientos a Excel para migración o backup.
    Columnas: id, tipo, sku, nombre, cantidad, motivo, canal, bodega,
              orden_id, usuario, origen, fecha_compra, fecha_importacion,
              stock_antes, stock_despues
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from inventario import get_conn

        conn = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT id, tipo, sku, nombre, cantidad, motivo, canal,
                   COALESCE(bodega_codigo,'CENTRAL'),
                   orden_id, usuario,
                   COALESCE(origen_registro,'sistema'),
                   fecha_compra_marketplace,
                   fecha_importacion,
                   stock_antes, stock_despues,
                   fecha
            FROM movimientos
            ORDER BY fecha DESC
        """)
        filas = cur.fetchall()
        cur.close(); conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        headers = ["id","tipo","sku","nombre","cantidad","motivo","canal",
                   "bodega","orden_id","usuario","origen_registro",
                   "fecha_compra_marketplace","fecha_importacion",
                   "stock_antes","stock_despues","fecha_sistema"]

        # Estilo header
        fill = PatternFill("solid", fgColor="1D4ED8")
        font = Font(color="FFFFFF", bold=True, size=11)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center")

        for row_i, fila in enumerate(filas, 2):
            for col_i, val in enumerate(fila, 1):
                ws.cell(row=row_i, column=col_i, value=val)

        # Anchos razonables
        anchos = [6,8,14,35,8,25,12,12,16,10,12,20,20,10,10,20]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        nombre_archivo = f"movimientos_lusync_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True,
                         download_name=nombre_archivo,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/importar_movimientos_excel", methods=["POST"])
def admin_importar_movimientos_excel():
    """Importa movimientos desde un Excel exportado por Lusync.
    - Solo inserta movimientos que no existan ya (detecta duplicados por orden_id+canal+sku+fecha)
    - Nunca borra movimientos existentes
    - Ideal para migración de cliente a nuevo servidor
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        import io, openpyxl
        from inventario import get_conn

        archivo = request.files.get("archivo")
        if not archivo:
            return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb.active

        # Leer headers de la primera fila
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        def col(row, nombre):
            try:
                idx = headers.index(nombre)
                return row[idx].value
            except (ValueError, IndexError):
                return None

        conn = get_conn(); cur = conn.cursor()
        # Asegurar columnas necesarias
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS fecha_compra_marketplace TIMESTAMP")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen_registro TEXT DEFAULT 'sistema'")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_antes INTEGER")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS stock_despues INTEGER")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS bodega_codigo TEXT DEFAULT 'CENTRAL'")
        cur.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS fecha_importacion TIMESTAMP")
        conn.commit()

        importados = 0
        duplicados = 0
        errores = 0

        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            try:
                tipo       = col(row, "tipo") or "salida"
                sku        = str(col(row, "sku") or "").strip()
                nombre     = col(row, "nombre") or sku
                cantidad   = int(col(row, "cantidad") or 0)
                motivo     = col(row, "motivo") or ""
                canal      = col(row, "canal") or "Sistema"
                bodega     = col(row, "bodega") or "CENTRAL"
                orden_id   = str(col(row, "orden_id") or "").strip() or None
                usuario    = col(row, "usuario") or "Importación"
                origen     = col(row, "origen_registro") or "import_excel"
                fecha_compra = col(row, "fecha_compra_marketplace")
                fecha_imp  = col(row, "fecha_importacion")
                stock_ant  = col(row, "stock_antes")
                stock_des  = col(row, "stock_despues")
                fecha_sis  = col(row, "fecha_sistema")

                if not sku:
                    errores += 1
                    continue

                # Detectar duplicado por orden_id + canal + sku (si tienen orden_id)
                if orden_id:
                    cur.execute("""
                        SELECT 1 FROM movimientos
                        WHERE orden_id=%s AND canal=%s AND sku=%s LIMIT 1
                    """, (orden_id, canal, sku))
                    if cur.fetchone():
                        duplicados += 1
                        continue

                cur.execute("""
                    INSERT INTO movimientos
                        (tipo, sku, nombre, cantidad, motivo, canal, bodega_codigo,
                         orden_id, usuario, origen_registro,
                         fecha_compra_marketplace, fecha_importacion,
                         stock_antes, stock_despues, fecha)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    tipo, sku, nombre, cantidad, motivo, canal, bodega,
                    orden_id, usuario, "import_excel",
                    fecha_compra, fecha_imp,
                    stock_ant, stock_des,
                    fecha_sis or __import__('datetime').datetime.now()
                ))
                importados += 1

            except Exception as e:
                errores += 1
                print(f"[Importar mov] fila error: {e}")

        conn.commit()
        cur.close(); conn.close()

        try:
            registrar_audit(session.get("usuario","Sistema"), request.remote_addr,
                            "importar_movimientos_excel", entidad="movimientos",
                            detalle=f"importados={importados} duplicados={duplicados} errores={errores}")
        except: pass

        return jsonify({
            "ok": True,
            "importados": importados,
            "duplicados": duplicados,
            "errores": errores,
            "mensaje": f"{importados} movimientos importados correctamente"
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/debug_falabella_items")
def admin_debug_falabella_items():
    """Muestra la respuesta RAW de GetOrderItems para una orden específica.
    Uso: /admin/debug_falabella_items?order_id=1152896462&token=XXX
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    token = request.args.get("token", "")
    if token != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    order_id = request.args.get("order_id", "1152896462")

    try:
        from falabella import llamar_api_falabella
        res = llamar_api_falabella(
            "GetOrderItems",
            params_extra={"OrderId": order_id},
            method="GET",
            formato="JSON"
        )
        return jsonify({
            "order_id": order_id,
            "ok": res.get("ok"),
            "error": res.get("error"),
            "raw_text": res.get("raw_text", "")[:3000],
            "data": res.get("data")
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/admin/normalizar_canales")
def admin_normalizar_canales():
    """Normaliza nombres de canal inconsistentes en la tabla movimientos.
    Convierte variantes como 'mercadolibre', 'MELI', 'woo', 'París' al
    nombre canónico usado en el dashboard.
    Uso: /admin/normalizar_canales?token=XXX
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    if request.args.get("token") != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn(); cur = conn.cursor()

        # Mapeo de variantes → nombre canónico
        mapeo = [
            # MercadoLibre
            ("mercadolibre", "MercadoLibre"),
            ("Mercadolibre", "MercadoLibre"),
            ("MERCADOLIBRE", "MercadoLibre"),
            ("meli", "MercadoLibre"),
            ("MELI", "MercadoLibre"),
            ("Meli", "MercadoLibre"),
            # Falabella
            ("falabella", "Falabella"),
            ("FALABELLA", "Falabella"),
            # Paris
            ("paris", "Paris"),
            ("París", "Paris"),
            ("PARIS", "Paris"),
            # Ripley
            ("ripley", "Ripley"),
            ("RIPLEY", "Ripley"),
            # Walmart
            ("walmart", "Walmart"),
            ("WALMART", "Walmart"),
            # Web / WooCommerce
            ("WooCommerce", "Web"),
            ("woocommerce", "Web"),
            ("Woocommerce", "Web"),
            ("woo", "Web"),
            ("Woo", "Web"),
            # Hites
            ("hites", "Hites"),
            ("HITES", "Hites"),
        ]

        total_actualizados = 0
        log = []
        for variante, canonico in mapeo:
            cur.execute(
                "UPDATE movimientos SET canal=%s WHERE canal=%s",
                (canonico, variante)
            )
            n = cur.rowcount
            if n > 0:
                log.append(f"'{variante}' → '{canonico}': {n} filas")
                total_actualizados += n

        conn.commit()
        cur.close(); conn.close()

        return jsonify({
            "ok": True,
            "total_actualizados": total_actualizados,
            "detalle": log,
            "mensaje": f"{total_actualizados} movimientos normalizados"
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/admin/test_flujo_venta")
def admin_test_flujo_venta():
    """
    Testea el flujo completo de una venta SIN tocar stock real.
    Simula: orden entra → stock se descuenta → se sincroniza a todos los canales.

    Uso: /admin/test_flujo_venta?sku=EDLABA001&canal=Walmart&token=XXX
    Parámetros:
      sku     = SKU Lusync a testear
      canal   = Canal que origina la venta (Walmart, Falabella, Paris, Ripley, MercadoLibre, Web)
      simular = 1 (default) solo simula sin tocar nada | 0 = ejecuta real (CUIDADO)
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    if request.args.get("token") != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    sku      = request.args.get("sku", "").strip()
    canal    = request.args.get("canal", "Walmart")
    simular  = request.args.get("simular", "1") != "0"

    if not sku:
        return jsonify({"error": "Falta parámetro sku"}), 400

    reporte = {
        "sku": sku,
        "canal": canal,
        "modo": "SIMULACIÓN (sin cambios reales)" if simular else "⚠️ EJECUCIÓN REAL",
        "pasos": []
    }

    def paso(nombre, estado, detalle, extra=None):
        p = {"paso": nombre, "estado": estado, "detalle": detalle}
        if extra:
            p["extra"] = extra
        reporte["pasos"].append(p)

    # ── PASO 1: Verificar que el SKU existe ──
    try:
        productos = cargar_productos()
        prod = next((p for p in productos if p["sku"] == sku), None)
        if not prod:
            paso("1. Verificar SKU", "❌ ERROR", f"SKU '{sku}' no existe en inventario")
            return jsonify(reporte)
        stock_actual = prod.get("stock", 0)
        paso("1. Verificar SKU", "✅ OK",
             f"Producto encontrado: {prod['nombre']}",
             {"stock_central": stock_actual, "nombre": prod["nombre"]})
    except Exception as e:
        paso("1. Verificar SKU", "❌ ERROR", str(e))
        return jsonify(reporte)

    # ── PASO 2: Verificar stock disponible ──
    if stock_actual <= 0:
        paso("2. Stock disponible", "⚠️ ADVERTENCIA",
             f"Stock en 0 — la venta se registraría pero no descontaría nada")
    else:
        paso("2. Stock disponible", "✅ OK",
             f"Stock Central = {stock_actual} unidades — suficiente para vender")

    # ── PASO 3: Verificar mapeo de canales ──
    try:
        from inventario import obtener_publicaciones_canal
        canales_a_verificar = ["mercadolibre", "falabella", "walmart", "paris", "ripley", "web"]
        mapeos = {}
        for c in canales_a_verificar:
            pubs = obtener_publicaciones_canal(sku, c) or []
            mapeos[c] = {
                "publicaciones": len(pubs),
                "detalle": [{"sku_canal": p.get("sku_canal"), "item_id": p.get("item_id_canal")} for p in pubs[:3]]
            }
        paso("3. Mapeo de publicaciones", "✅ OK",
             "Publicaciones encontradas por canal", mapeos)
    except Exception as e:
        paso("3. Mapeo de publicaciones", "❌ ERROR", str(e))

    # ── PASO 4: Simular/ejecutar descuento de stock ──
    if simular:
        stock_resultante = max(0, stock_actual - 1)
        paso("4. Descuento de stock", "✅ SIMULADO",
             f"Stock Central {stock_actual} → {stock_resultante} (−1 unidad)",
             {"stock_antes": stock_actual, "stock_despues": stock_resultante,
              "bodega": "CENTRAL", "nota": "No se ejecutó — modo simulación"})
    else:
        try:
            from bodegas_logic import descontar_venta
            resultado = descontar_venta(
                sku=sku, cantidad=1, canal=canal,
                fulfillment=False,
                orden_id=f"TEST-{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}",
                motivo=f"Test flujo venta — canal {canal}",
                usuario="Sistema (Test)",
                origen_registro="test"
            )
            stock_resultante = resultado.get("stock_despues", stock_actual - 1)
            paso("4. Descuento de stock", "✅ EJECUTADO",
                 f"Stock {stock_actual} → {stock_resultante}",
                 resultado)
        except Exception as e:
            paso("4. Descuento de stock", "❌ ERROR", str(e))
            stock_resultante = stock_actual - 1

    # ── PASO 5: Testear sync a cada canal (siempre en modo dry-run real) ──
    # Llamamos a cada función de actualización pero capturamos resultado sin afectar nada
    # en simulación, o ejecutamos real si simular=False
    try:
        from mercadolibre import actualizar_stock_meli
        from falabella import actualizar_stock_falabella_lusync
        from walmart import actualizar_stock_walmart_lusync
        from paris import actualizar_stock_paris
        from ripley import actualizar_stock_ripley
        from woo import actualizar_stock_woo

        canales_sync = [
            ("MercadoLibre", actualizar_stock_meli),
            ("Falabella",    actualizar_stock_falabella_lusync),
            ("Walmart",      actualizar_stock_walmart_lusync),
            ("Paris",        actualizar_stock_paris),
            ("Ripley",       actualizar_stock_ripley),
            ("Web/Woo",      actualizar_stock_woo),
        ]

        sync_resultados = {}
        for nombre, fn in canales_sync:
            if simular:
                # En simulación: verificar si tiene publicaciones mapeadas
                c_key = nombre.lower().replace("/woo","").replace("web","woocommerce").strip()
                pubs = mapeos.get(c_key, {}).get("publicaciones", 0) if 'mapeos' in dir() else 0
                if pubs > 0:
                    sync_resultados[nombre] = f"✅ LISTO para sync ({pubs} publicación/es mapeada/s)"
                else:
                    sync_resultados[nombre] = "⚠️ Sin publicaciones mapeadas — no se sincronizaría"
            else:
                try:
                    r = fn(sku, stock_resultante)
                    if isinstance(r, dict):
                        ok = r.get("exitosas", 0) > 0 or r.get("ok")
                        sync_resultados[nombre] = f"{'✅' if ok else '❌'} {r}"
                    else:
                        sync_resultados[nombre] = f"{'✅' if r else '❌'} resultado={r}"
                except Exception as e:
                    sync_resultados[nombre] = f"❌ Error: {str(e)[:100]}"

        paso("5. Sincronización a canales", "✅ OK" if simular else "✅ EJECUTADO",
             "Resultado por canal" if not simular else "Diagnóstico de publicaciones mapeadas",
             sync_resultados)
    except Exception as e:
        paso("5. Sincronización a canales", "❌ ERROR", str(e))

    # ── RESUMEN ──
    errores = [p for p in reporte["pasos"] if "❌" in p["estado"]]
    advertencias = [p for p in reporte["pasos"] if "⚠️" in p["estado"]]
    reporte["resumen"] = {
        "resultado": "❌ HAY ERRORES" if errores else ("⚠️ HAY ADVERTENCIAS" if advertencias else "✅ FLUJO OK"),
        "errores": len(errores),
        "advertencias": len(advertencias),
        "conclusion": (
            "El flujo tiene errores que impedirían la sincronización correcta" if errores
            else "El flujo funcionaría correctamente — la venta se descontaría y sincronizaría a todos los canales con publicaciones mapeadas" if not advertencias
            else "El flujo funciona pero algunos canales no se sincronizarían por falta de mapeo"
        )
    }

    return jsonify(reporte)



@app.route("/admin/autodescubrir_publicaciones")
def admin_autodescubrir_publicaciones():
    """
    Auto-descubrimiento de publicaciones por canal.
    Descarga el catálogo completo de cada marketplace, cruza con SKUs Lusync
    y genera un plan de acción para los que no cruzaron.

    Uso: /admin/autodescubrir_publicaciones?canal=meli&token=XXX
         /admin/autodescubrir_publicaciones?canal=todos&token=XXX
         /admin/autodescubrir_publicaciones?canal=meli&ejecutar=1&token=XXX

    Parámetros:
      canal   = meli | falabella | walmart | paris | ripley | todos
      ejecutar= 0 (default, solo diagnostica) | 1 (guarda los mapeos encontrados)
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    if request.args.get("token") != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    canal    = request.args.get("canal", "todos").lower().strip()
    ejecutar = request.args.get("ejecutar", "0") == "1"

    try:
        from inventario import get_conn, agregar_publicacion, obtener_publicaciones_canal

        # Cargar todos los SKUs Lusync actuales como índice de búsqueda
        productos = cargar_productos()
        skus_lusync = {p["sku"].upper().strip(): p["sku"] for p in productos}
        nombres_lusync = {p["sku"]: p.get("nombre", "") for p in productos}

        reporte_global = {
            "modo": "EJECUCIÓN REAL — mapeos guardados" if ejecutar else "SIMULACIÓN — no se guarda nada",
            "canales": {},
            "resumen_global": {},
            "plan_accion": []
        }

        def _normalizar_sku(s):
            return str(s or "").upper().strip().replace(" ", "")

        def _cruzar_y_guardar(canal_nombre, publicaciones_canal):
            """
            Cruza publicaciones del canal con SKUs Lusync.
            Retorna dict con cruzados, no_cruzados y plan de acción.

            REGLAS DE CRUCE (en orden de prioridad):
              1. SKU vacío → siempre no_cruzado, NUNCA se asigna automáticamente
              2. Match exacto SKU canal == SKU Lusync → cruzado confiable
              3. Match parcial con longitud mínima 6 chars → cruzado con advertencia
              4. Sin match → no_cruzado con plan de acción
            """
            cruzados = []
            no_cruzados = []
            ya_mapeados = []
            errores_guardado = []
            sin_sku = []

            for pub in publicaciones_canal:
                sku_canal   = str(pub.get("sku_canal") or "").strip()
                item_id     = pub.get("item_id_canal")
                nombre_pub  = pub.get("nombre", "")
                extra       = pub.get("extra", {})

                # FIX BUG 1: SKU vacío, None, "None", "nan" → NUNCA cruzar automáticamente
                # Antes cruzaba al primer SKU que encontraba — generaba falsos positivos masivos
                if not sku_canal or sku_canal.lower() in ("none", "nan", "null", "0", ""):
                    # Detectar si es publicación paraguas de catálogo MELI
                    # (tiene variantes con inventory_id pero sin seller_custom_field ni SELLER_SKU)
                    # En ese caso sus variantes individuales ya están mapeadas — no es una alerta real
                    extra_data = pub.get("extra", {})
                    variantes_raw = pub.get("variantes_raw", [])
                    tiene_inventory_id = any(
                        v.get("inventory_id") for v in variantes_raw
                    ) if variantes_raw else False

                    if tiene_inventory_id:
                        # Publicación paraguas — ignorar silenciosamente
                        continue

                    sin_sku.append({
                        "item_id": item_id,
                        "nombre": nombre_pub,
                        "extra": extra,
                        "razon": "Publicación sin SKU interno — debe asignarse manualmente"
                    })
                    continue

                sku_norm = _normalizar_sku(sku_canal)
                sku_lusync_match = skus_lusync.get(sku_norm)
                razon_match = "exacto"

                # FIX BUG 2: Match parcial solo si:
                #   - longitud mínima 6 caracteres (evita que "E10" cruce con "E10E11E12")
                #   - el SKU canal EMPIEZA CON el SKU Lusync (no al revés — evita falsos)
                #   - no es un prefijo genérico de 1-3 chars
                if not sku_lusync_match and len(sku_norm) >= 6:
                    for sk_up, sk_real in skus_lusync.items():
                        if len(sk_up) >= 6 and sku_norm.startswith(sk_up):
                            sku_lusync_match = sk_real
                            razon_match = "parcial"
                            break

                if sku_lusync_match:
                    # Verificar si ya está mapeado (por sku_canal O por item_id)
                    pubs_existentes = obtener_publicaciones_canal(sku_lusync_match, canal_nombre) or []
                    ya_existe = any(
                        (p.get("sku_canal") == sku_canal) or
                        (item_id and p.get("item_id_canal") == item_id)
                        for p in pubs_existentes
                    )

                    if ya_existe:
                        ya_mapeados.append({
                            "sku_lusync": sku_lusync_match,
                            "sku_canal": sku_canal,
                            "item_id": item_id,
                            "nombre": nombre_pub
                        })
                    else:
                        cruzados.append({
                            "sku_lusync": sku_lusync_match,
                            "sku_canal": sku_canal,
                            "item_id": item_id,
                            "nombre_pub": nombre_pub,
                            "nombre_lusync": nombres_lusync.get(sku_lusync_match, ""),
                            "razon_match": razon_match
                        })
                        if ejecutar:
                            try:
                                agregar_publicacion(
                                    sku_lusync=sku_lusync_match,
                                    canal=canal_nombre,
                                    sku_canal=sku_canal,
                                    item_id_canal=item_id,
                                    notas=f"auto_descubrimiento_{razon_match}"
                                )
                            except Exception as e:
                                errores_guardado.append(f"{sku_canal}: {e}")
                else:
                    no_cruzados.append({
                        "sku_canal": sku_canal,
                        "item_id": item_id,
                        "nombre": nombre_pub,
                        "extra": extra
                    })

            return {
                "cruzados": cruzados,
                "ya_mapeados": ya_mapeados,
                "no_cruzados": no_cruzados,
                "sin_sku": sin_sku,
                "errores_guardado": errores_guardado,
                "stats": {
                    "total_descargadas": len(publicaciones_canal),
                    "cruzadas_nuevas": len(cruzados),
                    "ya_estaban_mapeadas": len(ya_mapeados),
                    "no_cruzadas": len(no_cruzados),
                    "sin_sku_requieren_accion_manual": len(sin_sku),
                    "porcentaje_exito_confiable": round(
                        (len(cruzados) + len(ya_mapeados)) / max(len(publicaciones_canal), 1) * 100, 1
                    )
                }
            }

            return {
                "cruzados": cruzados,
                "ya_mapeados": ya_mapeados,
                "no_cruzados": no_cruzados,
                "errores_guardado": errores_guardado,
                "stats": {
                    "total_descargadas": len(publicaciones_canal),
                    "cruzadas_nuevas": len(cruzados),
                    "ya_estaban_mapeadas": len(ya_mapeados),
                    "no_cruzadas": len(no_cruzados),
                    "porcentaje_exito": round(
                        (len(cruzados) + len(ya_mapeados)) / max(len(publicaciones_canal), 1) * 100, 1
                    )
                }
            }

        canales_a_procesar = []
        if canal in ("todos", "meli", "mercadolibre"):
            canales_a_procesar.append("meli")
        if canal in ("todos", "falabella"):
            canales_a_procesar.append("falabella")
        if canal in ("todos", "walmart"):
            canales_a_procesar.append("walmart")
        if canal in ("todos", "paris"):
            canales_a_procesar.append("paris")
        if canal in ("todos", "ripley"):
            canales_a_procesar.append("ripley")

        # ── MELI ──
        if "meli" in canales_a_procesar:
            try:
                from mercadolibre import obtener_publicaciones_meli
                pubs_raw = []
                offset = 0
                while True:
                    lote = obtener_publicaciones_meli(limite=50, offset=offset)
                    if not lote or not lote.get("items"):
                        break
                    items = lote["items"]
                    for it in items:
                        item_id  = it.get("id") or it.get("item_id")
                        titulo   = it.get("title") or it.get("nombre") or ""
                        status   = it.get("status")
                        stock    = it.get("stock", 0)
                        sku_main = (it.get("sku_seller") or it.get("sku") or "").strip()
                        variantes = it.get("variantes_skus") or []

                        variantes_raw_list = it.get("variantes_raw", [])

                        if variantes:
                            # Publicación multivariante — generar una entrada por variante
                            skus_vistos = set()
                            for sku_var in variantes:
                                sku_var = sku_var.strip()
                                if sku_var and sku_var not in skus_vistos:
                                    skus_vistos.add(sku_var)
                                    pubs_raw.append({
                                        "sku_canal": sku_var,
                                        "item_id_canal": item_id,
                                        "nombre": titulo,
                                        "extra": {"status": status, "stock": stock},
                                        "variantes_raw": variantes_raw_list
                                    })
                        else:
                            # Publicación simple o paraguas sin SKU
                            pubs_raw.append({
                                "sku_canal": sku_main,
                                "item_id_canal": item_id,
                                "nombre": titulo,
                                "extra": {"status": status, "stock": stock},
                                "variantes_raw": variantes_raw_list
                            })
                    if len(items) < 50:
                        break
                    offset += 50
                    if offset > 2000:
                        break

                reporte_global["canales"]["MercadoLibre"] = _cruzar_y_guardar("mercadolibre", pubs_raw)
            except Exception as e:
                reporte_global["canales"]["MercadoLibre"] = {"error": str(e)}

        # ── FALABELLA ──
        if "falabella" in canales_a_procesar:
            try:
                from falabella import obtener_productos_falabella
                pubs_raw = []
                offset = 0
                while True:
                    lote = obtener_productos_falabella(limit=100, offset=offset, filter_status="all")
                    if not lote:
                        break
                    for p in lote:
                        skus_p = p.get("Skus", {}).get("Sku", [])
                        if isinstance(skus_p, dict):
                            skus_p = [skus_p]
                        for s in skus_p:
                            seller_sku = s.get("SellerSku") or s.get("ShopSku") or ""
                            pubs_raw.append({
                                "sku_canal": seller_sku,
                                "item_id_canal": None,
                                "nombre": p.get("PrimaryCategory") or p.get("name") or "",
                                "extra": {"status": s.get("Status")}
                            })
                    if len(lote) < 100:
                        break
                    offset += 100
                    if offset > 5000:
                        break
                reporte_global["canales"]["Falabella"] = _cruzar_y_guardar("falabella", pubs_raw)
            except Exception as e:
                reporte_global["canales"]["Falabella"] = {"error": str(e)}

        # ── WALMART ──
        if "walmart" in canales_a_procesar:
            try:
                from walmart import obtener_productos_walmart
                lote = obtener_productos_walmart(limit=50, max_paginas=20)
                pubs_raw = []
                for p in (lote or []):
                    sku = p.get("sku") or p.get("itemId") or ""
                    pubs_raw.append({
                        "sku_canal": str(sku),
                        "item_id_canal": None,
                        "nombre": p.get("productName") or p.get("itemDescription") or "",
                        "extra": {"status": p.get("publishedStatus"), "stock": p.get("availableInventory")}
                    })
                reporte_global["canales"]["Walmart"] = _cruzar_y_guardar("walmart", pubs_raw)
            except Exception as e:
                reporte_global["canales"]["Walmart"] = {"error": str(e)}

        # ── PARIS ──
        if "paris" in canales_a_procesar:
            try:
                from paris import obtener_productos_paris
                pubs_raw = []
                offset = 0
                while True:
                    lote = obtener_productos_paris(limite=25, offset=offset)
                    if not lote:
                        break
                    # Paris API puede devolver:
                    # {"products": [...]} o {"items": [...]} o lista directa o dict con "content"
                    if isinstance(lote, list):
                        items = lote
                    elif isinstance(lote, dict):
                        items = (
                            lote.get("products") or lote.get("items") or
                            lote.get("content") or lote.get("data") or []
                        )
                    else:
                        items = []

                    for p in (items or []):
                        # Paris puede usar sellerSku, sku, refId o identifier
                        seller_sku = (
                            p.get("sellerSku") or p.get("seller_sku") or
                            p.get("sku") or p.get("refId") or
                            p.get("identifier") or ""
                        )
                        if seller_sku:
                            pubs_raw.append({
                                "sku_canal": str(seller_sku).strip(),
                                "item_id_canal": p.get("id") or p.get("productId"),
                                "nombre": p.get("name") or p.get("title") or p.get("description") or "",
                                "extra": {"status": p.get("status") or p.get("state")}
                            })
                    if not items or len(items) < 25:
                        break
                    offset += 25
                    if offset > 2000:
                        break

                if not pubs_raw:
                    reporte_global["canales"]["Paris"] = {
                        "error": "API Paris devolvió 0 productos — puede ser error de credenciales o API no soporta listar catálogo",
                        "cruzados": [], "ya_mapeados": [], "no_cruzados": [], "sin_sku": [],
                        "stats": {"total_descargadas": 0, "cruzadas_nuevas": 0,
                                  "ya_estaban_mapeadas": 0, "no_cruzadas": 0,
                                  "sin_sku_requieren_accion_manual": 0, "porcentaje_exito_confiable": 0}
                    }
                else:
                    reporte_global["canales"]["Paris"] = _cruzar_y_guardar("paris", pubs_raw)
            except Exception as e:
                reporte_global["canales"]["Paris"] = {"error": str(e)}

        # ── RIPLEY ──
        if "ripley" in canales_a_procesar:
            try:
                import requests as _req
                from ripley import ripley_headers, RIPLEY_BASE_URL
                pubs_raw = []
                offset = 0
                # Ripley Mirakl: endpoint correcto es /api/offers (no /api/products)
                # /api/products da catálogo del marketplace, /api/offers son las ofertas del seller
                while True:
                    res = _req.get(
                        f"{RIPLEY_BASE_URL}/api/offers",
                        headers=ripley_headers(),
                        params={"max": 100, "offset": offset},
                        timeout=20
                    )
                    if res.status_code != 200:
                        # Intentar fallback con /api/products
                        res2 = _req.get(
                            f"{RIPLEY_BASE_URL}/api/products",
                            headers=ripley_headers(),
                            params={"max": 100, "offset": offset},
                            timeout=20
                        )
                        if res2.status_code != 200:
                            reporte_global["canales"]["Ripley"] = {
                                "error": f"API Ripley status {res.status_code}: {res.text[:200]}"
                            }
                            break
                        res = res2

                    data = res.json()
                    # Mirakl puede devolver: {"offers": [...]} o {"products": [...]} o lista directa
                    items = (
                        data.get("offers") or
                        data.get("products") or
                        data.get("items") or
                        (data if isinstance(data, list) else [])
                    )
                    for p in (items or []):
                        seller_sku = (
                            p.get("offer_sku") or p.get("shop_sku") or
                            p.get("offer_id") or p.get("sku") or ""
                        )
                        if seller_sku:
                            pubs_raw.append({
                                "sku_canal": str(seller_sku).strip(),
                                "item_id_canal": None,
                                "nombre": p.get("title") or p.get("description") or p.get("name") or "",
                                "extra": {"state": p.get("state_code") or p.get("status")}
                            })
                    if not items or len(items) < 100:
                        break
                    offset += 100
                    if offset > 5000:
                        break

                if pubs_raw or "Ripley" not in reporte_global["canales"]:
                    reporte_global["canales"]["Ripley"] = _cruzar_y_guardar("ripley", pubs_raw)
            except Exception as e:
                reporte_global["canales"]["Ripley"] = {"error": str(e)}

        # ── PLAN DE ACCIÓN para los no cruzados y sin SKU ──
        plan = []
        for canal_n, data in reporte_global["canales"].items():
            if "error" in data:
                plan.append({
                    "canal": canal_n,
                    "prioridad": "🔴 CRÍTICO",
                    "tipo": "error_api",
                    "problema": f"Error al conectar con la API: {data['error']}",
                    "accion": "Verificar credenciales del canal en Configuración → Canales",
                    "impacto": "Sin conexión no hay sincronización de stock ni órdenes"
                })
                continue

            # FIX 3A: Publicaciones sin SKU — ir al marketplace a agregar el SKU
            for ns in data.get("sin_sku", []):
                nombre_c = ns.get("nombre", "")
                item_id  = ns.get("item_id", "")
                plan.append({
                    "canal": canal_n,
                    "prioridad": "🔴 CRÍTICO",
                    "tipo": "sin_sku",
                    "sku_canal": "",
                    "item_id": item_id,
                    "nombre_publicacion": nombre_c,
                    "causa": "Publicación creada en el marketplace SIN SKU interno — Lusync no puede identificarla",
                    "accion": f"Ir a {canal_n} → editar publicación '{nombre_c}' (ID: {item_id}) → campo SKU/Código Interno → agregar el SKU Lusync → re-ejecutar auto-descubrimiento"
                })

            # FIX 3B: No cruzados — SKU existe en marketplace pero no en Lusync
            for nc in data.get("no_cruzados", []):
                sku_c    = nc.get("sku_canal", "")
                nombre_c = nc.get("nombre", "")
                item_id  = nc.get("item_id", "")

                if not sku_c:
                    continue  # ya manejado en sin_sku

                plan.append({
                    "canal": canal_n,
                    "prioridad": "🟠 IMPORTANTE",
                    "tipo": "sku_no_en_lusync",
                    "sku_canal": sku_c,
                    "item_id": item_id,
                    "nombre_publicacion": nombre_c,
                    "causa": f"SKU '{sku_c}' existe en {canal_n} pero no está cargado en Lusync",
                    "accion": f"OPCIÓN A (recomendada): Agregar '{nombre_c}' en Inventario con SKU '{sku_c}' → re-ejecutar auto-descubrimiento. OPCIÓN B: Mapeo SKUs → vincular '{sku_c}' al SKU Lusync que corresponda"
                })

        # Ordenar: 🔴 primero, 🟠 después
        reporte_global["plan_accion"] = sorted(
            plan,
            key=lambda x: (0 if "🔴" in x["prioridad"] else 1, x["canal"])
        )

        # Resumen del plan
        criticos  = sum(1 for p in plan if "🔴" in p["prioridad"])
        importantes = sum(1 for p in plan if "🟠" in p["prioridad"])
        reporte_global["resumen_plan"] = {
            "criticos_sin_sku": criticos,
            "importantes_sku_no_en_lusync": importantes,
            "total_acciones": len(plan)
        }

        # ── Resumen global ──
        total_desc  = sum(d.get("stats", {}).get("total_descargadas", 0) for d in reporte_global["canales"].values() if "stats" in d)
        total_cruz  = sum(d.get("stats", {}).get("cruzadas_nuevas", 0) for d in reporte_global["canales"].values() if "stats" in d)
        total_ya    = sum(d.get("stats", {}).get("ya_estaban_mapeadas", 0) for d in reporte_global["canales"].values() if "stats" in d)
        total_fail  = sum(d.get("stats", {}).get("no_cruzadas", 0) for d in reporte_global["canales"].values() if "stats" in d)
        total_sin_sku = sum(d.get("stats", {}).get("sin_sku_requieren_accion_manual", 0) for d in reporte_global["canales"].values() if "stats" in d)

        reporte_global["resumen_global"] = {
            "total_publicaciones_descargadas": total_desc,
            "cruzadas_automaticamente_nuevas": total_cruz,
            "ya_estaban_mapeadas": total_ya,
            "no_cruzadas_sku_no_en_lusync": total_fail,
            "sin_sku_requieren_accion_en_marketplace": total_sin_sku,
            "total_requieren_accion": total_fail + total_sin_sku,
            "porcentaje_exito_confiable": round((total_cruz + total_ya) / max(total_desc, 1) * 100, 1),
            "mapeos_guardados": ejecutar,
            "siguiente_paso": (
                "✅ Todo mapeado — sincronización activa en todos los canales"
                if (total_fail + total_sin_sku) == 0
                else f"⚠️ Revisa el plan_accion: {total_sin_sku} sin SKU en marketplace + {total_fail} SKU no cargados en Lusync"
            )
        }

        return jsonify(reporte_global)

    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/admin/debug_meli_item")
def admin_debug_meli_item():
    """
    Muestra la estructura RAW de un item de MELI para debug de variantes.
    Uso: /admin/debug_meli_item?item_id=MLC2749905118&token=XXX
    """
    bypass_token = os.environ.get("ADMIN_BYPASS_TOKEN", "lcTDX2fjcH3hiZFvv8apEwPd-eiCIqFdkKqJIVy1bVw")
    if request.args.get("token") != bypass_token and not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    item_id = request.args.get("item_id", "").strip()
    if not item_id:
        return jsonify({"error": "Falta item_id"}), 400

    try:
        import requests as _req
        from mercadolibre import meli_headers, MELI_API_URL

        # Pedir item completo SIN filtrar nada para ver todos los campos
        res = _req.get(
            f"{MELI_API_URL}/items/{item_id}",
            headers=meli_headers(),
            timeout=20
        )
        if res.status_code != 200:
            return jsonify({"error": f"MELI {res.status_code}", "detail": res.text[:500]}), 400

        data = res.json()

        # Mostrar TODOS los campos del item y variantes para debug
        resultado = {
            "item_id": data.get("id"),
            "title": data.get("title"),
            "seller_custom_field_item": data.get("seller_custom_field"),
            "todos_attributes_item": data.get("attributes", [])[:10],
            "num_variantes": len(data.get("variations") or []),
            "todos_los_campos_item": list(data.keys()),
            "variantes": []
        }

        for var in (data.get("variations") or [])[:5]:
            var_info = {
                "variation_id": var.get("id"),
                "todos_los_campos": list(var.keys()),
                "seller_custom_field": var.get("seller_custom_field"),
                "user_product_id": var.get("user_product_id"),
                "inventory_id": var.get("inventory_id"),
                "todos_attributes": var.get("attributes", []),
                "attribute_combinations": var.get("attribute_combinations", [])[:3],
            }
            resultado["variantes"].append(var_info)

        return jsonify(resultado)

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/admin/importar_excel_meli", methods=["GET", "POST"])
def admin_importar_excel_meli():
    """
    Importa Excel de MercadoLibre Seller Center.
    Devuelve 3 grupos: automáticos, requieren_alias, no_en_lusync.
    Con ejecutar=1 guarda los mapeos.
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    ejecutar = request.args.get("ejecutar", "0") == "1"

    if request.method != "POST" or "archivo" not in request.files:
        return jsonify({"error": "Envía el Excel con POST campo 'archivo'"}), 400

    try:
        try:
            import pandas as pd
        except ImportError:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "--break-system-packages", "-q"])
            import pandas as pd
        import io, json
        from inventario import get_conn, obtener_publicaciones_canal, agregar_publicacion, cargar_productos

        archivo = request.files["archivo"]
        contenido = archivo.read()
        alias_map = json.loads(request.form.get("alias_map", "{}"))

        df = pd.read_excel(io.BytesIO(contenido), sheet_name="Publicaciones", header=None)

        # ── Cargar inventario Lusync ──
        productos = cargar_productos()
        # sku_real: diccionario upper → real
        sku_real = {p["sku"].upper().strip(): p["sku"] for p in productos}
        lista_skus_lusync = sorted(sku_real.values())

        # ── Cargar mapeos ya existentes en MELI ──
        # Consulta sku_mapeo_canal (tabla principal multi-pub) + sku_mapeo (alias legacy)
        mapeados_existentes = {}  # sku_canal.upper → {sku_lusync, item_id}
        try:
            conn = get_conn()
            cur = conn.cursor()

            # Tabla principal: sku_mapeo_canal
            cur.execute("""
                SELECT sku_canal, sku_lusync, item_id_canal 
                FROM sku_mapeo_canal 
                WHERE canal='mercadolibre' AND activo=TRUE
            """)
            for (sc, sl, iid) in cur.fetchall():
                if sc:
                    mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid}
                if sl:
                    sku_real[sl.upper().strip()] = sl

            # Tabla legacy: sku_mapeo (aliases manuales por sku_mercadolibre)
            try:
                cur.execute("SELECT sku_lusync, sku_mercadolibre FROM sku_mapeo WHERE sku_mercadolibre IS NOT NULL AND sku_mercadolibre != ''")
                for (sl, sc_meli) in cur.fetchall():
                    if sc_meli:
                        mapeados_existentes[sc_meli.upper().strip()] = {"sku_lusync": sl, "item_id": None}
                    if sl:
                        sku_real[sl.upper().strip()] = sl
            except Exception:
                pass

            cur.close()
            conn.close()
        except Exception:
            pass

        # ── Procesar Excel ──
        data_rows = df.iloc[5:].copy()
        data_rows.columns = range(len(data_rows.columns))

        # Pre-análisis: items con SKU propio (no son hijos de paraguas)
        items_simples = set()
        for _, row in data_rows.iterrows():
            item_id = str(row[1]).strip() if pd.notna(row[1]) else ""
            var_id  = str(row[3]).strip() if pd.notna(row[3]) else ""
            sku     = str(row[4]).strip() if pd.notna(row[4]) else ""
            if item_id and item_id != "nan" and sku and sku != "nan" and (not var_id or var_id == "nan"):
                items_simples.add(item_id)

        automaticos    = []
        requieren_alias = []
        no_en_lusync   = []
        ya_mapeados    = []
        skus_vistos    = set()  # evitar duplicados en el reporte

        for _, row in data_rows.iterrows():
            item_id  = str(row[1]).strip() if pd.notna(row[1]) else ""
            var_id   = str(row[3]).strip() if pd.notna(row[3]) else ""
            sku_meli = str(row[4]).strip() if pd.notna(row[4]) else ""
            titulo   = str(row[5]).strip() if pd.notna(row[5]) else str(row[6]).strip() if pd.notna(row[6]) else ""

            if var_id and var_id != "nan" and item_id in items_simples:
                continue
            if not sku_meli or sku_meli.lower() in ("nan", "none", ""):
                continue
            if not item_id or item_id.lower() in ("nan", "none", ""):
                continue

            clave = sku_meli.upper() + "|" + item_id
            if clave in skus_vistos:
                continue
            skus_vistos.add(clave)

            entry = {"item_id": item_id, "sku_meli": sku_meli, "titulo": titulo}

            # ¿Ya está mapeado exactamente?
            existente = mapeados_existentes.get(sku_meli.upper())
            if existente and (not item_id.startswith("MLC") or existente.get("item_id") == item_id or existente.get("sku_lusync")):
                ya_mapeados.append({**entry, "sku_lusync": existente.get("sku_lusync", sku_meli)})
                continue

            # ¿Coincide exacto con Lusync?
            sku_lusync_exacto = sku_real.get(sku_meli.upper())
            if sku_lusync_exacto:
                automaticos.append({**entry, "sku_lusync": sku_lusync_exacto})
                continue

            # ¿Tiene alias definido manualmente?
            alias_key = sku_meli + "+" + item_id
            if alias_key in alias_map and alias_map[alias_key]:
                automaticos.append({**entry, "sku_lusync": alias_map[alias_key], "via_alias": True})
                continue

            # No coincide → requiere alias
            requieren_alias.append({**entry, "sku_lusync_sugerido": ""})

        # ── Ejecutar mapeos ──
        guardados = 0
        errores = []
        if ejecutar:
            todos = automaticos
            for m in todos:
                try:
                    agregar_publicacion(
                        sku_lusync=m["sku_lusync"],
                        canal="mercadolibre",
                        sku_canal=m["sku_meli"],
                        item_id_canal=m["item_id"] if m["item_id"].startswith("MLC") else None,
                        notas="importado_excel_meli"
                    )
                    guardados += 1
                except Exception as e:
                    errores.append(f"{m['sku_meli']}: {str(e)}")

        return jsonify({
            "automaticos":       automaticos,
            "requieren_alias":   requieren_alias,
            "no_en_lusync":      no_en_lusync,
            "ya_mapeados":       ya_mapeados,
            "skus_lusync_disponibles": lista_skus_lusync,
            "stats": {
                "total_con_sku":    len(automaticos) + len(requieren_alias) + len(no_en_lusync) + len(ya_mapeados),
                "automaticos":      len(automaticos),
                "requieren_alias":  len(requieren_alias),
                "no_en_lusync":     len(no_en_lusync),
                "ya_mapeados":      len(ya_mapeados),
                "nuevos_mapeados":  guardados,
                "errores":          len(errores),
            },
            "modo": "ejecutado" if ejecutar else "simulacion",
            "errores": errores
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500



@app.route("/admin/debug_skus_canal")
def admin_debug_skus_canal():
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT sku_lusync, sku_canal FROM publicaciones_canal WHERE canal='mercadolibre' LIMIT 20")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"rows": [{"sku_lusync": r[0], "sku_canal": r[1]} for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/admin/validar_meli")
def admin_validar_meli():
    """
    Valida el estado completo del mapeo de MercadoLibre.
    Compara publicaciones activas en MELI vs mapeos en Lusync.
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    try:
        from inventario import get_conn, cargar_productos
        from mercadolibre import obtener_publicaciones_meli

        # 1. Obtener publicaciones de MELI
        resultado_meli = obtener_publicaciones_meli()
        if not resultado_meli:
            return jsonify({"error": "No se pudo conectar con MercadoLibre"}), 500
        pubs_meli = resultado_meli.get("items", [])

        # 2. Cargar mapeos existentes en sku_mapeo_canal
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT sku_canal, sku_lusync, item_id_canal, activo
            FROM sku_mapeo_canal WHERE canal='mercadolibre'
        """)
        mapeos = {}
        for (sc, sl, iid, activo) in cur.fetchall():
            if sc: mapeos[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid, "activo": activo}
            if iid: mapeos[iid.upper().strip()] = {"sku_lusync": sl, "sku_canal": sc, "activo": activo}

        # 3. Cargar alias de sku_mapeo
        try:
            cur.execute("SELECT sku_lusync, sku_mercadolibre FROM sku_mapeo WHERE sku_mercadolibre IS NOT NULL AND sku_mercadolibre != ''")
            for (sl, sc) in cur.fetchall():
                if sc: mapeos[sc.upper().strip()] = {"sku_lusync": sl, "via": "sku_mapeo"}
        except Exception:
            pass
        cur.close()
        conn.close()

        # 4. Cruzar
        mapeadas        = []
        sin_mapeo       = []
        sin_sku_en_meli = []

        for pub in pubs_meli:
            item_id  = str(pub.get("item_id") or "").strip()
            sku_meli = str(pub.get("sku_seller") or "").strip()
            titulo   = str(pub.get("title") or "").strip()
            status   = str(pub.get("status") or "")
            stock    = int(pub.get("stock") or 0)
            variantes_skus = pub.get("variantes_skus") or []

            entry = {
                "item_id": item_id,
                "sku_meli": sku_meli,
                "titulo": titulo[:60],
                "status": status,
                "stock": stock
            }

            # Buscar mapeo por sku_meli, por item_id, o por cualquier variante
            mapeo = (mapeos.get(sku_meli.upper()) or 
                     mapeos.get(item_id.upper()) or
                     next((mapeos.get(v.upper()) for v in variantes_skus if mapeos.get(v.upper())), None))
            if mapeo:
                entry["sku_lusync"] = mapeo.get("sku_lusync")
                mapeadas.append(entry)
            elif not sku_meli:
                sin_sku_en_meli.append(entry)
            else:
                sin_mapeo.append(entry)

        # 5. Resumen
        activas_sin_mapeo = [p for p in sin_mapeo if p["status"] == "active" and p["stock"] > 0]

        return jsonify({
            "resumen": {
                "total_publicaciones_meli": len(pubs_meli),
                "mapeadas": len(mapeadas),
                "sin_mapeo": len(sin_mapeo),
                "sin_sku_en_meli": len(sin_sku_en_meli),
                "porcentaje_mapeo": round(len(mapeadas) / max(len(pubs_meli), 1) * 100, 1),
                "activas_sin_mapeo_con_stock": len(activas_sin_mapeo),
            },
            "alertas": {
                "critico": [
                    {"item_id": p["item_id"], "titulo": p["titulo"], "stock": p["stock"]}
                    for p in activas_sin_mapeo
                ],
                "sin_sku_en_meli": [
                    {"item_id": p["item_id"], "titulo": p["titulo"], "stock": p["stock"]}
                    for p in sin_sku_en_meli if p["status"] == "active"
                ]
            },
            "detalle": {
                "mapeadas": mapeadas[:10],
                "sin_mapeo": sin_mapeo[:20]
            }
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/importar_csv_walmart", methods=["POST"])
def admin_importar_csv_walmart():
    """
    Importa CSV de Walmart Chile (Mis productos → Exportar).
    Columnas: Imagen, Nombre del producto, SKU, Estado, Price, Inventario, Categoría, Item ID
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    ejecutar = request.args.get("ejecutar", "0") == "1"

    if "archivo" not in request.files:
        return jsonify({"error": "Envía el CSV con campo 'archivo'"}), 400

    try:
        import csv, io, json
        from inventario import get_conn, agregar_publicacion, cargar_productos

        archivo = request.files["archivo"]
        contenido = archivo.read().decode("utf-8-sig")
        alias_map = json.loads(request.form.get("alias_map", "{}"))

        reader = csv.DictReader(io.StringIO(contenido))
        filas = list(reader)

        # Cargar SKUs Lusync
        productos = cargar_productos()
        sku_real = {p["sku"].upper().strip(): p["sku"] for p in productos}
        lista_skus_lusync = sorted(sku_real.values())

        # Cargar mapeos ya existentes en Walmart
        mapeados_existentes = {}
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""
                SELECT sku_canal, sku_lusync, item_id_canal
                FROM sku_mapeo_canal WHERE canal='walmart' AND activo=TRUE
            """)
            for (sc, sl, iid) in cur.fetchall():
                if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid}
                if sl: sku_real[sl.upper().strip()] = sl
            # Alias de sku_mapeo legacy
            try:
                cur.execute("SELECT sku_lusync, sku_walmart FROM sku_mapeo WHERE sku_walmart IS NOT NULL AND sku_walmart != ''")
                for (sl, sc) in cur.fetchall():
                    if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": None}
            except Exception:
                pass
            cur.close()
            conn.close()
        except Exception:
            pass

        automaticos    = []
        requieren_alias = []
        no_en_lusync   = []
        ya_mapeados    = []
        skus_vistos    = set()

        for fila in filas:
            sku_wm  = (fila.get("SKU") or "").strip()
            titulo  = (fila.get("Nombre del producto") or "").strip()
            item_id = (fila.get("Item ID") or "").strip()
            stock   = int(fila.get("Inventario") or 0)
            estado  = (fila.get("Estado") or "").strip()

            if not sku_wm:
                continue
            if sku_wm.upper() in skus_vistos:
                continue
            skus_vistos.add(sku_wm.upper())

            entry = {"sku_canal": sku_wm, "titulo": titulo, "item_id": item_id,
                     "stock": stock, "estado": estado}

            # ¿Ya mapeado?
            existente = mapeados_existentes.get(sku_wm.upper())
            if existente:
                ya_mapeados.append({**entry, "sku_lusync": existente.get("sku_lusync", sku_wm)})
                continue

            # ¿Coincide exacto?
            sku_lusync_exacto = sku_real.get(sku_wm.upper())
            if sku_lusync_exacto:
                automaticos.append({**entry, "sku_lusync": sku_lusync_exacto})
                continue

            # ¿Tiene alias manual?
            alias_key = sku_wm + "+" + item_id
            if alias_key in alias_map and alias_map[alias_key]:
                automaticos.append({**entry, "sku_lusync": alias_map[alias_key], "via_alias": True})
                continue

            # Requiere alias
            requieren_alias.append({**entry, "sku_lusync_sugerido": ""})

        # Ejecutar
        guardados = 0
        errores = []
        if ejecutar:
            for m in automaticos:
                try:
                    agregar_publicacion(
                        sku_lusync=m["sku_lusync"],
                        canal="walmart",
                        sku_canal=m["sku_canal"],
                        item_id_canal=m.get("item_id") or None,
                        notas="importado_csv_walmart"
                    )
                    guardados += 1
                except Exception as e:
                    errores.append(f"{m['sku_canal']}: {str(e)}")

        return jsonify({
            "canal": "Walmart",
            "automaticos":       automaticos,
            "requieren_alias":   requieren_alias,
            "no_en_lusync":      no_en_lusync,
            "ya_mapeados":       ya_mapeados,
            "skus_lusync_disponibles": lista_skus_lusync,
            "stats": {
                "total": len(automaticos) + len(requieren_alias) + len(no_en_lusync) + len(ya_mapeados),
                "automaticos":     len(automaticos),
                "requieren_alias": len(requieren_alias),
                "no_en_lusync":    len(no_en_lusync),
                "ya_mapeados":     len(ya_mapeados),
                "nuevos_mapeados": guardados,
                "errores":         len(errores),
            },
            "modo": "ejecutado" if ejecutar else "simulacion",
            "errores": errores
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/importar_csv_ripley", methods=["POST"])
def admin_importar_csv_ripley():
    """
    Importa Excel de Ripley (offers-import.xlsx).
    Columnas: sku, product-id, product-id-type, description, ..., quantity, ..., state
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401

    ejecutar = request.args.get("ejecutar", "0") == "1"

    if "archivo" not in request.files:
        return jsonify({"error": "Envía el archivo con campo 'archivo'"}), 400

    try:
        import io, json, openpyxl
        from inventario import get_conn, agregar_publicacion, cargar_productos

        archivo = request.files["archivo"]
        contenido = archivo.read()
        alias_map = json.loads(request.form.get("alias_map", "{}"))

        wb = openpyxl.load_workbook(io.BytesIO(contenido))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return jsonify({"error": "Archivo vacío"}), 400

        # Headers en fila 1
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        def col(name):
            try: return headers.index(name)
            except ValueError: return None

        idx_sku   = col("sku")
        idx_desc  = col("description")
        idx_qty   = col("quantity")
        idx_pid   = col("product-id")
        idx_state = col("state")

        if idx_sku is None:
            return jsonify({"error": f"No se encontró columna 'sku'. Headers: {headers}"}), 400

        # Cargar SKUs Lusync
        productos = cargar_productos()
        sku_real = {p["sku"].upper().strip(): p["sku"] for p in productos}
        lista_skus_lusync = sorted(sku_real.values())

        # Mapeos existentes Ripley
        mapeados_existentes = {}
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""
                SELECT sku_canal, sku_lusync, item_id_canal
                FROM sku_mapeo_canal WHERE canal='ripley' AND activo=TRUE
            """)
            for (sc, sl, iid) in cur.fetchall():
                if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid}
                if sl: sku_real[sl.upper().strip()] = sl
            try:
                cur.execute("SELECT sku_lusync, sku_ripley FROM sku_mapeo WHERE sku_ripley IS NOT NULL AND sku_ripley != ''")
                for (sl, sc) in cur.fetchall():
                    if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl}
            except Exception:
                pass
            cur.close()
            conn.close()
        except Exception:
            pass

        automaticos     = []
        requieren_alias = []
        no_en_lusync    = []
        ya_mapeados     = []
        skus_vistos     = set()

        for row in rows[1:]:
            sku_rip = str(row[idx_sku] or "").strip() if idx_sku is not None else ""
            titulo  = str(row[idx_desc] or "").strip() if idx_desc is not None else ""
            stock   = int(row[idx_qty] or 0) if idx_qty is not None else 0
            item_id = str(row[idx_pid] or "").strip() if idx_pid is not None else ""
            estado  = str(row[idx_state] or "").strip() if idx_state is not None else ""

            if not sku_rip or sku_rip.upper() in skus_vistos:
                continue
            skus_vistos.add(sku_rip.upper())

            entry = {"sku_canal": sku_rip, "titulo": titulo, "item_id": item_id,
                     "stock": stock, "estado": estado}

            existente = mapeados_existentes.get(sku_rip.upper())
            if existente:
                ya_mapeados.append({**entry, "sku_lusync": existente.get("sku_lusync", sku_rip)})
                continue

            sku_lusync_exacto = sku_real.get(sku_rip.upper())
            if sku_lusync_exacto:
                automaticos.append({**entry, "sku_lusync": sku_lusync_exacto})
                continue

            alias_key = sku_rip + "+" + item_id
            if alias_key in alias_map and alias_map[alias_key]:
                automaticos.append({**entry, "sku_lusync": alias_map[alias_key], "via_alias": True})
                continue

            requieren_alias.append({**entry, "sku_lusync_sugerido": ""})

        guardados = 0
        errores = []
        if ejecutar:
            for m in automaticos:
                try:
                    agregar_publicacion(
                        sku_lusync=m["sku_lusync"],
                        canal="ripley",
                        sku_canal=m["sku_canal"],
                        item_id_canal=m.get("item_id") or None,
                        notas="importado_xlsx_ripley"
                    )
                    guardados += 1
                except Exception as e:
                    errores.append(f"{m['sku_canal']}: {str(e)}")

        return jsonify({
            "canal": "Ripley",
            "automaticos":        automaticos,
            "requieren_alias":    requieren_alias,
            "no_en_lusync":       no_en_lusync,
            "ya_mapeados":        ya_mapeados,
            "skus_lusync_disponibles": lista_skus_lusync,
            "stats": {
                "total":           len(automaticos)+len(requieren_alias)+len(no_en_lusync)+len(ya_mapeados),
                "automaticos":     len(automaticos),
                "requieren_alias": len(requieren_alias),
                "no_en_lusync":    len(no_en_lusync),
                "ya_mapeados":     len(ya_mapeados),
                "nuevos_mapeados": guardados,
                "errores":         len(errores),
            },
            "modo": "ejecutado" if ejecutar else "simulacion",
            "errores": errores
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/admin/importar_excel_falabella", methods=["POST"])
def admin_importar_excel_falabella():
    """
    Importa Excel de Falabella (SellerPriceTemplate.xlsx).
    Columnas: SellerSku, ShopSku, PriceFalabella, SalePriceFalabella, ..., Name
    Fila 1 = headers, fila 2+ = datos
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    ejecutar = request.args.get("ejecutar", "0") == "1"
    if "archivo" not in request.files:
        return jsonify({"error": "Envía el archivo con campo 'archivo'"}), 400
    try:
        import io, json, openpyxl
        from inventario import get_conn, agregar_publicacion, cargar_productos
        archivo = request.files["archivo"]
        alias_map = json.loads(request.form.get("alias_map", "{}"))
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Archivo vacío"}), 400
        headers = [str(h).strip() if h else "" for h in rows[0]]
        def col(name):
            for i,h in enumerate(headers):
                if h.lower() == name.lower(): return i
            return None
        idx_sku   = col("SellerSku")
        idx_shop  = col("ShopSku")
        idx_name  = col("Name")
        if idx_sku is None:
            return jsonify({"error": f"No se encontró columna 'SellerSku'. Headers: {headers}"}), 400
        productos = cargar_productos()
        sku_real = {p["sku"].upper().strip(): p["sku"] for p in productos}
        lista_skus_lusync = sorted(sku_real.values())
        mapeados_existentes = {}
        try:
            conn = get_conn(); cur = conn.cursor()
            cur.execute("SELECT sku_canal, sku_lusync, item_id_canal FROM sku_mapeo_canal WHERE canal='falabella' AND activo=TRUE")
            for (sc, sl, iid) in cur.fetchall():
                if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid}
                if sl: sku_real[sl.upper().strip()] = sl
            try:
                cur.execute("SELECT sku_lusync, sku_falabella FROM sku_mapeo WHERE sku_falabella IS NOT NULL AND sku_falabella != ''")
                for (sl, sc) in cur.fetchall():
                    if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl}
            except Exception: pass
            cur.close(); conn.close()
        except Exception: pass
        automaticos=[]; requieren_alias=[]; no_en_lusync=[]; ya_mapeados=[]; skus_vistos=set()
        for row in rows[1:]:
            sku_fal = str(row[idx_sku] or "").strip() if idx_sku is not None else ""
            titulo  = str(row[idx_name] or "").strip() if idx_name is not None else ""
            item_id = str(row[idx_shop] or "").strip() if idx_shop is not None else ""
            if not sku_fal or sku_fal.upper() in skus_vistos: continue
            skus_vistos.add(sku_fal.upper())
            entry = {"sku_canal": sku_fal, "titulo": titulo, "item_id": item_id, "stock": 0, "estado": ""}
            existente = mapeados_existentes.get(sku_fal.upper())
            if existente:
                ya_mapeados.append({**entry, "sku_lusync": existente.get("sku_lusync", sku_fal)}); continue
            sku_lusync_exacto = sku_real.get(sku_fal.upper())
            if sku_lusync_exacto:
                automaticos.append({**entry, "sku_lusync": sku_lusync_exacto}); continue
            alias_key = sku_fal + "+" + item_id
            if alias_key in alias_map and alias_map[alias_key]:
                automaticos.append({**entry, "sku_lusync": alias_map[alias_key], "via_alias": True}); continue
            requieren_alias.append({**entry, "sku_lusync_sugerido": ""})
        guardados=0; errores=[]
        if ejecutar:
            for m in automaticos:
                try:
                    agregar_publicacion(sku_lusync=m["sku_lusync"], canal="falabella",
                        sku_canal=m["sku_canal"], item_id_canal=m.get("item_id") or None,
                        notas="importado_excel_falabella")
                    guardados += 1
                except Exception as e: errores.append(f"{m['sku_canal']}: {str(e)}")
        return jsonify({"canal":"Falabella","automaticos":automaticos,"requieren_alias":requieren_alias,
            "no_en_lusync":no_en_lusync,"ya_mapeados":ya_mapeados,"skus_lusync_disponibles":lista_skus_lusync,
            "stats":{"total":len(automaticos)+len(requieren_alias)+len(no_en_lusync)+len(ya_mapeados),
                "automaticos":len(automaticos),"requieren_alias":len(requieren_alias),
                "no_en_lusync":len(no_en_lusync),"ya_mapeados":len(ya_mapeados),
                "nuevos_mapeados":guardados,"errores":len(errores)},
            "modo":"ejecutado" if ejecutar else "simulacion","errores":errores})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/admin/importar_excel_paris", methods=["POST"])
def admin_importar_excel_paris():
    """
    Importa Excel de precios Paris (export-price.xlsx).
    Hoja: marketplace. Fila 1=headers, col 0=SKU Paris, col 1=SELLER SKU, col 2=Nombre.
    El archivo tiene XML inválido — se parchea antes de leer.
    """
    if not session.get("logged"):
        return jsonify({"error": "no autorizado"}), 401
    ejecutar = request.args.get("ejecutar", "0") == "1"
    if "archivo" not in request.files:
        return jsonify({"error": "Envía el archivo con campo 'archivo'"}), 400
    try:
        import io, json, zipfile, re, openpyxl
        from inventario import get_conn, agregar_publicacion, cargar_productos

        archivo = request.files["archivo"]
        alias_map = json.loads(request.form.get("alias_map", "{}"))
        raw = archivo.read()

        # Parchear XML inválido del archivo Paris (colores RGB sin canal alpha y errorStyle inválido)
        def fix_color(m):
            rgb = m.group(1)
            if len(rgb) == 6: return f'rgb="FF{rgb}"' 
            elif len(rgb) == 7: return f'rgb="FF{rgb[:6]}"' 
            return m.group(0)

        with zipfile.ZipFile(io.BytesIO(raw)) as zin:
            styles = zin.read('xl/styles.xml').decode('utf-8')
            styles_fixed = re.sub(r'rgb="([0-9A-Fa-f]+)"', fix_color, styles)
            sheets_patched = {}
            for name in zin.namelist():
                if name.startswith('xl/worksheets/') and name.endswith('.xml'):
                    xml = zin.read(name).decode('utf-8')
                    xml = re.sub(r'errorStyle="[^"]*"', '', xml)
                    sheets_patched[name] = xml

        buf = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw)) as zin:
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for name in zin.namelist():
                    if name == 'xl/styles.xml':
                        zout.writestr(name, styles_fixed.encode('utf-8'))
                    elif name in sheets_patched:
                        zout.writestr(name, sheets_patched[name].encode('utf-8'))
                    else:
                        zout.writestr(name, zin.read(name))
        buf.seek(0)

        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)

        # Buscar hoja con datos — preferir 'marketplace', luego cualquiera con SELLER SKU en fila 1
        ws = None
        for sname in wb.sheetnames:
            if sname.lower() in ('marketplace', 'paris'):
                ws = wb[sname]; break
        if not ws:
            for sname in wb.sheetnames:
                if sname.endswith('data') or sname.lower() in ('configuration','easydata','parisdata','spiddata'): continue
                ws_try = wb[sname]
                rows_try = list(ws_try.iter_rows(values_only=True))
                if rows_try and any('seller' in str(c).lower() for c in rows_try[0] if c):
                    ws = ws_try; break
        if not ws:
            return jsonify({"error": f"No se encontró hoja de datos. Hojas: {wb.sheetnames}"}), 400

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Hoja vacía"}), 400

        # Headers en fila 1
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        def col(names):
            for name in names:
                for i, h in enumerate(headers):
                    if name in h: return i
            return None

        idx_sku_seller = col(['seller sku', 'sellersku', 'sku seller'])
        idx_sku_paris  = col(['sku(', 'sku *', 'sku(*']) 
        if idx_sku_paris is None: idx_sku_paris = 0
        idx_titulo     = col(['nombre', 'name', 'título']) or 2
        if idx_sku_seller is None: idx_sku_seller = 1

        # Cargar Lusync
        productos = cargar_productos()
        sku_real = {p["sku"].upper().strip(): p["sku"] for p in productos}
        lista_skus_lusync = sorted(sku_real.values())

        mapeados_existentes = {}
        try:
            conn = get_conn(); cur = conn.cursor()
            cur.execute("SELECT sku_canal, sku_lusync, item_id_canal FROM sku_mapeo_canal WHERE canal='paris' AND activo=TRUE")
            for (sc, sl, iid) in cur.fetchall():
                if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl, "item_id": iid}
                if sl: sku_real[sl.upper().strip()] = sl
            try:
                cur.execute("SELECT sku_lusync, sku_paris FROM sku_mapeo WHERE sku_paris IS NOT NULL AND sku_paris != ''")
                for (sl, sc) in cur.fetchall():
                    if sc: mapeados_existentes[sc.upper().strip()] = {"sku_lusync": sl}
            except Exception: pass
            cur.close(); conn.close()
        except Exception: pass

        automaticos=[]; requieren_alias=[]; no_en_lusync=[]; ya_mapeados=[]; skus_vistos=set()

        for row in rows[1:]:
            row = list(row)
            sku_seller = str(row[idx_sku_seller] or "").strip() if idx_sku_seller < len(row) else ""
            sku_paris  = str(row[idx_sku_paris]  or "").strip() if idx_sku_paris  < len(row) else ""
            titulo     = str(row[idx_titulo]     or "").strip() if idx_titulo     < len(row) else ""

            # Usar seller SKU como clave principal (es el SKU de Lusync)
            sku_usar = sku_seller or sku_paris
            if not sku_usar or sku_usar.upper() in skus_vistos: continue
            skus_vistos.add(sku_usar.upper())

            entry = {"sku_canal": sku_usar, "titulo": titulo, "item_id": sku_paris, "stock": 0, "estado": ""}

            existente = mapeados_existentes.get(sku_usar.upper())
            if existente:
                ya_mapeados.append({**entry, "sku_lusync": existente.get("sku_lusync", sku_usar)}); continue
            sku_lusync_exacto = sku_real.get(sku_usar.upper())
            if sku_lusync_exacto:
                automaticos.append({**entry, "sku_lusync": sku_lusync_exacto}); continue
            alias_key = sku_usar + "+" + sku_paris
            if alias_key in alias_map and alias_map[alias_key]:
                automaticos.append({**entry, "sku_lusync": alias_map[alias_key], "via_alias": True}); continue
            requieren_alias.append({**entry, "sku_lusync_sugerido": ""})

        guardados=0; errores=[]
        if ejecutar:
            for m in automaticos:
                try:
                    agregar_publicacion(sku_lusync=m["sku_lusync"], canal="paris",
                        sku_canal=m["sku_canal"], item_id_canal=m.get("item_id") or None,
                        notas="importado_excel_paris")
                    guardados += 1
                except Exception as e: errores.append(f"{m['sku_canal']}: {str(e)}")

        return jsonify({"canal":"Paris","automaticos":automaticos,"requieren_alias":requieren_alias,
            "no_en_lusync":no_en_lusync,"ya_mapeados":ya_mapeados,"skus_lusync_disponibles":lista_skus_lusync,
            "stats":{"total":len(automaticos)+len(requieren_alias)+len(no_en_lusync)+len(ya_mapeados),
                "automaticos":len(automaticos),"requieren_alias":len(requieren_alias),
                "no_en_lusync":len(no_en_lusync),"ya_mapeados":len(ya_mapeados),
                "nuevos_mapeados":guardados,"errores":len(errores)},
            "modo":"ejecutado" if ejecutar else "simulacion","errores":errores})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
